"""
recalc.py - Verify Excel formulas in the Equity Research workbook.
Parses formulas and checks for common errors without a full Excel engine.
"""

import openpyxl
import re
import sys

WORKBOOK = r"C:\Users\ryosuke0923\Core_Corporation_2359T_Equity_Research.xlsx"

def check_workbook(path):
    wb = openpyxl.load_workbook(path)
    errors = []
    warnings = []
    stats = {"sheets": 0, "cells_total": 0, "cells_formula": 0, "cells_value": 0}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        stats["sheets"] += 1
        print(f"\n{'='*60}")
        print(f"Sheet: {ws_name}")
        print(f"  Dimensions: {ws.dimensions}")
        print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

        formula_count = 0
        value_count = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                stats["cells_total"] += 1

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    stats["cells_formula"] += 1
                    formula = cell.value

                    # Check for common errors
                    # 1. Unmatched parentheses
                    if formula.count("(") != formula.count(")"):
                        errors.append(f"  [{ws_name}!{cell.coordinate}] Unmatched parentheses: {formula[:80]}")

                    # 2. References to non-existent sheets
                    sheet_refs = re.findall(r"'([^']+)'!", formula)
                    for ref_sheet in sheet_refs:
                        if ref_sheet not in wb.sheetnames:
                            errors.append(f"  [{ws_name}!{cell.coordinate}] Reference to non-existent sheet '{ref_sheet}'")

                    # 3. Empty formula
                    if formula.strip() == "=":
                        errors.append(f"  [{ws_name}!{cell.coordinate}] Empty formula")

                    # 4. Division by zero pattern (direct /0)
                    if re.search(r'/0[^.]', formula) or formula.endswith('/0'):
                        warnings.append(f"  [{ws_name}!{cell.coordinate}] Possible division by zero: {formula[:80]}")

                    # 5. Check cell references are within sheet bounds
                    local_refs = re.findall(r'(?<![\'!A-Z])([A-Z]{1,3})(\d+)(?!["\d])', formula)
                    for col_ref, row_ref in local_refs:
                        row_num = int(row_ref)
                        if row_num > 200:
                            warnings.append(f"  [{ws_name}!{cell.coordinate}] Large row reference {col_ref}{row_ref}")

                else:
                    value_count += 1
                    stats["cells_value"] += 1

        print(f"  Formulas: {formula_count}")
        print(f"  Values: {value_count}")

    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"  Sheets: {stats['sheets']}")
    print(f"  Total cells: {stats['cells_total']}")
    print(f"  Formula cells: {stats['cells_formula']}")
    print(f"  Value cells: {stats['cells_value']}")

    if errors:
        print(f"\n  ERRORS ({len(errors)}):")
        for e in errors:
            print(f"    {e}")
    else:
        print("\n  ERRORS: None found")

    if warnings:
        print(f"\n  WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"    {w}")
    else:
        print("\n  WARNINGS: None")

    # Check cross-sheet references are consistent
    print(f"\n{'='*60}")
    print("CROSS-SHEET REFERENCE CHECK")
    print(f"{'='*60}")

    # Verify key cells exist
    key_checks = [
        ("DCF Model", "C46", "PGM Implied Price"),
        ("DCF Model", "C55", "Exit Multiple Implied Price"),
        ("Comps Analysis", "C27", "Comps EV/EBITDA Implied"),
        ("Comps Analysis", "C28", "Comps PER Implied"),
        # LBO key cells (guarded by sheet existence)
        ("Transaction Assumptions", "D11", "LBO Equity Value"),
        ("Transaction Assumptions", "D15", "LBO Enterprise Value"),
        ("Transaction Assumptions", "H35", "LBO Total Uses"),
        ("Balance Sheet", "F49", "BS Check (Pro Forma)"),
        ("Balance Sheet", "G49", "BS Check (Year 1)"),
        ("Cash Flow Statement", "E55", "CF Consistency (Year 1)"),
        ("Debt Schedule", "D32", "Total Debt (Pro Forma)"),
        ("Returns Analysis", "D23", "Sponsor Equity Invested"),
    ]
    for sheet, cell_addr, desc in key_checks:
        if sheet not in wb.sheetnames:
            warnings.append(f"Sheet '{sheet}' not found, skipping {desc}")
            continue
        ws = wb[sheet]
        val = ws[cell_addr].value
        has_value = val is not None
        is_formula = isinstance(val, str) and val.startswith("=") if has_value else False
        status = "FORMULA" if is_formula else ("VALUE" if has_value else "EMPTY")
        icon = "OK" if has_value else "MISSING"
        print(f"  [{icon}] {sheet}!{cell_addr} ({desc}): {status}")
        if not has_value:
            errors.append(f"Key cell {sheet}!{cell_addr} ({desc}) is empty!")

    if errors:
        print(f"\nRESULT: FAIL - {len(errors)} error(s) found")
        return 1
    else:
        print(f"\nRESULT: PASS - All checks passed")
        return 0


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else WORKBOOK
    sys.exit(check_workbook(path))
