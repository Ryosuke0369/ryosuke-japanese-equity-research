"""
dcf_comps_build_v3.py - DCF / Comps Equity Research Excel Generator (V3)

V3 changes from V2:
  - Full PL waterfall: Revenue → COGS → Gross Profit → SGA → Operating Income
  - All historical data extracted from PDFs (COGS, SGA, OCF, Capex, BS)
  - DCF projected EBIT = Revenue - COGS - SGA (full waterfall)
  - Sensitivity: Revenue Growth vs Gross Margin %
  - Net debt extracted from BS (cash - debt)

Edit the config dict below, then run:  python kudan_4425_dcf_comps.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import subprocess, sys, os

try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False

# =====================================================================
# DYNAMIC PDF EXTRACTION
# =====================================================================
from pdf_parser_spatial import extract_all_financials

_script_dir = os.path.dirname(os.path.abspath(__file__))
_pdf_result = extract_all_financials(_script_dir)

# Fallback values if PDF extraction fails
_FALLBACK = {
    "hist_revenue": [271, 332, 490, 517],
    "hist_operating_income": [-433, -598, -527, -800],
    "hist_net_income": [-2237, -413, -69, -801],
    "hist_cogs": [147, 156, 52, 177],
    "hist_sga": [558, 775, 966, 1141],
    "hist_ocf": [-515, -619, -491, -815],
    "hist_capex": [137, 20, 433, 162],
    "hist_cash": [604, 852, 1720, 2594],
    "hist_debt": [None, 200, 200, 200],
    "latest_net_debt": -2394,
}

if _pdf_result is None:
    print("FATAL: PDF extraction failed. Using fallback values.")
    _pdf_result = _FALLBACK

# Helper to get extracted values with fallback
def _get(key):
    val = _pdf_result.get(key)
    if val is None:
        return _FALLBACK.get(key)
    return val

print(f"\n[Dynamic] hist_revenue:          {_get('hist_revenue')}")
print(f"[Dynamic] hist_operating_income: {_get('hist_operating_income')}")
print(f"[Dynamic] hist_net_income:       {_get('hist_net_income')}")
print(f"[Dynamic] hist_cogs:             {_get('hist_cogs')}")
print(f"[Dynamic] hist_sga:              {_get('hist_sga')}")
print(f"[Dynamic] hist_ocf:              {_get('hist_ocf')}")
print(f"[Dynamic] hist_capex:            {_get('hist_capex')}")
print(f"[Dynamic] latest_net_debt:       {_get('latest_net_debt')}")

# =====================================================================
# CONFIG — Edit this section for each company
# =====================================================================
config = {
    # ── Company Info ──
    "company_name": "Kudan Inc.",
    "ticker": "4425.T",
    "exchange": "TSE Growth",
    "sector": "Information & Communication",
    "current_price": 2153,
    "shares_outstanding": 11_286_767,
    "net_debt": _get("latest_net_debt") or -1917,  # JPY mn (negative = net cash), from BS

    # ── Historical Financials (JPY mn) — all from PDFs ──
    "hist_years": ["FY2022 (Mar-22)", "FY2023 (Mar-23)", "FY2024 (Mar-24)", "FY2025 (Mar-25)"],
    "hist_revenue":          _get("hist_revenue"),
    "hist_operating_income": _get("hist_operating_income"),
    "hist_net_income":       _get("hist_net_income"),
    "hist_cogs":             _get("hist_cogs"),
    "hist_sga":              _get("hist_sga"),
    "hist_ocf":              _get("hist_ocf"),
    "hist_capex":            _get("hist_capex"),
    "hist_cash":             _get("hist_cash"),
    "hist_debt":             _get("hist_debt"),

    # ── DCF Assumptions — Future Projections ──
    "revenue_growth": [0.80, 0.55, 0.40, 0.30, 0.20],   # per-year driver
    "cogs_pct": [0.34, 0.34, 0.30, 0.25, 0.25],           # per-year COGS %
    "capex_pct": 0.03,
    "da_pct": 0.015,
    "tax_rate": 0.30,
    "risk_free": 0.022,
    "beta": 1.75,
    "erp": 0.060,
    "size_premium": 0.050,
    "cost_of_debt_at": 0.015,
    "de_ratio": 0.05,
    "terminal_growth": 0.020,
    "exit_multiple": 15.0,
    "projection_years": 5,
    "base_year_revenue": (_get("hist_revenue") or [517])[-1],

    # ── SGA Projection (absolute values for loss-making companies) ──
    "proj_sga_absolute": [1200, 1300, 1400, 1500, 1600],  # adjusted for Year 3 profitability

    # ── Comparable Companies ──
    "comps": [
        {"name": "PKSHA Technology", "ticker": "3993.T", "mkt_cap": 85700, "ev": 70700, "revenue": 21800,
         "ebitda": 3500, "op_income": 3000, "net_income": 2000, "pbr": 4.5, "roe": 0.12},
        {"name": "Morpho", "ticker": "3653.T", "mkt_cap": 4300, "ev": 2300, "revenue": 3400,
         "ebitda": 100, "op_income": 50, "net_income": -10, "pbr": 1.2, "roe": -0.01},
        {"name": "DMP", "ticker": "3652.T", "mkt_cap": 8500, "ev": 5500, "revenue": 1200,
         "ebitda": -200, "op_income": -300, "net_income": -400, "pbr": 2.5, "roe": -0.15},
        {"name": "HEROZ", "ticker": "4382.T", "mkt_cap": 12000, "ev": 7000, "revenue": 2500,
         "ebitda": 400, "op_income": 300, "net_income": 200, "pbr": 3.0, "roe": 0.08},
        {"name": "Headwaters", "ticker": "4011.T", "mkt_cap": 8000, "ev": 6000, "revenue": 5500,
         "ebitda": 800, "op_income": 700, "net_income": 500, "pbr": 6.5, "roe": 0.20},
        {"name": "Neural Group", "ticker": "4056.T", "mkt_cap": 3500, "ev": 2500, "revenue": 1800,
         "ebitda": -100, "op_income": -150, "net_income": -200, "pbr": 4.0, "roe": -0.10},
    ],

    # ── Kudan Comps Data (for implied valuation) ──
    "core_ebitda": (_get("hist_operating_income") or [-800])[-1] + 8,
    "core_net_income": (_get("hist_net_income") or [-801])[-1],

    # ── Investment Thesis & Risks ──
    "investment_thesis": [
        "1. Global leader in Artificial Perception (SLAM) technology",
        "2. High leverage on revenue growth due to fixed-cost intensive IP licensing model",
        "3. Transitioning from R&D phase to commercial scaling phase",
    ],
    "key_risks": [
        "1. Prolonged losses and negative free cash flow burning cash runway",
        "2. Long sales cycles converting PoC (Proof of Concept) to commercial licenses",
        "3. High WACC (17.7%) depressing present value heavily",
    ],

    # ── Sensitivity Analysis Ranges ──
    "sens_wacc_range":       [0.13, 0.15, 0.170, 0.177, 0.19, 0.21, 0.23],
    "sens_tg_range":         [0.005, 0.01, 0.015, 0.02, 0.025, 0.03, 0.035],
    "sens_rev_growth_range": [0.30, 0.40, 0.50, 0.60, 0.70, 0.80],
    "sens_gm_range":         [0.50, 0.55, 0.60, 0.65, 0.70, 0.75],  # V3: Gross Margin %

    # ── V3 Settings ──
    "primary_multiple": "EV/Sales",  # "EV/EBITDA" or "EV/Sales"
}

# =====================================================================
# V2: DYNAMIC STOCK DATA FETCHING
# =====================================================================
def get_live_market_data(ticker_str, fallback_price, fallback_shares):
    if not YFINANCE_AVAILABLE:
        print("yfinance not installed. Using fallback market data.")
        return fallback_price, fallback_shares

    try:
        print(f"Fetching live data for {ticker_str} via yfinance...")
        tkr = yf.Ticker(ticker_str)
        info = tkr.info
        live_price = info.get("currentPrice") or info.get("regularMarketPrice") or fallback_price
        live_shares = info.get("sharesOutstanding") or fallback_shares
        print(f"Successfully fetched: Price={live_price}, Shares={live_shares}")
        return float(live_price), int(live_shares)
    except Exception as e:
        print(f"Warning: Failed to fetch live data ({e}). Using fallback market data.")
        return fallback_price, fallback_shares

config["current_price"], config["shares_outstanding"] = get_live_market_data(
    config.get("ticker", ""),
    config.get("current_price", 0),
    config.get("shares_outstanding", 0)
)

# =====================================================================
# V3: SETTINGS
# =====================================================================
USE_EV_SALES = (config.get("primary_multiple", "EV/EBITDA") == "EV/Sales")

# =====================================================================
# V3: ROW NUMBERS — Full waterfall, no SGA_OFFSET toggle
# =====================================================================
R_DRV_GROWTH   = 28  # driver row: Revenue Growth (YoY)
R_DRV_COGS     = 29  # driver row: COGS % of Revenue
R_DRV_SGA      = 30  # driver row: SGA Expense
R_REVENUE      = 31
R_COGS         = 32
R_GROSS_PROFIT = 33
R_GROSS_MARGIN = 34
R_SGA          = 35
R_OP_M_IMPL   = 36
R_EBIT         = 37
R_TAX          = 38
R_NOPAT        = 39
R_DA           = 40
R_CAPEX        = 41
R_UFCF         = 42
R_DISC         = 43
R_PV_FCF       = 44
# PGM section: R_PV_FCF + 2 gap
R_PGM_SEC    = R_PV_FCF + 2
R_SUM_PV     = R_PGM_SEC + 1
R_TV_PGM     = R_SUM_PV + 1
R_PV_TV_PGM  = R_TV_PGM + 1
R_EV_PGM     = R_PV_TV_PGM + 1
R_EQ_PGM     = R_EV_PGM + 1
R_PRICE_PGM  = R_EQ_PGM + 1
# Exit section: R_PRICE_PGM + 2 gap
R_EXIT_SEC   = R_PRICE_PGM + 2
R_SUM_PV_EX  = R_EXIT_SEC + 1
R_YR5_EBITDA = R_SUM_PV_EX + 1
R_TV_EXIT    = R_YR5_EBITDA + 1
R_PV_TV_EXIT = R_TV_EXIT + 1
R_EV_EXIT    = R_PV_TV_EXIT + 1
R_EQ_EXIT    = R_EV_EXIT + 1
R_PRICE_EXIT = R_EQ_EXIT + 1

# =====================================================================
# STYLE CONSTANTS
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="0000CC", bold=False)
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
GREEN_FONT  = Font(name="Arial", size=10, color="006600")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
GREY_FONT   = Font(name="Arial", size=9, italic=True, color="808080")

HEADER_FILL  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
LIGHT_FILL   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL   = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER   = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))
BOTTOM_DOUBLE = Border(bottom=Side(style="double"))
TOP_BOTTOM    = Border(top=Side(style="thin"), bottom=Side(style="double"))

FMT_YEN     = '#,##0;(#,##0)'
FMT_YEN_DEC = '#,##0.0;(#,##0.0)'
FMT_PCT     = '0.0%;(0.0%)'
FMT_PCT2    = '0.00%;(0.00%)'
FMT_RATIO   = '0.00"x"'
FMT_INT     = '#,##0'
FMT_EPS     = '#,##0.0;(#,##0.0)'

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c

def header_row(ws, row, col_start, col_end, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c

def col_letter(col_num):
    return get_column_letter(col_num)

# =====================================================================
# SENSITIVITY ANALYSIS HELPERS (V3: full waterfall)
# =====================================================================
def calc_dcf_pgm(rev_growth, gross_margin, wacc, tg, cfg):
    """Calculate implied share price using Perpetuity Growth Method.
    V3: Uses gross_margin (GM%) and absolute SGA to compute EBIT.
    """
    n = cfg["projection_years"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]
    sga_list = cfg["proj_sga_absolute"]

    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_fcf = 0
    for yr_idx, rev in enumerate(revenues):
        cogs = rev * (1 - gross_margin)
        gp = rev - cogs
        sga = sga_list[yr_idx]
        ebit = gp - sga
        # NOPAT floor: no tax benefit when EBIT < 0
        if ebit < 0:
            nopat = ebit
        else:
            nopat = ebit * (1 - tax)
        da = rev * da_pct
        capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_fcf = fcf

    tv = last_fcf * (1 + tg) / (wacc - tg)
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

def calc_dcf_exit(rev_growth, gross_margin, wacc, exit_mult, cfg):
    """Calculate implied share price using Exit Multiple Method.
    V3: Uses gross_margin (GM%) and absolute SGA.
    """
    n = cfg["projection_years"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]
    sga_list = cfg["proj_sga_absolute"]

    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_ebit = 0
    for yr_idx, rev in enumerate(revenues):
        cogs = rev * (1 - gross_margin)
        gp = rev - cogs
        sga = sga_list[yr_idx]
        ebit = gp - sga
        if ebit < 0:
            nopat = ebit
        else:
            nopat = ebit * (1 - tax)
        da = rev * da_pct
        capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_ebit = ebit

    yr5_ebitda = last_ebit + revenues[-1] * da_pct
    tv = yr5_ebitda * exit_mult
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

# =====================================================================
# Derived values for WACC
# =====================================================================
def calc_wacc(cfg):
    ke = cfg["risk_free"] + cfg["beta"] * cfg["erp"] + cfg["size_premium"]
    we = 1 / (1 + cfg["de_ratio"])
    wd = cfg["de_ratio"] / (1 + cfg["de_ratio"])
    return ke * we + cfg["cost_of_debt_at"] * wd

# =====================================================================
# BUILD WORKBOOK
# =====================================================================
wb = openpyxl.Workbook()
C = config  # shorthand

ticker_safe = C["ticker"].replace(".", "")
output_file = f"{ticker_safe}_Equity_Research_V3.xlsx"

# =====================================================================
# SHEET 1: Executive Summary
# =====================================================================
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "003366"

ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 22

# Disclaimer
set_cell(ws1, 1, 2,
    "DISCLAIMER: This is a sample analysis for demonstration purposes only. "
    "It does not constitute investment advice.",
    font=GREY_FONT)
ws1.merge_cells("B1:E1")

# Title
set_cell(ws1, 3, 2, f'{C["company_name"]} ({C["ticker"]})', font=TITLE_FONT)
ws1.merge_cells("B3:D3")
set_cell(ws1, 4, 2, "Equity Research Report", font=SUB_FONT)

# Company info
set_cell(ws1, 6, 2, "Company", font=BOLD_FONT)
set_cell(ws1, 6, 3, C["company_name"])
set_cell(ws1, 7, 2, "Ticker", font=BOLD_FONT)
set_cell(ws1, 7, 3, f'{C["ticker"]} ({C["exchange"]})')
set_cell(ws1, 8, 2, "Sector", font=BOLD_FONT)
set_cell(ws1, 8, 3, C["sector"])
set_cell(ws1, 9, 2, "Current Price", font=BOLD_FONT)
set_cell(ws1, 9, 3, C["current_price"], font=BLUE_FONT, fmt=FMT_YEN)

# Target Price = average of 4 methods (C16:C19)
set_cell(ws1, 10, 2, "Target Price (Mid)", font=BOLD_FONT)
set_cell(ws1, 10, 3, "=ROUND(AVERAGE(C16:C19),0)", font=BLACK_FONT, fmt=FMT_YEN)

# Recommendation
set_cell(ws1, 11, 2, "Recommendation", font=BOLD_FONT)
set_cell(ws1, 11, 3, '=IF(C12>0.15,"BUY",IF(C12>0.05,"HOLD","SELL"))', font=BLACK_FONT)

# Upside / Downside
set_cell(ws1, 12, 2, "Upside / Downside", font=BOLD_FONT)
set_cell(ws1, 12, 3, "=(C10-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Valuation Summary section
c = set_cell(ws1, 14, 2, "Valuation Summary", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=14, column=col_idx).fill = LIGHT_FILL

header_row(ws1, 15, 2, 4, ["Methodology", "Implied Value (JPY)", "vs Current Price"])

# DCF - Perpetuity Growth
set_cell(ws1, 16, 2, "DCF - Perpetuity Growth")
set_cell(ws1, 16, 3, f"='DCF Model'!C{R_PRICE_PGM}", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 16, 4, "=(C16-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# DCF - Exit Multiple
set_cell(ws1, 17, 2, "DCF - Exit Multiple")
set_cell(ws1, 17, 3, f"='DCF Model'!C{R_PRICE_EXIT}", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 17, 4, "=(C17-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Comps
if USE_EV_SALES:
    set_cell(ws1, 18, 2, "Comps - EV/Sales Median")
else:
    set_cell(ws1, 18, 2, "Comps - EV/EBITDA Median")
set_cell(ws1, 18, 3, "='Comps Analysis'!C27", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 18, 4, "=(C18-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

set_cell(ws1, 19, 2, "Comps - PER Median")
set_cell(ws1, 19, 3, "='Comps Analysis'!C28", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 19, 4, "=(C19-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Integrated Valuation Range
set_cell(ws1, 21, 2, "Integrated Valuation Range", font=BOLD_FONT)
set_cell(ws1, 21, 3, '=MIN(C16:C19)&" - "&MAX(C16:C19)', font=BLACK_FONT)

# Investment Thesis
c = set_cell(ws1, 23, 2, "Key Investment Thesis", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=23, column=col_idx).fill = LIGHT_FILL
for i, line in enumerate(C["investment_thesis"]):
    set_cell(ws1, 24 + i, 2, line)

# Key Risks
c = set_cell(ws1, 28, 2, "Key Risks", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=28, column=col_idx).fill = LIGHT_FILL
for i, line in enumerate(C["key_risks"]):
    set_cell(ws1, 29 + i, 2, line)

# =====================================================================
# SHEET 2: Financial Statements (V3 — Full PL Waterfall + BS Highlights)
# =====================================================================
ws2 = wb.create_sheet("Financial Statements")
ws2.sheet_properties.tabColor = "003366"

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 32
for letter in ["C", "D", "E", "F"]:
    ws2.column_dimensions[letter].width = 18

set_cell(ws2, 2, 2, f'{C["company_name"]} - Historical Financials (JPY mn)', font=TITLE_FONT)

n_hist = len(C["hist_years"])
header_row(ws2, 4, 3, 3 + n_hist - 1, C["hist_years"])

# ── Income Statement (V3 full waterfall) ──
section_title(ws2, 5, 2, "Income Statement")

set_cell(ws2, 6, 2, "Revenue", font=BOLD_FONT)
set_cell(ws2, 7, 2, "COGS", font=BOLD_FONT)
set_cell(ws2, 8, 2, "Gross Profit", font=BOLD_FONT)
set_cell(ws2, 9, 2, "Gross Margin")
set_cell(ws2, 10, 2, "SGA", font=BOLD_FONT)
set_cell(ws2, 11, 2, "Operating Income", font=BOLD_FONT)
set_cell(ws2, 12, 2, "Net Income", font=BOLD_FONT)
set_cell(ws2, 13, 2, "Operating Margin")
set_cell(ws2, 14, 2, "Net Margin")
set_cell(ws2, 15, 2, "Revenue Growth (YoY)")
set_cell(ws2, 16, 2, "Operating Income Growth (YoY)")

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)

    # Revenue
    set_cell(ws2, 6, col, C["hist_revenue"][i], font=BLUE_FONT, fmt=FMT_YEN)
    # COGS
    cogs_val = C["hist_cogs"][i] if C["hist_cogs"] and i < len(C["hist_cogs"]) else None
    set_cell(ws2, 7, col, cogs_val, font=BLUE_FONT, fmt=FMT_YEN)
    # Gross Profit = Revenue - COGS
    set_cell(ws2, 8, col, f"={cl}6-{cl}7", font=BLACK_FONT, fmt=FMT_YEN)
    # Gross Margin = GP / Revenue
    set_cell(ws2, 9, col, f"={cl}8/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    # SGA
    sga_val = C["hist_sga"][i] if C["hist_sga"] and i < len(C["hist_sga"]) else None
    set_cell(ws2, 10, col, sga_val, font=BLUE_FONT, fmt=FMT_YEN)
    # Operating Income
    set_cell(ws2, 11, col, C["hist_operating_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
    # Net Income
    set_cell(ws2, 12, col, C["hist_net_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
    # Operating Margin
    set_cell(ws2, 13, col, f"={cl}11/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    # Net Margin
    set_cell(ws2, 14, col, f"={cl}12/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    # YoY Growth
    if i == 0:
        set_cell(ws2, 15, col, "n/a")
        set_cell(ws2, 16, col, "n/a")
    else:
        prev_cl = col_letter(col - 1)
        set_cell(ws2, 15, col, f"=({cl}6-{prev_cl}6)/{prev_cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        set_cell(ws2, 16, col, f"=({cl}11-{prev_cl}11)/{prev_cl}11", font=BLACK_FONT, fmt=FMT_PCT)

# ── Cash Flow Statement ──
section_title(ws2, 18, 2, "Cash Flow Statement")
set_cell(ws2, 19, 2, "Operating Cash Flow", font=BOLD_FONT)
set_cell(ws2, 20, 2, "Capex", font=BOLD_FONT)
set_cell(ws2, 21, 2, "Free Cash Flow", font=BOLD_FONT)
set_cell(ws2, 22, 2, "FCF Margin")
set_cell(ws2, 23, 2, "Capex / Revenue")

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)
    ocf_val = C["hist_ocf"][i] if C["hist_ocf"] and i < len(C["hist_ocf"]) else None
    capex_val = C["hist_capex"][i] if C["hist_capex"] and i < len(C["hist_capex"]) else None
    set_cell(ws2, 19, col, ocf_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 20, col, capex_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 21, col, f"={cl}19-{cl}20", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws2, 22, col, f"={cl}21/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws2, 23, col, f"={cl}20/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)

# ── Balance Sheet Highlights ──
section_title(ws2, 25, 2, "Balance Sheet Highlights")
set_cell(ws2, 26, 2, "Cash & Deposits", font=BOLD_FONT)
set_cell(ws2, 27, 2, "Short-term Debt", font=BOLD_FONT)
set_cell(ws2, 28, 2, "Net Debt (Cash)", font=BOLD_FONT)

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)
    cash_val = C["hist_cash"][i] if C["hist_cash"] and i < len(C["hist_cash"]) else None
    debt_val = C["hist_debt"][i] if C["hist_debt"] and i < len(C["hist_debt"]) else None
    set_cell(ws2, 26, col, cash_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 27, col, debt_val if debt_val is not None else 0, font=BLUE_FONT, fmt=FMT_YEN)
    # Net Debt = Debt - Cash (negative = net cash)
    set_cell(ws2, 28, col, f"={cl}27-{cl}26", font=BLACK_FONT, fmt=FMT_YEN)

# =====================================================================
# SHEET 3: DCF Model (V3 — Full Waterfall)
# =====================================================================
ws3 = wb.create_sheet("DCF Model")
ws3.sheet_properties.tabColor = "006600"

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 32
for letter in ["C", "D", "E", "F", "G"]:
    ws3.column_dimensions[letter].width = 16

set_cell(ws3, 2, 2, f'DCF Valuation Model - {C["company_name"]}', font=TITLE_FONT)

# ── Assumptions ──
c = section_title(ws3, 4, 2, "Assumptions")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=4, column=col_idx).fill = LIGHT_FILL

assumptions = [
    # Revenue Growth Rate and COGS % removed — now in per-year driver rows
    ("Capex / Revenue",            C["capex_pct"],            FMT_PCT),      # C5
    ("Effective Tax Rate",         C["tax_rate"],             FMT_PCT),      # C6
    ("Risk-Free Rate",             C["risk_free"],            FMT_PCT),      # C7
    ("Beta",                       C["beta"],                 "0.00"),       # C8
    ("Equity Risk Premium",        C["erp"],                  FMT_PCT),      # C9
    ("Size Premium",               C["size_premium"],         FMT_PCT),      # C10
    ("After-tax Cost of Debt",     C["cost_of_debt_at"],      FMT_PCT),      # C11
    ("D/E Ratio",                  C["de_ratio"],             "0.000"),      # C12
    ("Terminal Growth Rate",       C["terminal_growth"],      FMT_PCT),      # C13
    ("Exit Multiple (EV/EBITDA)",  C["exit_multiple"],        FMT_RATIO),    # C14
    ("Fully Diluted Shares",       C["shares_outstanding"],   FMT_INT),      # C15
    ("Net Debt (JPY mn)",          C["net_debt"],             FMT_YEN),      # C16
    ("Base Year Revenue (JPY mn)", C["base_year_revenue"],    FMT_YEN),      # C17
    ("D&A / Revenue",              C["da_pct"],               FMT_PCT),      # C18
]
for i, (label, val, fmt) in enumerate(assumptions):
    r = 5 + i
    set_cell(ws3, r, 2, label, font=BOLD_FONT)
    set_cell(ws3, r, 3, val, font=BLUE_FONT, fmt=fmt)

# ── WACC Calculation ──
c = section_title(ws3, 20, 2, "WACC Calculation")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=20, column=col_idx).fill = LIGHT_FILL

set_cell(ws3, 21, 2, "Cost of Equity (Ke)", font=BOLD_FONT)
set_cell(ws3, 21, 3, "=C7+C8*C9+C10", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 22, 2, "Weight of Equity", font=BOLD_FONT)
set_cell(ws3, 22, 3, "=1/(1+C12)", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 23, 2, "Weight of Debt", font=BOLD_FONT)
set_cell(ws3, 23, 3, "=C12/(1+C12)", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 24, 2, "WACC", font=BOLD_FONT)
set_cell(ws3, 24, 3, "=C21*C22+C11*C23", font=BLACK_FONT, fmt=FMT_PCT2)

# ── Projected FCF (V3 Full Waterfall with Driver Rows) ──
c = section_title(ws3, 26, 2, "Projected Free Cash Flow")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=26, column=col_idx).fill = LIGHT_FILL

proj_years = C["projection_years"]
year_labels = [f"Year {y}" for y in range(1, proj_years + 1)]
header_row(ws3, 27, 3, 3 + proj_years - 1, year_labels)

# Driver row labels
row_labels_drv = [
    ("Revenue Growth (YoY)",          R_DRV_GROWTH),
    ("COGS % of Revenue",             R_DRV_COGS),
    ("SGA Expense",                   R_DRV_SGA),
]
for label, r in row_labels_drv:
    set_cell(ws3, r, 2, label, font=BOLD_FONT)

# V3: FCF row labels — full waterfall
row_labels_fcf = [
    ("Revenue",                       R_REVENUE),
    ("COGS",                          R_COGS),
    ("Gross Profit",                  R_GROSS_PROFIT),
    ("Gross Margin",                  R_GROSS_MARGIN),
    ("SGA Expense",                   R_SGA),
    ("Implied Operating Margin",      R_OP_M_IMPL),
    ("Operating Income (EBIT)",       R_EBIT),
    ("Less: Tax",                     R_TAX),
    ("NOPAT",                         R_NOPAT),
    ("Plus: D&A",                     R_DA),
    ("Less: Capex",                   R_CAPEX),
    ("Unlevered Free Cash Flow",      R_UFCF),
    ("Discount Factor",               R_DISC),
    ("PV of FCF",                     R_PV_FCF),
]
for label, r in row_labels_fcf:
    set_cell(ws3, r, 2, label, font=BOLD_FONT)

# V3: Year-by-year projection loop — full waterfall with driver rows
for yr in range(proj_years):
    col = 3 + yr
    cl = col_letter(col)
    prev_cl = col_letter(col - 1) if yr > 0 else None

    # ── Driver rows (input cells: yellow fill + blue font) ──
    set_cell(ws3, R_DRV_GROWTH, col, config["revenue_growth"][yr],
             font=BLUE_FONT, fmt=FMT_PCT, fill=INPUT_FILL)
    set_cell(ws3, R_DRV_COGS, col, config["cogs_pct"][yr],
             font=BLUE_FONT, fmt=FMT_PCT, fill=INPUT_FILL)
    set_cell(ws3, R_DRV_SGA, col, config["proj_sga_absolute"][yr],
             font=BLUE_FONT, fmt=FMT_YEN, fill=INPUT_FILL)

    # Revenue — references driver row
    if yr == 0:
        set_cell(ws3, R_REVENUE, col, f"=C17*(1+{cl}{R_DRV_GROWTH})", font=BLACK_FONT, fmt=FMT_YEN)
    else:
        set_cell(ws3, R_REVENUE, col, f"={prev_cl}{R_REVENUE}*(1+{cl}{R_DRV_GROWTH})", font=BLACK_FONT, fmt=FMT_YEN)

    # COGS = Revenue * COGS% driver
    set_cell(ws3, R_COGS, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_COGS}", font=BLACK_FONT, fmt=FMT_YEN)

    # Gross Profit = Revenue - COGS
    set_cell(ws3, R_GROSS_PROFIT, col, f"={cl}{R_REVENUE}-{cl}{R_COGS}", font=BLACK_FONT, fmt=FMT_YEN)

    # Gross Margin = GP / Revenue
    set_cell(ws3, R_GROSS_MARGIN, col, f"={cl}{R_GROSS_PROFIT}/{cl}{R_REVENUE}", font=BLACK_FONT, fmt=FMT_PCT)

    # SGA Expense — references driver row (calculated cell, black font)
    set_cell(ws3, R_SGA, col, f"={cl}{R_DRV_SGA}", font=BLACK_FONT, fmt=FMT_YEN)

    # Implied Operating Margin = (GP - SGA) / Revenue
    set_cell(ws3, R_OP_M_IMPL, col,
             f"=({cl}{R_GROSS_PROFIT}-{cl}{R_SGA})/{cl}{R_REVENUE}",
             font=BLACK_FONT, fmt=FMT_PCT)

    # EBIT = Gross Profit - SGA
    set_cell(ws3, R_EBIT, col, f"={cl}{R_GROSS_PROFIT}-{cl}{R_SGA}", font=BLACK_FONT, fmt=FMT_YEN)

    # Tax with NOPAT floor (no tax benefit when EBIT < 0)
    set_cell(ws3, R_TAX, col, f"=MAX(0,{cl}{R_EBIT}*C6)", font=BLACK_FONT, fmt=FMT_YEN)

    # NOPAT
    set_cell(ws3, R_NOPAT, col, f"={cl}{R_EBIT}-{cl}{R_TAX}", font=BLACK_FONT, fmt=FMT_YEN)
    # D&A
    set_cell(ws3, R_DA, col, f"={cl}{R_REVENUE}*C18", font=BLACK_FONT, fmt=FMT_YEN)
    # Capex
    set_cell(ws3, R_CAPEX, col, f"={cl}{R_REVENUE}*C5", font=BLACK_FONT, fmt=FMT_YEN)
    # UFCF
    set_cell(ws3, R_UFCF, col, f"={cl}{R_NOPAT}+{cl}{R_DA}-{cl}{R_CAPEX}", font=BLACK_FONT, fmt=FMT_YEN)
    # Discount Factor
    set_cell(ws3, R_DISC, col, f"=1/(1+C24)^{yr+1}", font=BLACK_FONT, fmt="0.0000")
    # PV of FCF
    set_cell(ws3, R_PV_FCF, col, f"={cl}{R_UFCF}*{cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

# ── Valuation - Perpetuity Growth Method ──
c = section_title(ws3, R_PGM_SEC, 2, "Valuation - Perpetuity Growth Method")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=R_PGM_SEC, column=col_idx).fill = LIGHT_GREEN

last_cl = col_letter(3 + proj_years - 1)  # G for 5 years

set_cell(ws3, R_SUM_PV, 2, "Sum of PV of FCFs", font=BOLD_FONT)
set_cell(ws3, R_SUM_PV, 3, f"=SUM(C{R_PV_FCF}:{last_cl}{R_PV_FCF})", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_TV_PGM, 2, "Terminal Value (PGM)", font=BOLD_FONT)
set_cell(ws3, R_TV_PGM, 3, f"={last_cl}{R_UFCF}*(1+C13)/(C24-C13)", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PV_TV_PGM, 2, "PV of Terminal Value", font=BOLD_FONT)
set_cell(ws3, R_PV_TV_PGM, 3, f"=C{R_TV_PGM}*{last_cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EV_PGM, 2, "Enterprise Value", font=BOLD_FONT)
set_cell(ws3, R_EV_PGM, 3, f"=C{R_SUM_PV}+C{R_PV_TV_PGM}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EQ_PGM, 2, "Equity Value", font=BOLD_FONT)
set_cell(ws3, R_EQ_PGM, 3, f"=C{R_EV_PGM}-C16", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PRICE_PGM, 2, "Implied Share Price (PGM)", font=BOLD_FONT)
set_cell(ws3, R_PRICE_PGM, 3, f"=ROUND(C{R_EQ_PGM}*1000000/C15,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# ── Valuation - Exit Multiple Method ──
c = section_title(ws3, R_EXIT_SEC, 2, "Valuation - Exit Multiple Method")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=R_EXIT_SEC, column=col_idx).fill = LIGHT_GREEN

set_cell(ws3, R_SUM_PV_EX, 2, "Sum of PV of FCFs", font=BOLD_FONT)
set_cell(ws3, R_SUM_PV_EX, 3, f"=C{R_SUM_PV}", font=BLACK_FONT, fmt=FMT_YEN)

# Year 5 EBITDA = EBIT + D&A
set_cell(ws3, R_YR5_EBITDA, 2, "Year 5 EBITDA", font=BOLD_FONT)
set_cell(ws3, R_YR5_EBITDA, 3, f"={last_cl}{R_EBIT}+{last_cl}{R_DA}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_TV_EXIT, 2, "Terminal Value (Exit Multiple)", font=BOLD_FONT)
set_cell(ws3, R_TV_EXIT, 3, f"=C{R_YR5_EBITDA}*C14", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PV_TV_EXIT, 2, "PV of Terminal Value", font=BOLD_FONT)
set_cell(ws3, R_PV_TV_EXIT, 3, f"=C{R_TV_EXIT}*{last_cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EV_EXIT, 2, "Enterprise Value", font=BOLD_FONT)
set_cell(ws3, R_EV_EXIT, 3, f"=C{R_SUM_PV_EX}+C{R_PV_TV_EXIT}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EQ_EXIT, 2, "Equity Value", font=BOLD_FONT)
set_cell(ws3, R_EQ_EXIT, 3, f"=C{R_EV_EXIT}-C16", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PRICE_EXIT, 2, "Implied Share Price (Exit Multiple)", font=BOLD_FONT)
set_cell(ws3, R_PRICE_EXIT, 3, f"=ROUND(C{R_EQ_EXIT}*1000000/C15,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# =====================================================================
# SHEET 4: Comps Analysis
# =====================================================================
ws4 = wb.create_sheet("Comps Analysis")
ws4.sheet_properties.tabColor = "006600"

ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 10
for letter in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
    ws4.column_dimensions[letter].width = 12

set_cell(ws4, 2, 2, "Comparable Company Analysis", font=TITLE_FONT)

# Header row
comp_headers = [
    "Company", "Ticker", "Mkt Cap\n(JPY mn)", "EV\n(JPY mn)",
    "Revenue\n(JPY mn)", "EBITDA\n(JPY mn)", "Op Income\n(JPY mn)",
    "Net Income\n(JPY mn)", "EV/EBITDA", "EV/Revenue", "PER",
    "PBR", "Op Margin", "ROE"
]
header_row(ws4, 4, 2, 15, comp_headers)

# Company data rows (rows 5-10 for 6 comps)
comps = C["comps"]
for i, comp in enumerate(comps):
    r = 5 + i

    set_cell(ws4, r, 2, comp["name"], font=BOLD_FONT)
    set_cell(ws4, r, 3, comp["ticker"])
    set_cell(ws4, r, 4, comp["mkt_cap"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 5, comp["ev"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 6, comp["revenue"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 7, comp["ebitda"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 8, comp["op_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 9, comp["net_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    if comp["ebitda"] <= 0:
        set_cell(ws4, r, 10, "N/A", font=BLACK_FONT, border=THIN_BORDER, alignment=Alignment(horizontal="right"))
    else:
        set_cell(ws4, r, 10, f"=E{r}/G{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

    if comp["revenue"] <= 0:
        set_cell(ws4, r, 11, "N/A", font=BLACK_FONT, border=THIN_BORDER, alignment=Alignment(horizontal="right"))
    else:
        set_cell(ws4, r, 11, f"=E{r}/F{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

    if comp["net_income"] <= 0:
        set_cell(ws4, r, 12, "N/A", font=BLACK_FONT, border=THIN_BORDER, alignment=Alignment(horizontal="right"))
    else:
        set_cell(ws4, r, 12, f"=D{r}/I{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)
    set_cell(ws4, r, 13, comp["pbr"], font=BLUE_FONT, fmt=FMT_RATIO, border=THIN_BORDER)
    set_cell(ws4, r, 14, f"=H{r}/F{r}", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws4, r, 15, comp["roe"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)

last_comp_row = 5 + len(comps) - 1

# ── Statistics ──
section_title(ws4, 14, 2, "Statistics")

stat_labels = ["25th Percentile", "Median (50th)", "75th Percentile"]
stat_rows = [15, 16, 17]

stat_col_map = [
    (4, 10),  # EV/EBITDA
    (5, 11),  # EV/Revenue
    (6, 12),  # PER
    (7, 13),  # PBR
    (8, 14),  # Op Margin
    (9, 15),  # ROE
]

for stat_idx, (label, r) in enumerate(zip(stat_labels, stat_rows)):
    set_cell(ws4, r, 2, label, font=BOLD_FONT)

    for dst_col, src_col in stat_col_map:
        src_letter = col_letter(src_col)
        rng = f"{src_letter}5:{src_letter}{last_comp_row}"

        if src_col in (14, 15):
            fmt = FMT_PCT
        else:
            fmt = FMT_RATIO

        if stat_idx == 0:
            formula = f"=PERCENTILE({rng},0.25)"
        elif stat_idx == 1:
            formula = f"=MEDIAN({rng})"
        else:
            formula = f"=PERCENTILE({rng},0.75)"

        set_cell(ws4, r, dst_col, formula, font=BLACK_FONT, fmt=fmt, border=THIN_BORDER)

# ── Implied Valuation ──
c = section_title(ws4, 19, 2, f'Implied Valuation for {C["company_name"]}')
c.fill = LIGHT_GREEN
for col_idx in range(3, 10):
    ws4.cell(row=19, column=col_idx).fill = LIGHT_GREEN

section_title(ws4, 20, 2, "Kudan Financials")

if USE_EV_SALES:
    set_cell(ws4, 21, 2, "Revenue (JPY mn)", font=BOLD_FONT)
    set_cell(ws4, 21, 3, C["base_year_revenue"], font=BLUE_FONT, fmt=FMT_YEN)
else:
    set_cell(ws4, 21, 2, "EBITDA (JPY mn)", font=BOLD_FONT)
    set_cell(ws4, 21, 3, C["core_ebitda"], font=BLUE_FONT, fmt=FMT_YEN)

set_cell(ws4, 22, 2, "Net Income (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 22, 3, C["core_net_income"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws4, 23, 2, "Shares Outstanding", font=BOLD_FONT)
set_cell(ws4, 23, 3, C["shares_outstanding"], font=BLUE_FONT, fmt=FMT_INT)
set_cell(ws4, 24, 2, "Net Debt (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 24, 3, C["net_debt"], font=BLUE_FONT, fmt=FMT_YEN)

section_title(ws4, 26, 2, "Implied Share Price (Median Multiples)")

if USE_EV_SALES:
    set_cell(ws4, 27, 2, "Via EV/Sales (Median)", font=BOLD_FONT)
    set_cell(ws4, 27, 3, "=ROUND((C21*E16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)
else:
    set_cell(ws4, 27, 2, "Via EV/EBITDA (Median)", font=BOLD_FONT)
    set_cell(ws4, 27, 3, "=ROUND((C21*D16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)

set_cell(ws4, 28, 2, "Via PER (Median)", font=BOLD_FONT)
set_cell(ws4, 28, 3, "=ROUND(C22*F16*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# =====================================================================
# SHEET 5: Sensitivity Analysis (V3 — Gross Margin instead of Op Margin)
# =====================================================================
ws5 = wb.create_sheet("Sensitivity Analysis")
ws5.sheet_properties.tabColor = "996600"

ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 24
for letter in ["C", "D", "E", "F", "G", "H", "I"]:
    ws5.column_dimensions[letter].width = 14

set_cell(ws5, 2, 2, "Sensitivity Analysis", font=TITLE_FONT)

# ── Table 1: WACC vs Terminal Growth Rate ──
section_title(ws5, 4, 2, "Table 1: WACC vs Terminal Growth Rate (Implied Share Price, JPY)")

wacc_range = C["sens_wacc_range"]
tg_range = C["sens_tg_range"]

set_cell(ws5, 5, 2, "WACC \\ Terminal g", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j, tg in enumerate(tg_range):
    set_cell(ws5, 5, 3 + j, tg, font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

base_wacc = calc_wacc(C)
base_gm = 1 - C["cogs_pct"][0]  # Base gross margin (Year 1)

for i, wacc in enumerate(wacc_range):
    r = 6 + i
    set_cell(ws5, r, 2, wacc, font=BOLD_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    for j, tg in enumerate(tg_range):
        price = calc_dcf_pgm(C["revenue_growth"][0], base_gm, wacc, tg, C)
        is_base = (abs(wacc - base_wacc) < 0.002 and abs(tg - C["terminal_growth"]) < 0.002)
        fill = LIGHT_YELLOW if is_base else None
        set_cell(ws5, r, 3 + j, price, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER,
                 fill=fill)

# ── Table 2: Revenue Growth vs Gross Margin % ──
table2_start = 6 + len(wacc_range) + 2
section_title(ws5, table2_start, 2,
              "Table 2: Revenue Growth vs Gross Margin % (Implied Share Price, JPY)")

rev_growth_range = C["sens_rev_growth_range"]
gm_range = C["sens_gm_range"]

hdr_row = table2_start + 1
set_cell(ws5, hdr_row, 2, "Rev Growth \\ Gross Margin", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j, gm in enumerate(gm_range):
    set_cell(ws5, hdr_row, 3 + j, gm, font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

for i, rg in enumerate(rev_growth_range):
    r = hdr_row + 1 + i
    set_cell(ws5, r, 2, rg, font=BOLD_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    for j, gm in enumerate(gm_range):
        price = calc_dcf_pgm(rg, gm, base_wacc, C["terminal_growth"], C)
        is_base = (abs(rg - C["revenue_growth"][0]) < 0.002 and abs(gm - base_gm) < 0.002)
        fill = LIGHT_YELLOW if is_base else None
        set_cell(ws5, r, 3 + j, price, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER,
                 fill=fill)

# =====================================================================
# SAVE & VERIFY
# =====================================================================
wb.save(output_file)
print(f"\nSaved: {output_file}")

# Run recalc.py for verification
recalc_script = os.path.join("scripts", "recalc.py")
if os.path.exists(recalc_script):
    print(f"\nRunning verification: python {recalc_script} {output_file}")
    result = subprocess.run([sys.executable, recalc_script, output_file],
                            capture_output=True, text=True)
    print(result.stdout)
    if result.stderr:
        print("STDERR:", result.stderr)
else:
    print(f"\nNote: {recalc_script} not found, skipping verification.")
