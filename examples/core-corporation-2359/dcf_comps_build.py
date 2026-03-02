"""
dcf_comps_build.py - DCF / Comps Equity Research Excel Generator
Edit the config dict below, then run:  python dcf_comps_build.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import subprocess, sys, os

# =====================================================================
# CONFIG — Edit this section for each company
# =====================================================================
config = {
    # ── Company Info ──
    "company_name": "Core Corporation",
    "ticker": "2359.T",
    "exchange": "TSE Prime",
    "sector": "Information Technology / GIS & System Integration",
    "current_price": 2240,
    "shares_outstanding": 14_844_000,
    "net_debt": -5937,  # JPY mn (negative = net cash)

    # ── Historical Financials (JPY mn) ──
    "hist_years": ["FY2022 (Mar-22)", "FY2023 (Mar-23)", "FY2024 (Mar-24)", "FY2025 (Mar-25)"],
    "hist_revenue":          [21798, 22848, 23999, 24599],
    "hist_operating_income": [2368,  2744,  3141,  3175],
    "hist_net_income":       [1679,  1885,  2095,  2115],
    "hist_ocf":              [1799,  1944,  2190,  2373],
    "hist_capex":            [136,   191,   133,   132],

    # ── DCF Assumptions ──
    "revenue_growth": 0.072,
    "op_margin": 0.1374,
    "capex_pct": 0.0054,
    "da_pct": 0.0054,
    "tax_rate": 0.3062,
    "risk_free": 0.022,
    "beta": 0.80,
    "erp": 0.06,
    "size_premium": 0.04,
    "cost_of_debt_at": 0.007,
    "de_ratio": 0.045,
    "terminal_growth": 0.015,
    "exit_multiple": 9.0,
    "projection_years": 5,
    "base_year_revenue": 24599,

    # ── Comparable Companies ──
    "comps": [
        {"name": "Systena",  "ticker": "2317.T", "mkt_cap": 151919, "ev": 128262, "revenue": 83621,
         "ebitda": 12528, "op_income": 12068, "net_income": 8480, "pbr": 7.5, "roe": 0.25},
        {"name": "TIS",      "ticker": "3626.T", "mkt_cap": 710806, "ev": 665004, "revenue": 571687,
         "ebitda": 88614, "op_income": 69048, "net_income": 50012, "pbr": 4.2, "roe": 0.18},
        {"name": "SCSK",     "ticker": "9719.T", "mkt_cap": 1771864, "ev": 2020332, "revenue": 596065,
         "ebitda": 90988, "op_income": 66122, "net_income": 45035, "pbr": 3.8, "roe": 0.17},
        {"name": "Obic",     "ticker": "4684.T", "mkt_cap": 1677958, "ev": 1480051, "revenue": 121240,
         "ebitda": 81065, "op_income": 78378, "net_income": 64621, "pbr": 12.0, "roe": 0.14},
        {"name": "BIPROGY", "ticker": "8056.T", "mkt_cap": 449598, "ev": 441483, "revenue": 404010,
         "ebitda": 56177, "op_income": 38236, "net_income": 26965, "pbr": 2.8, "roe": 0.15},
        {"name": "DTS",      "ticker": "9682.T", "mkt_cap": 177674, "ev": 151573, "revenue": 125908,
         "ebitda": 15626, "op_income": 14493, "net_income": 10635, "pbr": 3.0, "roe": 0.20},
        {"name": "Cresco",   "ticker": "4674.T", "mkt_cap": 64097, "ev": 50827, "revenue": 58761,
         "ebitda": 6611, "op_income": 5984, "net_income": 4406, "pbr": 2.5, "roe": 0.16},
        {"name": "NSD",      "ticker": "9759.T", "mkt_cap": 220942, "ev": 189008, "revenue": 107791,
         "ebitda": 19477, "op_income": 16856, "net_income": 11795, "pbr": 3.5, "roe": 0.18},
    ],

    # ── Core Corp Comps Data (for implied valuation) ──
    "core_ebitda": 3800,
    "core_net_income": 2115,

    # ── Investment Thesis & Risks ──
    "investment_thesis": [
        "1. Growth in government GIS business driven by Michibiki satellite system expansion",
        "2. Structural expansion of defense-related IT demand",
        "3. Solid cash flow generation and shareholder returns",
    ],
    "key_risks": [
        "1. High dependency on government budget cycles",
        "2. Intensifying competition for IT talent acquisition",
        "3. Revenue growth deceleration (recent: 2.5%)",
    ],

    # ── Sensitivity Analysis Ranges ──
    "sens_wacc_range":       [0.090, 0.095, 0.100, 0.105, 0.110, 0.115, 0.120, 0.125],
    "sens_tg_range":         [0.005, 0.010, 0.015, 0.020, 0.025],
    "sens_rev_growth_range": [0.05, 0.06, 0.07, 0.08, 0.09, 0.10],
    "sens_op_margin_range":  [0.11, 0.12, 0.13, 0.14, 0.15, 0.16],
}

# =====================================================================
# STYLE CONSTANTS (unified with M&A build.py)
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="0000CC", bold=False)
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
GREEN_FONT  = Font(name="Arial", size=10, color="006600")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
GREY_FONT   = Font(name="Arial", size=9, italic=True, color="808080")

HEADER_FILL  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
LIGHT_FILL   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL   = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER   = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))
BOTTOM_DOUBLE = Border(bottom=Side(style="double"))
TOP_BOTTOM    = Border(top=Side(style="thin"), bottom=Side(style="double"))

FMT_YEN     = '#,##0;(#,##0)'
FMT_YEN_DEC = '#,##0.0;(#,##0.0)'
FMT_PCT     = '0.0%;(0.0%)'
FMT_PCT2    = '0.00%;(0.00%)'
FMT_RATIO   = '0.00"x"'
FMT_INT     = '#,##0'
FMT_EPS     = '#,##0.0;(#,##0.0)'

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c

def header_row(ws, row, col_start, col_end, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c

def col_letter(col_num):
    return get_column_letter(col_num)

# =====================================================================
# SENSITIVITY ANALYSIS HELPERS (Python calculations for hardcoded values)
# =====================================================================
def calc_dcf_pgm(rev_growth, op_margin, wacc, tg, cfg):
    """Calculate implied share price using Perpetuity Growth Method."""
    n = cfg["projection_years"]
    base_rev = cfg["base_year_revenue"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]

    revenues = []
    for yr in range(1, n + 1):
        base_rev = base_rev * (1 + rev_growth) if yr == 1 else revenues[-1] * (1 + rev_growth)
        revenues.append(base_rev if yr == 1 else revenues[-1] * (1 + rev_growth))
    # Recalculate properly
    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_fcf = 0
    for yr_idx, rev in enumerate(revenues):
        ebit = rev * op_margin
        nopat = ebit * (1 - tax)
        da = rev * da_pct
        capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_fcf = fcf

    tv = last_fcf * (1 + tg) / (wacc - tg)
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

def calc_dcf_exit(rev_growth, op_margin, wacc, exit_mult, cfg):
    """Calculate implied share price using Exit Multiple Method."""
    n = cfg["projection_years"]
    base_rev = cfg["base_year_revenue"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]

    revenues = []
    rev = base_rev
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    for yr_idx, rev in enumerate(revenues):
        ebit = rev * op_margin
        nopat = ebit * (1 - tax)
        da = rev * da_pct
        capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df

    yr5_rev = revenues[-1]
    yr5_ebitda = yr5_rev * (op_margin + da_pct)
    tv = yr5_ebitda * exit_mult
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

# =====================================================================
# Derived values for WACC (used in sensitivity base case)
# =====================================================================
def calc_wacc(cfg):
    ke = cfg["risk_free"] + cfg["beta"] * cfg["erp"] + cfg["size_premium"]
    we = 1 / (1 + cfg["de_ratio"])
    wd = cfg["de_ratio"] / (1 + cfg["de_ratio"])
    return ke * we + cfg["cost_of_debt_at"] * wd

# =====================================================================
# BUILD WORKBOOK
# =====================================================================
wb = openpyxl.Workbook()
C = config  # shorthand

ticker_safe = C["ticker"].replace(".", "")
output_file = f"{ticker_safe}_Equity_Research.xlsx"

# =====================================================================
# SHEET 1: Executive Summary
# =====================================================================
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "003366"

ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 22

# Disclaimer
set_cell(ws1, 1, 2,
    "DISCLAIMER: This is a sample analysis for demonstration purposes only. "
    "It does not constitute investment advice.",
    font=GREY_FONT)
ws1.merge_cells("B1:E1")

# Title
set_cell(ws1, 3, 2, f'{C["company_name"]} ({C["ticker"]})', font=TITLE_FONT)
ws1.merge_cells("B3:D3")
set_cell(ws1, 4, 2, "Equity Research Report", font=SUB_FONT)

# Company info
set_cell(ws1, 6, 2, "Company", font=BOLD_FONT)
set_cell(ws1, 6, 3, C["company_name"])
set_cell(ws1, 7, 2, "Ticker", font=BOLD_FONT)
set_cell(ws1, 7, 3, f'{C["ticker"]} ({C["exchange"]})')
set_cell(ws1, 8, 2, "Sector", font=BOLD_FONT)
set_cell(ws1, 8, 3, C["sector"])
set_cell(ws1, 9, 2, "Current Price", font=BOLD_FONT)
set_cell(ws1, 9, 3, C["current_price"], font=BLUE_FONT, fmt=FMT_YEN)

# Target Price = average of 4 methods (C16:C19)
set_cell(ws1, 10, 2, "Target Price (Mid)", font=BOLD_FONT)
set_cell(ws1, 10, 3, "=ROUND(AVERAGE(C16:C19),0)", font=BLACK_FONT, fmt=FMT_YEN)

# Recommendation
set_cell(ws1, 11, 2, "Recommendation", font=BOLD_FONT)
set_cell(ws1, 11, 3, '=IF(C12>0.15,"BUY",IF(C12>0.05,"HOLD","SELL"))', font=BLACK_FONT)

# Upside / Downside
set_cell(ws1, 12, 2, "Upside / Downside", font=BOLD_FONT)
set_cell(ws1, 12, 3, "=(C10-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Valuation Summary section
c = set_cell(ws1, 14, 2, "Valuation Summary", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=14, column=col_idx).fill = LIGHT_FILL

header_row(ws1, 15, 2, 4, ["Methodology", "Implied Value (JPY)", "vs Current Price"])

# DCF - Perpetuity Growth
set_cell(ws1, 16, 2, "DCF - Perpetuity Growth")
set_cell(ws1, 16, 3, "='DCF Model'!C46", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 16, 4, "=(C16-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# DCF - Exit Multiple
set_cell(ws1, 17, 2, "DCF - Exit Multiple")
set_cell(ws1, 17, 3, "='DCF Model'!C55", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 17, 4, "=(C17-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Comps - EV/EBITDA Median
set_cell(ws1, 18, 2, "Comps - EV/EBITDA Median")
set_cell(ws1, 18, 3, "='Comps Analysis'!C27", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 18, 4, "=(C18-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Comps - PER Median
set_cell(ws1, 19, 2, "Comps - PER Median")
set_cell(ws1, 19, 3, "='Comps Analysis'!C28", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 19, 4, "=(C19-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

# Integrated Valuation Range
set_cell(ws1, 21, 2, "Integrated Valuation Range", font=BOLD_FONT)
set_cell(ws1, 21, 3, '=MIN(C16:C19)&" - "&MAX(C16:C19)', font=BLACK_FONT)

# Investment Thesis
c = set_cell(ws1, 23, 2, "Key Investment Thesis", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=23, column=col_idx).fill = LIGHT_FILL
for i, line in enumerate(C["investment_thesis"]):
    set_cell(ws1, 24 + i, 2, line)

# Key Risks
c = set_cell(ws1, 28, 2, "Key Risks", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=28, column=col_idx).fill = LIGHT_FILL
for i, line in enumerate(C["key_risks"]):
    set_cell(ws1, 29 + i, 2, line)

# =====================================================================
# SHEET 2: Financial Statements
# =====================================================================
ws2 = wb.create_sheet("Financial Statements")
ws2.sheet_properties.tabColor = "003366"

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 32
for letter in ["C", "D", "E", "F"]:
    ws2.column_dimensions[letter].width = 18

set_cell(ws2, 2, 2, f'{C["company_name"]} - Historical Financials (JPY mn)', font=TITLE_FONT)

n_hist = len(C["hist_years"])
header_row(ws2, 4, 3, 3 + n_hist - 1, C["hist_years"])

# Income Statement
section_title(ws2, 5, 2, "Income Statement")

set_cell(ws2, 6, 2, "Revenue", font=BOLD_FONT)
set_cell(ws2, 7, 2, "Operating Income", font=BOLD_FONT)
set_cell(ws2, 8, 2, "Net Income", font=BOLD_FONT)
set_cell(ws2, 9, 2, "Operating Margin")
set_cell(ws2, 10, 2, "Net Margin")
set_cell(ws2, 11, 2, "Revenue Growth (YoY)")
set_cell(ws2, 12, 2, "Operating Income Growth (YoY)")

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)
    set_cell(ws2, 6, col, C["hist_revenue"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 7, col, C["hist_operating_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 8, col, C["hist_net_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 9, col, f"={cl}7/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws2, 10, col, f"={cl}8/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)

    if i == 0:
        set_cell(ws2, 11, col, "n/a")
        set_cell(ws2, 12, col, "n/a")
    else:
        prev_cl = col_letter(col - 1)
        set_cell(ws2, 11, col, f"=({cl}6-{prev_cl}6)/{prev_cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        set_cell(ws2, 12, col, f"=({cl}7-{prev_cl}7)/{prev_cl}7", font=BLACK_FONT, fmt=FMT_PCT)

# Cash Flow Statement
section_title(ws2, 14, 2, "Cash Flow Statement")
set_cell(ws2, 15, 2, "Operating Cash Flow", font=BOLD_FONT)
set_cell(ws2, 16, 2, "Free Cash Flow", font=BOLD_FONT)
set_cell(ws2, 17, 2, "FCF Margin")
set_cell(ws2, 18, 2, "Capex", font=BOLD_FONT)
set_cell(ws2, 19, 2, "Capex / Revenue")

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)
    set_cell(ws2, 15, col, C["hist_ocf"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 18, col, C["hist_capex"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 16, col, f"={cl}15-{cl}18", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws2, 17, col, f"={cl}16/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws2, 19, col, f"={cl}18/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)

# =====================================================================
# SHEET 3: DCF Model
# =====================================================================
ws3 = wb.create_sheet("DCF Model")
ws3.sheet_properties.tabColor = "006600"

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 32
for letter in ["C", "D", "E", "F", "G"]:
    ws3.column_dimensions[letter].width = 16

set_cell(ws3, 2, 2, f'DCF Valuation Model - {C["company_name"]}', font=TITLE_FONT)

# ── Assumptions ──
c = section_title(ws3, 4, 2, "Assumptions")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=4, column=col_idx).fill = LIGHT_FILL

assumptions = [
    ("Revenue Growth Rate",        C["revenue_growth"],       FMT_PCT),
    ("Operating Margin",           C["op_margin"],            FMT_PCT),
    ("Capex / Revenue",            C["capex_pct"],            FMT_PCT),
    ("Effective Tax Rate",         C["tax_rate"],             FMT_PCT),
    ("Risk-Free Rate",             C["risk_free"],            FMT_PCT),
    ("Beta",                       C["beta"],                 "0.00"),
    ("Equity Risk Premium",        C["erp"],                  FMT_PCT),
    ("Size Premium",               C["size_premium"],         FMT_PCT),
    ("After-tax Cost of Debt",     C["cost_of_debt_at"],      FMT_PCT),
    ("D/E Ratio",                  C["de_ratio"],             "0.000"),
    ("Terminal Growth Rate",       C["terminal_growth"],      FMT_PCT),
    ("Exit Multiple (EV/EBITDA)",  C["exit_multiple"],        FMT_RATIO),
    ("Fully Diluted Shares",       C["shares_outstanding"],   FMT_INT),
    ("Net Debt (JPY mn)",          C["net_debt"],             FMT_YEN),
    ("Base Year Revenue (JPY mn)", C["base_year_revenue"],    FMT_YEN),
    ("D&A / Revenue",              C["da_pct"],               FMT_PCT),
]
for i, (label, val, fmt) in enumerate(assumptions):
    r = 5 + i
    set_cell(ws3, r, 2, label, font=BOLD_FONT)
    set_cell(ws3, r, 3, val, font=BLUE_FONT, fmt=fmt)

# ── WACC Calculation ──
c = section_title(ws3, 22, 2, "WACC Calculation")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=22, column=col_idx).fill = LIGHT_FILL

set_cell(ws3, 23, 2, "Cost of Equity (Ke)", font=BOLD_FONT)
set_cell(ws3, 23, 3, "=C9+C10*C11+C12", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 24, 2, "Weight of Equity", font=BOLD_FONT)
set_cell(ws3, 24, 3, "=1/(1+C14)", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 25, 2, "Weight of Debt", font=BOLD_FONT)
set_cell(ws3, 25, 3, "=C14/(1+C14)", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 26, 2, "WACC", font=BOLD_FONT)
set_cell(ws3, 26, 3, "=C23*C24+C13*C25", font=BLACK_FONT, fmt=FMT_PCT2)

# ── Projected FCF ──
c = section_title(ws3, 28, 2, "Projected Free Cash Flow")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=28, column=col_idx).fill = LIGHT_FILL

proj_years = C["projection_years"]
year_labels = [f"Year {y}" for y in range(1, proj_years + 1)]
header_row(ws3, 29, 3, 3 + proj_years - 1, year_labels)

row_labels_fcf = [
    ("Revenue",                       30),
    ("Operating Income (EBIT)",       31),
    ("Less: Tax",                     32),
    ("NOPAT",                         33),
    ("Plus: D&A",                     34),
    ("Less: Capex",                   35),
    ("Unlevered Free Cash Flow",      36),
    ("Discount Factor",               37),
    ("PV of FCF",                     38),
]
for label, r in row_labels_fcf:
    set_cell(ws3, r, 2, label, font=BOLD_FONT)

for yr in range(proj_years):
    col = 3 + yr
    cl = col_letter(col)
    prev_cl = col_letter(col - 1) if yr > 0 else None

    # Revenue
    if yr == 0:
        set_cell(ws3, 30, col, "=C19*(1+C5)", font=BLACK_FONT, fmt=FMT_YEN)
    else:
        set_cell(ws3, 30, col, f"={prev_cl}30*(1+C5)", font=BLACK_FONT, fmt=FMT_YEN)

    # EBIT
    set_cell(ws3, 31, col, f"={cl}30*C6", font=BLACK_FONT, fmt=FMT_YEN)
    # Tax
    set_cell(ws3, 32, col, f"={cl}31*C8", font=BLACK_FONT, fmt=FMT_YEN)
    # NOPAT
    set_cell(ws3, 33, col, f"={cl}31-{cl}32", font=BLACK_FONT, fmt=FMT_YEN)
    # D&A
    set_cell(ws3, 34, col, f"={cl}30*C20", font=BLACK_FONT, fmt=FMT_YEN)
    # Capex
    set_cell(ws3, 35, col, f"={cl}30*C7", font=BLACK_FONT, fmt=FMT_YEN)
    # UFCF
    set_cell(ws3, 36, col, f"={cl}33+{cl}34-{cl}35", font=BLACK_FONT, fmt=FMT_YEN)
    # Discount Factor
    set_cell(ws3, 37, col, f"=1/(1+C26)^{yr+1}", font=BLACK_FONT, fmt="0.0000")
    # PV of FCF
    set_cell(ws3, 38, col, f"={cl}36*{cl}37", font=BLACK_FONT, fmt=FMT_YEN)

# ── Valuation - Perpetuity Growth Method ──
c = section_title(ws3, 40, 2, "Valuation - Perpetuity Growth Method")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=40, column=col_idx).fill = LIGHT_GREEN

last_cl = col_letter(3 + proj_years - 1)  # G for 5 years

set_cell(ws3, 41, 2, "Sum of PV of FCFs", font=BOLD_FONT)
set_cell(ws3, 41, 3, f"=SUM(C38:{last_cl}38)", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 42, 2, "Terminal Value (PGM)", font=BOLD_FONT)
set_cell(ws3, 42, 3, f"={last_cl}36*(1+C15)/(C26-C15)", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 43, 2, "PV of Terminal Value", font=BOLD_FONT)
set_cell(ws3, 43, 3, f"=C42*{last_cl}37", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 44, 2, "Enterprise Value", font=BOLD_FONT)
set_cell(ws3, 44, 3, "=C41+C43", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 45, 2, "Equity Value", font=BOLD_FONT)
set_cell(ws3, 45, 3, "=C44-C18", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 46, 2, "Implied Share Price (PGM)", font=BOLD_FONT)
set_cell(ws3, 46, 3, "=ROUND(C45*1000000/C17,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# ── Valuation - Exit Multiple Method ──
c = section_title(ws3, 48, 2, "Valuation - Exit Multiple Method")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=48, column=col_idx).fill = LIGHT_GREEN

set_cell(ws3, 49, 2, "Sum of PV of FCFs", font=BOLD_FONT)
set_cell(ws3, 49, 3, "=C41", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 50, 2, "Year 5 EBITDA", font=BOLD_FONT)
set_cell(ws3, 50, 3, f"={last_cl}30*(C6+C20)", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 51, 2, "Terminal Value (Exit Multiple)", font=BOLD_FONT)
set_cell(ws3, 51, 3, "=C50*C16", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 52, 2, "PV of Terminal Value", font=BOLD_FONT)
set_cell(ws3, 52, 3, f"=C51*{last_cl}37", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 53, 2, "Enterprise Value", font=BOLD_FONT)
set_cell(ws3, 53, 3, "=C49+C52", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 54, 2, "Equity Value", font=BOLD_FONT)
set_cell(ws3, 54, 3, "=C53-C18", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, 55, 2, "Implied Share Price (Exit Multiple)", font=BOLD_FONT)
set_cell(ws3, 55, 3, "=ROUND(C54*1000000/C17,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# =====================================================================
# SHEET 4: Comps Analysis
# =====================================================================
ws4 = wb.create_sheet("Comps Analysis")
ws4.sheet_properties.tabColor = "006600"

ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 10
for letter in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
    ws4.column_dimensions[letter].width = 12

set_cell(ws4, 2, 2, "Comparable Company Analysis", font=TITLE_FONT)

# Header row
comp_headers = [
    "Company", "Ticker", "Mkt Cap\n(JPY mn)", "EV\n(JPY mn)",
    "Revenue\n(JPY mn)", "EBITDA\n(JPY mn)", "Op Income\n(JPY mn)",
    "Net Income\n(JPY mn)", "EV/EBITDA", "EV/Revenue", "PER",
    "PBR", "Op Margin", "ROE"
]
header_row(ws4, 4, 2, 15, comp_headers)

# Company data rows (rows 5-12)
comps = C["comps"]
for i, comp in enumerate(comps):
    r = 5 + i
    cl_d = col_letter(4)  # D = Mkt Cap
    cl_e = col_letter(5)  # E = EV
    cl_f = col_letter(6)  # F = Revenue
    cl_g = col_letter(7)  # G = EBITDA
    cl_h = col_letter(8)  # H = Op Income
    cl_i = col_letter(9)  # I = Net Income

    set_cell(ws4, r, 2, comp["name"], font=BOLD_FONT)
    set_cell(ws4, r, 3, comp["ticker"])
    set_cell(ws4, r, 4, comp["mkt_cap"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 5, comp["ev"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 6, comp["revenue"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 7, comp["ebitda"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 8, comp["op_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 9, comp["net_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # Calculated multiples (Excel formulas)
    set_cell(ws4, r, 10, f"=E{r}/G{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)  # EV/EBITDA
    set_cell(ws4, r, 11, f"=E{r}/F{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)  # EV/Revenue
    set_cell(ws4, r, 12, f"=D{r}/I{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)  # PER
    set_cell(ws4, r, 13, comp["pbr"], font=BLUE_FONT, fmt=FMT_RATIO, border=THIN_BORDER)      # PBR
    set_cell(ws4, r, 14, f"=H{r}/F{r}", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)     # Op Margin
    set_cell(ws4, r, 15, comp["roe"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)         # ROE

last_comp_row = 5 + len(comps) - 1  # row 12

# ── Statistics ──
section_title(ws4, 14, 2, "Statistics")

stat_labels = ["25th Percentile", "Median (50th)", "75th Percentile"]
stat_rows = [15, 16, 17]

# Statistics columns map: D->J, E->K, F->L, G->M, H->N, I->O
stat_col_map = [
    (4, 10),  # D15 -> EV/EBITDA stats from J column
    (5, 11),  # E15 -> EV/Revenue stats from K column
    (6, 12),  # F15 -> PER stats from L column
    (7, 13),  # G15 -> PBR stats from M column
    (8, 14),  # H15 -> Op Margin stats from N column
    (9, 15),  # I15 -> ROE stats from O column
]

for stat_idx, (label, r) in enumerate(zip(stat_labels, stat_rows)):
    set_cell(ws4, r, 2, label, font=BOLD_FONT)

    for dst_col, src_col in stat_col_map:
        src_letter = col_letter(src_col)
        rng = f"{src_letter}5:{src_letter}{last_comp_row}"

        # Determine format based on source column
        if src_col in (14, 15):  # Op Margin, ROE
            fmt = FMT_PCT
        else:  # multiples
            fmt = FMT_RATIO

        if stat_idx == 0:  # 25th percentile
            formula = f"=PERCENTILE({rng},0.25)"
        elif stat_idx == 1:  # Median
            formula = f"=MEDIAN({rng})"
        else:  # 75th percentile
            formula = f"=PERCENTILE({rng},0.75)"

        set_cell(ws4, r, dst_col, formula, font=BLACK_FONT, fmt=fmt, border=THIN_BORDER)

# ── Implied Valuation ──
c = section_title(ws4, 19, 2, f'Implied Valuation for {C["company_name"]}')
c.fill = LIGHT_GREEN
for col_idx in range(3, 10):
    ws4.cell(row=19, column=col_idx).fill = LIGHT_GREEN

section_title(ws4, 20, 2, "Core Corp Financials")
set_cell(ws4, 21, 2, "EBITDA (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 21, 3, C["core_ebitda"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws4, 22, 2, "Net Income (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 22, 3, C["core_net_income"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws4, 23, 2, "Shares Outstanding", font=BOLD_FONT)
set_cell(ws4, 23, 3, C["shares_outstanding"], font=BLUE_FONT, fmt=FMT_INT)
set_cell(ws4, 24, 2, "Net Debt (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 24, 3, C["net_debt"], font=BLUE_FONT, fmt=FMT_YEN)

section_title(ws4, 26, 2, "Implied Share Price (Median Multiples)")

# Via EV/EBITDA: EV = EBITDA * Median, Equity = EV - Net Debt, Price = Equity * 1M / Shares
# D16 has the Median EV/EBITDA
set_cell(ws4, 27, 2, "Via EV/EBITDA (Median)", font=BOLD_FONT)
set_cell(ws4, 27, 3, "=ROUND((C21*D16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# Via PER: Market Cap = NI * Median PER, Price = Market Cap * 1M / Shares
# F16 has the Median PER
set_cell(ws4, 28, 2, "Via PER (Median)", font=BOLD_FONT)
set_cell(ws4, 28, 3, "=ROUND(C22*F16*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# =====================================================================
# SHEET 5: Sensitivity Analysis
# =====================================================================
ws5 = wb.create_sheet("Sensitivity Analysis")
ws5.sheet_properties.tabColor = "996600"

ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 24
for letter in ["C", "D", "E", "F", "G", "H"]:
    ws5.column_dimensions[letter].width = 14

set_cell(ws5, 2, 2, "Sensitivity Analysis", font=TITLE_FONT)

# ── Table 1: WACC vs Terminal Growth Rate ──
section_title(ws5, 4, 2, "Table 1: WACC vs Terminal Growth Rate (Implied Share Price, JPY)")

wacc_range = C["sens_wacc_range"]
tg_range = C["sens_tg_range"]

# Header row
set_cell(ws5, 5, 2, "WACC \\ Terminal g", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j, tg in enumerate(tg_range):
    set_cell(ws5, 5, 3 + j, tg, font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

# Data rows
base_wacc = calc_wacc(C)
for i, wacc in enumerate(wacc_range):
    r = 6 + i
    set_cell(ws5, r, 2, wacc, font=BOLD_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    for j, tg in enumerate(tg_range):
        price = calc_dcf_pgm(C["revenue_growth"], C["op_margin"], wacc, tg, C)
        # Highlight base-case cell
        is_base = (abs(wacc - base_wacc) < 0.002 and abs(tg - C["terminal_growth"]) < 0.002)
        fill = LIGHT_YELLOW if is_base else None
        set_cell(ws5, r, 3 + j, price, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER,
                 fill=fill)

# ── Table 2: Revenue Growth vs Operating Margin ──
table2_start = 6 + len(wacc_range) + 2  # 2 rows gap
section_title(ws5, table2_start, 2,
              "Table 2: Revenue Growth vs Operating Margin (Implied Share Price, JPY)")

rev_growth_range = C["sens_rev_growth_range"]
op_margin_range = C["sens_op_margin_range"]

# Header row
hdr_row = table2_start + 1
set_cell(ws5, hdr_row, 2, "Rev Growth \\ Op Margin", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j, om in enumerate(op_margin_range):
    set_cell(ws5, hdr_row, 3 + j, om, font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

# Data rows
for i, rg in enumerate(rev_growth_range):
    r = hdr_row + 1 + i
    set_cell(ws5, r, 2, rg, font=BOLD_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    for j, om in enumerate(op_margin_range):
        price = calc_dcf_pgm(rg, om, base_wacc, C["terminal_growth"], C)
        is_base = (abs(rg - C["revenue_growth"]) < 0.002 and abs(om - C["op_margin"]) < 0.002)
        fill = LIGHT_YELLOW if is_base else None
        set_cell(ws5, r, 3 + j, price, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER,
                 fill=fill)

# =====================================================================
# SAVE & VERIFY
# =====================================================================
wb.save(output_file)
print(f"Saved: {output_file}")

# Run recalc.py for verification
recalc_script = os.path.join("scripts", "recalc.py")
if os.path.exists(recalc_script):
    print(f"\nRunning verification: python {recalc_script} {output_file}")
    result = subprocess.run([sys.executable, recalc_script, output_file],
                            capture_output=True, text=True)
    print(result.stdout)
    if result.stderr:
        print("STDERR:", result.stderr)
else:
    print(f"\nNote: {recalc_script} not found, skipping verification.")
