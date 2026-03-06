"""
patch_sensitivity.py - Rebuild the Sensitivity Analysis sheet with dynamic Excel formulas.

Opens the existing Excel file, preserves all other sheets (including hand-edited DCF Model),
replaces only the Sensitivity Analysis sheet with formula-driven sensitivity tables.

Headers are dynamic: centered on the current WACC / Terminal g / Exit Multiple
from the DCF Model sheet, spreading ± steps in each direction.

Tables:
  Table 1: WACC vs Terminal Growth Rate (PGM method)
  Table 2: WACC vs Exit Multiple (Exit Multiple method)
"""

import shutil
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURATION
# =====================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TARGET_FILE = os.path.join(SCRIPT_DIR, "Core_Corporation_2359T_Equity_Research.xlsx")
BACKUP_FILE = os.path.join(SCRIPT_DIR, "Core_Corporation_2359T_Equity_Research_backup.xlsx")

# Grid size: 7 rows x 7 columns, center at index 3 (4th position, 0-based)
N_GRID = 7
CENTER_IDX = 3  # 0-based; 4th from top/left

# Step sizes
WACC_STEP = 0.005       # 0.5%
TG_STEP   = 0.0025      # 0.25%
EXIT_STEP = 1.0         # 1.0x

# DCF Model sheet row constants (must match the generated file)
R_EBIT = 38
R_DA   = 41
R_UFCF = 45
PROJ_YEARS = 5
LAST_PROJ_COL = get_column_letter(3 + PROJ_YEARS - 1)  # "G"

# DCF Model fixed cell references
DCF = "'DCF Model'"
SHARES   = f"{DCF}!C15"
NET_DEBT = f"{DCF}!C16"

# UFCF cells for years 1-5
UFCF = [f"{DCF}!{get_column_letter(3 + yr)}{R_UFCF}" for yr in range(PROJ_YEARS)]

# =====================================================================
# STYLE CONSTANTS
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="0000CC")
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
NOTE_FONT   = Font(name="Arial", size=9, italic=True, color="808080")

HEADER_FILL  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
INPUT_FILL   = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

FMT_YEN   = '#,##0;(#,##0)'
FMT_PCT   = '0.0%;(0.0%)'
FMT_PCT2  = '0.00%;(0.00%)'
FMT_RATIO = '0.00"x"'

# =====================================================================
# Sensitivity sheet layout constants
# =====================================================================
# Row 3: current value anchors
# C3 = Current WACC (='DCF Model'!C24)
# F3 = Current Terminal g (='DCF Model'!C13)
# I3 = Current Exit Multiple (='DCF Model'!C14)
ANCHOR_WACC = "$C$3"
ANCHOR_TG   = "$F$3"
ANCHOR_EXIT = "$I$3"

T1_TITLE_ROW = 5
T1_HDR_ROW   = 6
T1_DATA_START = 7

# Table 2 starts after Table 1's N_GRID data rows + 2 gap rows
T2_TITLE_ROW  = T1_DATA_START + N_GRID + 2   # 16
T2_HDR_ROW    = T2_TITLE_ROW + 1              # 17
T2_DATA_START = T2_HDR_ROW + 1                # 18

# =====================================================================
# HELPERS
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c

def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c

def offset_formula(anchor, offset_val):
    """Build '=anchor + offset' or '=anchor - offset' formula."""
    if offset_val == 0:
        return f"={anchor}"
    elif offset_val > 0:
        return f"={anchor}+{offset_val}"
    else:
        return f"={anchor}-{abs(offset_val)}"

def build_pgm_formula(wacc_ref, tg_ref):
    """Build PGM implied share price formula referencing DCF Model sheet."""
    pv_parts = [f"{UFCF[yr]}/(1+{wacc_ref})^{yr+1}" for yr in range(PROJ_YEARS)]
    sum_pv = "+".join(pv_parts)
    last_ufcf = UFCF[PROJ_YEARS - 1]
    pv_tv = f"{last_ufcf}*(1+{tg_ref})/({wacc_ref}-{tg_ref})/(1+{wacc_ref})^{PROJ_YEARS}"
    return f"=IFERROR(ROUND(({sum_pv}+{pv_tv}-{NET_DEBT})*1000000/{SHARES},0),\"\")"

def build_exit_formula(wacc_ref, mult_ref):
    """Build Exit Multiple implied share price formula referencing DCF Model sheet."""
    pv_parts = [f"{UFCF[yr]}/(1+{wacc_ref})^{yr+1}" for yr in range(PROJ_YEARS)]
    sum_pv = "+".join(pv_parts)
    yr5_ebitda = f"({DCF}!{LAST_PROJ_COL}{R_EBIT}+{DCF}!{LAST_PROJ_COL}{R_DA})"
    pv_tv = f"{yr5_ebitda}*{mult_ref}/(1+{wacc_ref})^{PROJ_YEARS}"
    return f"=IFERROR(ROUND(({sum_pv}+{pv_tv}-{NET_DEBT})*1000000/{SHARES},0),\"\")"

# =====================================================================
# MAIN
# =====================================================================
def main():
    if not os.path.exists(TARGET_FILE):
        print(f"ERROR: {TARGET_FILE} not found.")
        return

    # 1. Backup
    shutil.copy2(TARGET_FILE, BACKUP_FILE)
    print(f"Backup created: {BACKUP_FILE}")

    # 2. Load existing workbook (NEVER Workbook() - preserves DCF Model data)
    wb = openpyxl.load_workbook(TARGET_FILE)
    print(f"Loaded: {TARGET_FILE}")
    print(f"Sheets: {wb.sheetnames}")

    # ── Sanity check: confirm DCF Model hand-edited WACC is intact ──
    dcf_ws = wb["DCF Model"]
    wacc_formula = dcf_ws["C24"].value
    print(f"DCF Model C24 (WACC formula): {wacc_formula}  [will NOT be touched]")

    # 3. Remove old Sensitivity Analysis sheet (preserve position)
    sheet_name = "Sensitivity Analysis"
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        wb.remove(wb[sheet_name])
        print(f"Removed old '{sheet_name}' sheet (was at index {idx}).")
    else:
        idx = len(wb.sheetnames)

    # 4. Create new Sensitivity Analysis sheet at same position
    ws = wb.create_sheet(sheet_name, index=idx)
    ws.sheet_properties.tabColor = "996600"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    for letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[letter].width = 14

    # ── Title ──
    set_cell(ws, 2, 2, "Sensitivity Analysis", font=TITLE_FONT)

    # ── Row 3: Anchor cells (current values from DCF Model) ──
    set_cell(ws, 3, 2, "Current WACC:", font=BOLD_FONT)
    set_cell(ws, 3, 3, f"={DCF}!C24", font=BLACK_FONT, fmt=FMT_PCT2)     # C3 = WACC anchor
    set_cell(ws, 3, 5, "Terminal g:", font=BOLD_FONT)
    set_cell(ws, 3, 6, f"={DCF}!C13", font=BLACK_FONT, fmt=FMT_PCT2)     # F3 = TG anchor
    set_cell(ws, 3, 8, "Exit Multiple:", font=BOLD_FONT)
    set_cell(ws, 3, 9, f"={DCF}!C14", font=BLACK_FONT, fmt=FMT_RATIO)    # I3 = Exit anchor

    # =================================================================
    # TABLE 1: WACC vs Terminal Growth Rate (PGM)
    # =================================================================
    section_title(ws, T1_TITLE_ROW, 2,
                  "Table 1: WACC vs Terminal Growth Rate (PGM - Implied Share Price, JPY)")

    # ── Column headers: Terminal g (dynamic, centered on F3) ──
    set_cell(ws, T1_HDR_ROW, 2, "WACC \\ Terminal g",
             font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center", wrap_text=True),
             border=THIN_BORDER)
    for j in range(N_GRID):
        col = 3 + j
        offset = (j - CENTER_IDX) * TG_STEP
        formula = offset_formula(ANCHOR_TG, round(offset, 6))
        set_cell(ws, T1_HDR_ROW, col, formula,
                 font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
                 alignment=Alignment(horizontal="center"),
                 border=THIN_BORDER)

    # ── Row headers: WACC (dynamic, centered on C3) + data formulas ──
    for i in range(N_GRID):
        r = T1_DATA_START + i
        offset = (i - CENTER_IDX) * WACC_STEP
        wacc_hdr_formula = offset_formula(ANCHOR_WACC, round(offset, 6))
        set_cell(ws, r, 2, wacc_hdr_formula,
                 font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER, fill=INPUT_FILL)

        for j in range(N_GRID):
            col = 3 + j
            cl = get_column_letter(col)
            wacc_ref = f"$B{r}"
            tg_ref = f"{cl}${T1_HDR_ROW}"
            formula = build_pgm_formula(wacc_ref, tg_ref)
            set_cell(ws, r, col, formula,
                     font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # =================================================================
    # TABLE 2: WACC vs Exit Multiple
    # =================================================================
    section_title(ws, T2_TITLE_ROW, 2,
                  "Table 2: WACC vs Exit Multiple (Exit Multiple - Implied Share Price, JPY)")

    # ── Column headers: Exit Multiple (dynamic, centered on I3) ──
    set_cell(ws, T2_HDR_ROW, 2, "WACC \\ Exit Multiple",
             font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center", wrap_text=True),
             border=THIN_BORDER)
    for j in range(N_GRID):
        col = 3 + j
        offset = (j - CENTER_IDX) * EXIT_STEP
        formula = offset_formula(ANCHOR_EXIT, round(offset, 6))
        set_cell(ws, T2_HDR_ROW, col, formula,
                 font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_RATIO,
                 alignment=Alignment(horizontal="center"),
                 border=THIN_BORDER)

    # ── Row headers: WACC (dynamic, same logic) + data formulas ──
    for i in range(N_GRID):
        r = T2_DATA_START + i
        offset = (i - CENTER_IDX) * WACC_STEP
        wacc_hdr_formula = offset_formula(ANCHOR_WACC, round(offset, 6))
        set_cell(ws, r, 2, wacc_hdr_formula,
                 font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER, fill=INPUT_FILL)

        for j in range(N_GRID):
            col = 3 + j
            cl = get_column_letter(col)
            wacc_ref = f"$B{r}"
            mult_ref = f"{cl}${T2_HDR_ROW}"
            formula = build_exit_formula(wacc_ref, mult_ref)
            set_cell(ws, r, col, formula,
                     font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # ── Note ──
    note_row = T2_DATA_START + N_GRID + 1
    set_cell(ws, note_row, 2,
             "All values dynamically linked to DCF Model. "
             "Headers auto-center on current WACC / Terminal g / Exit Multiple.",
             font=NOTE_FONT)
    ws.merge_cells(start_row=note_row, start_column=2,
                   end_row=note_row, end_column=9)

    # 5. Confirm DCF Model was NOT touched
    wacc_after = dcf_ws["C24"].value
    assert wacc_after == wacc_formula, "FATAL: DCF Model C24 was modified!"
    print(f"DCF Model C24 after patch: {wacc_after}  [CONFIRMED UNCHANGED]")

    # 6. Save
    wb.save(TARGET_FILE)
    print(f"\nSaved: {TARGET_FILE}")
    print("Sensitivity Analysis rebuilt with dynamic headers + formulas.")
    print("DCF Model sheet data preserved (untouched).")


if __name__ == "__main__":
    main()
