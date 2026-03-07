"""
generate_core_v4.py - Generate Core Corporation (2359.T) DCF/Comps model
with dedicated NWC Schedule (DSO/DIH/DPO) + Scenario Matrix.

Outputs: Core_Corporation_2359T_Equity_Research.xlsx
"""

import sys
import os

# ── Setup paths ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
sys.path.insert(0, os.path.join(PROJECT_ROOT, "scripts"))

from comps_fetcher import get_comps_data

# ── Core Corporation Config ──
config_core = {
    # ── Company Info ──
    "company_name": "Core Corporation",
    "ticker": "2359.T",
    "exchange": "TSE Prime",
    "sector": "Information & Communication",

    # ── Historical Financials (JPY mn) ──
    "hist_years": ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024E"],
    "hist_revenue":          [20904, 21919, 22534, 23554, 24500],
    "hist_cogs":             [16570, 17290, 17578, 18378, 19110],
    "hist_sga":              [2525, 2649, 2816, 2981, 3100],
    "hist_operating_income": [1808, 1979, 2139, 2194, 2290],
    "hist_net_income":       [1186, 1284, 1515, 1488, 1550],
    "hist_ocf":              [1323, 2118, 1172, 2302, 2000],
    "hist_capex":            [151, 144, 202, 107, 150],
    "hist_cash":             [7703, 9089, 9370, 10260, 10500],
    "hist_debt":             [38, 30, 8, 3, 0],
    "net_debt":              -10257,
    "base_year_revenue":     24500,
    "base_year_cogs":        19110,

    # ── NWC Base Year Actuals (JPY mn) ──
    "base_year_ar":   7719,   # 24500 * 115 / 365
    "base_year_inv":  314,    # 19110 * 6 / 365
    "base_year_ap":   1832,   # 19110 * 35 / 365

    # ── DCF Assumptions (Scenarios) ──
    "scenarios": {
        "Base": {
            "revenue_growth": [0.04, 0.04, 0.04, 0.04, 0.04],
            "cogs_pct": [0.78, 0.78, 0.78, 0.78, 0.78],
            "sga_pct": [0.13, 0.13, 0.13, 0.13, 0.13],
            "dso_days": [115, 115, 115, 115, 115],
            "dih_days": [6, 6, 6, 6, 6],
            "dpo_days": [35, 35, 35, 35, 35],
        },
        "Upside": {
            "revenue_growth": [0.06, 0.06, 0.06, 0.06, 0.06],
            "cogs_pct": [0.76, 0.76, 0.76, 0.76, 0.76],
            "sga_pct": [0.12, 0.12, 0.12, 0.12, 0.12],
            "dso_days": [110, 110, 110, 110, 110],
            "dih_days": [5, 5, 5, 5, 5],
            "dpo_days": [37, 37, 37, 37, 37],
        },
        "Management": {
            "revenue_growth": [0.05, 0.05, 0.05, 0.05, 0.05],
            "cogs_pct": [0.77, 0.77, 0.77, 0.77, 0.77],
            "sga_pct": [0.13, 0.13, 0.13, 0.13, 0.13],
            "dso_days": [115, 115, 115, 115, 115],
            "dih_days": [6, 6, 6, 6, 6],
            "dpo_days": [35, 35, 35, 35, 35],
        },
        "Downside 1": {
            "revenue_growth": [0.02, 0.02, 0.02, 0.02, 0.02],
            "cogs_pct": [0.80, 0.80, 0.80, 0.80, 0.80],
            "sga_pct": [0.14, 0.14, 0.14, 0.14, 0.14],
            "dso_days": [120, 120, 120, 120, 120],
            "dih_days": [7, 7, 7, 7, 7],
            "dpo_days": [33, 33, 33, 33, 33],
        },
        "Downside 2": {
            "revenue_growth": [0.00, 0.00, 0.00, 0.00, 0.00],
            "cogs_pct": [0.82, 0.82, 0.82, 0.82, 0.82],
            "sga_pct": [0.15, 0.15, 0.15, 0.15, 0.15],
            "dso_days": [125, 125, 125, 125, 125],
            "dih_days": [8, 8, 8, 8, 8],
            "dpo_days": [30, 30, 30, 30, 30],
        },
    },

    "capex_pct": 0.01,
    "da_pct": 0.01,
    "tax_rate": 0.306,
    "risk_free": 0.015,
    "beta": 0.85,
    "erp": 0.060,
    "size_premium": 0.020,
    "cost_of_debt_at": 0.01,
    "de_ratio": 0.0,
    "terminal_growth": 0.005,
    "exit_multiple": 10.0,
    "projection_years": 5,

    # ── Core Financials for Comps implied valuation ──
    "core_ebitda": 2290 + 245,
    "core_net_income": 1550,

    # ── Comps Data ──
    "comps_csv_path": os.path.join(SCRIPT_DIR, "comps_input_core_CSV.csv"),
    "primary_multiple": "EV/EBITDA",

    # ── Investment Thesis & Risks ──
    "investment_thesis": [
        "1. Stable IT services business with consistent revenue growth",
        "2. Strong net cash position (JPY 10.5bn) providing downside protection",
        "3. Expanding margins through operational efficiency improvements",
    ],
    "key_risks": [
        "1. Mature domestic IT services market with limited upside surprise",
        "2. Labor shortage in Japan's IT sector could pressure margins",
        "3. Low growth profile limits re-rating potential",
    ],

}

# ── Load comps from CSV ──
print("Loading comps data...")
config_core["comps"] = get_comps_data(config_core["comps_csv_path"])

# ── Restore flat arrays from Base scenario (for sensitivity Python calc fallback) ──
_base = config_core["scenarios"]["Base"]
config_core["revenue_growth"] = _base["revenue_growth"]
config_core["cogs_pct"]       = _base["cogs_pct"]
config_core["sga_pct"]        = _base["sga_pct"]

# ── Fetch live market data ──
try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False

def get_live_market_data(ticker_str, fallback_price=1000, fallback_shares=10_000_000):
    if not YFINANCE_AVAILABLE:
        print("yfinance not installed. Using fallback market data.")
        return fallback_price, fallback_shares
    try:
        print(f"Fetching live data for {ticker_str} via yfinance...")
        tkr = yf.Ticker(ticker_str)
        info = tkr.info
        live_price = info.get("currentPrice") or info.get("regularMarketPrice") or fallback_price
        live_shares = info.get("sharesOutstanding") or fallback_shares
        print(f"Successfully fetched: Price={live_price}, Shares={live_shares}")
        return float(live_price), int(live_shares)
    except Exception as e:
        print(f"Warning: Failed to fetch live data ({e}). Using fallback market data.")
        return fallback_price, fallback_shares

config_core["current_price"], config_core["shares_outstanding"] = get_live_market_data(
    config_core["ticker"]
)

# =====================================================================
# BUILD WORKBOOK
# =====================================================================

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Assign config shorthand ──
config = config_core
C = config

USE_EV_SALES = (config.get("primary_multiple", "EV/EBITDA") == "EV/Sales")

# =====================================================================
# DCF MODEL ROW NUMBERS (NWC % rows removed, 2 rows saved)
# =====================================================================
R_DRV_GROWTH   = 28
R_DRV_COGS     = 29
R_DRV_SGA      = 30
R_REVENUE      = 31
R_COGS         = 32
R_GROSS_PROFIT = 33
R_GROSS_MARGIN = 34
R_SGA          = 35
R_OP_M_IMPL   = 36
R_EBIT         = 37
R_TAX          = 38
R_NOPAT        = 39
R_DA           = 40
R_CAPEX        = 41
R_CHG_NWC      = 42
R_UFCF         = 43
R_DISC         = 44
R_PV_FCF       = 45

R_PGM_SEC    = R_PV_FCF + 2
R_SUM_PV     = R_PGM_SEC + 1
R_TV_PGM     = R_SUM_PV + 1
R_PV_TV_PGM  = R_TV_PGM + 1
R_EV_PGM     = R_PV_TV_PGM + 1
R_EQ_PGM     = R_EV_PGM + 1
R_PRICE_PGM  = R_EQ_PGM + 1

R_EXIT_SEC   = R_PRICE_PGM + 2
R_SUM_PV_EX  = R_EXIT_SEC + 1
R_YR5_EBITDA = R_SUM_PV_EX + 1
R_TV_EXIT    = R_YR5_EBITDA + 1
R_PV_TV_EXIT = R_TV_EXIT + 1
R_EV_EXIT    = R_PV_TV_EXIT + 1
R_EQ_EXIT    = R_EV_EXIT + 1
R_PRICE_EXIT = R_EQ_EXIT + 1

SCENARIO_NAMES = ["Base", "Upside", "Management", "Downside 1", "Downside 2"]
NUM_SCENARIOS  = 5

R_SCEN_SEC        = R_PRICE_EXIT + 2
R_SCEN_YEARS      = R_SCEN_SEC + 1
R_SCEN_BLK_GROWTH = R_SCEN_YEARS + 1
R_SCEN_BLK_COGS   = R_SCEN_BLK_GROWTH + 7
R_SCEN_BLK_SGA    = R_SCEN_BLK_COGS + 7

# =====================================================================
# NWC SCHEDULE ROW NUMBERS
# =====================================================================
NWC_R_DSO      = 5
NWC_R_DIH      = 6
NWC_R_DPO      = 7
NWC_R_REV      = 9
NWC_R_COGS     = 10
NWC_R_AR       = 12
NWC_R_INV      = 13
NWC_R_CA       = 14
NWC_R_AP       = 15
NWC_R_CL       = 16
NWC_R_NWC      = 18
NWC_R_CHG_NWC  = 19

NWC_R_SCEN_SEC      = 22
NWC_R_SCEN_YEARS    = 23
NWC_R_SCEN_BLK_DSO  = 24
NWC_R_SCEN_BLK_DIH  = NWC_R_SCEN_BLK_DSO + 7
NWC_R_SCEN_BLK_DPO  = NWC_R_SCEN_BLK_DIH + 7

# =====================================================================
# STYLE CONSTANTS
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="0000CC", bold=False)
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
GREEN_FONT  = Font(name="Arial", size=10, color="006600")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
GREY_FONT   = Font(name="Arial", size=9, italic=True, color="808080")

HEADER_FILL  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
LIGHT_FILL   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL   = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER   = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))
BOTTOM_DOUBLE = Border(bottom=Side(style="double"))
TOP_BOTTOM    = Border(top=Side(style="thin"), bottom=Side(style="double"))

FMT_YEN     = '#,##0;(#,##0)'
FMT_YEN_DEC = '#,##0.0;(#,##0.0)'
FMT_PCT     = '0.0%;(0.0%)'
FMT_PCT2    = '0.00%;(0.00%)'
FMT_RATIO   = '0.00"x"'
FMT_INT     = '#,##0'
FMT_EPS     = '#,##0.0;(#,##0.0)'
FMT_DAYS    = '#,##0'

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c

def header_row(ws, row, col_start, col_end, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c

def col_letter(col_num):
    return get_column_letter(col_num)

def choose_formula(block_start, cl):
    refs = [f"{cl}{block_start + 1 + s}" for s in range(NUM_SCENARIOS)]
    return f"=CHOOSE($D$25,{','.join(refs)})"

def nwc_choose_formula(block_start, cl):
    refs = [f"{cl}{block_start + 1 + s}" for s in range(NUM_SCENARIOS)]
    return f"=CHOOSE('DCF Model'!$D$25,{','.join(refs)})"

# =====================================================================
# SENSITIVITY ANALYSIS HELPERS (unused Python calc, kept for reference)
# =====================================================================
def calc_dcf_pgm(rev_growth, gross_margin, wacc, tg, cfg):
    n = cfg["projection_years"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]
    sga_pct_list = cfg["sga_pct"]

    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_fcf = 0
    for yr_idx, rev in enumerate(revenues):
        cogs = rev * (1 - gross_margin)
        gp = rev - cogs
        sga = rev * sga_pct_list[yr_idx]
        ebit = gp - sga
        nopat = ebit * (1 - tax) if ebit > 0 else ebit
        da = rev * da_pct
        capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_fcf = fcf

    tv = last_fcf * (1 + tg) / (wacc - tg)
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

def calc_dcf_exit(rev_growth, gross_margin, wacc, exit_mult, cfg):
    n = cfg["projection_years"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]
    sga_pct_list = cfg["sga_pct"]

    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_ebit = 0
    for yr_idx, rev in enumerate(revenues):
        cogs = rev * (1 - gross_margin)
        gp = rev - cogs
        sga = rev * sga_pct_list[yr_idx]
        ebit = gp - sga
        nopat = ebit * (1 - tax) if ebit > 0 else ebit
        da = rev * da_pct
        capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_ebit = ebit

    yr5_ebitda = last_ebit + revenues[-1] * da_pct
    tv = yr5_ebitda * exit_mult
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

def calc_wacc(cfg):
    ke = cfg["risk_free"] + cfg["beta"] * cfg["erp"] + cfg["size_premium"]
    we = 1 / (1 + cfg["de_ratio"])
    wd = cfg["de_ratio"] / (1 + cfg["de_ratio"])
    return ke * we + cfg["cost_of_debt_at"] * wd

# =====================================================================
# BUILD WORKBOOK
# =====================================================================
wb = openpyxl.Workbook()

output_file = os.path.join(SCRIPT_DIR, "Core_Corporation_2359T_Equity_Research.xlsx")

# =====================================================================
# SHEET 1: Executive Summary
# =====================================================================
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "003366"

ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 22

set_cell(ws1, 1, 2,
    "DISCLAIMER: This is a sample analysis for demonstration purposes only. "
    "It does not constitute investment advice.",
    font=GREY_FONT)
ws1.merge_cells("B1:E1")

set_cell(ws1, 3, 2, f'{C["company_name"]} ({C["ticker"]})', font=TITLE_FONT)
ws1.merge_cells("B3:D3")
set_cell(ws1, 4, 2, "Equity Research Report", font=SUB_FONT)

set_cell(ws1, 6, 2, "Company", font=BOLD_FONT)
set_cell(ws1, 6, 3, C["company_name"])
set_cell(ws1, 7, 2, "Ticker", font=BOLD_FONT)
set_cell(ws1, 7, 3, f'{C["ticker"]} ({C["exchange"]})')
set_cell(ws1, 8, 2, "Sector", font=BOLD_FONT)
set_cell(ws1, 8, 3, C["sector"])
set_cell(ws1, 9, 2, "Current Price", font=BOLD_FONT)
set_cell(ws1, 9, 3, C["current_price"], font=BLUE_FONT, fmt=FMT_YEN)

set_cell(ws1, 10, 2, "Target Price (Mid)", font=BOLD_FONT)
set_cell(ws1, 10, 3, "=ROUND(AVERAGE(C16:C19),0)", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws1, 11, 2, "Recommendation", font=BOLD_FONT)
set_cell(ws1, 11, 3, '=IF(C12>0.15,"BUY",IF(C12>0.05,"HOLD","SELL"))', font=BLACK_FONT)

set_cell(ws1, 12, 2, "Upside / Downside", font=BOLD_FONT)
set_cell(ws1, 12, 3, "=(C10-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

c = set_cell(ws1, 14, 2, "Valuation Summary", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=14, column=col_idx).fill = LIGHT_FILL

header_row(ws1, 15, 2, 4, ["Methodology", "Implied Value (JPY)", "vs Current Price"])

set_cell(ws1, 16, 2, "DCF - Perpetuity Growth")
set_cell(ws1, 16, 3, f"='DCF Model'!C{R_PRICE_PGM}", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 16, 4, "=(C16-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

set_cell(ws1, 17, 2, "DCF - Exit Multiple")
set_cell(ws1, 17, 3, f"='DCF Model'!C{R_PRICE_EXIT}", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 17, 4, "=(C17-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

if USE_EV_SALES:
    set_cell(ws1, 18, 2, "Comps - EV/Sales Median")
else:
    set_cell(ws1, 18, 2, "Comps - EV/EBITDA Median")
set_cell(ws1, 18, 3, "='Comps Analysis'!C27", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 18, 4, "=(C18-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

set_cell(ws1, 19, 2, "Comps - PER Median")
set_cell(ws1, 19, 3, "='Comps Analysis'!C28", font=GREEN_FONT, fmt=FMT_YEN)
set_cell(ws1, 19, 4, "=(C19-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

set_cell(ws1, 21, 2, "Integrated Valuation Range", font=BOLD_FONT)
set_cell(ws1, 21, 3, '=MIN(C16:C19)&" - "&MAX(C16:C19)', font=BLACK_FONT)

c = set_cell(ws1, 23, 2, "Key Investment Thesis", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=23, column=col_idx).fill = LIGHT_FILL
for i, line in enumerate(C["investment_thesis"]):
    set_cell(ws1, 24 + i, 2, line)

c = set_cell(ws1, 28, 2, "Key Risks", font=SUB_FONT)
c.fill = LIGHT_FILL
for col_idx in range(3, 5):
    ws1.cell(row=28, column=col_idx).fill = LIGHT_FILL
for i, line in enumerate(C["key_risks"]):
    set_cell(ws1, 29 + i, 2, line)

# =====================================================================
# SHEET 2: Financial Statements
# =====================================================================
ws2 = wb.create_sheet("Financial Statements")
ws2.sheet_properties.tabColor = "003366"

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 32
for letter in ["C", "D", "E", "F", "G"]:
    ws2.column_dimensions[letter].width = 18

set_cell(ws2, 2, 2, f'{C["company_name"]} - Historical Financials (JPY mn)', font=TITLE_FONT)

n_hist = len(C["hist_years"])
header_row(ws2, 4, 3, 3 + n_hist - 1, C["hist_years"])

section_title(ws2, 5, 2, "Income Statement")

set_cell(ws2, 6, 2, "Revenue", font=BOLD_FONT)
set_cell(ws2, 7, 2, "COGS", font=BOLD_FONT)
set_cell(ws2, 8, 2, "Gross Profit", font=BOLD_FONT)
set_cell(ws2, 9, 2, "Gross Margin")
set_cell(ws2, 10, 2, "SGA", font=BOLD_FONT)
set_cell(ws2, 11, 2, "Operating Income", font=BOLD_FONT)
set_cell(ws2, 12, 2, "Net Income", font=BOLD_FONT)
set_cell(ws2, 13, 2, "Operating Margin")
set_cell(ws2, 14, 2, "Net Margin")
set_cell(ws2, 15, 2, "Revenue Growth (YoY)")
set_cell(ws2, 16, 2, "Operating Income Growth (YoY)")

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)

    set_cell(ws2, 6, col, C["hist_revenue"][i], font=BLUE_FONT, fmt=FMT_YEN)
    cogs_val = C["hist_cogs"][i] if C["hist_cogs"] and i < len(C["hist_cogs"]) else None
    set_cell(ws2, 7, col, cogs_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 8, col, f"={cl}6-{cl}7", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws2, 9, col, f"={cl}8/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    sga_val = C["hist_sga"][i] if C["hist_sga"] and i < len(C["hist_sga"]) else None
    set_cell(ws2, 10, col, sga_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 11, col, C["hist_operating_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 12, col, C["hist_net_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 13, col, f"={cl}11/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws2, 14, col, f"={cl}12/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    if i == 0:
        set_cell(ws2, 15, col, "n/a")
        set_cell(ws2, 16, col, "n/a")
    else:
        prev_cl = col_letter(col - 1)
        set_cell(ws2, 15, col, f"=({cl}6-{prev_cl}6)/{prev_cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        set_cell(ws2, 16, col, f"=({cl}11-{prev_cl}11)/{prev_cl}11", font=BLACK_FONT, fmt=FMT_PCT)

section_title(ws2, 18, 2, "Cash Flow Statement")
set_cell(ws2, 19, 2, "Operating Cash Flow", font=BOLD_FONT)
set_cell(ws2, 20, 2, "Capex", font=BOLD_FONT)
set_cell(ws2, 21, 2, "Free Cash Flow", font=BOLD_FONT)
set_cell(ws2, 22, 2, "FCF Margin")
set_cell(ws2, 23, 2, "Capex / Revenue")

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)
    ocf_val = C["hist_ocf"][i] if C["hist_ocf"] and i < len(C["hist_ocf"]) else None
    capex_val = C["hist_capex"][i] if C["hist_capex"] and i < len(C["hist_capex"]) else None
    set_cell(ws2, 19, col, ocf_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 20, col, capex_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 21, col, f"={cl}19-{cl}20", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws2, 22, col, f"={cl}21/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws2, 23, col, f"={cl}20/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)

section_title(ws2, 25, 2, "Balance Sheet Highlights")
set_cell(ws2, 26, 2, "Cash & Deposits", font=BOLD_FONT)
set_cell(ws2, 27, 2, "Short-term Debt", font=BOLD_FONT)
set_cell(ws2, 28, 2, "Net Debt (Cash)", font=BOLD_FONT)

for i in range(n_hist):
    col = 3 + i
    cl = col_letter(col)
    cash_val = C["hist_cash"][i] if C["hist_cash"] and i < len(C["hist_cash"]) else None
    debt_val = C["hist_debt"][i] if C["hist_debt"] and i < len(C["hist_debt"]) else None
    set_cell(ws2, 26, col, cash_val, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 27, col, debt_val if debt_val is not None else 0, font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws2, 28, col, f"={cl}27-{cl}26", font=BLACK_FONT, fmt=FMT_YEN)

# =====================================================================
# SHEET 3: DCF Model
# =====================================================================
ws3 = wb.create_sheet("DCF Model")
ws3.sheet_properties.tabColor = "006600"

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 32
for letter in ["C", "D", "E", "F", "G"]:
    ws3.column_dimensions[letter].width = 16

set_cell(ws3, 2, 2, f'DCF Valuation Model - {C["company_name"]}', font=TITLE_FONT)

c = section_title(ws3, 4, 2, "Assumptions")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=4, column=col_idx).fill = LIGHT_FILL

assumptions = [
    ("Capex / Revenue",            C["capex_pct"],            FMT_PCT),
    ("Effective Tax Rate",         C["tax_rate"],             FMT_PCT),
    ("Risk-Free Rate",             C["risk_free"],            FMT_PCT),
    ("Beta",                       C["beta"],                 "0.00"),
    ("Equity Risk Premium",        C["erp"],                  FMT_PCT),
    ("Size Premium",               C["size_premium"],         FMT_PCT),
    ("After-tax Cost of Debt",     C["cost_of_debt_at"],      FMT_PCT),
    ("D/E Ratio",                  C["de_ratio"],             "0.000"),
    ("Terminal Growth Rate",       C["terminal_growth"],      FMT_PCT),
    ("Exit Multiple (EV/EBITDA)",  C["exit_multiple"],        FMT_RATIO),
    ("Fully Diluted Shares",       C["shares_outstanding"],   FMT_INT),
    ("Net Debt (JPY mn)",          C["net_debt"],             FMT_YEN),
    ("Base Year Revenue (JPY mn)", C["base_year_revenue"],    FMT_YEN),
    ("D&A / Revenue",              C["da_pct"],               FMT_PCT),
]
for i, (label, val, fmt) in enumerate(assumptions):
    r = 5 + i
    set_cell(ws3, r, 2, label, font=BOLD_FONT)
    set_cell(ws3, r, 3, val, font=BLUE_FONT, fmt=fmt)

c = section_title(ws3, 20, 2, "WACC Calculation")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=20, column=col_idx).fill = LIGHT_FILL

set_cell(ws3, 21, 2, "Cost of Equity (Ke)", font=BOLD_FONT)
set_cell(ws3, 21, 3, "=C7+C8*C9+C10", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 22, 2, "Weight of Equity", font=BOLD_FONT)
set_cell(ws3, 22, 3, "=1/(1+C12)", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 23, 2, "Weight of Debt", font=BOLD_FONT)
set_cell(ws3, 23, 3, "=C12/(1+C12)", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 24, 2, "WACC", font=BOLD_FONT)
set_cell(ws3, 24, 3, "=C21*C22+C11*C23", font=BLACK_FONT, fmt=FMT_PCT2)

set_cell(ws3, 25, 2, "Active Scenario", font=BOLD_FONT)
set_cell(ws3, 25, 3, "Base", font=BLUE_FONT, fill=INPUT_FILL)
set_cell(ws3, 25, 4,
         f"=MATCH(C25,B{R_SCEN_BLK_GROWTH + 1}:B{R_SCEN_BLK_GROWTH + NUM_SCENARIOS},0)",
         font=BLACK_FONT)

dv_scenario = DataValidation(
    type="list",
    formula1='"Base,Upside,Management,Downside 1,Downside 2"',
    allow_blank=False,
    showDropDown=False,
)
dv_scenario.add("C25")
ws3.add_data_validation(dv_scenario)

c = section_title(ws3, 26, 2, "Projected Free Cash Flow")
c.fill = LIGHT_FILL
for col_idx in range(3, 8):
    ws3.cell(row=26, column=col_idx).fill = LIGHT_FILL

proj_years = C["projection_years"]
year_labels = [f"Year {y}" for y in range(1, proj_years + 1)]
header_row(ws3, 27, 3, 3 + proj_years - 1, year_labels)

row_labels_drv = [
    ("Revenue Growth (YoY)",          R_DRV_GROWTH),
    ("COGS % of Revenue",             R_DRV_COGS),
    ("SGA % of Revenue",              R_DRV_SGA),
]
for label, r in row_labels_drv:
    set_cell(ws3, r, 2, label, font=BOLD_FONT)

row_labels_fcf = [
    ("Revenue",                       R_REVENUE),
    ("COGS",                          R_COGS),
    ("Gross Profit",                  R_GROSS_PROFIT),
    ("Gross Margin",                  R_GROSS_MARGIN),
    ("SGA Expense",                   R_SGA),
    ("Implied Operating Margin",      R_OP_M_IMPL),
    ("Operating Income (EBIT)",       R_EBIT),
    ("Less: Tax",                     R_TAX),
    ("NOPAT",                         R_NOPAT),
    ("Plus: D&A",                     R_DA),
    ("Less: Capex",                   R_CAPEX),
    ("Change in NWC",                 R_CHG_NWC),
    ("Unlevered Free Cash Flow",      R_UFCF),
    ("Discount Factor",               R_DISC),
    ("PV of FCF",                     R_PV_FCF),
]
for label, r in row_labels_fcf:
    set_cell(ws3, r, 2, label, font=BOLD_FONT)

for yr in range(proj_years):
    col = 3 + yr
    cl = col_letter(col)
    prev_cl = col_letter(col - 1) if yr > 0 else None

    set_cell(ws3, R_DRV_GROWTH, col, choose_formula(R_SCEN_BLK_GROWTH, cl),
             font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)
    set_cell(ws3, R_DRV_COGS, col, choose_formula(R_SCEN_BLK_COGS, cl),
             font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)
    set_cell(ws3, R_DRV_SGA, col, choose_formula(R_SCEN_BLK_SGA, cl),
             font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)

    if yr == 0:
        set_cell(ws3, R_REVENUE, col, f"=C17*(1+{cl}{R_DRV_GROWTH})", font=BLACK_FONT, fmt=FMT_YEN)
    else:
        set_cell(ws3, R_REVENUE, col, f"={prev_cl}{R_REVENUE}*(1+{cl}{R_DRV_GROWTH})", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_COGS, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_COGS}", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_GROSS_PROFIT, col, f"={cl}{R_REVENUE}-{cl}{R_COGS}", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_GROSS_MARGIN, col, f"={cl}{R_GROSS_PROFIT}/{cl}{R_REVENUE}", font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws3, R_SGA, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_SGA}", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_OP_M_IMPL, col,
             f"=({cl}{R_GROSS_PROFIT}-{cl}{R_SGA})/{cl}{R_REVENUE}",
             font=BLACK_FONT, fmt=FMT_PCT)
    set_cell(ws3, R_EBIT, col, f"={cl}{R_GROSS_PROFIT}-{cl}{R_SGA}", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_TAX, col, f"=MAX(0,{cl}{R_EBIT}*C6)", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_NOPAT, col, f"={cl}{R_EBIT}-{cl}{R_TAX}", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_DA, col, f"={cl}{R_REVENUE}*C18", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_CAPEX, col, f"={cl}{R_REVENUE}*C5", font=BLACK_FONT, fmt=FMT_YEN)

    # Change in NWC linked to NWC Schedule (offset: NWC col = DCF col + 1)
    nwc_col_letter = col_letter(col + 1)
    set_cell(ws3, R_CHG_NWC, col,
             f"='NWC Schedule'!{nwc_col_letter}{NWC_R_CHG_NWC}",
             font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_UFCF, col, f"={cl}{R_NOPAT}+{cl}{R_DA}-{cl}{R_CAPEX}-{cl}{R_CHG_NWC}", font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws3, R_DISC, col, f"=1/(1+C24)^{yr+1}", font=BLACK_FONT, fmt="0.0000")
    set_cell(ws3, R_PV_FCF, col, f"={cl}{R_UFCF}*{cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

# ── Valuation - Perpetuity Growth Method ──
c = section_title(ws3, R_PGM_SEC, 2, "Valuation - Perpetuity Growth Method")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=R_PGM_SEC, column=col_idx).fill = LIGHT_GREEN

last_cl = col_letter(3 + proj_years - 1)

set_cell(ws3, R_SUM_PV, 2, "Sum of PV of FCFs", font=BOLD_FONT)
set_cell(ws3, R_SUM_PV, 3, f"=SUM(C{R_PV_FCF}:{last_cl}{R_PV_FCF})", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_TV_PGM, 2, "Terminal Value (PGM)", font=BOLD_FONT)
set_cell(ws3, R_TV_PGM, 3, f"={last_cl}{R_UFCF}*(1+C13)/(C24-C13)", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PV_TV_PGM, 2, "PV of Terminal Value", font=BOLD_FONT)
set_cell(ws3, R_PV_TV_PGM, 3, f"=C{R_TV_PGM}*{last_cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EV_PGM, 2, "Enterprise Value", font=BOLD_FONT)
set_cell(ws3, R_EV_PGM, 3, f"=C{R_SUM_PV}+C{R_PV_TV_PGM}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EQ_PGM, 2, "Equity Value", font=BOLD_FONT)
set_cell(ws3, R_EQ_PGM, 3, f"=C{R_EV_PGM}-C16", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PRICE_PGM, 2, "Implied Share Price (PGM)", font=BOLD_FONT)
set_cell(ws3, R_PRICE_PGM, 3, f"=ROUND(C{R_EQ_PGM}*1000000/C15,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# ── Valuation - Exit Multiple Method ──
c = section_title(ws3, R_EXIT_SEC, 2, "Valuation - Exit Multiple Method")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=R_EXIT_SEC, column=col_idx).fill = LIGHT_GREEN

set_cell(ws3, R_SUM_PV_EX, 2, "Sum of PV of FCFs", font=BOLD_FONT)
set_cell(ws3, R_SUM_PV_EX, 3, f"=C{R_SUM_PV}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_YR5_EBITDA, 2, "Year 5 EBITDA", font=BOLD_FONT)
set_cell(ws3, R_YR5_EBITDA, 3, f"={last_cl}{R_EBIT}+{last_cl}{R_DA}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_TV_EXIT, 2, "Terminal Value (Exit Multiple)", font=BOLD_FONT)
set_cell(ws3, R_TV_EXIT, 3, f"=C{R_YR5_EBITDA}*C14", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PV_TV_EXIT, 2, "PV of Terminal Value", font=BOLD_FONT)
set_cell(ws3, R_PV_TV_EXIT, 3, f"=C{R_TV_EXIT}*{last_cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EV_EXIT, 2, "Enterprise Value", font=BOLD_FONT)
set_cell(ws3, R_EV_EXIT, 3, f"=C{R_SUM_PV_EX}+C{R_PV_TV_EXIT}", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_EQ_EXIT, 2, "Equity Value", font=BOLD_FONT)
set_cell(ws3, R_EQ_EXIT, 3, f"=C{R_EV_EXIT}-C16", font=BLACK_FONT, fmt=FMT_YEN)

set_cell(ws3, R_PRICE_EXIT, 2, "Implied Share Price (Exit Multiple)", font=BOLD_FONT)
set_cell(ws3, R_PRICE_EXIT, 3, f"=ROUND(C{R_EQ_EXIT}*1000000/C15,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# ── Scenario Input Matrix (Revenue Growth, COGS, SGA only; NWC moved to NWC Schedule) ──
c = section_title(ws3, R_SCEN_SEC, 2, "Scenario Input Matrix")
c.fill = LIGHT_GREEN
for col_idx in range(3, 8):
    ws3.cell(row=R_SCEN_SEC, column=col_idx).fill = LIGHT_GREEN

for yr in range(proj_years):
    set_cell(ws3, R_SCEN_YEARS, 3 + yr, f"Year {yr + 1}",
             font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center"))

driver_blocks = [
    ("Revenue Growth (YoY)", "revenue_growth",    FMT_PCT, R_SCEN_BLK_GROWTH),
    ("COGS % of Revenue",    "cogs_pct",          FMT_PCT, R_SCEN_BLK_COGS),
    ("SGA % of Revenue",     "sga_pct",           FMT_PCT, R_SCEN_BLK_SGA),
]

for drv_label, drv_key, drv_fmt, blk_start in driver_blocks:
    section_title(ws3, blk_start, 2, drv_label)
    for s, scen_name in enumerate(SCENARIO_NAMES):
        r = blk_start + 1 + s
        set_cell(ws3, r, 2, scen_name, font=BOLD_FONT)
        scen_data = config["scenarios"][scen_name][drv_key]
        for yr in range(proj_years):
            set_cell(ws3, r, 3 + yr, scen_data[yr],
                     font=BLUE_FONT, fmt=drv_fmt, fill=INPUT_FILL)

# =====================================================================
# SHEET 4: NWC Schedule (NEW)
# =====================================================================
ws_nwc = wb.create_sheet("NWC Schedule")
ws_nwc.sheet_properties.tabColor = "CC6600"

ws_nwc.column_dimensions["A"].width = 3
ws_nwc.column_dimensions["B"].width = 28
ws_nwc.column_dimensions["C"].width = 16
for letter in ["D", "E", "F", "G", "H"]:
    ws_nwc.column_dimensions[letter].width = 16

set_cell(ws_nwc, 2, 2, f'NWC Schedule - {C["company_name"]}', font=TITLE_FONT)

# ── Headers: Base Year + Year 1-5 ──
nwc_headers = ["Base Year"] + [f"Year {y}" for y in range(1, proj_years + 1)]
header_row(ws_nwc, 4, 3, 3 + proj_years, nwc_headers)

# ── Working Capital Drivers ──
c = section_title(ws_nwc, NWC_R_DSO - 1, 2, "Working Capital Drivers (Days)")
c.fill = LIGHT_FILL
for col_idx in range(3, 3 + proj_years + 1):
    ws_nwc.cell(row=NWC_R_DSO - 1, column=col_idx).fill = LIGHT_FILL

set_cell(ws_nwc, NWC_R_DSO, 2, "DSO (Days Sales Outstanding)", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_DIH, 2, "DIH (Days Inventory Held)", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_DPO, 2, "DPO (Days Payable Outstanding)", font=BOLD_FONT)

# Base Year DSO/DIH/DPO (computed from actuals)
set_cell(ws_nwc, NWC_R_DSO, 3, f"=C{NWC_R_AR}/C{NWC_R_REV}*365",
         font=BLACK_FONT, fmt=FMT_DAYS)
set_cell(ws_nwc, NWC_R_DIH, 3, f"=C{NWC_R_INV}/C{NWC_R_COGS}*365",
         font=BLACK_FONT, fmt=FMT_DAYS)
set_cell(ws_nwc, NWC_R_DPO, 3, f"=C{NWC_R_AP}/C{NWC_R_COGS}*365",
         font=BLACK_FONT, fmt=FMT_DAYS)

# Projected DSO/DIH/DPO (CHOOSE from scenario matrix)
for yr in range(proj_years):
    nwc_col = 4 + yr  # D=Year1, E=Year2, ...
    cl = col_letter(nwc_col)
    set_cell(ws_nwc, NWC_R_DSO, nwc_col,
             nwc_choose_formula(NWC_R_SCEN_BLK_DSO, cl),
             font=BLACK_FONT, fmt=FMT_DAYS, fill=LIGHT_FILL)
    set_cell(ws_nwc, NWC_R_DIH, nwc_col,
             nwc_choose_formula(NWC_R_SCEN_BLK_DIH, cl),
             font=BLACK_FONT, fmt=FMT_DAYS, fill=LIGHT_FILL)
    set_cell(ws_nwc, NWC_R_DPO, nwc_col,
             nwc_choose_formula(NWC_R_SCEN_BLK_DPO, cl),
             font=BLACK_FONT, fmt=FMT_DAYS, fill=LIGHT_FILL)

# ── Revenue & COGS (linked from DCF Model) ──
c = section_title(ws_nwc, NWC_R_REV - 1, 2, "P&L Reference (JPY mn)")
c.fill = LIGHT_FILL
for col_idx in range(3, 3 + proj_years + 1):
    ws_nwc.cell(row=NWC_R_REV - 1, column=col_idx).fill = LIGHT_FILL

set_cell(ws_nwc, NWC_R_REV, 2, "Revenue", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_COGS, 2, "COGS", font=BOLD_FONT)

# Base Year
set_cell(ws_nwc, NWC_R_REV, 3, C["base_year_revenue"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws_nwc, NWC_R_COGS, 3, C["base_year_cogs"], font=BLUE_FONT, fmt=FMT_YEN)

# Projected (linked to DCF Model; NWC col D = DCF col C, offset +1)
for yr in range(proj_years):
    nwc_col = 4 + yr
    dcf_col_letter = col_letter(3 + yr)
    set_cell(ws_nwc, NWC_R_REV, nwc_col,
             f"='DCF Model'!{dcf_col_letter}{R_REVENUE}",
             font=BLACK_FONT, fmt=FMT_YEN)
    set_cell(ws_nwc, NWC_R_COGS, nwc_col,
             f"='DCF Model'!{dcf_col_letter}{R_COGS}",
             font=BLACK_FONT, fmt=FMT_YEN)

# ── Working Capital Items ──
c = section_title(ws_nwc, NWC_R_AR - 1, 2, "Working Capital Items (JPY mn)")
c.fill = LIGHT_FILL
for col_idx in range(3, 3 + proj_years + 1):
    ws_nwc.cell(row=NWC_R_AR - 1, column=col_idx).fill = LIGHT_FILL

set_cell(ws_nwc, NWC_R_AR, 2, "Accounts Receivable", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_INV, 2, "Inventory", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_CA, 2, "Current Assets (AR + Inv)", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_AP, 2, "Accounts Payable", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_CL, 2, "Current Liabilities (AP)", font=BOLD_FONT)

# Base Year actuals
set_cell(ws_nwc, NWC_R_AR, 3, C["base_year_ar"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws_nwc, NWC_R_INV, 3, C["base_year_inv"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws_nwc, NWC_R_CA, 3, f"=C{NWC_R_AR}+C{NWC_R_INV}", font=BLACK_FONT, fmt=FMT_YEN)
set_cell(ws_nwc, NWC_R_AP, 3, C["base_year_ap"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws_nwc, NWC_R_CL, 3, f"=C{NWC_R_AP}", font=BLACK_FONT, fmt=FMT_YEN)

# Projected WC items
for yr in range(proj_years):
    nwc_col = 4 + yr
    cl = col_letter(nwc_col)

    # AR = Revenue * DSO / 365
    set_cell(ws_nwc, NWC_R_AR, nwc_col,
             f"={cl}{NWC_R_REV}*{cl}{NWC_R_DSO}/365",
             font=BLACK_FONT, fmt=FMT_YEN)
    # Inv = COGS * DIH / 365
    set_cell(ws_nwc, NWC_R_INV, nwc_col,
             f"={cl}{NWC_R_COGS}*{cl}{NWC_R_DIH}/365",
             font=BLACK_FONT, fmt=FMT_YEN)
    # CA = AR + Inv
    set_cell(ws_nwc, NWC_R_CA, nwc_col,
             f"={cl}{NWC_R_AR}+{cl}{NWC_R_INV}",
             font=BLACK_FONT, fmt=FMT_YEN)
    # AP = COGS * DPO / 365
    set_cell(ws_nwc, NWC_R_AP, nwc_col,
             f"={cl}{NWC_R_COGS}*{cl}{NWC_R_DPO}/365",
             font=BLACK_FONT, fmt=FMT_YEN)
    # CL = AP
    set_cell(ws_nwc, NWC_R_CL, nwc_col,
             f"={cl}{NWC_R_AP}",
             font=BLACK_FONT, fmt=FMT_YEN)

# ── NWC Summary ──
c = section_title(ws_nwc, NWC_R_NWC - 1, 2, "Net Working Capital (JPY mn)")
c.fill = LIGHT_GREEN
for col_idx in range(3, 3 + proj_years + 1):
    ws_nwc.cell(row=NWC_R_NWC - 1, column=col_idx).fill = LIGHT_GREEN

set_cell(ws_nwc, NWC_R_NWC, 2, "Net Working Capital", font=BOLD_FONT)
set_cell(ws_nwc, NWC_R_CHG_NWC, 2, "Change in NWC", font=BOLD_FONT)

# Base Year NWC
set_cell(ws_nwc, NWC_R_NWC, 3, f"=C{NWC_R_CA}-C{NWC_R_CL}",
         font=BLACK_FONT, fmt=FMT_YEN, border=BOTTOM_BORDER)
set_cell(ws_nwc, NWC_R_CHG_NWC, 3, "n/a", font=BLACK_FONT)

# Projected NWC & Change
for yr in range(proj_years):
    nwc_col = 4 + yr
    cl = col_letter(nwc_col)
    prev_cl = col_letter(nwc_col - 1)

    set_cell(ws_nwc, NWC_R_NWC, nwc_col,
             f"={cl}{NWC_R_CA}-{cl}{NWC_R_CL}",
             font=BLACK_FONT, fmt=FMT_YEN, border=BOTTOM_BORDER)
    set_cell(ws_nwc, NWC_R_CHG_NWC, nwc_col,
             f"={cl}{NWC_R_NWC}-{prev_cl}{NWC_R_NWC}",
             font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

# ── Scenario Input Matrix (DSO, DIH, DPO) ──
c = section_title(ws_nwc, NWC_R_SCEN_SEC, 2, "Scenario Input Matrix (Working Capital Days)")
c.fill = LIGHT_GREEN
for col_idx in range(3, 3 + proj_years + 1):
    ws_nwc.cell(row=NWC_R_SCEN_SEC, column=col_idx).fill = LIGHT_GREEN

# Year headers for scenario matrix (columns D-H, matching projected years)
for yr in range(proj_years):
    set_cell(ws_nwc, NWC_R_SCEN_YEARS, 4 + yr, f"Year {yr + 1}",
             font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center"))

nwc_driver_blocks = [
    ("DSO (Days)", "dso_days", FMT_DAYS, NWC_R_SCEN_BLK_DSO),
    ("DIH (Days)", "dih_days", FMT_DAYS, NWC_R_SCEN_BLK_DIH),
    ("DPO (Days)", "dpo_days", FMT_DAYS, NWC_R_SCEN_BLK_DPO),
]

for drv_label, drv_key, drv_fmt, blk_start in nwc_driver_blocks:
    section_title(ws_nwc, blk_start, 2, drv_label)
    for s, scen_name in enumerate(SCENARIO_NAMES):
        r = blk_start + 1 + s
        set_cell(ws_nwc, r, 2, scen_name, font=BOLD_FONT)
        scen_data = config["scenarios"][scen_name][drv_key]
        for yr in range(proj_years):
            set_cell(ws_nwc, r, 4 + yr, scen_data[yr],
                     font=BLUE_FONT, fmt=drv_fmt, fill=INPUT_FILL)

# =====================================================================
# SHEET 5: Comps Analysis
# =====================================================================
ws4 = wb.create_sheet("Comps Analysis")
ws4.sheet_properties.tabColor = "006600"

ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 10
for letter in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
    ws4.column_dimensions[letter].width = 12

set_cell(ws4, 2, 2, "Comparable Company Analysis", font=TITLE_FONT)

comp_headers = [
    "Company", "Ticker", "Mkt Cap\n(JPY mn)", "EV\n(JPY mn)",
    "Revenue\n(JPY mn)", "EBITDA\n(JPY mn)", "Op Income\n(JPY mn)",
    "Net Income\n(JPY mn)", "EV/EBITDA", "EV/Revenue", "PER",
    "PBR", "Op Margin", "ROE"
]
header_row(ws4, 4, 2, 15, comp_headers)

comps = C["comps"]
for i, comp in enumerate(comps):
    r = 5 + i

    _na = lambda ws, r, c: set_cell(ws, r, c, "N/A", font=BLACK_FONT, border=THIN_BORDER,
                                     alignment=Alignment(horizontal="right"))

    set_cell(ws4, r, 2, comp["name"], font=BOLD_FONT)
    set_cell(ws4, r, 3, comp["ticker"])

    if comp["mkt_cap"] is None:
        _na(ws4, r, 4)
    else:
        set_cell(ws4, r, 4, comp["mkt_cap"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    if comp["ev"] is None:
        _na(ws4, r, 5)
    else:
        set_cell(ws4, r, 5, comp["ev"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    set_cell(ws4, r, 6, comp["revenue"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 7, comp["ebitda"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 8, comp["op_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws4, r, 9, comp["net_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    if comp["ev"] is None or comp["ebitda"] <= 0:
        _na(ws4, r, 10)
    else:
        set_cell(ws4, r, 10, f"=E{r}/G{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

    if comp["ev"] is None or comp["revenue"] <= 0:
        _na(ws4, r, 11)
    else:
        set_cell(ws4, r, 11, f"=E{r}/F{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

    if comp["mkt_cap"] is None or comp["net_income"] <= 0:
        _na(ws4, r, 12)
    else:
        set_cell(ws4, r, 12, f"=D{r}/I{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

    if comp["pbr"] is None:
        _na(ws4, r, 13)
    else:
        set_cell(ws4, r, 13, comp["pbr"], font=BLUE_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

    set_cell(ws4, r, 14, f"=H{r}/F{r}", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

    if comp["roe"] is None:
        _na(ws4, r, 15)
    else:
        set_cell(ws4, r, 15, comp["roe"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)

last_comp_row = 5 + len(comps) - 1

section_title(ws4, 14, 2, "Statistics")

stat_labels = ["25th Percentile", "Median (50th)", "75th Percentile"]
stat_rows = [15, 16, 17]

stat_col_map = [
    (4, 10),  # EV/EBITDA
    (5, 11),  # EV/Revenue
    (6, 12),  # PER
    (7, 13),  # PBR
    (8, 14),  # Op Margin
    (9, 15),  # ROE
]

for stat_idx, (label, r) in enumerate(zip(stat_labels, stat_rows)):
    set_cell(ws4, r, 2, label, font=BOLD_FONT)

    for dst_col, src_col in stat_col_map:
        src_letter = col_letter(src_col)
        rng = f"{src_letter}5:{src_letter}{last_comp_row}"

        if src_col in (14, 15):
            fmt = FMT_PCT
        else:
            fmt = FMT_RATIO

        if stat_idx == 0:
            formula = f"=PERCENTILE({rng},0.25)"
        elif stat_idx == 1:
            formula = f"=MEDIAN({rng})"
        else:
            formula = f"=PERCENTILE({rng},0.75)"

        set_cell(ws4, r, dst_col, formula, font=BLACK_FONT, fmt=fmt, border=THIN_BORDER)

# ── Implied Valuation ──
c = section_title(ws4, 19, 2, f'Implied Valuation for {C["company_name"]}')
c.fill = LIGHT_GREEN
for col_idx in range(3, 10):
    ws4.cell(row=19, column=col_idx).fill = LIGHT_GREEN

section_title(ws4, 20, 2, "Core Corporation Financials")

if USE_EV_SALES:
    set_cell(ws4, 21, 2, "Revenue (JPY mn)", font=BOLD_FONT)
    set_cell(ws4, 21, 3, C["base_year_revenue"], font=BLUE_FONT, fmt=FMT_YEN)
else:
    set_cell(ws4, 21, 2, "EBITDA (JPY mn)", font=BOLD_FONT)
    set_cell(ws4, 21, 3, C["core_ebitda"], font=BLUE_FONT, fmt=FMT_YEN)

set_cell(ws4, 22, 2, "Net Income (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 22, 3, C["core_net_income"], font=BLUE_FONT, fmt=FMT_YEN)
set_cell(ws4, 23, 2, "Shares Outstanding", font=BOLD_FONT)
set_cell(ws4, 23, 3, C["shares_outstanding"], font=BLUE_FONT, fmt=FMT_INT)
set_cell(ws4, 24, 2, "Net Debt (JPY mn)", font=BOLD_FONT)
set_cell(ws4, 24, 3, C["net_debt"], font=BLUE_FONT, fmt=FMT_YEN)

section_title(ws4, 26, 2, "Implied Share Price (Median Multiples)")

if USE_EV_SALES:
    set_cell(ws4, 27, 2, "Via EV/Sales (Median)", font=BOLD_FONT)
    set_cell(ws4, 27, 3, "=ROUND((C21*E16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)
else:
    set_cell(ws4, 27, 2, "Via EV/EBITDA (Median)", font=BOLD_FONT)
    set_cell(ws4, 27, 3, "=ROUND((C21*D16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)

set_cell(ws4, 28, 2, "Via PER (Median)", font=BOLD_FONT)
set_cell(ws4, 28, 3, "=ROUND(C22*F16*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
         border=TOP_BOTTOM)

# =====================================================================
# SHEET 6: Sensitivity Analysis (Dynamic Excel formulas)
# =====================================================================
ws5 = wb.create_sheet("Sensitivity Analysis")
ws5.sheet_properties.tabColor = "996600"

ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 24
for letter in ["C", "D", "E", "F", "G", "H", "I"]:
    ws5.column_dimensions[letter].width = 14

set_cell(ws5, 2, 2, "Sensitivity Analysis", font=TITLE_FONT)

# ── Current values reference ──
set_cell(ws5, 3, 2, "Current WACC:", font=BOLD_FONT)
set_cell(ws5, 3, 3, "='DCF Model'!C24", font=BLACK_FONT, fmt=FMT_PCT2)
set_cell(ws5, 3, 5, "Terminal g:", font=BOLD_FONT)
set_cell(ws5, 3, 6, "='DCF Model'!C13", font=BLACK_FONT, fmt=FMT_PCT2)
set_cell(ws5, 3, 8, "Exit Multiple:", font=BOLD_FONT)
set_cell(ws5, 3, 9, "='DCF Model'!C14", font=BLACK_FONT, fmt=FMT_RATIO)

# ── Dynamic formula builders ──
_DCF = "'DCF Model'"
_SHARES = f"{_DCF}!C15"
_NET_DEBT = f"{_DCF}!C16"
_last_cl = col_letter(3 + proj_years - 1)
_ufcf_cells = [f"{_DCF}!{col_letter(3 + yr)}{R_UFCF}" for yr in range(proj_years)]

def _build_pgm_formula(wacc_ref, tg_ref):
    pv_parts = [f"{_ufcf_cells[yr]}/(1+{wacc_ref})^{yr+1}" for yr in range(proj_years)]
    last_ufcf = _ufcf_cells[proj_years - 1]
    pv_tv = f"{last_ufcf}*(1+{tg_ref})/({wacc_ref}-{tg_ref})/(1+{wacc_ref})^{proj_years}"
    return f'=IFERROR(ROUND(({"+".join(pv_parts)}+{pv_tv}-{_NET_DEBT})*1000000/{_SHARES},0),"")'

def _build_exit_formula(wacc_ref, mult_ref):
    pv_parts = [f"{_ufcf_cells[yr]}/(1+{wacc_ref})^{yr+1}" for yr in range(proj_years)]
    yr5_ebitda = f"({_DCF}!{_last_cl}{R_EBIT}+{_DCF}!{_last_cl}{R_DA})"
    pv_tv = f"{yr5_ebitda}*{mult_ref}/(1+{wacc_ref})^{proj_years}"
    return f'=IFERROR(ROUND(({"+".join(pv_parts)}+{pv_tv}-{_NET_DEBT})*1000000/{_SHARES},0),"")'

# ── Dynamic header helpers ──
_N_GRID = 7
_CENTER_IDX = 3
_WACC_STEP = 0.005
_TG_STEP   = 0.0025
_EXIT_STEP = 1.0
_ANCHOR_WACC = "$C$3"
_ANCHOR_TG   = "$F$3"
_ANCHOR_EXIT = "$I$3"

def _offset_formula(anchor, offset_val):
    if offset_val == 0:
        return f"={anchor}"
    elif offset_val > 0:
        return f"={anchor}+{offset_val}"
    else:
        return f"={anchor}-{abs(offset_val)}"

# ── Table 1: WACC vs Terminal Growth Rate (PGM) ──
T1_TITLE = 5
T1_HDR = 6
T1_DATA = 7

section_title(ws5, T1_TITLE, 2,
              "Table 1: WACC vs Terminal Growth Rate (PGM - Implied Share Price, JPY)")

set_cell(ws5, T1_HDR, 2, "WACC \\ Terminal g", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j in range(_N_GRID):
    offset = round((j - _CENTER_IDX) * _TG_STEP, 6)
    set_cell(ws5, T1_HDR, 3 + j, _offset_formula(_ANCHOR_TG, offset),
             font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

for i in range(_N_GRID):
    r = T1_DATA + i
    offset = round((i - _CENTER_IDX) * _WACC_STEP, 6)
    set_cell(ws5, r, 2, _offset_formula(_ANCHOR_WACC, offset),
             font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER, fill=INPUT_FILL)
    for j in range(_N_GRID):
        col = 3 + j
        cl = col_letter(col)
        formula = _build_pgm_formula(f"$B{r}", f"{cl}${T1_HDR}")
        set_cell(ws5, r, col, formula, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# ── Table 2: WACC vs Exit Multiple ──
T2_TITLE = T1_DATA + _N_GRID + 2
T2_HDR = T2_TITLE + 1
T2_DATA = T2_HDR + 1

section_title(ws5, T2_TITLE, 2,
              "Table 2: WACC vs Exit Multiple (Exit Multiple - Implied Share Price, JPY)")

set_cell(ws5, T2_HDR, 2, "WACC \\ Exit Multiple", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j in range(_N_GRID):
    offset = round((j - _CENTER_IDX) * _EXIT_STEP, 6)
    set_cell(ws5, T2_HDR, 3 + j, _offset_formula(_ANCHOR_EXIT, offset),
             font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_RATIO,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

for i in range(_N_GRID):
    r = T2_DATA + i
    offset = round((i - _CENTER_IDX) * _WACC_STEP, 6)
    set_cell(ws5, r, 2, _offset_formula(_ANCHOR_WACC, offset),
             font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER, fill=INPUT_FILL)
    for j in range(_N_GRID):
        col = 3 + j
        cl = col_letter(col)
        formula = _build_exit_formula(f"$B{r}", f"{cl}${T2_HDR}")
        set_cell(ws5, r, col, formula, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# ── Note ──
_note_row = T2_DATA + _N_GRID + 1
set_cell(ws5, _note_row, 2,
         "All values dynamically linked to DCF Model. "
         "Headers auto-center on current WACC / Terminal g / Exit Multiple.",
         font=Font(name="Arial", size=9, italic=True, color="808080"))
ws5.merge_cells(start_row=_note_row, start_column=2,
                end_row=_note_row, end_column=9)

# =====================================================================
# SAVE
# =====================================================================
wb.save(output_file)
print(f"\nSaved: {output_file}")
print("Done! Open the file in Excel to verify formulas and scenario dropdown.")
print(f"Sheets: {wb.sheetnames}")
