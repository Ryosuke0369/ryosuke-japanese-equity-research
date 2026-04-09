"""
sotp_template.py - SOTP (Sum-of-the-Parts) Valuation Excel Generator

Generates a 6-sheet Excel workbook:
  1. Cover & Thesis
  2. Segment Data
  3. Peer Comps
  4. SOTP Valuation
  5. Sensitivity
  6. D&A Allocation

All calculations use Excel formulas (not hardcoded Python values).
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# =====================================================================
# STYLE CONSTANTS (matching dcf_comps_template.py)
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="0000FF", bold=False)
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
GREEN_FONT  = Font(name="Arial", size=10, color="006600")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
GREY_FONT   = Font(name="Arial", size=9, italic=True, color="808080")
TICKER_FONT = Font(name="Arial", size=10, color="808080")
NOTE_FONT   = Font(name="Arial", size=9, color="808080")
EXCLUDED_FONT      = Font(name="Arial", size=10, color="808080", italic=True)
EXCLUDED_NOTE_FONT = Font(name="Arial", size=9, color="808080", italic=True)

HEADER_FILL    = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
LIGHT_FILL     = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_BLUE_BG  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBTOTAL_FILL  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER     = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))
SECTION_BOTTOM  = Border(bottom=Side(style="thin"))
SUBTOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="thin"))
TOP_BOTTOM      = Border(top=Side(style="thin"), bottom=Side(style="double"))
_GRAY_SIDE      = Side(style="thin", color="B0B0B0")
INPUT_BORDER    = Border(left=_GRAY_SIDE, right=_GRAY_SIDE,
                          top=_GRAY_SIDE, bottom=_GRAY_SIDE)

FMT_YEN     = '#,##0;(#,##0)'
FMT_PCT     = '0.0%;(0.0%)'
FMT_PCT2    = '0.00%;(0.00%)'
FMT_RATIO   = '0.00"x"'
FMT_INT     = '#,##0'
FMT_PRICE   = '¥#,##0;(¥#,##0)'

CENTER = Alignment(horizontal="center")
LEFT   = Alignment(horizontal="left")
RIGHT  = Alignment(horizontal="right")
WRAP   = Alignment(horizontal="center", wrap_text=True)


# =====================================================================
# HELPER
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c


def header_row(ws, row, col_start, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.font = font
        c.fill = fill
        c.alignment = WRAP
        c.border = SECTION_BOTTOM


def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c


def cl(col_num):
    return get_column_letter(col_num)


# =====================================================================
# SHEET 1: COVER & THESIS
# =====================================================================
def build_cover_sheet(wb, sotp):
    ws = wb.active
    ws.title = "Cover & Thesis"
    ws.sheet_properties.tabColor = "000080"

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # Title
    set_cell(ws, 2, 2, "SOTP Valuation Model", font=TITLE_FONT)
    set_cell(ws, 3, 2, f"Valuation Date: {sotp.get('valuation_date', '')}", font=BLACK_FONT)

    # Thesis
    set_cell(ws, 5, 2, "Investment Thesis", font=SUB_FONT)
    set_cell(ws, 6, 2, sotp.get("thesis", ""), font=BLACK_FONT)
    ws.merge_cells("B6:F8")
    ws.cell(row=6, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Cross-method comparison table
    r = 10
    set_cell(ws, r, 2, "Valuation Cross-Check", font=SUB_FONT)
    r += 1
    header_row(ws, r, 2, ["Method", "Fair Value (¥)", "vs SOTP Base"])
    r += 1

    xcheck = sotp.get("dcf_crosscheck", {})
    methods = [
        ("SOTP (Base Case)", None, None),  # patched after SOTP Valuation sheet is built
        ("DCF — Perpetuity Growth", xcheck.get("pgm_fair_value"), None),
        ("DCF — Exit Multiple", xcheck.get("exit_fair_value"), None),
        ("Comps — EV/EBITDA", xcheck.get("comps_ev_ebitda"), None),
        ("Comps — PER", xcheck.get("comps_per"), None),
    ]

    sotp_val_cell = f"$C${r}"  # Row of SOTP base case value

    for i, (method, val, _) in enumerate(methods):
        row = r + i
        set_cell(ws, row, 2, method, font=BLACK_FONT, border=THIN_BORDER)
        if i == 0:
            # SOTP: link to SOTP Valuation sheet
            set_cell(ws, row, 3, val, font=GREEN_FONT, fmt=FMT_PRICE, border=THIN_BORDER)
        else:
            set_cell(ws, row, 3, val, font=BLUE_FONT, fmt=FMT_PRICE, border=THIN_BORDER)
        if i == 0:
            set_cell(ws, row, 4, "—", font=GREY_FONT, border=THIN_BORDER, alignment=CENTER)
        else:
            # vs SOTP = (method - SOTP) / SOTP
            set_cell(ws, row, 4, f"=(C{row}-{sotp_val_cell})/{sotp_val_cell}",
                     font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER, alignment=CENTER)


# =====================================================================
# SHEET 2: SEGMENT DATA
# =====================================================================
def build_segment_data_sheet(wb, sotp):
    ws = wb.create_sheet("Segment Data")
    ws.sheet_properties.tabColor = "006600"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35

    segments = sotp["segments"]
    fiscal_years = segments[0].get("fiscal_years", ["FY23", "FY24", "FY25E"])
    n_years = len(fiscal_years)

    # Set FY column widths
    for i in range(n_years):
        ws.column_dimensions[cl(3 + i)].width = 16

    set_cell(ws, 1, 2, "Segment Financial Data", font=TITLE_FONT)
    set_cell(ws, 2, 2, "(¥ millions)", font=GREY_FONT)

    r = 4
    # Header row
    headers = ["Segment"] + fiscal_years
    header_row(ws, r, 2, headers)
    r += 1

    seg_start_rows = {}  # key -> {"rev_row": ..., "op_row": ..., "opm_row": ...}

    for seg in segments:
        key = seg["key"]
        # Segment name
        set_cell(ws, r, 2, f"{seg['label']}  ({seg['label_jp']})", font=BOLD_FONT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + n_years)
        r += 1

        # Revenue
        rev_row = r
        set_cell(ws, r, 2, "  Revenue", font=BLACK_FONT)
        for j, val in enumerate(seg["revenue"]):
            set_cell(ws, r, 3 + j, val, font=BLUE_FONT, fmt=FMT_YEN, border=INPUT_BORDER)
        r += 1

        # Operating Profit
        op_row = r
        set_cell(ws, r, 2, "  Operating Profit", font=BLACK_FONT)
        for j, val in enumerate(seg["op"]):
            set_cell(ws, r, 3 + j, val, font=BLUE_FONT, fmt=FMT_YEN, border=INPUT_BORDER)
        r += 1

        # OPM (formula)
        opm_row = r
        set_cell(ws, r, 2, "  OPM %", font=BLACK_FONT)
        for j in range(n_years):
            c = cl(3 + j)
            set_cell(ws, r, 3 + j,
                     f"=IF({c}{rev_row}=0,0,{c}{op_row}/{c}{rev_row})",
                     font=BLACK_FONT, fmt=FMT_PCT, border=INPUT_BORDER)
        r += 1

        seg_start_rows[key] = {"rev_row": rev_row, "op_row": op_row, "opm_row": opm_row}
        r += 1  # blank row between segments

    # Consolidated totals
    set_cell(ws, r, 2, "Consolidated Total", font=BOLD_FONT, fill=SUBTOTAL_FILL)
    r += 1

    total_rev_row = r
    set_cell(ws, r, 2, "  Total Revenue", font=BOLD_FONT, fill=SUBTOTAL_FILL)
    for j in range(n_years):
        c = cl(3 + j)
        refs = [f"{c}{seg_start_rows[seg['key']]['rev_row']}" for seg in segments]
        set_cell(ws, r, 3 + j, f"={'+'.join(refs)}",
                 font=GREEN_FONT, fmt=FMT_YEN, fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
    r += 1

    total_op_row = r
    set_cell(ws, r, 2, "  Total Operating Profit", font=BOLD_FONT, fill=SUBTOTAL_FILL)
    for j in range(n_years):
        c = cl(3 + j)
        refs = [f"{c}{seg_start_rows[seg['key']]['op_row']}" for seg in segments]
        set_cell(ws, r, 3 + j, f"={'+'.join(refs)}",
                 font=GREEN_FONT, fmt=FMT_YEN, fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
    r += 1

    # Total OPM
    set_cell(ws, r, 2, "  Total OPM %", font=BOLD_FONT, fill=SUBTOTAL_FILL)
    for j in range(n_years):
        c = cl(3 + j)
        set_cell(ws, r, 3 + j,
                 f"=IF({c}{total_rev_row}=0,0,{c}{total_op_row}/{c}{total_rev_row})",
                 font=GREEN_FONT, fmt=FMT_PCT, fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)

    return seg_start_rows


# =====================================================================
# SHEET 3: PEER COMPS
# =====================================================================
def build_peer_comps_sheet(wb, sotp):
    ws = wb.create_sheet("Peer Comps")
    ws.sheet_properties.tabColor = "800000"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 30

    set_cell(ws, 1, 2, "Peer Comparable Analysis", font=TITLE_FONT)

    r = 3
    segments = sotp["segments"]

    # Track where selected multiples end up for SOTP Valuation sheet references
    selected_multiple_cells = {}  # key -> cell reference e.g. "D15"

    for seg in segments:
        key = seg["key"]
        peers = seg.get("peers", [])
        included_peers = [p for p in peers if not p.get("excluded", False)]
        excluded_peers = [p for p in peers if p.get("excluded", False)]

        # Segment header
        set_cell(ws, r, 2, f"{seg['label']}", font=SUB_FONT, fill=LIGHT_FILL)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        for col in range(2, 8):
            ws.cell(row=r, column=col).fill = LIGHT_FILL
        r += 1

        # Column headers
        header_row(ws, r, 2, ["Company", "Ticker", "EV/EBITDA", "OPM %", "Mkt Cap ($B)", "Note"])
        r += 1

        # Included peers
        ev_ebitda_cells = []
        for p in included_peers:
            set_cell(ws, r, 2, p["name"], font=BLACK_FONT, border=THIN_BORDER)
            set_cell(ws, r, 3, p["ticker"], font=TICKER_FONT, border=THIN_BORDER)
            set_cell(ws, r, 4, p["ev_ebitda"], font=BLUE_FONT, fmt=FMT_RATIO, border=THIN_BORDER)
            set_cell(ws, r, 5, p["opm"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
            set_cell(ws, r, 6, p.get("market_cap_usd_b", ""), font=BLUE_FONT, fmt=FMT_INT, border=THIN_BORDER)
            set_cell(ws, r, 7, p.get("note", ""), font=NOTE_FONT, border=THIN_BORDER)
            ev_ebitda_cells.append(f"D{r}")
            r += 1

        # Excluded peers (gray italic to visually indicate exclusion)
        for p in excluded_peers:
            set_cell(ws, r, 2, p["name"], font=EXCLUDED_FONT, border=THIN_BORDER)
            set_cell(ws, r, 3, p["ticker"], font=EXCLUDED_FONT, border=THIN_BORDER)
            set_cell(ws, r, 4, p["ev_ebitda"], font=EXCLUDED_FONT, fmt=FMT_RATIO, border=THIN_BORDER)
            set_cell(ws, r, 5, p["opm"], font=EXCLUDED_FONT, fmt=FMT_PCT, border=THIN_BORDER)
            set_cell(ws, r, 6, p.get("market_cap_usd_b", ""), font=EXCLUDED_FONT, fmt=FMT_INT, border=THIN_BORDER)
            set_cell(ws, r, 7, p.get("note", ""), font=EXCLUDED_NOTE_FONT, border=THIN_BORDER)
            r += 1

        # Median row
        set_cell(ws, r, 2, "Median", font=BOLD_FONT, fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
        set_cell(ws, r, 3, "", fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
        if ev_ebitda_cells:
            median_formula = f"=MEDIAN({','.join(ev_ebitda_cells)})"
        else:
            median_formula = 0
        set_cell(ws, r, 4, median_formula, font=BOLD_FONT, fmt=FMT_RATIO,
                 fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
        for col in [5, 6, 7]:
            set_cell(ws, r, col, "", fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
        r += 1

        # Selected multiple
        set_cell(ws, r, 2, "Selected Multiple", font=BOLD_FONT, fill=LIGHT_YELLOW, border=SUBTOTAL_BORDER)
        set_cell(ws, r, 3, seg.get("multiple_source", ""), font=NOTE_FONT, fill=LIGHT_YELLOW, border=SUBTOTAL_BORDER)
        set_cell(ws, r, 4, seg["selected_multiple"], font=BLUE_FONT, fmt=FMT_RATIO,
                 fill=LIGHT_YELLOW, border=SUBTOTAL_BORDER)
        selected_multiple_cells[key] = f"D{r}"
        for col in [5, 6, 7]:
            set_cell(ws, r, col, "", fill=LIGHT_YELLOW, border=SUBTOTAL_BORDER)
        r += 2  # blank row

    return selected_multiple_cells


# =====================================================================
# SHEET 4: SOTP VALUATION
# =====================================================================
def build_sotp_valuation_sheet(wb, sotp, seg_data_rows, peer_mult_cells):
    ws = wb.create_sheet("SOTP Valuation")
    ws.sheet_properties.tabColor = "0000FF"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    set_cell(ws, 1, 2, "SOTP Valuation", font=TITLE_FONT)
    set_cell(ws, 2, 2, "(¥ millions unless noted)", font=GREY_FONT)

    segments = sotp["segments"]
    consol = sotp["consolidated"]
    n_seg = len(segments)
    fiscal_years = segments[0].get("fiscal_years", ["FY23", "FY24", "FY25E"])
    # Use last FY for valuation
    val_fy_idx = len(fiscal_years) - 1
    val_fy_col_letter = cl(3 + val_fy_idx)  # column in Segment Data sheet

    # ── EBITDA Buildup ──
    r = 4
    header_row(ws, r, 2, ["Segment", "OP (FY25E)", "D&A Alloc %", "D&A", "EBITDA", "Multiple (x)", "Segment EV"])
    r += 1

    seg_rows = {}  # key -> row number

    for i, seg in enumerate(segments):
        key = seg["key"]
        row = r + i
        seg_rows[key] = row

        # Segment name
        set_cell(ws, row, 2, seg["label"], font=BLACK_FONT, border=THIN_BORDER)

        # OP = link to Segment Data
        op_ref = f"='Segment Data'!{val_fy_col_letter}{seg_data_rows[key]['op_row']}"
        set_cell(ws, row, 3, op_ref, font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)

        # D&A allocation % (input)
        set_cell(ws, row, 4, seg["da_allocation_pct"], font=BLUE_FONT, fmt=FMT_PCT, border=INPUT_BORDER)

        # D&A = consolidated D&A total × allocation %
        # consolidated D&A total is in a named cell below
        da_total_cell = f"B{r + n_seg + 3}"  # will be placed below segment rows
        set_cell(ws, row, 5, f"={da_total_cell}*D{row}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

        # EBITDA = OP + D&A
        set_cell(ws, row, 6, f"=C{row}+E{row}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

        # Multiple = link to Peer Comps selected multiple
        mult_ref = f"='Peer Comps'!{peer_mult_cells[key]}"
        set_cell(ws, row, 7, mult_ref, font=GREEN_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

        # Segment EV = EBITDA × Multiple
        set_cell(ws, row, 8, f"=F{row}*G{row}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # Equity method investments row
    r_eq_method = r + n_seg
    set_cell(ws, r_eq_method, 2, "Equity Method Investments", font=BLACK_FONT, border=THIN_BORDER)
    eq_op = consol.get("equity_method_op", 0)
    eq_mult = consol.get("equity_method_multiple", 10.0)
    set_cell(ws, r_eq_method, 3, eq_op, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws, r_eq_method, 4, "", border=THIN_BORDER)
    set_cell(ws, r_eq_method, 5, "", border=THIN_BORDER)
    set_cell(ws, r_eq_method, 6, f"=C{r_eq_method}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws, r_eq_method, 7, eq_mult, font=BLUE_FONT, fmt=FMT_RATIO, border=THIN_BORDER)
    set_cell(ws, r_eq_method, 8, f"=F{r_eq_method}*G{r_eq_method}",
             font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # Total Segment EV
    r_total_ev = r_eq_method + 1
    set_cell(ws, r_total_ev, 2, "Total Segment EV", font=BOLD_FONT, fill=SUBTOTAL_FILL, border=TOP_BOTTOM)
    ev_refs = [f"H{seg_rows[seg['key']]}" for seg in segments] + [f"H{r_eq_method}"]
    set_cell(ws, r_total_ev, 8, f"={'+'.join(ev_refs)}",
             font=BOLD_FONT, fmt=FMT_YEN, fill=SUBTOTAL_FILL, border=TOP_BOTTOM)
    for col in range(3, 8):
        set_cell(ws, r_total_ev, col, "", fill=SUBTOTAL_FILL, border=TOP_BOTTOM)

    # ── Consolidated D&A (reference cell) ──
    r_da = r_total_ev + 2
    set_cell(ws, r_da, 2, consol["da_total"], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws, r_da, 3, "← Consolidated D&A (input)", font=GREY_FONT)
    # Fix the da_total_cell reference used above
    # We need to update it — the actual cell is B{r_da}
    # Since we already wrote formulas referencing da_total_cell, let's verify
    expected_da_cell = f"B{r + n_seg + 3}"
    actual_da_row = r_da
    # If mismatch, we need to adjust. Let's place it exactly where expected.
    # Actually, let's recalculate: r + n_seg + 3 = r + n_seg + 3
    # r_eq_method = r + n_seg, r_total_ev = r + n_seg + 1, r_da = r + n_seg + 3
    # So B{r + n_seg + 3} = B{r_da} ✓

    # ── EV Composition ── (for reference)
    r_comp = r_da + 2
    set_cell(ws, r_comp, 2, "EV Composition %", font=SUB_FONT)
    r_comp += 1
    for seg in segments:
        key = seg["key"]
        set_cell(ws, r_comp, 2, f"  {seg['label']}", font=BLACK_FONT)
        set_cell(ws, r_comp, 3, f"=H{seg_rows[key]}/H{r_total_ev}",
                 font=BLACK_FONT, fmt=FMT_PCT)
        r_comp += 1

    # ── Equity Bridge ──
    r_bridge = r_comp + 1
    set_cell(ws, r_bridge, 2, "Equity Bridge", font=SUB_FONT, fill=LIGHT_FILL)
    ws.merge_cells(start_row=r_bridge, start_column=2, end_row=r_bridge, end_column=3)
    for col in [2, 3]:
        ws.cell(row=r_bridge, column=col).fill = LIGHT_FILL
    r_bridge += 1

    # Total EV (link)
    r_tev = r_bridge
    set_cell(ws, r_tev, 2, "Total Enterprise Value", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_tev, 3, f"=H{r_total_ev}", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    r_bridge += 1

    # Conglomerate discount
    r_disc = r_bridge
    discount_base = sotp.get("conglomerate_discount", {}).get("base", 0.0)
    set_cell(ws, r_disc, 2, "Conglomerate Discount", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_disc, 3, discount_base, font=BLUE_FONT, fmt=FMT_PCT, border=INPUT_BORDER,
             alignment=CENTER)
    r_bridge += 1

    # Adjusted EV
    r_adj_ev = r_bridge
    set_cell(ws, r_adj_ev, 2, "Adjusted EV", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_adj_ev, 3, f"=C{r_tev}*(1-C{r_disc})", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    r_bridge += 1

    # Less: Net Debt
    r_nd = r_bridge
    set_cell(ws, r_nd, 2, "Less: Net Debt", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_nd, 3, consol["net_debt"], font=BLUE_FONT, fmt=FMT_YEN, border=INPUT_BORDER)
    r_bridge += 1

    # Less: Minority Interest
    r_mi = r_bridge
    set_cell(ws, r_mi, 2, "Less: Minority Interest", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_mi, 3, consol["minority_interest"], font=BLUE_FONT, fmt=FMT_YEN, border=INPUT_BORDER)
    r_bridge += 1

    # Equity Value
    r_eq = r_bridge
    set_cell(ws, r_eq, 2, "Equity Value", font=BOLD_FONT, fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
    set_cell(ws, r_eq, 3, f"=C{r_adj_ev}-C{r_nd}-C{r_mi}",
             font=BOLD_FONT, fmt=FMT_YEN, fill=SUBTOTAL_FILL, border=SUBTOTAL_BORDER)
    r_bridge += 1

    # Shares outstanding
    r_shares = r_bridge
    split_ratio = sotp.get("stock_split_ratio", 1)
    set_cell(ws, r_shares, 2, "Shares Outstanding (thousands)", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_shares, 3, consol["shares_outstanding"], font=BLUE_FONT, fmt=FMT_INT, border=INPUT_BORDER)
    r_bridge += 1

    # Stock split ratio
    r_split = r_bridge
    set_cell(ws, r_split, 2, "Stock Split Ratio", font=BLACK_FONT, border=THIN_BORDER)
    set_cell(ws, r_split, 3, split_ratio, font=BLUE_FONT, fmt=FMT_INT, border=INPUT_BORDER)
    r_bridge += 1

    # Fair Value Per Share (post-split)
    # Equity Value (¥M) × 1,000,000 / (shares × 1,000) / split_ratio
    # = Equity Value × 1000 / shares / split_ratio
    r_fv = r_bridge
    set_cell(ws, r_fv, 2, "Fair Value Per Share (post-split)", font=BOLD_FONT,
             fill=LIGHT_BLUE_BG, border=TOP_BOTTOM)
    # Formula: Equity Value (in ¥M) * 1000000 / (shares_outstanding * 1000) / split_ratio
    # = C{r_eq} * 1000 / C{r_shares} / C{r_split}
    set_cell(ws, r_fv, 3,
             f"=C{r_eq}*1000/C{r_shares}/C{r_split}",
             font=Font(name="Arial", size=12, bold=True, color="000080"),
             fmt=FMT_PRICE, fill=LIGHT_BLUE_BG, border=TOP_BOTTOM)

    # Store key cell references for other sheets
    return {
        "seg_rows": seg_rows,
        "r_total_ev": r_total_ev,
        "r_da_total": r_da,
        "r_disc": r_disc,
        "r_tev": r_tev,
        "r_adj_ev": r_adj_ev,
        "r_nd": r_nd,
        "r_mi": r_mi,
        "r_eq": r_eq,
        "r_shares": r_shares,
        "r_split": r_split,
        "r_fv": r_fv,
        "n_seg": n_seg,
        "seg_ev_col": 8,  # column H
    }


# =====================================================================
# SHEET 5: SENSITIVITY
# =====================================================================
def build_sensitivity_sheet(wb, sotp, val_refs):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "FF6600"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38

    set_cell(ws, 1, 2, "Sensitivity Analysis", font=TITLE_FONT)

    sens = sotp.get("sensitivity", {})
    cong_disc = sotp.get("conglomerate_discount", {})
    segments = sotp["segments"]
    consol = sotp["consolidated"]
    seg_rows = val_refs["seg_rows"]

    # ── TABLE 1: Primary Segment Multiple × Conglomerate Discount ──
    r = 3
    primary_key = sens.get("primary_segment_key", "aero")
    primary_label = ""
    for seg in segments:
        if seg["key"] == primary_key:
            primary_label = seg["label"]
            break

    primary_mults = sens.get("primary_multiples", [14, 16, 18, 20, 22, 25])
    disc_range = cong_disc.get("sensitivity_range", [0.0, 0.05, 0.10, 0.15, 0.20])

    set_cell(ws, r, 2, f"Table 1: {primary_label} Multiple × Conglomerate Discount", font=SUB_FONT)
    set_cell(ws, r + 1, 2, "Fair Value (¥/share, post-split)", font=GREY_FONT)
    r += 2

    # Input cells for the sensitivity calculation
    # We'll place the base assumptions in a reference area
    r_inputs = r
    set_cell(ws, r, 2, "Base Assumptions (do not edit)", font=GREY_FONT)
    r += 1

    # Place all segment OPs and D&A allocations as input references
    input_start = r
    for i, seg in enumerate(segments):
        key = seg["key"]
        set_cell(ws, r, 2, f"{seg['label']} OP", font=GREY_FONT)
        set_cell(ws, r, 3, f"='SOTP Valuation'!C{seg_rows[key]}", font=GREEN_FONT, fmt=FMT_YEN)
        set_cell(ws, r, 4, f"='SOTP Valuation'!E{seg_rows[key]}", font=GREEN_FONT, fmt=FMT_YEN)
        # EBITDA
        set_cell(ws, r, 5, f"=C{r}+D{r}", font=BLACK_FONT, fmt=FMT_YEN)
        # Base multiple
        set_cell(ws, r, 6, f"='SOTP Valuation'!G{seg_rows[key]}", font=GREEN_FONT, fmt=FMT_RATIO)
        r += 1

    # Equity method
    r_eq_method_sens = r
    set_cell(ws, r, 2, "Equity Method EV", font=GREY_FONT)
    r_eq_method_val = val_refs["seg_rows"][segments[0]["key"]] + len(segments)  # r_eq_method row in SOTP sheet
    # Actually need the equity method row from SOTP — it's n_seg rows after first seg row
    eq_method_sotp_row = list(seg_rows.values())[0] + len(segments)
    set_cell(ws, r, 3, f"='SOTP Valuation'!H{eq_method_sotp_row}", font=GREEN_FONT, fmt=FMT_YEN)
    r += 1

    # Net debt and minority
    r_nd_sens = r
    set_cell(ws, r, 2, "Net Debt", font=GREY_FONT)
    set_cell(ws, r, 3, f"='SOTP Valuation'!C{val_refs['r_nd']}", font=GREEN_FONT, fmt=FMT_YEN)
    r += 1
    r_mi_sens = r
    set_cell(ws, r, 2, "Minority Interest", font=GREY_FONT)
    set_cell(ws, r, 3, f"='SOTP Valuation'!C{val_refs['r_mi']}", font=GREEN_FONT, fmt=FMT_YEN)
    r += 1
    r_shares_sens = r
    set_cell(ws, r, 2, "Shares (thousands)", font=GREY_FONT)
    set_cell(ws, r, 3, f"='SOTP Valuation'!C{val_refs['r_shares']}", font=GREEN_FONT, fmt=FMT_INT)
    r += 1
    r_split_sens = r
    set_cell(ws, r, 2, "Split Ratio", font=GREY_FONT)
    set_cell(ws, r, 3, f"='SOTP Valuation'!C{val_refs['r_split']}", font=GREEN_FONT, fmt=FMT_INT)
    r += 2

    # ── Sensitivity Table 1 ──
    table1_row = r
    set_cell(ws, r, 2, f"{primary_label} EV/EBITDA →", font=BOLD_FONT)
    n_mults = len(primary_mults)
    n_discs = len(disc_range)

    # Column headers (multiples)
    for j, mult in enumerate(primary_mults):
        set_cell(ws, r, 3 + j, mult, font=HEADER_FONT, fmt=FMT_RATIO,
                 fill=HEADER_FILL, alignment=CENTER)
        ws.column_dimensions[cl(3 + j)].width = 14
    r += 1

    # Find the primary segment's index in the input area
    primary_idx = None
    for idx, seg in enumerate(segments):
        if seg["key"] == primary_key:
            primary_idx = idx
            break

    primary_ebitda_cell = f"$E${input_start + primary_idx}"

    # Build references for non-primary segment EVs (they stay fixed)
    other_ev_parts = []
    for idx, seg in enumerate(segments):
        if seg["key"] != primary_key:
            ebitda_cell = f"E{input_start + idx}"
            mult_cell = f"F{input_start + idx}"
            other_ev_parts.append(f"{ebitda_cell}*{mult_cell}")
    # Add equity method EV
    other_ev_parts.append(f"C{r_eq_method_sens}")
    other_ev_formula = "+".join(other_ev_parts)

    for i, disc in enumerate(disc_range):
        row = r + i
        set_cell(ws, row, 2, disc, font=BOLD_FONT, fmt=FMT_PCT,
                 fill=LIGHT_FILL, alignment=CENTER)
        for j, mult in enumerate(primary_mults):
            col = 3 + j
            # Fair Value = ((primary_EBITDA * mult + other_EVs) * (1 - disc) - net_debt - minority) * 1000 / shares / split
            formula = (
                f"=(({primary_ebitda_cell}*{mult}+{other_ev_formula})"
                f"*(1-$B{row})"
                f"-$C${r_nd_sens}-$C${r_mi_sens})"
                f"*1000/$C${r_shares_sens}/$C${r_split_sens}"
            )
            is_base = (mult == sotp["segments"][primary_idx]["selected_multiple"]
                       and disc == cong_disc.get("base", 0.0))
            fill = LIGHT_BLUE_BG if is_base else None
            set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=FMT_PRICE,
                     border=THIN_BORDER, fill=fill, alignment=CENTER)

    r += n_discs + 2

    # ── TABLE 2: Two-segment sensitivity ──
    t2 = sens.get("table2", {})
    if t2:
        row_key = t2.get("row_segment_key", "industrial")
        col_key = t2.get("col_segment_key", "energy")
        row_mults = t2.get("row_multiples", [7, 8, 10, 12, 14])
        col_mults = t2.get("col_multiples", [6, 8, 10, 12])

        row_label = ""
        col_label = ""
        row_idx = col_idx = None
        for idx, seg in enumerate(segments):
            if seg["key"] == row_key:
                row_label = seg["label"]
                row_idx = idx
            if seg["key"] == col_key:
                col_label = seg["label"]
                col_idx = idx

        set_cell(ws, r, 2, f"Table 2: {row_label} × {col_label} Multiple Sensitivity", font=SUB_FONT)
        set_cell(ws, r + 1, 2, "Fair Value (¥/share, post-split) — Base discount, base Aero multiple",
                 font=GREY_FONT)
        r += 2

        # Column headers
        set_cell(ws, r, 2, f"{row_label} ↓ / {col_label} →", font=BOLD_FONT)
        for j, cm in enumerate(col_mults):
            set_cell(ws, r, 3 + j, cm, font=HEADER_FONT, fmt=FMT_RATIO,
                     fill=HEADER_FILL, alignment=CENTER)
        r += 1

        # For table 2, primary segment and all other segments except row_key and col_key stay fixed
        fixed_ev_parts = []
        for idx, seg in enumerate(segments):
            if seg["key"] not in (row_key, col_key):
                ebitda_cell = f"$E${input_start + idx}"
                mult_cell = f"$F${input_start + idx}"
                fixed_ev_parts.append(f"{ebitda_cell}*{mult_cell}")
        fixed_ev_parts.append(f"$C${r_eq_method_sens}")
        fixed_ev_formula = "+".join(fixed_ev_parts)

        row_ebitda_cell = f"$E${input_start + row_idx}"
        col_ebitda_cell = f"$E${input_start + col_idx}"
        base_disc = cong_disc.get("base", 0.0)

        for i, rm in enumerate(row_mults):
            row_n = r + i
            set_cell(ws, row_n, 2, rm, font=BOLD_FONT, fmt=FMT_RATIO,
                     fill=LIGHT_FILL, alignment=CENTER)
            for j, cm in enumerate(col_mults):
                col_n = 3 + j
                # Fair Value with these two multiples varied, everything else fixed
                formula = (
                    f"=(({row_ebitda_cell}*$B{row_n}+{col_ebitda_cell}*{cl(col_n)}${r-1}"
                    f"+{fixed_ev_formula})"
                    f"*(1-{base_disc})"
                    f"-$C${r_nd_sens}-$C${r_mi_sens})"
                    f"*1000/$C${r_shares_sens}/$C${r_split_sens}"
                )
                # Check if base case
                is_base = (rm == segments[row_idx]["selected_multiple"]
                           and cm == segments[col_idx]["selected_multiple"])
                fill = LIGHT_BLUE_BG if is_base else None
                set_cell(ws, row_n, col_n, formula, font=BLACK_FONT, fmt=FMT_PRICE,
                         border=THIN_BORDER, fill=fill, alignment=CENTER)


# =====================================================================
# SHEET 6: D&A ALLOCATION
# =====================================================================
def build_da_allocation_sheet(wb, sotp, val_refs):
    ws = wb.create_sheet("D&A Allocation")
    ws.sheet_properties.tabColor = "808080"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    set_cell(ws, 1, 2, "D&A Allocation Methodology", font=TITLE_FONT)

    segments = sotp["segments"]
    consol = sotp["consolidated"]

    r = 3
    set_cell(ws, r, 2, "IHI does not disclose segment-level D&A.", font=BLACK_FONT)
    r += 1
    set_cell(ws, r, 2, "Allocation is estimated based on fixed asset intensity by segment.", font=BLACK_FONT)
    r += 2

    # Consolidated D&A
    set_cell(ws, r, 2, "Consolidated D&A (FY25E)", font=BOLD_FONT)
    set_cell(ws, r, 3, consol["da_total"], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws, r, 4, "¥M (input)", font=GREY_FONT)
    r_da_total = r
    r += 2

    # Allocation table
    header_row(ws, r, 2, ["Segment", "Allocation %", "D&A (¥M)", "Check"])
    r += 1

    alloc_rows = []
    for seg in segments:
        set_cell(ws, r, 2, seg["label"], font=BLACK_FONT, border=THIN_BORDER)
        set_cell(ws, r, 3, seg["da_allocation_pct"], font=BLUE_FONT, fmt=FMT_PCT, border=INPUT_BORDER)
        set_cell(ws, r, 4, f"=$C${r_da_total}*C{r}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
        alloc_rows.append(r)
        r += 1

    # Total check row
    set_cell(ws, r, 2, "Total", font=BOLD_FONT, fill=SUBTOTAL_FILL, border=TOP_BOTTOM)
    alloc_pct_refs = [f"C{ar}" for ar in alloc_rows]
    set_cell(ws, r, 3, f"={'+'.join(alloc_pct_refs)}",
             font=BOLD_FONT, fmt=FMT_PCT, fill=SUBTOTAL_FILL, border=TOP_BOTTOM)
    alloc_da_refs = [f"D{ar}" for ar in alloc_rows]
    set_cell(ws, r, 4, f"={'+'.join(alloc_da_refs)}",
             font=BOLD_FONT, fmt=FMT_YEN, fill=SUBTOTAL_FILL, border=TOP_BOTTOM)
    # Check = should equal 100%
    set_cell(ws, r, 5, f'=IF(ABS(C{r}-1)<0.001,"OK","CHECK")',
             font=BOLD_FONT, fill=SUBTOTAL_FILL, border=TOP_BOTTOM)

    r += 2
    set_cell(ws, r, 2, "Notes:", font=SUB_FONT)
    r += 1
    notes = [
        "• Aero segment has highest capital intensity (engine test facilities, MRO hangars)",
        "• Energy segment includes boiler/turbine manufacturing → moderate fixed assets",
        "• Industrial includes turbocharger plants (Niigata, overseas) → moderate",
        "• Social Infrastructure: project-based, lower owned fixed assets",
        "• Allocation sums to 100% — residual is implicitly in corporate/eliminations",
    ]
    for note in notes:
        set_cell(ws, r, 2, note, font=GREY_FONT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1


# =====================================================================
# MAIN ENTRY POINT
# =====================================================================
def generate_sotp_excel(sotp, output_path):
    """Generate the 6-sheet SOTP Excel workbook."""
    wb = openpyxl.Workbook()

    # Sheet 1: Cover & Thesis
    build_cover_sheet(wb, sotp)

    # Sheet 2: Segment Data
    seg_data_rows = build_segment_data_sheet(wb, sotp)

    # Sheet 3: Peer Comps
    peer_mult_cells = build_peer_comps_sheet(wb, sotp)

    # Sheet 4: SOTP Valuation
    val_refs = build_sotp_valuation_sheet(wb, sotp, seg_data_rows, peer_mult_cells)

    # Patch Cover sheet: link SOTP base case to the dynamic fair value row
    ws_cover = wb["Cover & Thesis"]
    r_fv = val_refs["r_fv"]
    ws_cover.cell(row=12, column=3).value = f"='SOTP Valuation'!C{r_fv}"

    # Sheet 5: Sensitivity
    build_sensitivity_sheet(wb, sotp, val_refs)

    # Sheet 6: D&A Allocation
    build_da_allocation_sheet(wb, sotp, val_refs)

    wb.save(output_path)
    print(f"Generated SOTP workbook: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
