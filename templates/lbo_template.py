"""
lbo_build.py - LBO Analysis Excel Generator for KFC Holdings Japan (9873)
Edit the config dict below, then run:  python lbo_build.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import subprocess, sys, os

# =====================================================================
# CONFIG — Edit this section for each deal
# =====================================================================
config = {
    "company_name": "KFC Holdings Japan, Ltd.",
    "ticker": "9873",
    "sector": "Quick-Service Restaurant (QSR)",
    "brand": "Kentucky Fried Chicken",
    "franchisor": "Yum! Brands, Inc.",
    "major_shareholder": "Mitsubishi Corporation 35.12%",
    "acquirer": "The Carlyle Group",
    "announcement_date": "May 20, 2024",
    "closing_date": "September 2024",
    "fye": "March 31",
    "units": "JPY millions (March 31 fiscal year-end)",
    "prepared_by": "Ryosuke Sato",
    "github_url": "github.com/Ryosuke0369/ryosuke-japanese-equity-research",

    "tob_price": 6500,
    "pre_tob_price": 5398,
    "shares_outstanding": 22354,  # thousands
    "existing_debt": 2845,
    "existing_cash": 24636,
    "transaction_fees": 3000,
    "financing_fees": 1500,

    "tla_amount": 50000,
    "tla_rate": 0.025,
    "tla_maturity": 7,
    "tla_type": "Amortizing",
    "tlb_amount": 25000,
    "tlb_rate": 0.035,
    "tlb_maturity": 7,
    "tlb_type": "Bullet",

    "hist_years": ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024"],
    "hist_revenue":    [79634, 89652, 97520, 99926, 110685],
    "hist_cogs":       [44674, 50954, 56610, 59395, 65367],
    "hist_sga":        [30173, 32344, 34802, 36909, 39454],
    "hist_da":         [1955,  1477,  1463,  1924,  2411],
    "hist_interest":   [10,    5,     5,     11,    55],
    "hist_tax":        [1854,  2316,  1952,  1757,  2646],
    "hist_net_income": [2923,  4033,  4151,  1854,  3163],
    "hist_capex":      [1867,  1740,  1853,  4360,  2550],
    "hist_wc_change":  [500,   800,   600,   1200,  1000],

    "proj_years": 5,
    "proj_revenue_growth": [0.04, 0.035, 0.03, 0.03, 0.03],
    "proj_cogs_pct":       [0.59, 0.588, 0.586, 0.585, 0.585],
    "proj_sga_pct":        [0.345, 0.34, 0.335, 0.33, 0.328],
    "proj_da_pct":         [0.022, 0.022, 0.021, 0.021, 0.02],
    "proj_capex_pct":      [0.025, 0.025, 0.024, 0.023, 0.023],
    "proj_tax_rate":       0.35,
    "proj_wc_change":      500,

    "bs_cash": 24636,
    "bs_ar": 12805,
    "bs_inventory": 3200,
    "bs_other_ca": 2500,
    "bs_ppe": 8500,
    "bs_intangibles": 1200,
    "bs_goodwill_pre": 0,
    "bs_other_nca": 5900,
    "bs_ap": 9800,
    "bs_accrued": 8500,
    "bs_other_cl": 3700,
    "bs_other_ncl": 3200,
    "bs_share_capital": 7769,
    "bs_retained_earnings": 23000,
    "bs_other_equity": -73,

    "dso": 42,
    "dio": 18,
    "dpo": 55,
    "other_ca_pct": 0.023,
    "accrued_pct": 0.077,

    "cash_to_balance_sheet": -4500,

    "exit_ev_ebitda": 15,
    "exit_cash_assumption": 10000,

    "sens_exit_multiples": [12, 13, 14, 15, 16, 17],
    "sens_exit_years": [3, 4, 5],

    "investment_thesis": [
        ("1. Brand Power", "KFC is a dominant QSR brand in Japan with 55+ years of history. Unique cultural tie to Christmas season drives strong seasonal demand."),
        ("2. Margin Expansion", "Operating margin improvement from ~5% to 7-8% through menu diversification (lunch/afternoon), digitization, and operational efficiency."),
        ("3. Store Expansion", "Carlyle's capital enables accelerated store openings in untapped domestic markets across Japan."),
        ("4. Exit Optionality", "Multiple exit paths: IPO re-listing, strategic sale (Yum! Brands direct acquisition), or secondary sale to another PE fund."),
        ("5. Downside Protection", "Stable cash flow generation, low pre-deal leverage, and strong brand equity provide resilience."),
    ],
}

# =====================================================================
# STYLE CONSTANTS (unified with dcf_comps_build.py)
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="0000CC", bold=False)
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
GREEN_FONT  = Font(name="Arial", size=10, color="006600")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
GREY_FONT   = Font(name="Arial", size=9, italic=True, color="808080")
CHECK_FONT  = Font(name="Arial", size=10, color="006600", bold=True)

HEADER_FILL  = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
LIGHT_FILL   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL   = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

THIN_BORDER   = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))
BOTTOM_DOUBLE = Border(bottom=Side(style="double"))
TOP_BOTTOM    = Border(top=Side(style="thin"), bottom=Side(style="double"))

FMT_YEN     = '#,##0;(#,##0)'
FMT_YEN_DEC = '#,##0.0;(#,##0.0)'
FMT_PCT     = '0.0%;(0.0%)'
FMT_PCT2    = '0.00%;(0.00%)'
FMT_RATIO   = '0.00"x"'
FMT_INT     = '#,##0'
FMT_MULT    = '0.0"x"'

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c

def header_row(ws, row, col_start, col_end, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c

def fill_row(ws, row, col_start, col_end, fill=LIGHT_FILL):
    for cc in range(col_start, col_end + 1):
        ws.cell(row=row, column=cc).fill = fill

def CL(col_num):
    return get_column_letter(col_num)

# =====================================================================
# PYTHON CALCULATORS (for sensitivity hardcoding)
# =====================================================================
def calc_projections(cfg):
    """Calculate projected financials."""
    n = cfg["proj_years"]
    rev = cfg["hist_revenue"][-1]
    results = []
    tla = cfg["tla_amount"]
    tla_annual = tla / cfg["tla_maturity"]
    cum_tla = tla_annual  # Pro Forma already takes one repayment

    for yr in range(n):
        rev = rev * (1 + cfg["proj_revenue_growth"][yr])
        cogs = rev * cfg["proj_cogs_pct"][yr]
        gp = rev - cogs
        sga = rev * cfg["proj_sga_pct"][yr]
        ebit = gp - sga
        da = rev * cfg["proj_da_pct"][yr]
        ebitda = ebit + da
        capex = rev * cfg["proj_capex_pct"][yr]

        tla_begin = tla - cum_tla
        tla_end = tla_begin - tla_annual
        cum_tla += tla_annual

        tla_int = (tla_begin + tla_end) / 2 * cfg["tla_rate"]
        tlb_int = cfg["tlb_amount"] * cfg["tlb_rate"]
        total_int = tla_int + tlb_int

        ebt = ebit - total_int
        tax = ebt * cfg["proj_tax_rate"]
        ni = ebt - tax
        total_debt = tla_end + cfg["tlb_amount"]

        results.append({
            "year": yr + 1, "revenue": rev, "ebitda": ebitda, "ebit": ebit,
            "da": da, "interest": total_int, "ni": ni, "capex": capex,
            "total_debt": total_debt, "tla_end": tla_end,
        })
    return results

def calc_returns(exit_mult, exit_year, cfg):
    """Calculate MOIC and IRR for given exit multiple and exit year."""
    proj = calc_projections(cfg)
    eq_val = cfg["tob_price"] * cfg["shares_outstanding"] / 1000
    total_debt_new = cfg["tla_amount"] + cfg["tlb_amount"]
    sponsor = eq_val + cfg["existing_debt"] + cfg["transaction_fees"] + cfg["financing_fees"] - total_debt_new

    p = proj[exit_year - 1]
    exit_ev = p["ebitda"] * exit_mult
    exit_equity = exit_ev - p["total_debt"] + cfg["exit_cash_assumption"]
    moic = exit_equity / sponsor if sponsor > 0 else 0
    irr = moic ** (1 / exit_year) - 1 if moic > 0 else 0
    return moic, irr

# =====================================================================
# BUILD WORKBOOK
# =====================================================================
wb = openpyxl.Workbook()
C = config

output_file = "9873_KFC_Japan_LBO.xlsx"

# Pre-compute key values
equity_value = C["tob_price"] * C["shares_outstanding"] / 1000
enterprise_value = equity_value + C["existing_debt"] - C["existing_cash"]
total_new_debt = C["tla_amount"] + C["tlb_amount"]
total_uses = equity_value + C["existing_debt"] + C["transaction_fees"] + C["financing_fees"]
sponsor_equity = total_uses - total_new_debt
pre_deal_equity = C["bs_share_capital"] + C["bs_retained_earnings"] + C["bs_other_equity"]
goodwill_plug = equity_value - pre_deal_equity
proj = calc_projections(C)

n_hist = len(C["hist_years"])
n_proj = C["proj_years"]
hist_start = 4   # col D
proj_start = 10  # col J (I is spacer)

# =====================================================================
# SHEET 1: Cover
# =====================================================================
ws1 = wb.active
ws1.title = "Cover"
ws1.sheet_properties.tabColor = "003366"

ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 40
ws1.column_dimensions["C"].width = 5
ws1.column_dimensions["D"].width = 30
ws1.column_dimensions["E"].width = 25

set_cell(ws1, 1, 2,
    "DISCLAIMER: This is a sample analysis for demonstration purposes only. "
    "It does not constitute investment advice.",
    font=GREY_FONT)
ws1.merge_cells("B1:E1")

set_cell(ws1, 3, 2, "CONFIDENTIAL", font=Font(name="Arial", size=12, bold=True, color="CC0000"))
set_cell(ws1, 5, 2, "Leveraged Buyout Analysis", font=TITLE_FONT)
set_cell(ws1, 7, 2, f'{C["company_name"]} ({C["ticker"]})', font=SUB_FONT)
set_cell(ws1, 8, 2, f'{C["brand"]} \u2014 {C["sector"]}', font=BLACK_FONT)
set_cell(ws1, 9, 2, f'Acquirer: {C["acquirer"]}', font=BLACK_FONT)

ltm_ebitda = C["hist_revenue"][-1] - C["hist_cogs"][-1] - C["hist_sga"][-1] + C["hist_da"][-1]

set_cell(ws1, 11, 2, "Transaction Overview", font=SUB_FONT)
info = [
    ("Announcement Date", C["announcement_date"]),
    ("Closing Date", C["closing_date"]),
    ("TOB Price", f'\u00a5{C["tob_price"]:,}/share'),
    ("Equity Value", f'\u00a5{equity_value:,.0f} mn'),
    ("Enterprise Value", f'\u00a5{enterprise_value:,.0f} mn'),
    ("EV/LTM EBITDA", f'{enterprise_value / ltm_ebitda:.1f}x'),
    ("Sponsor Equity", f'\u00a5{sponsor_equity:,.0f} mn'),
]
for i, (lbl, val) in enumerate(info):
    set_cell(ws1, 12 + i, 2, lbl, font=BOLD_FONT)
    set_cell(ws1, 12 + i, 4, val, font=BLACK_FONT)

set_cell(ws1, 20, 2, f'Prepared by: {C["prepared_by"]}', font=GREY_FONT)
set_cell(ws1, 21, 2, C["units"], font=GREY_FONT)
set_cell(ws1, 22, 2, C["github_url"], font=GREY_FONT)

# =====================================================================
# SHEET 2: Transaction Assumptions
# =====================================================================
ws2 = wb.create_sheet("Transaction Assumptions")
ws2.sheet_properties.tabColor = "003366"

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 5
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 3
ws2.column_dimensions["G"].width = 35
ws2.column_dimensions["H"].width = 18
ws2.column_dimensions["I"].width = 14

set_cell(ws2, 2, 2, "Transaction Assumptions", font=TITLE_FONT)

# ── Transaction Structure ──
c = section_title(ws2, 4, 2, "Transaction Structure")
c.fill = LIGHT_FILL
fill_row(ws2, 4, 3, 5)

set_cell(ws2, 6, 2, "TOB Price (JPY/share)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 6, 4, C["tob_price"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 7, 2, "Pre-TOB Closing Price (JPY/share)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 7, 4, C["pre_tob_price"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 8, 2, "Acquisition Premium", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 8, 4, "=(D6-D7)/D7", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 10, 2, "Shares Outstanding (thousands)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 10, 4, C["shares_outstanding"], font=BLUE_FONT, fmt=FMT_INT, border=THIN_BORDER)

set_cell(ws2, 11, 2, "Equity Value (JPY mm)", font=BOLD_FONT, border=TOP_BOTTOM)
set_cell(ws2, 11, 4, "=D6*D10/1000", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

set_cell(ws2, 13, 2, "(+) Total Debt (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 13, 4, C["existing_debt"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 14, 2, "(-) Cash & Equivalents (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 14, 4, C["existing_cash"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 15, 2, "Enterprise Value (JPY mm)", font=BOLD_FONT, border=TOP_BOTTOM)
set_cell(ws2, 15, 4, "=D11+D13-D14", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

# ── Valuation Multiples ──
c = section_title(ws2, 17, 2, "Valuation Multiples")
c.fill = LIGHT_FILL
fill_row(ws2, 17, 3, 5)

set_cell(ws2, 19, 2, "LTM EBITDA (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 19, 4, ltm_ebitda, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 20, 2, "EV / LTM EBITDA", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 20, 4, "=D15/D19", font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER)

ltm_ebit = C["hist_revenue"][-1] - C["hist_cogs"][-1] - C["hist_sga"][-1]
set_cell(ws2, 21, 2, "LTM EBIT (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 21, 4, ltm_ebit, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 22, 2, "EV / LTM EBIT", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 22, 4, "=D15/D21", font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER)

set_cell(ws2, 23, 2, "LTM Revenue (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 23, 4, C["hist_revenue"][-1], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 24, 2, "EV / LTM Revenue", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 24, 4, "=D15/D23", font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER)

# ── Sources & Uses (side by side) ──
c = section_title(ws2, 26, 2, "Sources & Uses")
c.fill = LIGHT_FILL
fill_row(ws2, 26, 3, 5)
ws2.cell(row=26, column=7).fill = LIGHT_FILL
fill_row(ws2, 26, 8, 9)

header_row(ws2, 27, 2, 5, ["Sources", "", "Amount", "% Total"])
header_row(ws2, 27, 7, 9, ["Uses", "Amount", "% Total"])

set_cell(ws2, 28, 2, "Term Loan A", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 28, 4, C["tla_amount"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 28, 5, "=D28/D31", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 29, 2, "Term Loan B", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 29, 4, C["tlb_amount"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 29, 5, "=D29/D31", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 30, 2, "Sponsor Equity", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 30, 4, "=H35-D28-D29", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 30, 5, "=D30/D31", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 31, 2, "Total Sources", font=BOLD_FONT, border=TOP_BOTTOM)
set_cell(ws2, 31, 4, "=SUM(D28:D30)", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
set_cell(ws2, 31, 5, "=SUM(E28:E30)", font=BLACK_FONT, fmt=FMT_PCT, border=TOP_BOTTOM)

set_cell(ws2, 28, 7, "Equity Purchase", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 28, 8, "=D11", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 28, 9, "=H28/H35", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 29, 7, "Refinance Existing Debt", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 29, 8, "=D13", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 29, 9, "=H29/H35", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 30, 7, "Transaction Fees", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 30, 8, C["transaction_fees"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 30, 9, "=H30/H35", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 31, 7, "Financing Fees", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 31, 8, C["financing_fees"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 31, 9, "=H31/H35", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 35, 7, "Total Uses", font=BOLD_FONT, border=TOP_BOTTOM)
set_cell(ws2, 35, 8, "=SUM(H28:H31)", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
set_cell(ws2, 35, 9, "=SUM(I28:I31)", font=BLACK_FONT, fmt=FMT_PCT, border=TOP_BOTTOM)

# ── Debt Terms ──
c = section_title(ws2, 38, 2, "Debt Terms")
c.fill = LIGHT_FILL
fill_row(ws2, 38, 3, 5)

header_row(ws2, 39, 2, 5, ["", "", "Term Loan A", "Term Loan B"])

set_cell(ws2, 40, 2, "Amount (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 40, 4, "=D28", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws2, 40, 5, "=D29", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws2, 41, 2, "Interest Rate", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 41, 4, C["tla_rate"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
set_cell(ws2, 41, 5, C["tlb_rate"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 42, 2, "Maturity (years)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 42, 4, C["tla_maturity"], font=BLUE_FONT, fmt='0', border=THIN_BORDER)
set_cell(ws2, 42, 5, C["tlb_maturity"], font=BLUE_FONT, fmt='0', border=THIN_BORDER)

set_cell(ws2, 43, 2, "Repayment Type", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 43, 4, C["tla_type"], font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws2, 43, 5, C["tlb_type"], font=BLACK_FONT, border=THIN_BORDER)

set_cell(ws2, 45, 2, "Total Leverage (Debt/EBITDA)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 45, 4, "=(D28+D29)/D19", font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER)

set_cell(ws2, 46, 2, "Equity / Total Capitalization", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 46, 4, "=D30/D31", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

set_cell(ws2, 47, 2, "Cash to Balance Sheet (JPY mm)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws2, 47, 4, abs(C["cash_to_balance_sheet"]), font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# =====================================================================
# SHEET 3: Income Statement
# =====================================================================
ws3 = wb.create_sheet("Income Statement")
ws3.sheet_properties.tabColor = "006600"

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 5
for i in range(n_hist):
    ws3.column_dimensions[CL(hist_start + i)].width = 14
ws3.column_dimensions["I"].width = 2
for i in range(n_proj):
    ws3.column_dimensions[CL(proj_start + i)].width = 14

set_cell(ws3, 2, 2, f'{C["company_name"]} \u2014 Income Statement ({C["units"]})', font=TITLE_FONT)

header_row(ws3, 4, hist_start, hist_start + n_hist - 1, C["hist_years"])
header_row(ws3, 4, proj_start, proj_start + n_proj - 1, [f"FY{2025+i}E" for i in range(n_proj)])

for i in range(n_hist):
    set_cell(ws3, 5, hist_start + i, "Actual", font=Font(name="Arial", size=9, italic=True),
             alignment=Alignment(horizontal="center"))
for i in range(n_proj):
    set_cell(ws3, 5, proj_start + i, "Projected", font=Font(name="Arial", size=9, italic=True, color="006600"),
             alignment=Alignment(horizontal="center"))

# Row labels
for r, lbl in [(7,"Revenue"),(8,"Revenue Growth"),(10,"COGS"),(11,"COGS % Revenue"),
               (13,"Gross Profit"),(14,"Gross Margin"),(16,"SG&A"),(17,"SG&A % Revenue"),
               (19,"Operating Income (EBIT)"),(20,"Operating Margin"),
               (22,"D&A"),(23,"D&A % Revenue"),(25,"EBITDA"),(26,"EBITDA Margin"),
               (28,"Interest Expense"),(30,"EBT"),(32,"Income Taxes"),(33,"Effective Tax Rate"),
               (35,"Net Income"),(36,"Net Margin"),
               (39,"Free Cash Flow Summary"),(40,"EBITDA"),(41,"(-) Capex"),
               (42,"(-) Cash Taxes"),(43,"(-) Changes in WC"),(44,"Levered Free Cash Flow"),
               (46,"Capex"),(47,"Capex % Revenue")]:
    margin_rows = [8,11,14,17,20,23,26,33,36,47]
    set_cell(ws3, r, 2, lbl, font=BLACK_FONT if r in margin_rows else BOLD_FONT)

# Historical data
for r, data in [(7, C["hist_revenue"]), (10, C["hist_cogs"]), (16, C["hist_sga"]),
                (22, C["hist_da"]), (28, C["hist_interest"]), (32, C["hist_tax"]),
                (35, C["hist_net_income"]), (46, C["hist_capex"])]:
    for i, val in enumerate(data):
        set_cell(ws3, r, hist_start + i, val, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# Historical formulas
for i in range(n_hist):
    col = hist_start + i
    cl = CL(col)
    set_cell(ws3, 13, col, f"={cl}7-{cl}10", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 19, col, f"={cl}13-{cl}16", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 25, col, f"={cl}19+{cl}22", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 30, col, f"={cl}19-{cl}28", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    if i > 0:
        set_cell(ws3, 8, col, f"=({cl}7-{CL(col-1)}7)/{CL(col-1)}7", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    else:
        set_cell(ws3, 8, col, "n/a", font=GREY_FONT, border=THIN_BORDER)

    for mr, num, den in [(11,10,7),(14,13,7),(17,16,7),(20,19,7),(23,22,7),(26,25,7),(33,32,30),(36,35,7),(47,46,7)]:
        set_cell(ws3, mr, col, f"={cl}{num}/{cl}{den}", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

    # FCF summary
    set_cell(ws3, 40, col, f"={cl}25", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 41, col, f"=-{cl}46", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 42, col, f"=-{cl}32", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 43, col, -C["hist_wc_change"][i], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 44, col, f"={cl}40+{cl}41+{cl}42+{cl}43", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# Projected
last_hist_cl = CL(hist_start + n_hist - 1)
for i in range(n_proj):
    col = proj_start + i
    cl = CL(col)
    pcl = CL(col - 1) if i > 0 else last_hist_cl
    debt_cl = CL(5 + i)

    set_cell(ws3, 7, col, f"={pcl}7*(1+{C['proj_revenue_growth'][i]})", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 8, col, C["proj_revenue_growth"][i], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 10, col, f"={cl}7*{C['proj_cogs_pct'][i]}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 11, col, C["proj_cogs_pct"][i], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 13, col, f"={cl}7-{cl}10", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 14, col, f"={cl}13/{cl}7", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 16, col, f"={cl}7*{C['proj_sga_pct'][i]}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 17, col, C["proj_sga_pct"][i], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 19, col, f"={cl}13-{cl}16", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 20, col, f"={cl}19/{cl}7", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 22, col, f"={cl}7*{C['proj_da_pct'][i]}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 23, col, C["proj_da_pct"][i], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 25, col, f"={cl}19+{cl}22", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 26, col, f"={cl}25/{cl}7", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 28, col, f"='Debt Schedule'!{debt_cl}34", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 30, col, f"={cl}19-{cl}28", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 32, col, f"={cl}30*{C['proj_tax_rate']}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 33, col, C["proj_tax_rate"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
    set_cell(ws3, 35, col, f"={cl}30-{cl}32", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 36, col, f"={cl}35/{cl}7", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

    set_cell(ws3, 40, col, f"={cl}25", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 41, col, f"=-{cl}46", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 42, col, f"=-{cl}32", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 43, col, -C["proj_wc_change"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 44, col, f"={cl}40+{cl}41+{cl}42+{cl}43", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    set_cell(ws3, 46, col, f"={cl}7*{C['proj_capex_pct'][i]}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws3, 47, col, C["proj_capex_pct"][i], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)

# =====================================================================
# SHEET 4: Transaction Summary
# =====================================================================
ws4 = wb.create_sheet("Transaction Summary")
ws4.sheet_properties.tabColor = "996600"

ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 35
ws4.column_dimensions["C"].width = 5
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 16

set_cell(ws4, 2, 2, "Transaction Summary", font=TITLE_FONT)

c = section_title(ws4, 4, 2, "Key Metrics (Cross-Reference)")
c.fill = LIGHT_FILL
fill_row(ws4, 4, 3, 5)

for r, lbl, ref, fmt in [
    (6, "Equity Value", "='Transaction Assumptions'!D11", FMT_YEN),
    (7, "Enterprise Value", "='Transaction Assumptions'!D15", FMT_YEN),
    (8, "Total New Debt", "='Transaction Assumptions'!D28+'Transaction Assumptions'!D29", FMT_YEN),
    (9, "Sponsor Equity", "='Transaction Assumptions'!D30", FMT_YEN),
    (10, "EV / LTM EBITDA", "='Transaction Assumptions'!D20", FMT_MULT),
    (11, "Total Leverage", "='Transaction Assumptions'!D45", FMT_MULT),
    (12, "Equity / Total Cap", "='Transaction Assumptions'!D46", FMT_PCT),
]:
    set_cell(ws4, r, 2, lbl, font=BOLD_FONT, border=THIN_BORDER)
    set_cell(ws4, r, 4, ref, font=GREEN_FONT, fmt=fmt, border=THIN_BORDER)

# =====================================================================
# SHEET 7: Debt Schedule (built before BS/CF)
# =====================================================================
ws7 = wb.create_sheet("Debt Schedule")
ws7.sheet_properties.tabColor = "CC0000"

ws7.column_dimensions["A"].width = 3
ws7.column_dimensions["B"].width = 30
ws7.column_dimensions["C"].width = 5
ws7.column_dimensions["D"].width = 16
for i in range(n_proj):
    ws7.column_dimensions[CL(5 + i)].width = 14

set_cell(ws7, 2, 2, f'Debt Schedule ({C["units"]})', font=TITLE_FONT)
header_row(ws7, 4, 4, 4 + n_proj, ["Pro Forma"] + [f"Year {y}" for y in range(1, n_proj + 1)])

# TLA
c = section_title(ws7, 6, 2, "Term Loan A"); c.fill = LIGHT_FILL; fill_row(ws7, 6, 3, 4+n_proj)
set_cell(ws7, 7, 2, "Rate", font=BOLD_FONT)
set_cell(ws7, 7, 4, "='Transaction Assumptions'!D41", font=GREEN_FONT, fmt=FMT_PCT)
for r, lbl in [(9,"Beginning Balance"),(10,"Scheduled Repayment"),(11,"Optional Prepayment"),(13,"Ending Balance"),(16,"Interest Expense")]:
    set_cell(ws7, r, 2, lbl, font=BOLD_FONT, border=TOP_BOTTOM if r==13 else THIN_BORDER)

tla_annual = C["tla_amount"] / C["tla_maturity"]
set_cell(ws7, 9, 4, "='Transaction Assumptions'!D28", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws7, 10, 4, round(tla_annual), font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws7, 11, 4, 0, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws7, 13, 4, "=D9-D10-D11", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
set_cell(ws7, 16, 4, 0, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

for i in range(n_proj):
    col = 5 + i; cl = CL(col); pcl = CL(col-1)
    set_cell(ws7, 9, col, f"={pcl}13", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 10, col, round(tla_annual), font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 11, col, 0, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 13, col, f"={cl}9-{cl}10-{cl}11", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws7, 16, col, f"=({cl}9+{cl}13)/2*$D$7", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# TLB
c = section_title(ws7, 18, 2, "Term Loan B"); c.fill = LIGHT_FILL; fill_row(ws7, 18, 3, 4+n_proj)
set_cell(ws7, 19, 2, "Rate", font=BOLD_FONT)
set_cell(ws7, 19, 4, "='Transaction Assumptions'!E41", font=GREEN_FONT, fmt=FMT_PCT)
for r, lbl in [(21,"Beginning Balance"),(22,"Scheduled Repayment"),(24,"Ending Balance"),(27,"Interest Expense")]:
    set_cell(ws7, r, 2, lbl, font=BOLD_FONT, border=TOP_BOTTOM if r==24 else THIN_BORDER)

set_cell(ws7, 21, 4, "='Transaction Assumptions'!D29", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws7, 22, 4, 0, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws7, 24, 4, "=D21-D22", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
set_cell(ws7, 27, 4, 0, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

for i in range(n_proj):
    col = 5 + i; cl = CL(col); pcl = CL(col-1)
    set_cell(ws7, 21, col, f"={pcl}24", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 22, col, 0, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 24, col, f"={cl}21-{cl}22", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws7, 27, col, f"=({cl}21+{cl}24)/2*$D$19", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# Total Debt Summary
c = section_title(ws7, 29, 2, "Total Debt Summary"); c.fill = LIGHT_FILL; fill_row(ws7, 29, 3, 4+n_proj)
for r, lbl in [(30,"TLA Ending"),(31,"TLB Ending"),(32,"Total Debt"),(34,"Total Interest Expense")]:
    set_cell(ws7, r, 2, lbl, font=BOLD_FONT, border=TOP_BOTTOM if r in [32,34] else THIN_BORDER)

for col in range(4, 4+n_proj+1):
    cl = CL(col)
    set_cell(ws7, 30, col, f"={cl}13", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 31, col, f"={cl}24", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws7, 32, col, f"={cl}30+{cl}31", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws7, 34, col, f"={cl}16+{cl}27", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

# Credit Metrics
c = section_title(ws7, 36, 2, "Credit Metrics"); c.fill = LIGHT_FILL; fill_row(ws7, 36, 3, 4+n_proj)
set_cell(ws7, 38, 2, "EBITDA", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws7, 38, 4, "='Transaction Assumptions'!D19", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
for i in range(n_proj):
    set_cell(ws7, 38, 5+i, f"='Income Statement'!{CL(proj_start+i)}25", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)

for r, lbl in [(39,"Total Debt / EBITDA"),(40,"Interest Coverage (EBITDA/Int)"),(41,"Debt Paydown (cumulative)")]:
    set_cell(ws7, r, 2, lbl, font=BOLD_FONT, border=THIN_BORDER)

for col in range(4, 4+n_proj+1):
    cl = CL(col)
    set_cell(ws7, 39, col, f"={cl}32/{cl}38", font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER)
    set_cell(ws7, 40, col, f"=IF({cl}34=0,0,{cl}38/{cl}34)", font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER)

set_cell(ws7, 41, 4, 0, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
for i in range(n_proj):
    set_cell(ws7, 41, 5+i, f"=D32-{CL(5+i)}32", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

# =====================================================================
# SHEET 5: Balance Sheet
# =====================================================================
ws5 = wb.create_sheet("Balance Sheet")
ws5.sheet_properties.tabColor = "006600"

ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 30
ws5.column_dimensions["C"].width = 5
ws5.column_dimensions["D"].width = 14
ws5.column_dimensions["E"].width = 14
ws5.column_dimensions["F"].width = 14
for i in range(n_proj):
    ws5.column_dimensions[CL(7+i)].width = 14

set_cell(ws5, 2, 2, f'{C["company_name"]} \u2014 Balance Sheet ({C["units"]})', font=TITLE_FONT)
header_row(ws5, 4, 4, 4+2+n_proj, ["Pre-Deal","Adj","Pro Forma"]+[f"Year {y}" for y in range(1,n_proj+1)])

# Row labels
for r, lbl, fnt in [
    (6,"Assets",SUB_FONT),(7,"Current Assets",BOLD_FONT),
    (9,"Cash & Equivalents",BLACK_FONT),(10,"Accounts Receivable",BLACK_FONT),
    (11,"Inventory",BLACK_FONT),(12,"Other Current Assets",BLACK_FONT),
    (13,"Total Current Assets",BOLD_FONT),
    (15,"Non-Current Assets",BOLD_FONT),
    (16,"PP&E (net)",BLACK_FONT),(17,"Intangible Assets",BLACK_FONT),
    (18,"Goodwill",BLACK_FONT),(19,"Other Non-Current Assets",BLACK_FONT),
    (20,"Total Non-Current Assets",BOLD_FONT),
    (22,"Total Assets",BOLD_FONT),
    (24,"Liabilities & Equity",SUB_FONT),(25,"Current Liabilities",BOLD_FONT),
    (26,"Accounts Payable",BLACK_FONT),(27,"Accrued Expenses",BLACK_FONT),
    (28,"Other Current Liabilities",BLACK_FONT),(29,"Current Portion of LTD",BLACK_FONT),
    (30,"Total Current Liabilities",BOLD_FONT),
    (32,"Non-Current Liabilities",BOLD_FONT),
    (33,"Long-Term Debt",BLACK_FONT),(34,"Other NC Liabilities",BLACK_FONT),
    (35,"Total Non-Current Liabilities",BOLD_FONT),
    (37,"Total Liabilities",BOLD_FONT),
    (39,"Shareholders' Equity",BOLD_FONT),
    (40,"Share Capital",BLACK_FONT),(41,"Retained Earnings",BLACK_FONT),(42,"Other Equity",BLACK_FONT),
    (43,"Total Equity",BOLD_FONT),
    (45,"Total Liabilities & Equity",BOLD_FONT),
    (48,"Balance Check (Assets - L&E)",BOLD_FONT),(49,"Status",CHECK_FONT),
]:
    set_cell(ws5, r, 2, lbl, font=fnt)

# Pre-Deal (D)
for r, val in [(9,C["bs_cash"]),(10,C["bs_ar"]),(11,C["bs_inventory"]),(12,C["bs_other_ca"]),
               (16,C["bs_ppe"]),(17,C["bs_intangibles"]),(18,C["bs_goodwill_pre"]),(19,C["bs_other_nca"]),
               (26,C["bs_ap"]),(27,C["bs_accrued"]),(28,C["bs_other_cl"]),(29,0),
               (33,C["existing_debt"]),(34,C["bs_other_ncl"]),
               (40,C["bs_share_capital"]),(41,C["bs_retained_earnings"]),(42,C["bs_other_equity"])]:
    set_cell(ws5, r, 4, val, font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

for r, formula in [(13,"=SUM(D9:D12)"),(20,"=SUM(D16:D19)"),(22,"=D13+D20"),
                   (30,"=SUM(D26:D29)"),(35,"=D33+D34"),(37,"=D30+D35"),
                   (43,"=SUM(D40:D42)"),(45,"=D37+D43"),(48,"=D22-D45")]:
    set_cell(ws5, r, 4, formula, font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM if r not in [48] else THIN_BORDER)
set_cell(ws5, 49, 4, '=IF(ABS(D48)<1,"\u2713 Balance","X Imbalance")', font=CHECK_FONT, border=THIN_BORDER)

# Adjustments (E)
pf_cash = abs(C["cash_to_balance_sheet"])
cash_adj = pf_cash - C["bs_cash"]
pf_current_ltd = round(tla_annual)
pf_total_debt = total_new_debt - round(tla_annual)  # after Pro Forma repayment
new_lt_debt = pf_total_debt - pf_current_ltd
lt_debt_adj = new_lt_debt - C["existing_debt"]

for r, val in [(9,cash_adj),(10,0),(11,0),(12,0),
               (16,0),(17,0),(18,goodwill_plug),(19,0),
               (26,0),(27,0),(28,0),(29,pf_current_ltd),
               (33,lt_debt_adj),(34,0),
               (40,-C["bs_share_capital"]),(41,sponsor_equity-C["bs_retained_earnings"]),(42,-C["bs_other_equity"])]:
    set_cell(ws5, r, 5, val, font=BLUE_FONT if r==18 else BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

for r, formula in [(13,"=SUM(E9:E12)"),(20,"=SUM(E16:E19)"),(22,"=E13+E20"),
                   (30,"=SUM(E26:E29)"),(35,"=E33+E34"),(37,"=E30+E35"),
                   (43,"=SUM(E40:E42)"),(45,"=E37+E43"),(48,"=E22-E45")]:
    set_cell(ws5, r, 5, formula, font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM if r not in [48] else THIN_BORDER)
set_cell(ws5, 49, 5, '=IF(ABS(E48)<1,"\u2713 Balance","X Imbalance")', font=CHECK_FONT, border=THIN_BORDER)

# Pro Forma (F)
for r in [9,10,11,12,16,17,18,19,26,27,28,29,33,34,40,41,42]:
    set_cell(ws5, r, 6, f"=D{r}+E{r}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
for r, formula in [(13,"=SUM(F9:F12)"),(20,"=SUM(F16:F19)"),(22,"=F13+F20"),
                   (30,"=SUM(F26:F29)"),(35,"=F33+F34"),(37,"=F30+F35"),
                   (43,"=SUM(F40:F42)"),(45,"=F37+F43"),(48,"=F22-F45")]:
    set_cell(ws5, r, 6, formula, font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM if r not in [48] else THIN_BORDER)
set_cell(ws5, 49, 6, '=IF(ABS(F48)<1,"\u2713 Balance","X Imbalance")', font=CHECK_FONT, border=THIN_BORDER)

# Projected (G-K)
for i in range(n_proj):
    col = 7+i; cl = CL(col); pcl = CL(col-1)
    is_cl = CL(proj_start+i); debt_cl = CL(5+i); bs_pcl = pcl; cf_cl = CL(5+i)

    set_cell(ws5, 9, col, f"='Cash Flow Statement'!{cf_cl}47", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 10, col, f"='Income Statement'!{is_cl}7*{C['dso']}/365", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 11, col, f"='Income Statement'!{is_cl}10*{C['dio']}/365", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 12, col, f"='Income Statement'!{is_cl}7*{C['other_ca_pct']}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 13, col, f"=SUM({cl}9:{cl}12)", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws5, 16, col, f"={pcl}16+'Income Statement'!{is_cl}46-'Income Statement'!{is_cl}22", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 17, col, f"={pcl}17", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 18, col, f"={pcl}18", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 19, col, f"={pcl}19", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 20, col, f"=SUM({cl}16:{cl}19)", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws5, 22, col, f"={cl}13+{cl}20", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws5, 26, col, f"='Income Statement'!{is_cl}10*{C['dpo']}/365", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 27, col, f"='Income Statement'!{is_cl}7*{C['accrued_pct']}", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 28, col, f"={pcl}28", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 29, col, f"='Debt Schedule'!{debt_cl}10", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 30, col, f"=SUM({cl}26:{cl}29)", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws5, 33, col, f"='Debt Schedule'!{debt_cl}32-{cl}29", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 34, col, f"={pcl}34", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 35, col, f"={cl}33+{cl}34", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws5, 37, col, f"={cl}30+{cl}35", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws5, 40, col, f"={pcl}40", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 41, col, f"={pcl}41+'Income Statement'!{is_cl}35", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 42, col, f"={pcl}42", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 43, col, f"=SUM({cl}40:{cl}42)", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws5, 45, col, f"={cl}37+{cl}43", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws5, 48, col, f"={cl}22-{cl}45", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws5, 49, col, f'=IF(ABS({cl}48)<1,"\u2713 Balance","X Imbalance")', font=CHECK_FONT, border=THIN_BORDER)

# =====================================================================
# SHEET 6: Cash Flow Statement
# =====================================================================
ws6 = wb.create_sheet("Cash Flow Statement")
ws6.sheet_properties.tabColor = "006600"

ws6.column_dimensions["A"].width = 3
ws6.column_dimensions["B"].width = 30
ws6.column_dimensions["C"].width = 5
ws6.column_dimensions["D"].width = 14
for i in range(n_proj):
    ws6.column_dimensions[CL(5+i)].width = 14

set_cell(ws6, 2, 2, f'{C["company_name"]} \u2014 Cash Flow Statement ({C["units"]})', font=TITLE_FONT)
header_row(ws6, 4, 4, 4+n_proj, ["Pro Forma"]+[f"Year {y}" for y in range(1,n_proj+1)])

c = section_title(ws6, 6, 2, "Operating Cash Flow"); c.fill = LIGHT_FILL; fill_row(ws6, 6, 3, 4+n_proj)
for r, lbl in [(9,"Net Income"),(10,"Add: Interest Expense"),(11,"Add: D&A"),
               (13,"Changes in Working Capital:"),(15,"  Change in AR"),(16,"  Change in Inventory"),
               (17,"  Change in AP"),(18,"  Change in Accrued"),(19,"  Change in Other WC"),
               (20,"Total WC Change"),(22,"Total Operating Cash Flow")]:
    set_cell(ws6, r, 2, lbl, font=BOLD_FONT if r in [9,13,20,22] else BLACK_FONT,
             border=THIN_BORDER if r != 13 else None)

c = section_title(ws6, 24, 2, "Investing Cash Flow"); c.fill = LIGHT_FILL; fill_row(ws6, 24, 3, 4+n_proj)
set_cell(ws6, 27, 2, "Capital Expenditures", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws6, 28, 2, "Total Investing Cash Flow", font=BOLD_FONT, border=TOP_BOTTOM)

c = section_title(ws6, 30, 2, "Financing Cash Flow"); c.fill = LIGHT_FILL; fill_row(ws6, 30, 3, 4+n_proj)
for r, lbl in [(34,"Debt Repayment (TLA)"),(35,"Debt Repayment (TLB)"),(37,"Interest Paid"),
               (39,"Total Financing Cash Flow"),(41,"Net Change in Cash"),
               (45,"Beginning Cash"),(47,"Ending Cash")]:
    set_cell(ws6, r, 2, lbl, font=BOLD_FONT if r in [39,41,45,47] else BLACK_FONT,
             border=TOP_BOTTOM if r in [39,41,47] else THIN_BORDER)

set_cell(ws6, 52, 2, "Consistency Check", font=SUB_FONT)
set_cell(ws6, 53, 2, "Ending Cash (CF)", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws6, 54, 2, "Cash on BS", font=BLACK_FONT, border=THIN_BORDER)
set_cell(ws6, 55, 2, "Status", font=CHECK_FONT, border=THIN_BORDER)

# Pro Forma (D)
for r in [9,10,11,15,16,17,18,19,20,22,27,28,34,35,37,39,41]:
    set_cell(ws6, r, 4, 0, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws6, 45, 4, "='Balance Sheet'!F9", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
set_cell(ws6, 47, 4, "=D45+D41", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

# Projected (E-I)
for i in range(n_proj):
    col = 5+i; cl = CL(col); pcl = CL(col-1)
    is_cl = CL(proj_start+i); debt_cl = CL(5+i); bs_cl = CL(7+i); bs_pcl = CL(6+i)

    set_cell(ws6, 9, col, f"='Income Statement'!{is_cl}35", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 10, col, f"='Debt Schedule'!{debt_cl}34", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 11, col, f"='Income Statement'!{is_cl}22", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    set_cell(ws6, 15, col, f"=-('Balance Sheet'!{bs_cl}10-'Balance Sheet'!{bs_pcl}10)", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 16, col, f"=-('Balance Sheet'!{bs_cl}11-'Balance Sheet'!{bs_pcl}11)", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 17, col, f"='Balance Sheet'!{bs_cl}26-'Balance Sheet'!{bs_pcl}26", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 18, col, f"='Balance Sheet'!{bs_cl}27-'Balance Sheet'!{bs_pcl}27", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 19, col, f"=-('Balance Sheet'!{bs_cl}12-'Balance Sheet'!{bs_pcl}12)+('Balance Sheet'!{bs_cl}28-'Balance Sheet'!{bs_pcl}28)", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 20, col, f"=SUM({cl}15:{cl}19)", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 22, col, f"={cl}9+{cl}10+{cl}11+{cl}20", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws6, 27, col, f"=-'Income Statement'!{is_cl}46", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 28, col, f"={cl}27", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws6, 34, col, f"=-'Debt Schedule'!{debt_cl}10", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 35, col, f"=-'Debt Schedule'!{debt_cl}22", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 37, col, f"=-'Debt Schedule'!{debt_cl}34", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 39, col, f"={cl}34+{cl}35+{cl}37", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    set_cell(ws6, 41, col, f"={cl}22+{cl}28+{cl}39", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws6, 45, col, f"={pcl}47", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 47, col, f"={cl}45+{cl}41", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

    set_cell(ws6, 53, col, f"={cl}47", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 54, col, f"='Balance Sheet'!{bs_cl}9", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)
    set_cell(ws6, 55, col, f'=IF(ABS({cl}53-{cl}54)<1,"\u2713 Match","X Mismatch")', font=CHECK_FONT, border=THIN_BORDER)

# =====================================================================
# SHEET 8: Returns Analysis
# =====================================================================
ws8 = wb.create_sheet("Returns Analysis")
ws8.sheet_properties.tabColor = "996600"

ws8.column_dimensions["A"].width = 3
ws8.column_dimensions["B"].width = 30
ws8.column_dimensions["C"].width = 5
ws8.column_dimensions["D"].width = 16
for i in range(n_proj):
    ws8.column_dimensions[CL(5+i)].width = 14

set_cell(ws8, 2, 2, "Returns Analysis", font=TITLE_FONT)
c = section_title(ws8, 4, 2, "Exit Analysis"); c.fill = LIGHT_FILL; fill_row(ws8, 4, 3, 4+n_proj)
header_row(ws8, 5, 4, 4+n_proj, ["Entry"]+[f"Year {y}" for y in range(1,n_proj+1)])

set_cell(ws8, 7, 2, "Exit Assumptions", font=BOLD_FONT)
set_cell(ws8, 9, 2, "Exit EV/EBITDA Multiple", font=BOLD_FONT, border=THIN_BORDER)
for i in range(n_proj):
    set_cell(ws8, 9, 5+i, C["exit_ev_ebitda"], font=BLUE_FONT, fmt=FMT_MULT, border=THIN_BORDER)

set_cell(ws8, 11, 2, "EBITDA (Exit Year)", font=BOLD_FONT, border=THIN_BORDER)
for i in range(n_proj):
    set_cell(ws8, 11, 5+i, f"='Income Statement'!{CL(proj_start+i)}25", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws8, 13, 2, "Exit Enterprise Value", font=BOLD_FONT, border=THIN_BORDER)
for i in range(n_proj):
    cl = CL(5+i)
    set_cell(ws8, 13, 5+i, f"={cl}9*{cl}11", font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws8, 15, 2, "(-) Exit Debt", font=BOLD_FONT, border=THIN_BORDER)
for i in range(n_proj):
    set_cell(ws8, 15, 5+i, f"='Debt Schedule'!{CL(5+i)}32", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws8, 16, 2, "(+) Exit Cash", font=BOLD_FONT, border=THIN_BORDER)
for i in range(n_proj):
    set_cell(ws8, 16, 5+i, C["exit_cash_assumption"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws8, 18, 2, "Exit Equity Value", font=BOLD_FONT, border=TOP_BOTTOM)
for i in range(n_proj):
    cl = CL(5+i)
    set_cell(ws8, 18, 5+i, f"={cl}13-{cl}15+{cl}16", font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

c = section_title(ws8, 20, 2, "Sponsor Returns"); c.fill = LIGHT_GREEN; fill_row(ws8, 20, 3, 4+n_proj, LIGHT_GREEN)

set_cell(ws8, 23, 2, "Sponsor Equity (Investment)", font=BOLD_FONT, border=THIN_BORDER)
set_cell(ws8, 23, 4, "='Transaction Assumptions'!D30", font=GREEN_FONT, fmt=FMT_YEN, border=THIN_BORDER)

set_cell(ws8, 25, 2, "MOIC", font=BOLD_FONT, border=TOP_BOTTOM)
for i in range(n_proj):
    set_cell(ws8, 25, 5+i, f"={CL(5+i)}18/$D$23", font=BLACK_FONT, fmt=FMT_MULT, border=TOP_BOTTOM)

set_cell(ws8, 27, 2, "IRR", font=BOLD_FONT, border=TOP_BOTTOM)
for i in range(n_proj):
    _, irr = calc_returns(C["exit_ev_ebitda"], i+1, C)
    set_cell(ws8, 27, 5+i, round(irr, 4), font=BLACK_FONT, fmt=FMT_PCT2, border=TOP_BOTTOM)

for r in [25,27]:
    for cc in range(4, 4+n_proj+1):
        ws8.cell(row=r, column=cc).fill = LIGHT_YELLOW

# Sensitivity MOIC
exit_mults = C["sens_exit_multiples"]
exit_years = C["sens_exit_years"]

c = section_title(ws8, 30, 2, "Sensitivity: MOIC (Exit Multiple x Exit Year)"); c.fill = LIGHT_FILL
set_cell(ws8, 31, 2, "Exit Mult \\ Exit Year", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j, ey in enumerate(exit_years):
    set_cell(ws8, 31, 4+j, f"Year {ey}", font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

for i, em in enumerate(exit_mults):
    r = 32+i
    set_cell(ws8, r, 2, em, font=BOLD_FONT, fmt=FMT_MULT, border=THIN_BORDER)
    for j, ey in enumerate(exit_years):
        moic, _ = calc_returns(em, ey, C)
        fill = LIGHT_YELLOW if (em == C["exit_ev_ebitda"] and ey == 5) else None
        set_cell(ws8, r, 4+j, round(moic, 2), font=BLACK_FONT, fmt=FMT_MULT, border=THIN_BORDER, fill=fill)

# Sensitivity IRR
irr_start = 32 + len(exit_mults) + 1
c = section_title(ws8, irr_start, 2, "Sensitivity: IRR (Exit Multiple x Exit Year)"); c.fill = LIGHT_FILL
set_cell(ws8, irr_start+1, 2, "Exit Mult \\ Exit Year", font=HEADER_FONT, fill=HEADER_FILL,
         alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
for j, ey in enumerate(exit_years):
    set_cell(ws8, irr_start+1, 4+j, f"Year {ey}", font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center"), border=THIN_BORDER)

for i, em in enumerate(exit_mults):
    r = irr_start+2+i
    set_cell(ws8, r, 2, em, font=BOLD_FONT, fmt=FMT_MULT, border=THIN_BORDER)
    for j, ey in enumerate(exit_years):
        _, irr = calc_returns(em, ey, C)
        fill = LIGHT_YELLOW if (em == C["exit_ev_ebitda"] and ey == 5) else None
        set_cell(ws8, r, 4+j, round(irr, 4), font=BLACK_FONT, fmt=FMT_PCT2, border=THIN_BORDER, fill=fill)

# Investment Thesis
thesis_start = irr_start + 2 + len(exit_mults) + 1
c = section_title(ws8, thesis_start, 2, "Investment Thesis"); c.fill = LIGHT_FILL; fill_row(ws8, thesis_start, 3, 7)

for i, (title, desc) in enumerate(C["investment_thesis"]):
    r = thesis_start + 1 + i
    set_cell(ws8, r, 2, title, font=BOLD_FONT, border=THIN_BORDER)
    set_cell(ws8, r, 4, desc, font=BLACK_FONT, border=THIN_BORDER)

# =====================================================================
# PRINT SETTINGS & SHEET ORDER
# =====================================================================
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

wb.move_sheet("Debt Schedule", offset=2)

# =====================================================================
# SAVE & VERIFY
# =====================================================================
wb.save(output_file)
print(f"Saved: {output_file}")
print("Sheets:", [s.title for s in wb.worksheets])

recalc_script = os.path.join("scripts", "recalc.py")
if os.path.exists(recalc_script):
    print(f"\nRunning verification: python {recalc_script} {output_file}")
    result = subprocess.run([sys.executable, recalc_script, output_file],
                            capture_output=True, text=True)
    print(result.stdout)
    if result.stderr:
        print("STDERR:", result.stderr)
else:
    print(f"\nNote: {recalc_script} not found, skipping verification.")
