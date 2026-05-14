"""
market_analysis_template.py - Market Perception Analysis Excel Generator

Generates a 2-sheet Excel workbook for analyzing market perception:
1. Implied Growth Analysis - Reverse DCF to extract market's growth expectations
2. Market Scorecard - 4-factor weighted scorecard for buy/sell judgment

Generic template for any ticker. All values dynamically extracted from a
companion DCF Excel (produced by dcf_comps_template.py).

Sister templates:
  - dcf_comps_template.py: DCF/Comps valuation
  - sotp_template.py:      Sum-of-the-parts valuation
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

# ---------------------------------------------------------------------------
# Style constants (kept aligned with dcf_comps_template.py / sotp_template.py)
# ---------------------------------------------------------------------------
TITLE_FONT      = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
TITLE_FILL      = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT     = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL     = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
SUBHEADER_FONT  = Font(name='Calibri', size=10, bold=True)
SUBHEADER_FILL  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
INPUT_FILL      = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
OUTPUT_FILL     = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
HIGHLIGHT_FILL  = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
VERDICT_FILL    = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
WARNING_FILL    = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

THIN_SIDE = Side(border_style='thin', color='808080')
BORDER    = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)

# 18-point alpha grid spanning Deep-Pessimism through Bullish
ALPHA_VALUES = [-0.5, -0.3, -0.1, 0.0, 0.1, 0.2, 0.3, 0.4, 0.5,
                 0.6,  0.7,  0.8, 0.9, 1.0, 1.1, 1.2, 1.3, 1.5]


# ---------------------------------------------------------------------------
# DCF Excel data extraction
# ---------------------------------------------------------------------------
def _find_last_actual_column(ws_seg):
    """Identify the last non-blank actual-revenue column in 'Segment Analysis'.

    Row 21 is the consolidated 'Total Revenue' row by template convention. We
    scan F backward to A and pick the rightmost column that holds a numeric
    revenue value -- that column is the last historical actual.
    """
    for col_letter in 'FEDCBA':
        v = ws_seg[f'{col_letter}21'].value
        if isinstance(v, (int, float)) and v > 0:
            return col_letter
    return 'F'


def _get_forecast_columns(ws_seg, last_actual_col):
    """Return the 5 forecast column letters that follow the last actual."""
    actual_idx = column_index_from_string(last_actual_col)
    return [get_column_letter(actual_idx + 1 + i) for i in range(5)]


def extract_dcf_data(dcf_excel_path, segment_layout):
    """Pull every value the market-analysis sheets need from a DCF Excel.

    The Base scenario growth/OPM rows are read from 'Segment Analysis' (rows
    34/41/48/55/62/69 by template convention) regardless of the workbook's
    currently-active scenario. D&A / Capex / DeltaNWC use the active scenario's
    DCF Model row 42/43/44 -- documented limitation of the spec.
    """
    if not os.path.exists(dcf_excel_path):
        raise FileNotFoundError(f"DCF Excel not found: {dcf_excel_path}")

    wb = openpyxl.load_workbook(dcf_excel_path, data_only=True)
    if 'Segment Analysis' not in wb.sheetnames or 'DCF Model' not in wb.sheetnames:
        raise ValueError(
            f"DCF Excel missing required sheets (Segment Analysis / DCF Model): {dcf_excel_path}"
        )
    ws_seg = wb['Segment Analysis']
    ws_dcf = wb['DCF Model']

    # ── DCF macro assumptions ──
    wacc            = float(ws_dcf['C26'].value or 0)
    terminal_growth = float(ws_dcf['C13'].value or 0)
    tax_rate        = float(ws_dcf['C6'].value  or 0)
    net_debt        = float(ws_dcf['C16'].value or 0)
    shares          = float(ws_dcf['C15'].value or 0)
    raw_stub        = ws_dcf['C19'].value
    stub_fraction   = float(raw_stub) if raw_stub is not None else 0.0

    # Stub Fraction sanity check: 0 means the DCF discounts year-1 with no stub,
    # which is fine when the projection starts at fiscal-year boundary, but if
    # a partial-year stub is expected (typical mid-year valuation date), 0 is
    # likely wrong. Warn so the user can verify.
    import warnings as _w
    if raw_stub is None or stub_fraction == 0:
        _w.warn(
            f"Stub Fraction is {raw_stub} in DCF Model!C19. If this DCF was "
            f"built mid-fiscal-year, expected ~0.25; if at fiscal-year start, "
            f"0 is correct. Verify alignment.",
            UserWarning, stacklevel=2,
        )
    elif stub_fraction < 0 or stub_fraction > 1:
        _w.warn(
            f"Stub Fraction {stub_fraction} is outside [0, 1]; using as-is.",
            UserWarning, stacklevel=2,
        )

    last_actual_col = _find_last_actual_column(ws_seg)
    forecast_cols   = _get_forecast_columns(ws_seg, last_actual_col)

    # ── Consolidated FY-actual figures (from Segment Analysis row 21/22) ──
    fy_actual_total_rev = float(ws_seg[f'{last_actual_col}21'].value or 0)
    fy_actual_total_op  = float(ws_seg[f'{last_actual_col}22'].value or 0)
    fy_actual_opm = (fy_actual_total_op / fy_actual_total_rev) if fy_actual_total_rev else 0.0

    # ── Per-segment Base growth + OPM (independent of active scenario) ──
    segments_data = []
    for seg in segment_layout['segments']:
        fy_actual = float(ws_seg[seg['dcf_fy26_cell']].value or 0)
        growth = [float(ws_seg[f'{c}{seg["dcf_growth_base_row"]}'].value or 0) for c in forecast_cols]
        opm    = [float(ws_seg[f'{c}{seg["dcf_opm_base_row"]}'].value or 0) for c in forecast_cols]

        rev_path = [fy_actual]
        for g in growth:
            rev_path.append(rev_path[-1] * (1 + g))

        segments_data.append({
            'name': seg['name'],
            'fy_actual': fy_actual,
            'growth': growth,
            'opm': opm,
            'rev_path': rev_path,
        })

    # Aggregate to consolidated Base trajectory
    base_total_rev_path = [
        sum(s['rev_path'][t] for s in segments_data) for t in range(6)
    ] if segments_data else [fy_actual_total_rev] + [0] * 5

    base_total_op = [
        sum(s['rev_path'][yr+1] * s['opm'][yr] for s in segments_data) for yr in range(5)
    ]
    base_total_opm = [
        (base_total_op[yr] / base_total_rev_path[yr+1]) if base_total_rev_path[yr+1] else 0.0
        for yr in range(5)
    ]
    base_total_growth = [
        (base_total_rev_path[yr+1] / base_total_rev_path[yr] - 1) if base_total_rev_path[yr] else 0.0
        for yr in range(5)
    ]

    # ── DCF Model active-scenario WC items (rows 42/43/44, columns C-G) ──
    da         = [float(ws_dcf[f'{c}42'].value or 0) for c in 'CDEFG']
    capex      = [float(ws_dcf[f'{c}43'].value or 0) for c in 'CDEFG']
    nwc_change = [float(ws_dcf[f'{c}44'].value or 0) for c in 'CDEFG']

    # "My prediction" = Base scenario FY1 growth (used in Scorecard Factor 4)
    my_rev_growth = base_total_growth[0]
    my_op_growth  = (base_total_op[0] / fy_actual_total_op - 1) if fy_actual_total_op else 0.0

    base_pgm  = float(ws_dcf['C55'].value or 0)
    base_exit = float(ws_dcf['C64'].value or 0)

    # Compute a Base-consistent PGM target in Python so Block 1 can show it
    # without depending on whichever scenario was active when the DCF saved.
    fcfs = []
    for yr in range(5):
        nopat = max(0.0, base_total_op[yr] * (1 - tax_rate))
        fcfs.append(nopat + da[yr] - capex[yr] - nwc_change[yr])
    sum_pv = sum(fcfs[yr] / (1 + wacc) ** (stub_fraction + yr) for yr in range(5))
    if wacc > terminal_growth:
        tv = fcfs[-1] * (1 + terminal_growth) / (wacc - terminal_growth)
    else:
        tv = 0
    pv_tv = tv / (1 + wacc) ** (stub_fraction + 4)
    ev_base = sum_pv + pv_tv
    equity_base = ev_base - net_debt
    base_pgm_consistent = (equity_base * 1_000_000 / shares) if shares else 0.0

    return {
        'wacc': wacc,
        'terminal_growth': terminal_growth,
        'tax_rate': tax_rate,
        'net_debt': net_debt,
        'shares': shares,
        'stub_fraction': stub_fraction,
        'fy_actual_total_rev': fy_actual_total_rev,
        'fy_actual_total_op': fy_actual_total_op,
        'fy_actual_opm': fy_actual_opm,
        'segments': segments_data,
        'base_total_rev_path': base_total_rev_path,
        'base_total_op': base_total_op,
        'base_total_opm': base_total_opm,
        'base_total_growth': base_total_growth,
        'da': da,
        'capex': capex,
        'nwc_change': nwc_change,
        'my_op_growth': my_op_growth,
        'my_rev_growth': my_rev_growth,
        'base_pgm_target': base_pgm,                 # active-scenario PGM (legacy)
        'base_exit_target': base_exit,
        'base_pgm_consistent': base_pgm_consistent,  # Python-recomputed Base-scenario PGM
        'forecast_cols': forecast_cols,
        'last_actual_col': last_actual_col,
    }


# ---------------------------------------------------------------------------
# Sheet 1: Implied Growth Analysis
# ---------------------------------------------------------------------------
def _build_implied_growth_sheet(wb, config, dd):
    ws = wb.create_sheet('Implied Growth Analysis')

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 38
    for col in 'CDEFGHIJKLMNOPQRST':
        ws.column_dimensions[col].width = 13

    # Title
    ws['B2'] = f"Implied Growth Analysis - {config['company_name']} ({config['ticker']})"
    ws['B2'].font = TITLE_FONT
    ws['B2'].fill = TITLE_FILL
    ws.merge_cells('B2:T2')

    shares   = dd['shares']
    net_debt = dd['net_debt']
    wacc     = dd['wacc']
    tax      = dd['tax_rate']
    stub     = dd['stub_fraction']
    tg       = dd['terminal_growth']

    # ────────── Block 1: Market Inputs (price points) ──────────
    ws['B5'] = 'Block 1: Market Inputs (Price Points)'
    ws['B5'].font = HEADER_FONT
    ws['B5'].fill = HEADER_FILL
    ws.merge_cells('B5:G5')

    headers_b1 = ['Label', 'Price (JPY)', 'Equity (JPY mn)', 'EV (JPY mn)',
                  'Required PV (JPY mn)', 'Timing']
    for i, h in enumerate(headers_b1):
        c = chr(ord('B') + i)
        ws[f'{c}6'] = h
        ws[f'{c}6'].font = SUBHEADER_FONT
        ws[f'{c}6'].fill = SUBHEADER_FILL

    # Price points: prefer config, otherwise auto-build.
    # Default order keeps Current Price first (Block 4/5 reference its alpha).
    if config.get('price_points'):
        price_points = list(config['price_points'])
    else:
        price_points = [
            {'label': 'Current Price', 'price': config['current_price'], 'timing': 'now'},
        ]
        if config.get('entry_price') is not None:
            price_points.append(
                {'label': 'Entry Price', 'price': config['entry_price'], 'timing': 'cost basis'}
            )
        price_points.extend([
            {'label': 'DCF Base PGM Target',      'price': dd['base_pgm_consistent'], 'timing': 'Base scenario'},
            {'label': 'DCF Exit-Multiple Target', 'price': dd['base_exit_target'],    'timing': 'active scenario'},
        ])
        if config.get('price_low') is not None:
            price_points.append({'label': 'Recent Low',  'price': config['price_low'],  'timing': 'recent'})
        if config.get('price_high') is not None:
            price_points.append({'label': 'Recent High', 'price': config['price_high'], 'timing': 'recent'})

    pp_rows = []
    for i, pp in enumerate(price_points[:7]):  # cap at 7 rows (rows 7..13)
        r = 7 + i
        ws[f'B{r}'] = pp['label']
        ws[f'C{r}'] = pp['price']
        ws[f'C{r}'].number_format = '#,##0'
        ws[f'D{r}'] = f'=C{r}*{shares}/1000000'
        ws[f'D{r}'].number_format = '#,##0'
        ws[f'E{r}'] = f'=D{r}+{net_debt}'
        ws[f'E{r}'].number_format = '#,##0'
        ws[f'F{r}'] = f'=E{r}'
        ws[f'F{r}'].number_format = '#,##0'
        ws[f'G{r}'] = pp.get('timing', '')
        for c in 'BCDEFG':
            ws[f'{c}{r}'].border = BORDER
        pp_rows.append((r, pp['label']))

    # ────────── Block 2: Reverse DCF Engine (alpha-scan) ──────────
    ws['B14'] = 'Block 2: Reverse DCF Engine (alpha-scan)'
    ws['B14'].font = HEADER_FONT
    ws['B14'].fill = HEADER_FILL
    ws.merge_cells('B14:T14')

    ws['B16'] = 'alpha (Base growth multiplier)'
    ws['B16'].font = SUBHEADER_FONT
    ws['B16'].fill = SUBHEADER_FILL
    ws['B17'] = 'alpha value'
    for i, alpha in enumerate(ALPHA_VALUES):
        c = get_column_letter(3 + i)  # C=3
        ws[f'{c}17'] = alpha
        ws[f'{c}17'].number_format = '0.00'
        ws[f'{c}17'].font = SUBHEADER_FONT
        ws[f'{c}17'].fill = HIGHLIGHT_FILL

    ws['B19'] = 'Calculations (Year 1 to Year 5)'
    ws['B19'].font = SUBHEADER_FONT

    # Helper region (auxiliary inputs embedded as values)
    HELPER_START = 47
    n_seg = len(dd['segments'])
    HELPER_HEADER_ROW   = HELPER_START          # 47
    FY_REV_ROW          = HELPER_START + 1      # 48
    FY_OP_ROW           = HELPER_START + 2      # 49
    FY_OPM_ROW          = HELPER_START + 3      # 50
    SEG_PATH_HEADER_ROW = HELPER_START + 4      # 51
    SEG_PATH_FIRST_ROW  = HELPER_START + 5      # 52
    TOTAL_REV_ROW       = SEG_PATH_FIRST_ROW + n_seg
    GROWTH_ROW          = TOTAL_REV_ROW + 1
    OPM_ROW             = TOTAL_REV_ROW + 2
    DA_ROW              = TOTAL_REV_ROW + 3
    CAPEX_ROW           = TOTAL_REV_ROW + 4
    NWC_ROW             = TOTAL_REV_ROW + 5

    # Years label header for helpers
    fc = dd['forecast_cols']  # e.g. ['G','H','I','J','K']
    last_actual = dd['last_actual_col']

    ws[f'B{HELPER_HEADER_ROW}'] = '── Auxiliary inputs (extracted from DCF) ──'
    ws[f'B{HELPER_HEADER_ROW}'].font = SUBHEADER_FONT
    ws[f'B{HELPER_HEADER_ROW}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'B{HELPER_HEADER_ROW}:T{HELPER_HEADER_ROW}')

    ws[f'B{FY_REV_ROW}'] = f'FY Actual Total Revenue (mn) [{last_actual}]'
    ws[f'C{FY_REV_ROW}'] = dd['fy_actual_total_rev']
    ws[f'C{FY_REV_ROW}'].number_format = '#,##0'

    ws[f'B{FY_OP_ROW}'] = 'FY Actual Total OP (mn)'
    ws[f'C{FY_OP_ROW}'] = dd['fy_actual_total_op']
    ws[f'C{FY_OP_ROW}'].number_format = '#,##0'

    ws[f'B{FY_OPM_ROW}'] = 'FY Actual Total OPM'
    ws[f'C{FY_OPM_ROW}'] = dd['fy_actual_opm']
    ws[f'C{FY_OPM_ROW}'].number_format = '0.00%'

    ws[f'B{SEG_PATH_HEADER_ROW}'] = 'Base scenario - per-segment revenue path (mn)'
    ws[f'B{SEG_PATH_HEADER_ROW}'].font = SUBHEADER_FONT

    # Year header row (D-H = Year 1..5)
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{SEG_PATH_HEADER_ROW}'] = f'Year {yr+1}'
        ws[f'{col}{SEG_PATH_HEADER_ROW}'].font = SUBHEADER_FONT

    # Per-segment Base revenue paths
    for i, seg in enumerate(dd['segments']):
        r = SEG_PATH_FIRST_ROW + i
        ws[f'B{r}'] = seg['name']
        ws[f'C{r}'] = seg['fy_actual']
        ws[f'C{r}'].number_format = '#,##0'
        for yr in range(5):
            col = get_column_letter(4 + yr)
            ws[f'{col}{r}'] = seg['rev_path'][yr+1]
            ws[f'{col}{r}'].number_format = '#,##0'

    # Total Base revenue, growth, OPM (consolidated)
    ws[f'B{TOTAL_REV_ROW}'] = 'Total Base Revenue (mn)'
    ws[f'B{TOTAL_REV_ROW}'].font = SUBHEADER_FONT
    ws[f'C{TOTAL_REV_ROW}'] = dd['fy_actual_total_rev']
    ws[f'C{TOTAL_REV_ROW}'].number_format = '#,##0'
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{TOTAL_REV_ROW}'] = dd['base_total_rev_path'][yr+1]
        ws[f'{col}{TOTAL_REV_ROW}'].number_format = '#,##0'

    ws[f'B{GROWTH_ROW}'] = 'Total Base Growth (YoY)'
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{GROWTH_ROW}'] = dd['base_total_growth'][yr]
        ws[f'{col}{GROWTH_ROW}'].number_format = '0.00%'

    ws[f'B{OPM_ROW}'] = 'Total Base OPM (weighted)'
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{OPM_ROW}'] = dd['base_total_opm'][yr]
        ws[f'{col}{OPM_ROW}'].number_format = '0.00%'

    ws[f'B{DA_ROW}'] = 'D&A (DCF active scenario)'
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{DA_ROW}'] = dd['da'][yr]
        ws[f'{col}{DA_ROW}'].number_format = '#,##0'

    ws[f'B{CAPEX_ROW}'] = 'Capex (DCF active scenario)'
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{CAPEX_ROW}'] = dd['capex'][yr]
        ws[f'{col}{CAPEX_ROW}'].number_format = '#,##0'

    ws[f'B{NWC_ROW}'] = 'DeltaNWC (DCF active scenario)'
    for yr in range(5):
        col = get_column_letter(4 + yr)
        ws[f'{col}{NWC_ROW}'] = dd['nwc_change'][yr]
        ws[f'{col}{NWC_ROW}'].number_format = '#,##0'

    # Block 2 main calculations (rows 20-45 per spec)
    for yr in range(5):
        ws[f'B{20+yr}'] = f'  Year {yr+1} growth (alpha x Base)'
    for yr in range(5):
        ws[f'B{25+yr}'] = f'  Year {yr+1} Revenue (mn)'
    for yr in range(5):
        ws[f'B{30+yr}'] = f'  Year {yr+1} EBIT (mn)'
    for yr in range(5):
        ws[f'B{35+yr}'] = f'  Year {yr+1} FCF (mn)'

    for i, alpha in enumerate(ALPHA_VALUES):
        col = get_column_letter(3 + i)

        # Row 20-24: scaled growth
        for yr in range(5):
            r = 20 + yr
            base_g_col = get_column_letter(4 + yr)
            ws[f'{col}{r}'] = f'={col}17*{base_g_col}{GROWTH_ROW}'
            ws[f'{col}{r}'].number_format = '0.00%'

        # Row 25-29: revenue
        for yr in range(5):
            r = 25 + yr
            growth_cell = f'{col}{20+yr}'
            if yr == 0:
                ws[f'{col}{r}'] = f'=C{FY_REV_ROW}*(1+{growth_cell})'
            else:
                ws[f'{col}{r}'] = f'={col}{r-1}*(1+{growth_cell})'
            ws[f'{col}{r}'].number_format = '#,##0'

        # Row 30-34: EBIT, with OPM scaling: FY_OPM + alpha * (Base_OPM - FY_OPM)
        for yr in range(5):
            r = 30 + yr
            base_opm_col = get_column_letter(4 + yr)
            opm_formula = (f'(C{FY_OPM_ROW}+{col}17*'
                           f'({base_opm_col}{OPM_ROW}-C{FY_OPM_ROW}))')
            ws[f'{col}{r}'] = f'={col}{25+yr}*{opm_formula}'
            ws[f'{col}{r}'].number_format = '#,##0'

        # Row 35-39: FCF = MAX(0, EBIT*(1-tax)) + DA - Capex - dNWC
        for yr in range(5):
            r = 35 + yr
            ebit = f'{col}{30+yr}'
            yr_col = get_column_letter(4 + yr)
            ws[f'{col}{r}'] = (
                f'=MAX(0,{ebit}*(1-{tax}))+{yr_col}{DA_ROW}-{yr_col}{CAPEX_ROW}-{yr_col}{NWC_ROW}'
            )
            ws[f'{col}{r}'].number_format = '#,##0'

        # Row 40: Sum PV of explicit FCFs
        terms = [f'{col}{35+yr}/(1+{wacc})^({stub}+{yr})' for yr in range(5)]
        ws[f'{col}40'] = '=' + '+'.join(terms)
        ws[f'{col}40'].number_format = '#,##0'

        # Row 41: Terminal Value (PGM)
        ws[f'{col}41'] = f'={col}39*(1+{tg})/({wacc}-{tg})'
        ws[f'{col}41'].number_format = '#,##0'

        # Row 42: PV of TV
        ws[f'{col}42'] = f'={col}41/(1+{wacc})^({stub}+4)'
        ws[f'{col}42'].number_format = '#,##0'

        # Row 43: EV
        ws[f'{col}43'] = f'={col}40+{col}42'
        ws[f'{col}43'].number_format = '#,##0'

        # Row 44: Equity Value
        ws[f'{col}44'] = f'={col}43-{net_debt}'
        ws[f'{col}44'].number_format = '#,##0'

        # Row 45: Implied Share Price
        ws[f'{col}45'] = f'={col}44*1000000/{shares}'
        ws[f'{col}45'].number_format = '#,##0'
        ws[f'{col}45'].fill = OUTPUT_FILL

    # Section labels for the calc rows
    ws['B40'] = 'Sum PV of explicit FCFs (mn)'
    ws['B40'].font = SUBHEADER_FONT
    ws['B41'] = 'Terminal Value, PGM (mn)'
    ws['B42'] = 'PV of Terminal Value (mn)'
    ws['B43'] = 'Enterprise Value (mn)'
    ws['B43'].font = SUBHEADER_FONT
    ws['B44'] = 'Equity Value (mn)'
    ws['B45'] = 'Implied Share Price (JPY)'
    ws['B45'].font = SUBHEADER_FONT
    ws['B45'].fill = OUTPUT_FILL

    # ─────────────────────────────────────────────────────────────────────
    # Compute downstream block row positions dynamically so they never
    # collide with the helper region (which scales with segment count).
    #   - Block 3 starts >= 62 and >= helper_end + 2
    #   - Block 4 / 5 chain after Block 3 with a 1-row gap
    #   - "How to use" goes after Block 5 with a 6-row buffer (>= row 95)
    # All cross-block references (Gap, Market View alpha) use these vars,
    # so adding/removing price points or segments stays consistent.
    # ─────────────────────────────────────────────────────────────────────
    block3_header_row    = max(62, NWC_ROW + 2)
    block3_colheader_row = block3_header_row + 1
    block3_data_start    = block3_header_row + 2
    block3_data_end      = block3_data_start + len(pp_rows) - 1  # inclusive
    market_alpha_d_row   = block3_data_start  # Current Price is always first

    block4_header_row    = block3_data_end + 2
    block4_colheader_row = block4_header_row + 1
    block4_data_start    = block4_header_row + 2

    block5_header_row    = block4_data_start + 5 + 1  # 5 scenarios + 1 gap
    block5_market_row    = block5_header_row + 1
    block5_my_row        = block5_header_row + 2
    block5_gap_row       = block5_header_row + 3
    block5_action_row    = block5_header_row + 4

    howto_start = max(95, block5_action_row + 6)

    # ────────── Block 3: Implied alpha per price point ──────────
    ws[f'B{block3_header_row}'] = 'Block 3: Implied alpha per Price Point (local linear interpolation)'
    ws[f'B{block3_header_row}'].font = HEADER_FONT
    ws[f'B{block3_header_row}'].fill = HEADER_FILL
    ws.merge_cells(f'B{block3_header_row}:F{block3_header_row}')

    for i, h in enumerate(['Label', 'Price (JPY)', 'Implied alpha',
                           'Interpretation', 'Market Pricing']):
        c = chr(ord('B') + i)
        ws[f'{c}{block3_colheader_row}'] = h
        ws[f'{c}{block3_colheader_row}'].font = SUBHEADER_FONT
        ws[f'{c}{block3_colheader_row}'].fill = SUBHEADER_FILL

    # Local linear interpolation: alpha vs implied-price is convex (TV scales
    # exponentially), so a global least-squares line (FORECAST) systematically
    # over/under-shoots in low-alpha regions. We instead bracket each input
    # price between two adjacent grid points and interpolate just those two.
    alpha_row = 17
    price_row = 45
    alpha_range = f'$C${alpha_row}:$T${alpha_row}'
    impl_price_range = f'$C${price_row}:$T${price_row}'

    def _interp_formula(price_cell):
        # MATCH(price, prices, 1) returns the index of the largest grid price
        # <= the input. Combine with INDEX to recover (alpha_lo, price_lo) and
        # (alpha_hi, price_hi) for the two-point linear blend.
        m = f'MATCH({price_cell},{impl_price_range},1)'
        a_lo = f'INDEX({alpha_range},{m})'
        a_hi = f'INDEX({alpha_range},{m}+1)'
        p_lo = f'INDEX({impl_price_range},{m})'
        p_hi = f'INDEX({impl_price_range},{m}+1)'
        return f'={a_lo}+({price_cell}-{p_lo})/({p_hi}-{p_lo})*({a_hi}-{a_lo})'

    for i, (r_pp, _label) in enumerate(pp_rows):
        r = block3_data_start + i
        # Reference Block 1 label & price by formula so edits propagate.
        ws[f'B{r}'] = f'=B{r_pp}'
        ws[f'C{r}'] = f'=C{r_pp}'
        ws[f'C{r}'].number_format = '#,##0'
        ws[f'D{r}'] = _interp_formula(f'C{r}')
        ws[f'D{r}'].number_format = '0.000'
        ws[f'E{r}'] = (f'=IF(D{r}<-0.3,"Deep Pessimism",'
                      f'IF(D{r}<0,"Pessimistic",'
                      f'IF(D{r}<0.5,"Neutral",'
                      f'IF(D{r}<1,"Mild Optimism","Bullish"))))')
        ws[f'F{r}'] = (f'=IF(D{r}<-0.3,"Base growth halved or worse",'
                      f'IF(D{r}<0,"Below-Base, conservative",'
                      f'IF(D{r}<0.5,"0-50% of Base growth priced in",'
                      f'IF(D{r}<1,"50-100% of Base priced in","Above-Base optimism"))))')
        for c in 'BCDEF':
            ws[f'{c}{r}'].border = BORDER

    # ────────── Block 4: Comparison Matrix ──────────
    ws[f'B{block4_header_row}'] = 'Block 4: Comparison Matrix (5-scenario vs Market)'
    ws[f'B{block4_header_row}'].font = HEADER_FONT
    ws[f'B{block4_header_row}'].fill = HEADER_FILL
    ws.merge_cells(f'B{block4_header_row}:G{block4_header_row}')

    for i, h in enumerate(['Scenario', 'alpha', '5Y CAGR (Rev)', 'Avg OPM',
                           'Gap vs Market', 'Verdict']):
        c = chr(ord('B') + i)
        ws[f'{c}{block4_colheader_row}'] = h
        ws[f'{c}{block4_colheader_row}'].font = SUBHEADER_FONT
        ws[f'{c}{block4_colheader_row}'].fill = SUBHEADER_FILL

    scenarios = [
        ('Base',       1.00),
        ('Upside',     1.50),
        ('Management', 0.85),
        ('Downside 1', 0.40),
        ('Downside 2', 0.15),
    ]
    for i, (name, alpha) in enumerate(scenarios):
        r = block4_data_start + i
        ws[f'B{r}'] = name
        ws[f'C{r}'] = alpha
        ws[f'C{r}'].number_format = '0.00'

        cagr_terms = []
        opm_terms = []
        for yr in range(5):
            yr_col = get_column_letter(4 + yr)
            cagr_terms.append(f'(1+C{r}*{yr_col}{GROWTH_ROW})')
            opm_terms.append(f'{yr_col}{OPM_ROW}')
        ws[f'D{r}'] = '=(' + '*'.join(cagr_terms) + ')^(1/5)-1'
        ws[f'D{r}'].number_format = '0.00%'
        ws[f'E{r}'] = (f'=C{FY_OPM_ROW}+(AVERAGE({",".join(opm_terms)})-C{FY_OPM_ROW})*C{r}')
        ws[f'E{r}'].number_format = '0.00%'
        # Gap references the Block 3 Current Price row dynamically.
        ws[f'F{r}'] = f'=C{r}-D{market_alpha_d_row}'
        ws[f'F{r}'].number_format = '+0.00;-0.00;0.00'
        ws[f'G{r}'] = (f'=IF(F{r}>0.5,"BULLISH (Buy candidate)",'
                      f'IF(F{r}>0.2,"Mildly Bullish",'
                      f'IF(F{r}>=-0.2,"Aligned",'
                      f'IF(F{r}>=-0.5,"Mildly Bearish","BEARISH (Sell candidate)"))))')
        for c in 'BCDEFG':
            ws[f'{c}{r}'].border = BORDER

    # ────────── Block 5: Final Verdict ──────────
    ws[f'B{block5_header_row}'] = 'Block 5: Final Verdict'
    ws[f'B{block5_header_row}'].font = HEADER_FONT
    ws[f'B{block5_header_row}'].fill = HEADER_FILL
    ws.merge_cells(f'B{block5_header_row}:G{block5_header_row}')

    ws[f'B{block5_market_row}'] = 'Market View alpha (current price)'
    ws[f'C{block5_market_row}'] = f'=D{market_alpha_d_row}'
    ws[f'C{block5_market_row}'].number_format = '0.000'
    ws[f'B{block5_my_row}'] = 'My View alpha (editable)'
    ws[f'C{block5_my_row}'] = 1.0
    ws[f'C{block5_my_row}'].number_format = '0.000'
    ws[f'C{block5_my_row}'].fill = INPUT_FILL
    ws[f'B{block5_gap_row}'] = 'Gap (My - Market)'
    ws[f'C{block5_gap_row}'] = f'=C{block5_my_row}-C{block5_market_row}'
    ws[f'C{block5_gap_row}'].number_format = '+0.000;-0.000;0.000'
    ws[f'B{block5_action_row}'] = 'Action Recommendation'
    ws[f'C{block5_action_row}'] = (
        f'=IF(C{block5_gap_row}>0.5,"BUY (upside surprise expected)",'
        f'IF(C{block5_gap_row}>0.2,"BUY/HOLD",'
        f'IF(C{block5_gap_row}>=-0.2,"HOLD",'
        f'IF(C{block5_gap_row}>=-0.5,"HOLD/SELL","SELL (disappointment risk)"))))'
    )
    ws[f'C{block5_action_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'C{block5_action_row}'].fill = VERDICT_FILL
    for r in range(block5_market_row, block5_action_row + 1):
        for c in 'BC':
            ws[f'{c}{r}'].border = BORDER

    # ────────── How to use (separated from any calc cells) ──────────
    notes = [
        '── How to use ──',
        '1. Block 1: market reference price points',
        '2. Block 2: alpha-scan implied share prices',
        '3. Block 3: implied alpha per price (local linear interpolation)',
        '4. Block 4: 5-scenario comparison',
        f'5. Block 5: final verdict (edit C{block5_my_row} to set your own alpha)',
        '',
        '── alpha interpretation ──',
        'alpha = 1.0 -> full Base scenario priced in',
        'alpha = 0.0 -> zero growth priced in (FY actual flat)',
        'alpha < 0   -> contraction priced in',
        'alpha > 1.0 -> above-Base optimism priced in',
    ]
    for i, txt in enumerate(notes):
        r = howto_start + i
        ws[f'B{r}'] = txt
        if txt.startswith('──'):
            ws[f'B{r}'].font = SUBHEADER_FONT

    # Expose the Market View alpha cell so the Scorecard sheet can reference it.
    return {
        'sheet': ws,
        'market_alpha_cell': f'C{block5_market_row}',
        'block3_data_start': block3_data_start,
    }


# ---------------------------------------------------------------------------
# Sheet 2: Market Scorecard
# ---------------------------------------------------------------------------
def _build_market_scorecard_sheet(wb, config, dd, market_alpha_cell='C90'):
    """Build the Market Scorecard sheet.

    market_alpha_cell: address of the Market View alpha cell on the
    'Implied Growth Analysis' sheet. Defaults to legacy 'C90' for backward
    compatibility but should be passed from _build_implied_growth_sheet so
    the reference tracks layout changes (segments, price-point count).
    """
    ws = wb.create_sheet('Market Scorecard')

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 32

    ws['B2'] = f"Market Scorecard - {config['company_name']} ({config['ticker']})"
    ws['B2'].font = TITLE_FONT
    ws['B2'].fill = TITLE_FILL
    ws.merge_cells('B2:G2')

    # ────────── Block 1: Scorecard Summary ──────────
    ws['B5'] = 'Block 1: Scorecard Summary'
    ws['B5'].font = HEADER_FONT
    ws['B5'].fill = HEADER_FILL
    ws.merge_cells('B5:G5')

    for i, h in enumerate(['Factor', 'Measure', 'Score (-2..+2)', 'Weight',
                           'Contribution', 'Verdict']):
        c = chr(ord('B') + i)
        ws[f'{c}6'] = h
        ws[f'{c}6'].font = SUBHEADER_FONT
        ws[f'{c}6'].fill = SUBHEADER_FILL

    # Factor 1: Implied Growth (40%)
    # Note: spec literally says "=1 - 'Implied Growth Analysis'!C91" but C91 is
    # the *user's* alpha (editable, default 1.0) -- which would always make the
    # measure 0. The intent (per the verdict labels "市場大幅悲観 = 強気") is
    # clearly to compare 1 vs the *market* alpha, which lives in C90. We use C90.
    ws['B7'] = 'Implied Growth (1 - market alpha)'
    ws['C7'] = f"=1-'Implied Growth Analysis'!{market_alpha_cell}"
    ws['C7'].number_format = '0.000'
    ws['D7'] = '=IF(C7>=1,2,IF(C7>=0.5,1,IF(C7>=-0.2,0,IF(C7>=-0.5,-1,-2))))'
    ws['E7'] = 0.40
    ws['E7'].number_format = '0%'
    ws['F7'] = '=D7*E7'
    ws['F7'].number_format = '+0.00;-0.00;0.00'
    ws['G7'] = ('=IF(D7=2,"市場大幅悲観 = 強気",'
                'IF(D7=1,"市場やや悲観 = 買い候補",'
                'IF(D7=0,"市場と認識一致",'
                'IF(D7=-1,"市場やや楽観 = 警戒","市場大幅楽観 = 失望リスク"))))')

    # Factor 2: Price Momentum (25%) -- 3M change
    ws['B8'] = 'Price Momentum (3M change)'
    ws['C8'] = '=IFERROR((C19-C20)/C20,0)'
    ws['C8'].number_format = '+0.00%;-0.00%;0.00%'
    ws['D8'] = '=IF(C8<=-0.20,2,IF(C8<=-0.10,1,IF(C8<=0.10,0,IF(C8<=0.20,-1,-2))))'
    ws['E8'] = 0.25
    ws['E8'].number_format = '0%'
    ws['F8'] = '=D8*E8'
    ws['F8'].number_format = '+0.00;-0.00;0.00'
    ws['G8'] = ('=IF(D8=2,"過剰悲観 = リバウンド余地大",'
                'IF(D8=1,"悲観織込み = サプライズ余地",'
                'IF(D8=0,"期待ニュートラル",'
                'IF(D8=-1,"期待やや過熱","期待過熱 = 失望売りリスク"))))')

    # Factor 3: Margin Balance (20%) -- 信用倍率 + 売残充実度
    ws['B9'] = 'Margin Balance (credit ratio)'
    ws['C9'] = '=IFERROR(C34/C33,0)'  # short ratio (sell/buy) for display
    ws['C9'].number_format = '0.00'
    ws['D9'] = ('=IF(AND(C36<=1,C34/C35>=0.5),2,'
                'IF(AND(C36<=3,C34/C35>=0.3),1,'
                'IF(AND(C36<=10,C34/C35>=0.1),0,'
                'IF(AND(C36<=20,C34/C35>=0.05),-1,-2))))')
    ws['E9'] = 0.20
    ws['E9'].number_format = '0%'
    ws['F9'] = '=D9*E9'
    ws['F9'].number_format = '+0.00;-0.00;0.00'
    ws['G9'] = ('=IF(D9=2,"踏み上げ余地大",'
                'IF(D9=1,"需給良好",'
                'IF(D9=0,"ニュートラル",'
                'IF(D9=-1,"戻り売り圧力","需給最悪"))))')

    # Factor 4: Forecast Gap (15%)
    ws['B10'] = 'Forecast Gap (vs Company)'
    ws['C10'] = '=AVERAGE(E44:E45)'
    ws['C10'].number_format = '+0.00%;-0.00%;0.00%'
    ws['D10'] = '=ROUND(AVERAGE(F44:F45),0)'
    ws['E10'] = 0.15
    ws['E10'].number_format = '0%'
    ws['F10'] = '=D10*E10'
    ws['F10'].number_format = '+0.00;-0.00;0.00'
    ws['G10'] = ('=IF(D10=2,"会社予想を大幅上回る",'
                'IF(D10=1,"会社予想を上回る",'
                'IF(D10=0,"会社予想と一致",'
                'IF(D10=-1,"会社予想を下回る","会社予想を大幅下回る"))))')

    # Total Score
    ws['B12'] = 'Total Score'
    ws['B12'].font = SUBHEADER_FONT
    ws['F12'] = '=SUM(F7:F10)'
    ws['F12'].number_format = '+0.00;-0.00;0.00'
    ws['F12'].font = Font(name='Calibri', size=12, bold=True)
    ws['F12'].fill = HIGHLIGHT_FILL
    ws['G12'] = ('=IF(F12>=1,"STRONG BUY",'
                 'IF(F12>=0.5,"BUY",'
                 'IF(F12>=-0.5,"HOLD",'
                 'IF(F12>=-1,"CAUTION","AVOID"))))')
    ws['G12'].font = Font(name='Calibri', size=12, bold=True)
    ws['G12'].fill = VERDICT_FILL

    for r in range(7, 11):
        for c in 'BCDEFG':
            ws[f'{c}{r}'].border = BORDER
    for c in 'BCDEFG':
        ws[f'{c}12'].border = BORDER

    # ────────── Block 2: Price Momentum Inputs ──────────
    ws['B17'] = 'Block 2: Price Momentum Inputs'
    ws['B17'].font = HEADER_FONT
    ws['B17'].fill = HEADER_FILL
    ws.merge_cells('B17:G17')

    momentum_rows = [
        ('Current Price',    config.get('current_price'), 19),
        ('Price 3M ago',     config.get('price_3m_ago'),  20),
        ('Price 1M ago',     config.get('price_1m_ago'),  21),
        ('Recent High',      config.get('price_high'),    22),
        ('Recent Low',       config.get('price_low'),     23),
    ]
    for label, val, r in momentum_rows:
        ws[f'B{r}'] = label
        ws[f'C{r}'] = val if val is not None else ''
        ws[f'C{r}'].fill = INPUT_FILL
        ws[f'C{r}'].number_format = '#,##0'
        ws[f'B{r}'].border = BORDER
        ws[f'C{r}'].border = BORDER

    ws['B25'] = '── Derived metrics ──'
    ws['B25'].font = SUBHEADER_FONT
    ws['B26'] = '3M Change'
    ws['C26'] = '=IFERROR((C19-C20)/C20,"")'
    ws['C26'].number_format = '+0.00%;-0.00%;0.00%'
    ws['B27'] = '1M Change'
    ws['C27'] = '=IFERROR((C19-C21)/C21,"")'
    ws['C27'].number_format = '+0.00%;-0.00%;0.00%'
    ws['B28'] = 'High-to-Current'
    ws['C28'] = '=IFERROR((C19-C22)/C22,"")'
    ws['C28'].number_format = '+0.00%;-0.00%;0.00%'

    # ────────── Block 3: Margin (信用残) Inputs ──────────
    ws['B31'] = 'Block 3: Margin Trading Inputs'
    ws['B31'].font = HEADER_FONT
    ws['B31'].fill = HEADER_FILL
    ws.merge_cells('B31:G31')

    margin_rows = [
        ('信用買残 (千株 / 1k shares)',          config.get('margin_buy'),         33),
        ('信用売残 (千株 / 1k shares)',          config.get('margin_sell'),        34),
        ('過去6ヶ月 売残ピーク (千株)',           config.get('margin_sell_peak_6m'), 35),
    ]
    for label, val, r in margin_rows:
        ws[f'B{r}'] = label
        ws[f'C{r}'] = val if val is not None else ''
        ws[f'C{r}'].fill = INPUT_FILL
        ws[f'C{r}'].number_format = '#,##0'
        ws[f'B{r}'].border = BORDER
        ws[f'C{r}'].border = BORDER

    # Computed metrics
    ws['B36'] = '信用倍率 (買 / 売)'
    ws['C36'] = '=IFERROR(C33/C34,99)'
    ws['C36'].number_format = '0.00'
    ws['C36'].fill = OUTPUT_FILL
    ws['B37'] = '売残充実度 (現在 / ピーク)'
    ws['C37'] = '=IFERROR(C34/C35,0)'
    ws['C37'].number_format = '0.00'
    ws['C37'].fill = OUTPUT_FILL
    for r in (36, 37):
        ws[f'B{r}'].border = BORDER
        ws[f'C{r}'].border = BORDER

    # ────────── Block 4: Forecast Gap Inputs ──────────
    ws['B41'] = 'Block 4: Forecast Gap Inputs'
    ws['B41'].font = HEADER_FONT
    ws['B41'].fill = HEADER_FILL
    ws.merge_cells('B41:G41')

    for i, h in enumerate(['Item', 'Self (DCF Base)', 'Company forecast', 'Gap', 'Score']):
        c = chr(ord('B') + i)
        ws[f'{c}43'] = h
        ws[f'{c}43'].font = SUBHEADER_FONT
        ws[f'{c}43'].fill = SUBHEADER_FILL

    # Row 44: 営業利益成長率
    ws['B44'] = '営業利益成長率'
    ws['C44'] = dd['my_op_growth']  # embedded value
    ws['C44'].number_format = '+0.00%;-0.00%;0.00%'
    ws['C44'].fill = OUTPUT_FILL
    ws['D44'] = config.get('company_op_growth', 0)
    ws['D44'].number_format = '+0.00%;-0.00%;0.00%'
    ws['D44'].fill = INPUT_FILL
    ws['E44'] = '=C44-D44'
    ws['E44'].number_format = '+0.00%;-0.00%;0.00%'
    ws['F44'] = '=IF(E44>=0.10,2,IF(E44>=0.05,1,IF(E44>=-0.05,0,IF(E44>=-0.10,-1,-2))))'

    # Row 45: 売上成長率
    ws['B45'] = '売上成長率'
    ws['C45'] = dd['my_rev_growth']
    ws['C45'].number_format = '+0.00%;-0.00%;0.00%'
    ws['C45'].fill = OUTPUT_FILL
    ws['D45'] = config.get('company_rev_growth', 0)
    ws['D45'].number_format = '+0.00%;-0.00%;0.00%'
    ws['D45'].fill = INPUT_FILL
    ws['E45'] = '=C45-D45'
    ws['E45'].number_format = '+0.00%;-0.00%;0.00%'
    ws['F45'] = '=IF(E45>=0.10,2,IF(E45>=0.05,1,IF(E45>=-0.05,0,IF(E45>=-0.10,-1,-2))))'

    for r in (44, 45):
        for c in 'BCDEF':
            ws[f'{c}{r}'].border = BORDER

    # Optional analyst consensus rows
    if config.get('analyst_consensus_op') is not None:
        ws['B47'] = 'Analyst Consensus OP growth'
        ws['C47'] = config['analyst_consensus_op']
        ws['C47'].number_format = '+0.00%;-0.00%;0.00%'
        ws['C47'].fill = INPUT_FILL
    if config.get('analyst_consensus_rev') is not None:
        ws['B48'] = 'Analyst Consensus Rev growth'
        ws['C48'] = config['analyst_consensus_rev']
        ws['C48'].number_format = '+0.00%;-0.00%;0.00%'
        ws['C48'].fill = INPUT_FILL

    # ────────── Block 5: Interpretation Guide ──────────
    ws['B53'] = 'Block 5: Interpretation Guide'
    ws['B53'].font = HEADER_FONT
    ws['B53'].fill = HEADER_FILL
    ws.merge_cells('B53:G53')

    guide = [
        ('STRONG BUY (Total >= +1.0)',       '4要素のうち複数が大悲観/サプライズ余地あり。市場との乖離が大きく、買いシグナル強い。'),
        ('BUY        (+0.5 .. +1.0)',        '買い候補。ファンダメンタルズが市場の織り込みを上回る公算。'),
        ('HOLD       (-0.5 .. +0.5)',        '市場とおおむね整合。様子見。'),
        ('CAUTION    (-1.0 .. -0.5)',        '市場過熱気味。下げリスクあり、保有量を減らす検討。'),
        ('AVOID      (Total < -1.0)',        '市場の楽観が過度。失望リスク高い。'),
    ]
    for i, (lbl, desc) in enumerate(guide):
        r = 55 + i
        ws[f'B{r}'] = lbl
        ws[f'B{r}'].font = SUBHEADER_FONT
        ws[f'C{r}'] = desc
        ws[f'C{r}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'C{r}:G{r}')

    ws['B62'] = '── 使い方 / How to use ──'
    ws['B62'].font = SUBHEADER_FONT
    usage = [
        '1. Block 2 / 3 / 4 の黄色入力欄に最新の市場データを入力する。',
        f"2. Implied Growth Analysis シートで Block 5 の {market_alpha_cell.replace('C', 'C')} 行近くで My View alpha を入力する。",
        '3. Block 1 (this sheet) Total Score と判定を確認する。',
        '4. 複数銘柄で同じテンプレートを使用し、横比較で買い候補を絞り込む。',
    ]
    for i, txt in enumerate(usage):
        ws[f'B{63+i}'] = txt

    return ws


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def generate_market_analysis_excel(config, output_path, dcf_excel_path=None):
    """Generate a 2-sheet Market Analysis Excel for any ticker.

    Args:
        config: dict with required keys (ticker, company_name, current_price,
                segment_layout) and optional momentum/margin/forecast inputs.
        output_path: target .xlsx path.
        dcf_excel_path: optional path to a DCF Excel from dcf_comps_template.py.
                        When provided every macro and segment value is read from
                        it and embedded as a value (no external references).

    Returns:
        str: output_path
    """
    if not config.get('ticker') or not config.get('company_name'):
        raise ValueError("config requires 'ticker' and 'company_name'.")
    if 'segment_layout' not in config or not config['segment_layout'].get('segments'):
        raise ValueError("config['segment_layout']['segments'] is required.")

    # Pull DCF data (all values, not formulas)
    if dcf_excel_path and os.path.exists(dcf_excel_path):
        dd = extract_dcf_data(dcf_excel_path, config['segment_layout'])
    else:
        if dcf_excel_path:
            print(f"  [warn] DCF Excel not found: {dcf_excel_path} -- using config-supplied fallbacks.")
        dd = _build_fallback_data(config)

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    iga_info = _build_implied_growth_sheet(wb, config, dd)
    market_alpha_cell = iga_info['market_alpha_cell']
    _build_market_scorecard_sheet(wb, config, dd, market_alpha_cell=market_alpha_cell)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or '.', exist_ok=True)
    wb.save(output_path)
    print(f"  Wrote: {output_path}")

    # Sanity check: alpha=1.0 implied price should be close to the DCF Base
    # PGM target. Large discrepancy usually points to inconsistent NWC handling
    # or a mismatched stub fraction.
    _verify_alpha_one_matches_base_pgm(dd)

    return output_path


def _verify_alpha_one_matches_base_pgm(dd, threshold_pct=5.0):
    """Recompute the alpha=1 implied share price in Python (same formulas as
    the Excel reverse DCF) and compare against the DCF Base PGM target.

    We do this in Python (not by re-opening the saved workbook) because
    openpyxl returns formulas, not computed values, and LibreOffice/Excel
    may not be available for headless recalc on every machine.
    """
    fy_rev = dd['fy_actual_total_rev']
    fy_opm = dd['fy_actual_opm']
    rev = fy_rev
    revs = []
    for yr in range(5):
        rev *= (1 + 1.0 * dd['base_total_growth'][yr])
        revs.append(rev)
    ops = [revs[yr] * (fy_opm + 1.0 * (dd['base_total_opm'][yr] - fy_opm)) for yr in range(5)]
    fcfs = [
        max(0.0, ops[yr] * (1 - dd['tax_rate'])) + dd['da'][yr] - dd['capex'][yr] - dd['nwc_change'][yr]
        for yr in range(5)
    ]
    sum_pv = sum(fcfs[yr] / (1 + dd['wacc'])**(dd['stub_fraction'] + yr) for yr in range(5))
    if dd['wacc'] > dd['terminal_growth']:
        tv = fcfs[-1] * (1 + dd['terminal_growth']) / (dd['wacc'] - dd['terminal_growth'])
    else:
        tv = 0
    pv_tv = tv / (1 + dd['wacc'])**(dd['stub_fraction'] + 4)
    ev = sum_pv + pv_tv
    equity = ev - dd['net_debt']
    alpha_one_price = equity * 1_000_000 / dd['shares'] if dd['shares'] else 0.0

    base_pgm = dd['base_pgm_target']
    if base_pgm <= 0:
        return
    diff_pct = (alpha_one_price - base_pgm) / base_pgm * 100

    print(f"\n  === Verification: alpha=1.0 vs DCF Base PGM ===")
    print(f"    Implied Price (alpha=1.0): JPY {alpha_one_price:,.0f}")
    print(f"    DCF Base PGM Target:       JPY {base_pgm:,.0f}")
    print(f"    Difference:                {diff_pct:+.2f}%")

    if abs(diff_pct) > threshold_pct:
        import warnings as _w
        _w.warn(
            f"alpha=1.0 implied price (JPY {alpha_one_price:,.0f}) differs from "
            f"DCF Base PGM target (JPY {base_pgm:,.0f}) by {diff_pct:+.2f}% "
            f"(>{threshold_pct}% threshold). Likely cause: DCF active scenario "
            f"!= Base, NWC scaling mismatch, or stub fraction inconsistency.",
            UserWarning, stacklevel=2,
        )


def _build_fallback_data(config):
    """Construct a minimal data dict when no DCF Excel is available.

    Looks for these optional config keys: wacc, terminal_growth, tax_rate,
    net_debt, shares_outstanding, stub_fraction. Segments must still come
    from segment_layout but with config-supplied per-segment fields.
    """
    seg_layout = config['segment_layout']
    n_seg = len(seg_layout['segments'])
    # Without a DCF file we cannot compute Base scenario projections; bail
    # gracefully unless the caller provided full overrides.
    raise NotImplementedError(
        "Fallback path without dcf_excel_path requires a richer config "
        "schema (per-segment Base growth/OPM, D&A/Capex/NWC arrays, etc.). "
        "Pass dcf_excel_path to use the auto-extraction path instead."
    )


# ---------------------------------------------------------------------------
# Standalone demo: 株式会社コア (2359) post-Vision2029
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(here, '..'))

    config_2359 = {
        "ticker": "2359.T",
        "company_name": "株式会社コア",
        "current_price": 2006,
        "entry_price":   2127,
        "segment_layout": {
            "segments": [
                {"name": "未来社会ソリューション",
                 "dcf_fy26_cell": "F6",
                 "dcf_growth_base_row": 34,
                 "dcf_opm_base_row": 41},
                {"name": "産業技術ソリューション",
                 "dcf_fy26_cell": "F11",
                 "dcf_growth_base_row": 48,
                 "dcf_opm_base_row": 55},
                {"name": "顧客共創ビジネス",
                 "dcf_fy26_cell": "F16",
                 "dcf_growth_base_row": 62,
                 "dcf_opm_base_row": 69},
            ]
        },
        "price_3m_ago": 1964,
        "price_1m_ago": 2330,
        "price_high":   2330,
        "price_low":    1960,
        "margin_buy":          50000,
        "margin_sell":          5100,
        "margin_sell_peak_6m": 169300,
        "company_op_growth":  0.113,
        "company_rev_growth": 0.093,
    }

    dcf_path = os.path.join(project_root, "models", "2359_DCF_Model_20260509.xlsx")
    out_path = os.path.join(project_root, "reports", "2359_market_analysis_20260509_v2.xlsx")

    generate_market_analysis_excel(config_2359, out_path, dcf_excel_path=dcf_path)
