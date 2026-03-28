"""
dcf_comps_build_v3.py - DCF / Comps Equity Research Excel Generator (V3)

Can be used standalone (with PDF extraction) or imported as a library:
  - Standalone: python templates/dcf_comps_template.py
  - Library: from dcf_comps_template import generate_dcf_workbook
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import subprocess, sys, os

try:
    import yfinance as yf
    YFINANCE_AVAILABLE = True
except ImportError:
    YFINANCE_AVAILABLE = False

# =====================================================================
# V3: ROW NUMBERS — Full waterfall, no SGA_OFFSET toggle
# =====================================================================
R_DRV_GROWTH   = 30  # driver row: Revenue Growth (YoY)
R_DRV_COGS     = 31  # driver row: COGS % of Revenue
R_DRV_SGA      = 32  # driver row: SGA Expense
R_REVENUE      = 33
R_COGS         = 34
R_GROSS_PROFIT = 35
R_GROSS_MARGIN = 36
R_SGA          = 37
R_OP_M_IMPL   = 38
R_EBIT         = 39
R_TAX          = 40
R_NOPAT        = 41
R_DA           = 42
R_CAPEX        = 43
R_CHG_NWC      = 44  # Change in NWC (linked from NWC Schedule)
R_UFCF         = 45
R_DISC         = 46
R_PV_FCF       = 47
# Stub period assumption rows
R_STUB_FRACTION = 19
R_LTM_REVENUE   = 20

# PGM section: R_PV_FCF + 2 gap
R_PGM_SEC    = R_PV_FCF + 2
R_SUM_PV     = R_PGM_SEC + 1
R_TV_PGM     = R_SUM_PV + 1
R_PV_TV_PGM  = R_TV_PGM + 1
R_EV_PGM     = R_PV_TV_PGM + 1
R_EQ_PGM     = R_EV_PGM + 1
R_PRICE_PGM  = R_EQ_PGM + 1
# Exit section: R_PRICE_PGM + 2 gap
R_EXIT_SEC   = R_PRICE_PGM + 2
R_SUM_PV_EX  = R_EXIT_SEC + 1
R_YR5_EBITDA = R_SUM_PV_EX + 1
R_TV_EXIT    = R_YR5_EBITDA + 1
R_PV_TV_EXIT = R_TV_EXIT + 1
R_EV_EXIT    = R_PV_TV_EXIT + 1
R_EQ_EXIT    = R_EV_EXIT + 1
R_PRICE_EXIT = R_EQ_EXIT + 1

# Scenario Matrix Section (below Exit valuation)
SCENARIO_NAMES = ["Base", "Upside", "Management", "Downside 1", "Downside 2"]
NUM_SCENARIOS  = 5

R_SCEN_SEC        = R_PRICE_EXIT + 2       # Section header
R_SCEN_YEARS      = R_SCEN_SEC + 1         # Year 1-5 column headers

# Each block: sub-header 1 row + 5 scenario rows + 1 blank = 7 rows
R_SCEN_BLK_GROWTH = R_SCEN_YEARS + 1
R_SCEN_BLK_COGS   = R_SCEN_BLK_GROWTH + 7
R_SCEN_BLK_SGA    = R_SCEN_BLK_COGS + 7

# ── NWC Schedule Row Numbers ──
NWC_R_DSO      = 5
NWC_R_DIH      = 6
NWC_R_DPO      = 7
NWC_R_REV      = 9
NWC_R_COGS     = 10
NWC_R_AR       = 12
NWC_R_INV      = 13
NWC_R_CA       = 14
NWC_R_AP       = 15
NWC_R_CL       = 16
NWC_R_NWC      = 18
NWC_R_CHG_NWC  = 19

NWC_R_SCEN_SEC      = 22
NWC_R_SCEN_YEARS    = 23
NWC_R_SCEN_BLK_DSO  = 24
NWC_R_SCEN_BLK_DIH  = NWC_R_SCEN_BLK_DSO + 7
NWC_R_SCEN_BLK_DPO  = NWC_R_SCEN_BLK_DIH + 7

# =====================================================================
# STYLE CONSTANTS
# =====================================================================
BLUE_FONT   = Font(name="Arial", size=10, color="000000", bold=False)  # unified to black
BLACK_FONT  = Font(name="Arial", size=10, color="000000")
GREEN_FONT  = Font(name="Arial", size=10, color="006600")
BOLD_FONT   = Font(name="Arial", size=10, bold=True)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Arial", size=14, bold=True)
SUB_FONT    = Font(name="Arial", size=11, bold=True)
GREY_FONT   = Font(name="Arial", size=9, italic=True, color="808080")

HEADER_FILL    = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
LIGHT_FILL     = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_GREEN    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SUBTOTAL_FILL  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Borders
THIN_BORDER     = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))
_GRAY_SIDE      = Side(style="thin", color="B0B0B0")
LIGHT_GRAY_BORDER = Border(bottom=Side(style="dotted", color="B0B0B0"))
_GRAY_HAIR       = Side(style="hair", color="B0B0B0")
NWC_DATA_BORDER  = Border(left=_GRAY_HAIR, right=_GRAY_HAIR,
                           top=_GRAY_HAIR, bottom=_GRAY_HAIR)
SECTION_BOTTOM  = Border(bottom=Side(style="thin"))
SUBTOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="thin"))
TOP_BOTTOM      = Border(top=Side(style="thin"), bottom=Side(style="double"))
INPUT_BORDER    = Border(left=_GRAY_SIDE, right=_GRAY_SIDE,
                          top=_GRAY_SIDE, bottom=_GRAY_SIDE)

FMT_YEN     = '#,##0;(#,##0)'
FMT_YEN_DEC = '#,##0.0;(#,##0.0)'
FMT_PCT     = '0.0%;(0.0%)'
FMT_PCT2    = '0.00%;(0.00%)'
FMT_RATIO   = '0.00"x"'
FMT_INT     = '#,##0'
FMT_EPS     = '#,##0.0;(#,##0.0)'
FMT_DAYS    = '#,##0'

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fmt:       c.number_format = fmt
    if fill:      c.fill = fill
    if border:    c.border = border
    if alignment: c.alignment = alignment
    return c

def header_row(ws, row, col_start, col_end, labels, fill=HEADER_FILL, font=HEADER_FONT):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = SECTION_BOTTOM

def section_title(ws, row, col, text, font=SUB_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font
    return c

def col_letter(col_num):
    return get_column_letter(col_num)

def choose_formula(block_start, cl):
    """Generate CHOOSE formula referencing 5 scenario rows in the matrix."""
    refs = [f"{cl}{block_start + 1 + s}" for s in range(NUM_SCENARIOS)]
    return f"=CHOOSE($D$27,{','.join(refs)})"

def nwc_choose_formula(block_start, cl):
    """Generate CHOOSE formula for NWC Schedule referencing DCF Model scenario index."""
    refs = [f"{cl}{block_start + 1 + s}" for s in range(NUM_SCENARIOS)]
    return f"=CHOOSE('DCF Model'!$D$27,{','.join(refs)})"

def seg_choose_formula(scenario_rows, cl):
    """Generate CHOOSE formula for Segment Analysis referencing DCF Model scenario index.

    Args:
        scenario_rows: list of 5 row numbers (one per scenario: Base, Upside, Mgmt, DS1, DS2)
        cl: column letter
    Returns:
        Excel CHOOSE formula string
    """
    refs = [f"{cl}{r}" for r in scenario_rows]
    return f"=CHOOSE('DCF Model'!$D$27,{','.join(refs)})"

# =====================================================================
# SENSITIVITY ANALYSIS HELPERS (V3: full waterfall)
# =====================================================================
def calc_dcf_pgm(rev_growth, gross_margin, wacc, tg, cfg):
    """Calculate implied share price using Perpetuity Growth Method.
    V3: Uses gross_margin (GM%) and absolute SGA to compute EBIT.
    """
    n = cfg["projection_years"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]
    sga_pct_list = cfg["sga_pct"]
    _capex_method = cfg.get("capex_method", "revenue_pct")
    _da_method = cfg.get("da_method", "revenue_pct")
    _capex_direct = cfg.get("capex_direct", {}).get("projections", []) if _capex_method == "direct" else []
    _da_direct = cfg.get("da_direct", {}).get("projections", []) if _da_method == "direct" else []

    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_fcf = 0
    for yr_idx, rev in enumerate(revenues):
        cogs = rev * (1 - gross_margin)
        gp = rev - cogs
        sga = rev * sga_pct_list[yr_idx]
        ebit = gp - sga
        nopat = ebit * (1 - tax) if ebit > 0 else ebit
        if _da_method == "direct" and yr_idx < len(_da_direct) and _da_direct[yr_idx] is not None:
            da = _da_direct[yr_idx]
        else:
            da = rev * da_pct
        if _capex_method == "direct" and yr_idx < len(_capex_direct) and _capex_direct[yr_idx] is not None:
            capex = _capex_direct[yr_idx]
        else:
            capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_fcf = fcf

    tv = last_fcf * (1 + tg) / (wacc - tg)
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

def calc_dcf_exit(rev_growth, gross_margin, wacc, exit_mult, cfg):
    """Calculate implied share price using Exit Multiple Method."""
    n = cfg["projection_years"]
    capex_pct = cfg["capex_pct"]
    da_pct = cfg["da_pct"]
    tax = cfg["tax_rate"]
    net_debt = cfg["net_debt"]
    shares = cfg["shares_outstanding"]
    sga_pct_list = cfg["sga_pct"]
    _capex_method = cfg.get("capex_method", "revenue_pct")
    _da_method = cfg.get("da_method", "revenue_pct")
    _capex_direct = cfg.get("capex_direct", {}).get("projections", []) if _capex_method == "direct" else []
    _da_direct = cfg.get("da_direct", {}).get("projections", []) if _da_method == "direct" else []

    revenues = []
    rev = cfg["base_year_revenue"]
    for _ in range(n):
        rev = rev * (1 + rev_growth)
        revenues.append(rev)

    sum_pv_fcf = 0
    last_ebit = 0
    for yr_idx, rev in enumerate(revenues):
        cogs = rev * (1 - gross_margin)
        gp = rev - cogs
        sga = rev * sga_pct_list[yr_idx]
        ebit = gp - sga
        nopat = ebit * (1 - tax) if ebit > 0 else ebit
        if _da_method == "direct" and yr_idx < len(_da_direct) and _da_direct[yr_idx] is not None:
            da = _da_direct[yr_idx]
        else:
            da = rev * da_pct
        if _capex_method == "direct" and yr_idx < len(_capex_direct) and _capex_direct[yr_idx] is not None:
            capex = _capex_direct[yr_idx]
        else:
            capex = rev * capex_pct
        fcf = nopat + da - capex
        df = 1 / (1 + wacc) ** (yr_idx + 1)
        sum_pv_fcf += fcf * df
        last_ebit = ebit

    if _da_method == "direct" and len(_da_direct) >= n and _da_direct[n - 1] is not None:
        yr5_da = _da_direct[n - 1]
    else:
        yr5_da = revenues[-1] * da_pct
    yr5_ebitda = last_ebit + yr5_da
    tv = yr5_ebitda * exit_mult
    pv_tv = tv / (1 + wacc) ** n
    ev = sum_pv_fcf + pv_tv
    equity = ev - net_debt
    price = round(equity * 1_000_000 / shares)
    return price

# =====================================================================
# Derived values for WACC
# =====================================================================
def calc_wacc(cfg):
    ke = cfg["risk_free"] + cfg["beta"] * cfg["erp"] + cfg["size_premium"]
    we = 1 / (1 + cfg["de_ratio"])
    wd = cfg["de_ratio"] / (1 + cfg["de_ratio"])
    return ke * we + cfg["cost_of_debt_at"] * wd

# =====================================================================
# BUILD WORKBOOK
# =====================================================================

# V2: DYNAMIC STOCK DATA FETCHING
# =====================================================================
def get_live_market_data(ticker_str, fallback_price, fallback_shares):
    if not YFINANCE_AVAILABLE:
        print("yfinance not installed. Using fallback market data.")
        return fallback_price, fallback_shares, 1.0

    try:
        print(f"Fetching live data for {ticker_str} via yfinance...")
        tkr = yf.Ticker(ticker_str)
        info = tkr.info
        live_price = info.get("currentPrice") or info.get("regularMarketPrice") or fallback_price
        live_shares = info.get("sharesOutstanding") or fallback_shares
        raw_beta = info.get("beta")
        if raw_beta and 0.6 <= raw_beta <= 1.5:
            live_beta = raw_beta
        else:
            live_beta = 1.0  # sector-standard fallback
            print(f"  Beta {raw_beta} outside [0.6, 1.5] range - using fallback 1.0")
        print(f"Successfully fetched: Price={live_price}, Shares={live_shares}, Beta={live_beta}")
        return float(live_price), int(live_shares), float(live_beta)
    except Exception as e:
        print(f"Warning: Failed to fetch live data ({str(e).encode('ascii', 'replace').decode()}). Using fallback market data.")
        return fallback_price, fallback_shares, 1.0


# =====================================================================
# SEGMENT ANALYSIS SHEET
# =====================================================================
def _create_segment_sheet(wb, C, segments, proj_years, year_labels):
    """Generate Segment Analysis sheet from overrides JSON segments data.

    v2: Revenue uses YoY growth rates (not absolute values).
    Revenue display = Base Year × (1+growth) chain via CHOOSE formulas.
    Includes Consolidated Inputs section (SGA%, NWC%) after segment blocks.

    Returns:
        dict with total_rev_row, total_op_row, n_hist,
              sga_scenario_rows, nwc_scenario_rows for cross-sheet references.
    """
    ws = wb.create_sheet("Segment Analysis")
    ws.sheet_properties.tabColor = "8B0000"  # Dark red

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34

    # Determine historical years from segment data
    n_hist = 0
    for seg in segments:
        hist_rev = seg.get("historical", {}).get("revenue", [])
        if len(hist_rev) > n_hist:
            n_hist = len(hist_rev)

    n_data_cols = n_hist + proj_years
    for ci in range(n_data_cols):
        ws.column_dimensions[col_letter(3 + ci)].width = 16

    # Check if segments have revenue_growth (v2) or revenue only (v1 fallback)
    _has_growth = any(
        seg.get("projections", {}).get("revenue_growth")
        for seg in segments
    )
    if not _has_growth:
        print("WARNING: segments use absolute 'revenue' without 'revenue_growth'. "
              "Falling back to v1 absolute revenue mode.")

    # Check for consolidated-level scenario data
    _scenarios = C.get("scenarios", {})
    _has_nwc_pct = (
        C.get("nwc_method") == "revenue_pct"
        and all("nwc_pct" in _scenarios[sn] for sn in SCENARIO_NAMES if sn in _scenarios)
    )

    # Title
    set_cell(ws, 2, 2, f'Segment Analysis - {C["company_name"]}', font=TITLE_FONT)

    # Header row: historical FY labels + projection year labels
    hist_labels = []
    if n_hist >= 3:
        hist_labels = ["FY-2", "FY-1", "FY0 (Base)"]
    elif n_hist == 2:
        hist_labels = ["FY-1", "FY0 (Base)"]
    elif n_hist == 1:
        hist_labels = ["FY0 (Base)"]

    # If projection_start_fy is available, use actual FY labels
    if C.get("projection_start_fy"):
        import re as _re
        _m = _re.search(r"FY(\d+)", C["projection_start_fy"])
        if _m:
            _base_fy = int(_m.group(1))
            hist_labels = [f"FY{_base_fy - n_hist + 1 + i}" for i in range(n_hist)]

    all_headers = hist_labels + year_labels
    header_row(ws, 4, 3, 3 + len(all_headers) - 1, all_headers)

    # ─────────────────────────────────────────────────────────────────
    # FIRST PASS: Build Segment Scenario Input Matrix to know row numbers
    # before writing display area (display area CHOOSE formulas need them).
    # ─────────────────────────────────────────────────────────────────

    # Pre-calculate display area row count to determine matrix start row.
    # Per segment: 1 header + 1 rev + 1 op + 1 opm + 1 blank = 5 rows
    # Total block: 1 header + 1 rev + 1 op + 1 opm + 1 blank = 5 rows
    # Reconciliation: 1 header + 3 rows + 1 blank = 5 rows
    display_rows = len(segments) * 5 + 5 + 5
    matrix_start = 5 + display_rows + 1  # +1 for extra gap

    # Build scenario input matrix row map
    # v2: Revenue Growth (%) + OP Margin blocks per segment
    # Each segment block: header(1) + 5 scenario rows + blank(1) = 7 rows for Revenue Growth
    #                     header(1) + 5 scenario rows + blank(1) = 7 rows for OPM
    # Total per segment: 14 rows
    seg_matrix_info = []
    matrix_cur = matrix_start + 2  # +2 for section header + year headers

    for seg_idx, seg in enumerate(segments):
        # Revenue Growth block (v2)
        rev_sub_header = matrix_cur
        rev_scenario_rows = [matrix_cur + 1 + s for s in range(NUM_SCENARIOS)]
        matrix_cur += 1 + NUM_SCENARIOS + 1  # sub-header + 5 rows + blank

        # OP Margin block
        opm_sub_header = matrix_cur
        opm_scenario_rows = [matrix_cur + 1 + s for s in range(NUM_SCENARIOS)]
        matrix_cur += 1 + NUM_SCENARIOS + 1  # sub-header + 5 rows + blank

        seg_matrix_info.append({
            "rev_sub_header": rev_sub_header,
            "rev_scenario_rows": rev_scenario_rows,
            "opm_sub_header": opm_sub_header,
            "opm_scenario_rows": opm_scenario_rows,
        })

    # ── Consolidated Inputs section (COGS%, SGA%, NWC%) after segment blocks ──
    consolidated_start = matrix_cur + 1  # gap row

    # Check if cogs_pct exists in scenarios
    _has_cogs_pct = any("cogs_pct" in _scenarios.get(sn, {}) for sn in SCENARIO_NAMES)

    # COGS% block (if present): sub-header(1) + 5 rows + blank(1) = 7 rows
    if _has_cogs_pct:
        cogs_sub_header_row = consolidated_start + 2  # after section header + year headers
        cogs_scenario_rows = [cogs_sub_header_row + 1 + s for s in range(NUM_SCENARIOS)]
        sga_sub_header_row = cogs_sub_header_row + 1 + NUM_SCENARIOS + 1  # after COGS block + blank
    else:
        cogs_sub_header_row = None
        cogs_scenario_rows = None
        sga_sub_header_row = consolidated_start + 2  # original position

    sga_scenario_rows = [sga_sub_header_row + 1 + s for s in range(NUM_SCENARIOS)]
    nwc_scenario_rows = None
    if _has_nwc_pct:
        nwc_sub_header_row = sga_sub_header_row + 1 + NUM_SCENARIOS + 1  # after SGA block + blank
        nwc_scenario_rows = [nwc_sub_header_row + 1 + s for s in range(NUM_SCENARIOS)]

    # ─────────────────────────────────────────────────────────────────
    # DISPLAY AREA: Per-segment blocks with CHOOSE formulas
    # ─────────────────────────────────────────────────────────────────
    cur_row = 5  # Start of segment blocks

    seg_rev_rows = []   # Track revenue row for each segment (for Total SUM)
    seg_op_rows = []    # Track OP row for each segment

    for seg_idx, seg in enumerate(segments):
        seg_name = seg.get("name", f"Segment {seg_idx + 1}")
        seg_name_jp = seg.get("name_jp", "")
        display_name = f"{seg_name}" + (f" ({seg_name_jp})" if seg_name_jp else "")

        hist = seg.get("historical", {})
        hist_rev = hist.get("revenue", [None] * n_hist)
        hist_op = hist.get("op", [None] * n_hist)

        # Pad historical arrays to n_hist
        while len(hist_rev) < n_hist:
            hist_rev.insert(0, None)
        while len(hist_op) < n_hist:
            hist_op.insert(0, None)

        mi = seg_matrix_info[seg_idx]

        # ── Segment header ──
        c = section_title(ws, cur_row, 2, display_name)
        c.fill = LIGHT_FILL
        for ci in range(3, 3 + n_data_cols):
            ws.cell(row=cur_row, column=ci).fill = LIGHT_FILL
        cur_row += 1

        # ── Revenue row ──
        rev_row = cur_row
        seg_rev_rows.append(rev_row)
        set_cell(ws, rev_row, 2, "Revenue", font=BOLD_FONT)

        # Historical revenue
        for i in range(n_hist):
            col = 3 + i
            val = hist_rev[i]
            if val is not None:
                set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                         border=NWC_DATA_BORDER)
            else:
                set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

        # Projected revenue — v2: Base Year × (1+growth) chain via CHOOSE
        if _has_growth:
            base_year_col = col_letter(3 + n_hist - 1)  # last historical column
            for yr in range(proj_years):
                col = 3 + n_hist + yr
                cl = col_letter(col)
                # Growth rate CHOOSE references from scenario matrix
                growth_refs = [f"{cl}{r}" for r in mi["rev_scenario_rows"]]
                growth_choose = f"CHOOSE('DCF Model'!$D$27,{','.join(growth_refs)})"
                if yr == 0:
                    # Y1: =BaseYearRev × (1 + CHOOSE(growth))
                    formula = f"={base_year_col}{rev_row}*(1+{growth_choose})"
                else:
                    prev_cl = col_letter(col - 1)
                    # Y2+: =PrevYearRev × (1 + CHOOSE(growth))
                    formula = f"={prev_cl}{rev_row}*(1+{growth_choose})"
                set_cell(ws, rev_row, col, formula, font=BLACK_FONT, fmt=FMT_YEN,
                         border=NWC_DATA_BORDER)
        else:
            # v1 fallback: absolute revenue CHOOSE
            for yr in range(proj_years):
                col = 3 + n_hist + yr
                cl = col_letter(col)
                formula = seg_choose_formula(mi["rev_scenario_rows"], cl)
                set_cell(ws, rev_row, col, formula, font=BLACK_FONT, fmt=FMT_YEN,
                         border=NWC_DATA_BORDER)
        cur_row += 1

        # ── Operating Profit row ──
        op_row = cur_row
        seg_op_rows.append(op_row)
        set_cell(ws, op_row, 2, "Operating Profit", font=BOLD_FONT)

        # Historical OP
        for i in range(n_hist):
            col = 3 + i
            val = hist_op[i]
            if val is not None:
                set_cell(ws, op_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                         border=NWC_DATA_BORDER)
            else:
                set_cell(ws, op_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

        # Projected OP = Revenue × CHOOSE(OP Margin from scenario matrix)
        for yr in range(proj_years):
            col = 3 + n_hist + yr
            cl = col_letter(col)
            opm_refs = [f"{cl}{r}" for r in mi["opm_scenario_rows"]]
            margin_choose = f"CHOOSE('DCF Model'!$D$27,{','.join(opm_refs)})"
            set_cell(ws, op_row, col,
                     f"={cl}{rev_row}*{margin_choose}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
        cur_row += 1

        # ── OP Margin row ──
        opm_row = cur_row
        set_cell(ws, opm_row, 2, "OP Margin", font=Font(name="Arial", size=10,
                 italic=True, color="808080"))

        for i in range(n_hist):
            col = 3 + i
            cl = col_letter(col)
            if hist_rev[i] is not None and hist_op[i] is not None:
                set_cell(ws, opm_row, col, f"={cl}{op_row}/{cl}{rev_row}",
                         font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
            else:
                set_cell(ws, opm_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

        # Projected OPM — CHOOSE from scenario matrix (display only)
        for yr in range(proj_years):
            col = 3 + n_hist + yr
            cl = col_letter(col)
            formula = seg_choose_formula(mi["opm_scenario_rows"], cl)
            set_cell(ws, opm_row, col, formula,
                     font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
        cur_row += 2  # blank row between segments

    # ═══════════════════════════════════════════════════════════════
    # TOTAL BLOCK
    # ═══════════════════════════════════════════════════════════════
    c = section_title(ws, cur_row, 2, "Consolidated Total")
    c.fill = LIGHT_GREEN
    for ci in range(3, 3 + n_data_cols):
        ws.cell(row=cur_row, column=ci).fill = LIGHT_GREEN
    cur_row += 1

    # Total Revenue
    total_rev_row = cur_row
    set_cell(ws, total_rev_row, 2, "Total Revenue", font=BOLD_FONT)
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        refs = [f"{cl}{r}" for r in seg_rev_rows]
        formula = f"={'+'.join(refs)}"
        set_cell(ws, total_rev_row, col, formula, font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # Total OP
    total_op_row = cur_row
    set_cell(ws, total_op_row, 2, "Total Operating Profit", font=BOLD_FONT)
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        refs = [f"{cl}{r}" for r in seg_op_rows]
        formula = f"={'+'.join(refs)}"
        set_cell(ws, total_op_row, col, formula, font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # Total OP Margin
    total_opm_row = cur_row
    set_cell(ws, total_opm_row, 2, "Total OP Margin", font=Font(name="Arial", size=10,
             italic=True, color="808080"))
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        set_cell(ws, total_opm_row, col,
                 f"=IFERROR({cl}{total_op_row}/{cl}{total_rev_row},\"—\")",
                 font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    cur_row += 2

    # ═══════════════════════════════════════════════════════════════
    # RECONCILIATION CHECK (projected years only)
    # ═══════════════════════════════════════════════════════════════
    c = section_title(ws, cur_row, 2, "Reconciliation vs DCF Model")
    c.fill = LIGHT_YELLOW
    for ci in range(3, 3 + n_data_cols):
        ws.cell(row=cur_row, column=ci).fill = LIGHT_YELLOW
    cur_row += 1

    set_cell(ws, cur_row, 2, "DCF Model Revenue", font=BOLD_FONT)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        dcf_col_letter = col_letter(3 + yr)
        set_cell(ws, cur_row, col,
                 f"='DCF Model'!{dcf_col_letter}{R_REVENUE}",
                 font=GREEN_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
    cur_row += 1

    set_cell(ws, cur_row, 2, "Segment Total Revenue", font=BOLD_FONT)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, cur_row, col,
                 f"={cl}{total_rev_row}",
                 font=BLACK_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
    cur_row += 1

    recon_diff_row = cur_row
    set_cell(ws, recon_diff_row, 2, "Difference", font=BOLD_FONT)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, recon_diff_row, col,
                 f"={cl}{cur_row - 1}-{cl}{cur_row - 2}",
                 font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)
    cur_row += 2

    # ═══════════════════════════════════════════════════════════════
    # SEGMENT SCENARIO INPUT MATRIX
    # ═══════════════════════════════════════════════════════════════
    # Section header
    c = section_title(ws, matrix_start, 2, "Segment Scenario Input Matrix")
    c.fill = PatternFill(start_color="E6CCE6", end_color="E6CCE6", fill_type="solid")
    for ci in range(3, 3 + n_data_cols):
        ws.cell(row=matrix_start, column=ci).fill = PatternFill(
            start_color="E6CCE6", end_color="E6CCE6", fill_type="solid")

    # Year headers (projection columns only)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        set_cell(ws, matrix_start + 1, col, year_labels[yr],
                 font=HEADER_FONT, fill=HEADER_FILL,
                 alignment=Alignment(horizontal="center"))

    # Per-segment scenario blocks
    for seg_idx, seg in enumerate(segments):
        seg_name = seg.get("name", f"Segment {seg_idx + 1}")
        proj = seg.get("projections", {})
        scenario_proj = seg.get("scenario_projections", {})

        mi = seg_matrix_info[seg_idx]

        # ── Revenue Growth block (v2) or Revenue absolute block (v1 fallback) ──
        if _has_growth:
            section_title(ws, mi["rev_sub_header"], 2,
                          f"{seg_name} - Revenue Growth (YoY)")
            for s, scen_name in enumerate(SCENARIO_NAMES):
                r = mi["rev_scenario_rows"][s]
                set_cell(ws, r, 2, scen_name, font=BOLD_FONT)

                # Get scenario-specific revenue_growth; fallback to Base projections
                if (scen_name in scenario_proj
                        and "revenue_growth" in scenario_proj[scen_name]):
                    scen_data = scenario_proj[scen_name]["revenue_growth"]
                else:
                    scen_data = proj.get("revenue_growth", [])

                for yr in range(proj_years):
                    col = 3 + n_hist + yr
                    val = scen_data[yr] if yr < len(scen_data) else None
                    if val is not None:
                        set_cell(ws, r, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                                 border=INPUT_BORDER)
                    else:
                        set_cell(ws, r, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
        else:
            # v1 fallback: absolute revenue
            section_title(ws, mi["rev_sub_header"], 2, f"{seg_name} - Revenue")
            for s, scen_name in enumerate(SCENARIO_NAMES):
                r = mi["rev_scenario_rows"][s]
                set_cell(ws, r, 2, scen_name, font=BOLD_FONT)

                if scen_name in scenario_proj and "revenue" in scenario_proj[scen_name]:
                    scen_rev = scenario_proj[scen_name]["revenue"]
                else:
                    scen_rev = proj.get("revenue", [])

                for yr in range(proj_years):
                    col = 3 + n_hist + yr
                    val = scen_rev[yr] if yr < len(scen_rev) else None
                    if val is not None:
                        set_cell(ws, r, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                                 border=INPUT_BORDER)
                    else:
                        set_cell(ws, r, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

        # ── OP Margin block (unchanged) ──
        section_title(ws, mi["opm_sub_header"], 2, f"{seg_name} - OP Margin")
        for s, scen_name in enumerate(SCENARIO_NAMES):
            r = mi["opm_scenario_rows"][s]
            set_cell(ws, r, 2, scen_name, font=BOLD_FONT)

            if scen_name in scenario_proj and "op_margin" in scenario_proj[scen_name]:
                scen_opm = scenario_proj[scen_name]["op_margin"]
            else:
                scen_opm = proj.get("op_margin", [])

            for yr in range(proj_years):
                col = 3 + n_hist + yr
                val = scen_opm[yr] if yr < len(scen_opm) else None
                if val is not None:
                    set_cell(ws, r, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                             border=INPUT_BORDER)
                else:
                    set_cell(ws, r, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

    # ═══════════════════════════════════════════════════════════════
    # CONSOLIDATED INPUTS (SGA%, NWC%) — v2 addition
    # ═══════════════════════════════════════════════════════════════
    CONSOL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    c = section_title(ws, consolidated_start, 2, "Consolidated Inputs (連結レベル)")
    c.fill = CONSOL_FILL
    for ci in range(3, 3 + n_data_cols):
        ws.cell(row=consolidated_start, column=ci).fill = CONSOL_FILL

    # Year headers
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        set_cell(ws, consolidated_start + 1, col, year_labels[yr],
                 font=HEADER_FONT, fill=HEADER_FILL,
                 alignment=Alignment(horizontal="center"))

    # ── COGS % of Revenue (only when cogs_pct exists in scenarios) ──
    if _has_cogs_pct and cogs_scenario_rows is not None:
        section_title(ws, cogs_sub_header_row, 2, "COGS % of Revenue")
        for s, scen_name in enumerate(SCENARIO_NAMES):
            r = cogs_scenario_rows[s]
            set_cell(ws, r, 2, scen_name, font=BOLD_FONT)
            scen_data = _scenarios.get(scen_name, {}).get("cogs_pct", [])
            for yr in range(proj_years):
                col = 3 + n_hist + yr
                val = scen_data[yr] if yr < len(scen_data) else None
                if val is not None:
                    set_cell(ws, r, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                             border=INPUT_BORDER)
                else:
                    set_cell(ws, r, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

    # ── SGA % of Revenue ──
    section_title(ws, sga_sub_header_row, 2, "SGA % of Revenue")
    for s, scen_name in enumerate(SCENARIO_NAMES):
        r = sga_scenario_rows[s]
        set_cell(ws, r, 2, scen_name, font=BOLD_FONT)
        scen_data = _scenarios.get(scen_name, {}).get("sga_pct", [])
        for yr in range(proj_years):
            col = 3 + n_hist + yr
            val = scen_data[yr] if yr < len(scen_data) else None
            if val is not None:
                set_cell(ws, r, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                         border=INPUT_BORDER)
            else:
                set_cell(ws, r, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

    # ── NWC % of Revenue (only when nwc_pct exists in scenarios) ──
    if _has_nwc_pct and nwc_scenario_rows is not None:
        section_title(ws, nwc_sub_header_row, 2, "NWC % of Revenue")
        for s, scen_name in enumerate(SCENARIO_NAMES):
            r = nwc_scenario_rows[s]
            set_cell(ws, r, 2, scen_name, font=BOLD_FONT)
            scen_data = _scenarios.get(scen_name, {}).get("nwc_pct", [])
            for yr in range(proj_years):
                col = 3 + n_hist + yr
                val = scen_data[yr] if yr < len(scen_data) else None
                if val is not None:
                    set_cell(ws, r, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                             border=INPUT_BORDER)
                else:
                    set_cell(ws, r, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)

    # Freeze pane
    ws.freeze_panes = "C5"

    return {
        "total_rev_row": total_rev_row,
        "total_op_row": total_op_row,
        "n_hist": n_hist,
        "cogs_scenario_rows": cogs_scenario_rows,
        "sga_scenario_rows": sga_scenario_rows,
        "nwc_scenario_rows": nwc_scenario_rows,
    }


# =====================================================================
# DRIVER ANALYSIS SHEET
# =====================================================================
def _create_driver_sheet(wb, C, segments, proj_years, year_labels):
    """Generate Driver Analysis sheet with driver_type-specific sections.

    Supported driver_types:
      - 'backlog': Order backlog → revenue recognition (equipment makers)
      - 'manmonth': Headcount × utilization × unit price (IT services)
      - 'growth_rate': Revenue growth rate based (generic)
      - 'manual': Direct revenue input (no driver decomposition)
    """
    ws = wb.create_sheet("Driver Analysis")
    ws.sheet_properties.tabColor = "4B0082"  # Indigo

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34

    # Determine historical years (same logic as segment sheet)
    n_hist = 0
    for seg in segments:
        hist_rev = seg.get("historical", {}).get("revenue", [])
        if len(hist_rev) > n_hist:
            n_hist = len(hist_rev)

    n_data_cols = n_hist + proj_years
    for ci in range(n_data_cols):
        ws.column_dimensions[col_letter(3 + ci)].width = 16

    set_cell(ws, 2, 2, f'Driver Analysis - {C["company_name"]}', font=TITLE_FONT)

    # Header row
    hist_labels = []
    if n_hist >= 3:
        hist_labels = ["FY-2", "FY-1", "FY0 (Base)"]
    elif n_hist == 2:
        hist_labels = ["FY-1", "FY0 (Base)"]
    elif n_hist == 1:
        hist_labels = ["FY0 (Base)"]

    if C.get("projection_start_fy"):
        import re as _re
        _m = _re.search(r"FY(\d+)", C["projection_start_fy"])
        if _m:
            _base_fy = int(_m.group(1))
            hist_labels = [f"FY{_base_fy - n_hist + 1 + i}" for i in range(n_hist)]

    all_headers = hist_labels + year_labels
    header_row(ws, 4, 3, 3 + len(all_headers) - 1, all_headers)

    cur_row = 5

    for seg_idx, seg in enumerate(segments):
        seg_name = seg.get("name", f"Segment {seg_idx + 1}")
        seg_name_jp = seg.get("name_jp", "")
        display_name = f"{seg_name}" + (f" ({seg_name_jp})" if seg_name_jp else "")
        driver_type = seg.get("driver_type", "manual")

        hist = seg.get("historical", {})
        proj = seg.get("projections", {})

        # ── Segment header ──
        c = section_title(ws, cur_row, 2, f"{display_name} [{driver_type}]")
        c.fill = LIGHT_FILL
        for ci in range(3, 3 + n_data_cols):
            ws.cell(row=cur_row, column=ci).fill = LIGHT_FILL
        cur_row += 1

        # ══════════════════════════════════════════════════════════
        if driver_type == "backlog":
            cur_row = _driver_backlog(ws, seg, hist, proj, n_hist, proj_years, cur_row)

        elif driver_type == "manmonth":
            cur_row = _driver_manmonth(ws, seg, hist, proj, n_hist, proj_years, cur_row)

        elif driver_type == "growth_rate":
            cur_row = _driver_growth_rate(ws, seg, hist, proj, n_hist, proj_years, cur_row)

        elif driver_type == "retail":
            cur_row = _driver_retail(ws, seg, hist, proj, n_hist, proj_years, cur_row)

        elif driver_type == "subscription":
            cur_row = _driver_subscription(ws, seg, hist, proj, n_hist, proj_years, cur_row)

        else:  # "manual" or unknown
            cur_row = _driver_manual(ws, seg, hist, proj, n_hist, proj_years, cur_row)

        cur_row += 1  # blank row between segments

    # Freeze pane
    ws.freeze_panes = "C5"


# ── Driver sub-functions ──

def _driver_backlog(ws, seg, hist, proj, n_hist, proj_years, cur_row):
    """Backlog-based driver: Beginning Backlog + Orders - Revenue = Ending Backlog."""

    hist_rev = hist.get("revenue", [None] * n_hist)
    hist_orders = hist.get("orders", [None] * n_hist)
    hist_backlog = hist.get("backlog_end", [None] * n_hist)
    proj_rev = proj.get("revenue", [None] * proj_years)
    proj_orders = proj.get("orders", [None] * proj_years)

    # Pad to n_hist
    while len(hist_rev) < n_hist:
        hist_rev.insert(0, None)
    while len(hist_orders) < n_hist:
        hist_orders.insert(0, None)
    while len(hist_backlog) < n_hist:
        hist_backlog.insert(0, None)

    n_data_cols = n_hist + proj_years

    # Row: Beginning Backlog
    bb_row = cur_row
    set_cell(ws, bb_row, 2, "Beginning Backlog", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        if i > 0 and hist_backlog[i - 1] is not None:
            set_cell(ws, bb_row, col, hist_backlog[i - 1], font=BLUE_FONT,
                     fmt=FMT_YEN, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, bb_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        prev_cl = col_letter(col - 1)
        # Beginning Backlog = previous period's Ending Backlog
        eb_row = bb_row + 3  # Ending backlog row (calculated below)
        if yr == 0 and hist_backlog[-1] is not None:
            set_cell(ws, bb_row, col, hist_backlog[-1], font=BLUE_FONT,
                     fmt=FMT_YEN, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, bb_row, col, f"={prev_cl}{eb_row}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
    cur_row += 1

    # Row: + New Orders
    ord_row = cur_row
    set_cell(ws, ord_row, 2, "+ New Orders", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_orders[i]
        if val is not None:
            set_cell(ws, ord_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=NWC_DATA_BORDER)
        else:
            set_cell(ws, ord_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_orders[yr] if yr < len(proj_orders) else None
        if val is not None:
            set_cell(ws, ord_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, ord_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Row: - Revenue (Recognized)
    rev_row = cur_row
    set_cell(ws, rev_row, 2, "- Revenue (Recognized)", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_rev[i]
        if val is not None:
            set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=NWC_DATA_BORDER)
        else:
            set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_rev[yr] if yr < len(proj_rev) else None
        if val is not None:
            set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Row: = Ending Backlog (formula: BB + Orders - Revenue)
    eb_row = cur_row
    set_cell(ws, eb_row, 2, "Ending Backlog", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_backlog[i]
        if val is not None:
            set_cell(ws, eb_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        else:
            set_cell(ws, eb_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, eb_row, col,
                 f"={cl}{bb_row}+{cl}{ord_row}-{cl}{rev_row}",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # Row: Book-to-Bill ratio
    btb_row = cur_row
    set_cell(ws, btb_row, 2, "Book-to-Bill",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        set_cell(ws, btb_row, col,
                 f"=IFERROR({cl}{ord_row}/{cl}{rev_row},\"—\")",
                 font=BLACK_FONT, fmt=FMT_RATIO, border=NWC_DATA_BORDER)
    cur_row += 1

    return cur_row


def _driver_manmonth(ws, seg, hist, proj, n_hist, proj_years, cur_row):
    """Man-month driver: HC × Utilization × Unit Price × 12 = Layer 1, + Layer 2."""

    hist_rev = hist.get("revenue", [None] * n_hist)
    hist_l2 = hist.get("layer2_revenue", [None] * n_hist)
    proj_hc = proj.get("headcount", [None] * proj_years)
    proj_util = proj.get("utilization", [None] * proj_years)
    proj_price = proj.get("unit_price_monthly", [None] * proj_years)
    proj_l2 = proj.get("layer2_revenue", [None] * proj_years)

    while len(hist_rev) < n_hist:
        hist_rev.insert(0, None)
    while len(hist_l2) < n_hist:
        hist_l2.insert(0, None)

    n_data_cols = n_hist + proj_years

    # Headcount
    hc_row = cur_row
    set_cell(ws, hc_row, 2, "Headcount", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, hc_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_hc[yr] if yr < len(proj_hc) else None
        if val is not None:
            set_cell(ws, hc_row, col, val, font=BLUE_FONT, fmt=FMT_INT,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, hc_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Utilization
    util_row = cur_row
    set_cell(ws, util_row, 2, "× Utilization", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, util_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_util[yr] if yr < len(proj_util) else None
        if val is not None:
            set_cell(ws, util_row, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, util_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Unit Price (M/month)
    price_row = cur_row
    set_cell(ws, price_row, 2, "× Unit Price (JPY mn/month)", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, price_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_price[yr] if yr < len(proj_price) else None
        if val is not None:
            set_cell(ws, price_row, col, val, font=BLUE_FONT, fmt=FMT_YEN_DEC,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, price_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Layer 1 Revenue = HC × Util × Price × 12
    l1_row = cur_row
    set_cell(ws, l1_row, 2, "Layer 1 Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, l1_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, l1_row, col,
                 f"={cl}{hc_row}*{cl}{util_row}*{cl}{price_row}*12",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # Layer 2 (Solution) Revenue
    l2_row = cur_row
    set_cell(ws, l2_row, 2, "+ Layer 2 (Solution) Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_l2[i]
        if val is not None:
            set_cell(ws, l2_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=NWC_DATA_BORDER)
        else:
            set_cell(ws, l2_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_l2[yr] if yr < len(proj_l2) else None
        if val is not None:
            set_cell(ws, l2_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, l2_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Segment Revenue = L1 + L2
    seg_rev_row = cur_row
    set_cell(ws, seg_rev_row, 2, "Segment Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_rev[i]
        if val is not None:
            set_cell(ws, seg_rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        else:
            set_cell(ws, seg_rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, seg_rev_row, col,
                 f"={cl}{l1_row}+{cl}{l2_row}",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # Layer 2 Mix %
    mix_row = cur_row
    set_cell(ws, mix_row, 2, "Layer 2 Mix %",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        set_cell(ws, mix_row, col,
                 f"=IFERROR({cl}{l2_row}/{cl}{seg_rev_row},\"—\")",
                 font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    cur_row += 1

    return cur_row


def _driver_growth_rate(ws, seg, hist, proj, n_hist, proj_years, cur_row):
    """Growth rate driver: Revenue = Prior × (1 + g)."""

    hist_rev = hist.get("revenue", [None] * n_hist)
    proj_growth = proj.get("revenue_growth", [None] * proj_years)

    while len(hist_rev) < n_hist:
        hist_rev.insert(0, None)

    n_data_cols = n_hist + proj_years

    # Revenue Growth row
    g_row = cur_row
    set_cell(ws, g_row, 2, "Revenue Growth (YoY)", font=BOLD_FONT)
    # Historical: compute from data
    for i in range(n_hist):
        col = 3 + i
        cl = col_letter(col)
        if i == 0 or hist_rev[i] is None or hist_rev[i - 1] is None:
            set_cell(ws, g_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
        else:
            prev_cl = col_letter(col - 1)
            rev_row = cur_row + 1  # revenue row is next
            set_cell(ws, g_row, col,
                     f"=({cl}{rev_row}-{prev_cl}{rev_row})/{prev_cl}{rev_row}",
                     font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_growth[yr] if yr < len(proj_growth) else None
        if val is not None:
            set_cell(ws, g_row, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, g_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # Revenue row
    rev_row = cur_row
    set_cell(ws, rev_row, 2, "Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_rev[i]
        if val is not None:
            set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=NWC_DATA_BORDER)
        else:
            set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        prev_cl = col_letter(col - 1)
        set_cell(ws, rev_row, col,
                 f"={prev_cl}{rev_row}*(1+{cl}{g_row})",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    return cur_row


def _driver_manual(ws, seg, hist, proj, n_hist, proj_years, cur_row):
    """Manual driver: Revenue directly input, no decomposition."""

    hist_rev = hist.get("revenue", [None] * n_hist)
    proj_rev = proj.get("revenue", [None] * proj_years)

    while len(hist_rev) < n_hist:
        hist_rev.insert(0, None)

    n_data_cols = n_hist + proj_years

    # Revenue row (direct input)
    rev_row = cur_row
    set_cell(ws, rev_row, 2, "Revenue (Direct Input)", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_rev[i]
        if val is not None:
            set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=NWC_DATA_BORDER)
        else:
            set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_rev[yr] if yr < len(proj_rev) else None
        if val is not None:
            set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # YoY Growth (computed)
    g_row = cur_row
    set_cell(ws, g_row, 2, "YoY Growth",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        if ci == 0:
            set_cell(ws, g_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
        else:
            prev_cl = col_letter(col - 1)
            set_cell(ws, g_row, col,
                     f"=IFERROR(({cl}{rev_row}-{prev_cl}{rev_row})/{prev_cl}{rev_row},\"—\")",
                     font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    cur_row += 1

    return cur_row


def _driver_retail(ws, seg, hist, proj, n_hist, proj_years, cur_row):
    """Retail driver: Store count rollforward + SSSG + new store contribution."""

    hist_rev = hist.get("revenue", [None] * n_hist)
    hist_store = hist.get("store_count", [None] * n_hist)
    proj_new_stores = proj.get("new_stores", [None] * proj_years)
    proj_closures = proj.get("closures", [None] * proj_years)
    proj_sssg = proj.get("sssg", [None] * proj_years)
    new_store_months = proj.get("new_store_months", 6)

    while len(hist_rev) < n_hist:
        hist_rev.insert(0, None)
    while len(hist_store) < n_hist:
        hist_store.insert(0, None)

    n_data_cols = n_hist + proj_years

    # R1: Beginning Store Count
    beg_row = cur_row
    set_cell(ws, beg_row, 2, "Beginning Store Count", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        if i > 0 and hist_store[i - 1] is not None:
            set_cell(ws, beg_row, col, hist_store[i - 1], font=BLUE_FONT,
                     fmt=FMT_INT, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, beg_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        prev_cl = col_letter(col - 1)
        end_row = beg_row + 3  # Ending Store Count row
        if yr == 0 and hist_store[-1] is not None:
            set_cell(ws, beg_row, col, hist_store[-1], font=BLUE_FONT,
                     fmt=FMT_INT, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, beg_row, col, f"={prev_cl}{end_row}",
                     font=BLACK_FONT, fmt=FMT_INT, border=NWC_DATA_BORDER)
    cur_row += 1

    # R2: + New Stores
    new_row = cur_row
    set_cell(ws, new_row, 2, "+ New Stores", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, new_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_new_stores[yr] if yr < len(proj_new_stores) else None
        if val is not None:
            set_cell(ws, new_row, col, val, font=BLUE_FONT, fmt=FMT_INT,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, new_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # R3: - Closures
    close_row = cur_row
    set_cell(ws, close_row, 2, "- Closures", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, close_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_closures[yr] if yr < len(proj_closures) else None
        if val is not None:
            set_cell(ws, close_row, col, val, font=BLUE_FONT, fmt=FMT_INT,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, close_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # R4: Ending Store Count = Beginning + New - Closures
    end_sc_row = cur_row
    set_cell(ws, end_sc_row, 2, "Ending Store Count", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_store[i]
        if val is not None:
            set_cell(ws, end_sc_row, col, val, font=BLUE_FONT, fmt=FMT_INT,
                     border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        else:
            set_cell(ws, end_sc_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, end_sc_row, col,
                 f"={cl}{beg_row}+{cl}{new_row}-{cl}{close_row}",
                 font=BLACK_FONT, fmt=FMT_INT,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # R5: SSSG
    sssg_row = cur_row
    set_cell(ws, sssg_row, 2, "SSSG", font=BOLD_FONT)
    # Historical: reverse-calc from per-store revenue YoY
    avg_row = beg_row + 8  # Avg Store Count row (R9)
    seg_rev_row = beg_row + 7  # Segment Revenue row (R8)
    for i in range(n_hist):
        col = 3 + i
        cl = col_letter(col)
        prev_cl = col_letter(col - 1)
        if i == 0:
            set_cell(ws, sssg_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, sssg_row, col,
                     f"=IFERROR(({cl}{seg_rev_row}/{cl}{avg_row})/({prev_cl}{seg_rev_row}/{prev_cl}{avg_row})-1,\"—\")",
                     font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_sssg[yr] if yr < len(proj_sssg) else None
        if val is not None:
            set_cell(ws, sssg_row, col, val, font=BLUE_FONT, fmt=FMT_PCT,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, sssg_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # R6: Existing Store Revenue = Prior Segment Revenue × (1 + SSSG)
    exist_row = cur_row
    set_cell(ws, exist_row, 2, "Existing Store Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, exist_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        prev_cl = col_letter(col - 1)
        set_cell(ws, exist_row, col,
                 f"={prev_cl}{seg_rev_row}*(1+{cl}{sssg_row})",
                 font=BLACK_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
    cur_row += 1

    # R7: + New Store Revenue = New Stores × (Prior Rev / Prior Ending SC) × (months/12)
    new_rev_row = cur_row
    set_cell(ws, new_rev_row, 2, "+ New Store Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, new_rev_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    month_frac = round(new_store_months / 12, 6)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        prev_cl = col_letter(col - 1)
        set_cell(ws, new_rev_row, col,
                 f"={cl}{new_row}*({prev_cl}{seg_rev_row}/{prev_cl}{end_sc_row})*{month_frac}",
                 font=BLACK_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
    cur_row += 1

    # R8: Segment Revenue
    seg_rev_row_actual = cur_row
    assert seg_rev_row_actual == seg_rev_row, "Row layout mismatch for Segment Revenue"
    set_cell(ws, seg_rev_row, 2, "Segment Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_rev[i]
        if val is not None:
            set_cell(ws, seg_rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        else:
            set_cell(ws, seg_rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, seg_rev_row, col,
                 f"={cl}{exist_row}+{cl}{new_rev_row}",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # R9: Avg Store Count (KPI)
    avg_row_actual = cur_row
    assert avg_row_actual == avg_row, "Row layout mismatch for Avg Store Count"
    set_cell(ws, avg_row, 2, "Avg Store Count",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        set_cell(ws, avg_row, col,
                 f"=IFERROR(({cl}{beg_row}+{cl}{end_sc_row})/2,\"—\")",
                 font=BLACK_FONT, fmt=FMT_INT, border=NWC_DATA_BORDER)
    cur_row += 1

    return cur_row


def _driver_subscription(ws, seg, hist, proj, n_hist, proj_years, cur_row):
    """Subscription driver: ARR bridge (churn/expansion/new) with NRR fallback."""

    hist_rev = hist.get("revenue", [None] * n_hist)
    hist_arr = hist.get("arr_end", [None] * n_hist)
    proj_nrr = proj.get("nrr", [None] * proj_years)
    proj_churn_rate = proj.get("churn_rate", None)
    proj_new_arr = proj.get("new_arr", [None] * proj_years)

    while len(hist_rev) < n_hist:
        hist_rev.insert(0, None)
    while len(hist_arr) < n_hist:
        hist_arr.insert(0, None)

    n_data_cols = n_hist + proj_years

    # ── NRR fallback: pre-compute churn/expansion values ──
    proj_beg_arr = []
    proj_churned = []
    proj_expansion = []
    proj_end_arr = []

    for yr in range(proj_years):
        if yr == 0:
            beg = ([v for v in hist_arr if v is not None] or [0])[-1]
        else:
            beg = proj_end_arr[yr - 1]
        proj_beg_arr.append(beg)

        nrr = proj_nrr[yr] if yr < len(proj_nrr) and proj_nrr[yr] is not None else 1.0
        if proj_churn_rate and yr < len(proj_churn_rate) and proj_churn_rate[yr] is not None:
            churn_r = proj_churn_rate[yr]
        else:
            churn_r = 0.0
        expansion_r = nrr - (1 - churn_r)

        churned = beg * churn_r
        expansion = beg * expansion_r
        new_arr_val = proj_new_arr[yr] if yr < len(proj_new_arr) and proj_new_arr[yr] is not None else 0

        proj_churned.append(churned)
        proj_expansion.append(expansion)
        proj_end_arr.append(beg - churned + expansion + new_arr_val)

    # S1: Beginning ARR
    beg_arr_row = cur_row
    set_cell(ws, beg_arr_row, 2, "Beginning ARR", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        if i > 0 and hist_arr[i - 1] is not None:
            set_cell(ws, beg_arr_row, col, hist_arr[i - 1], font=BLUE_FONT,
                     fmt=FMT_YEN, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, beg_arr_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    end_arr_row = beg_arr_row + 4  # S5: Ending ARR
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        prev_cl = col_letter(col - 1)
        if yr == 0 and hist_arr[-1] is not None:
            set_cell(ws, beg_arr_row, col, hist_arr[-1], font=BLUE_FONT,
                     fmt=FMT_YEN, border=NWC_DATA_BORDER)
        else:
            set_cell(ws, beg_arr_row, col, f"={prev_cl}{end_arr_row}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=NWC_DATA_BORDER)
    cur_row += 1

    # S2: - Churned ARR (positive value, subtracted in S5 formula)
    churn_row = cur_row
    set_cell(ws, churn_row, 2, "- Churned ARR", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, churn_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        set_cell(ws, churn_row, col, proj_churned[yr], font=BLUE_FONT, fmt=FMT_YEN,
                 border=INPUT_BORDER)
    cur_row += 1

    # S3: + Expansion ARR
    exp_row = cur_row
    set_cell(ws, exp_row, 2, "+ Expansion ARR", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, exp_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        set_cell(ws, exp_row, col, proj_expansion[yr], font=BLUE_FONT, fmt=FMT_YEN,
                 border=INPUT_BORDER)
    cur_row += 1

    # S4: + New ARR
    new_arr_row = cur_row
    set_cell(ws, new_arr_row, 2, "+ New ARR", font=BOLD_FONT)
    for i in range(n_hist):
        set_cell(ws, new_arr_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        val = proj_new_arr[yr] if yr < len(proj_new_arr) and proj_new_arr[yr] is not None else None
        if val is not None:
            set_cell(ws, new_arr_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=INPUT_BORDER)
        else:
            set_cell(ws, new_arr_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    cur_row += 1

    # S5: Ending ARR = Beginning - Churned + Expansion + New
    end_arr_row_actual = cur_row
    assert end_arr_row_actual == end_arr_row, "Row layout mismatch for Ending ARR"
    set_cell(ws, end_arr_row, 2, "Ending ARR", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_arr[i]
        if val is not None:
            set_cell(ws, end_arr_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        else:
            set_cell(ws, end_arr_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, end_arr_row, col,
                 f"={cl}{beg_arr_row}-{cl}{churn_row}+{cl}{exp_row}+{cl}{new_arr_row}",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # S6: Revenue = (Beginning ARR + Ending ARR) / 2
    rev_row = cur_row
    set_cell(ws, rev_row, 2, "Revenue", font=BOLD_FONT)
    for i in range(n_hist):
        col = 3 + i
        val = hist_rev[i]
        if val is not None:
            set_cell(ws, rev_row, col, val, font=BLUE_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        else:
            set_cell(ws, rev_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, rev_row, col,
                 f"=({cl}{beg_arr_row}+{cl}{end_arr_row})/2",
                 font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
    cur_row += 1

    # S7: NRR % (KPI)
    nrr_row = cur_row
    set_cell(ws, nrr_row, 2, "NRR %",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for i in range(n_hist):
        set_cell(ws, nrr_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, nrr_row, col,
                 f"=IFERROR(({cl}{end_arr_row}-{cl}{new_arr_row})/{cl}{beg_arr_row},\"—\")",
                 font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    cur_row += 1

    # S8: Gross Churn % (KPI)
    gc_row = cur_row
    set_cell(ws, gc_row, 2, "Gross Churn %",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for i in range(n_hist):
        set_cell(ws, gc_row, 3 + i, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
    for yr in range(proj_years):
        col = 3 + n_hist + yr
        cl = col_letter(col)
        set_cell(ws, gc_row, col,
                 f"=IFERROR({cl}{churn_row}/{cl}{beg_arr_row},\"—\")",
                 font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    cur_row += 1

    # S9: ARR YoY Growth (KPI)
    yoy_row = cur_row
    set_cell(ws, yoy_row, 2, "ARR YoY Growth",
             font=Font(name="Arial", size=10, italic=True, color="808080"))
    for ci in range(n_data_cols):
        col = 3 + ci
        cl = col_letter(col)
        if ci == 0:
            set_cell(ws, yoy_row, col, "—", font=GREY_FONT, border=NWC_DATA_BORDER)
        else:
            prev_cl = col_letter(col - 1)
            set_cell(ws, yoy_row, col,
                     f"=IFERROR({cl}{end_arr_row}/{prev_cl}{end_arr_row}-1,\"—\")",
                     font=BLACK_FONT, fmt=FMT_PCT, border=NWC_DATA_BORDER)
    cur_row += 1

    return cur_row


# =====================================================================
# BUILD WORKBOOK
# =====================================================================
def generate_dcf_workbook(config, output_path=None):
    """Generate DCF/Comps Excel workbook from config dict.

    Args:
        config: Dict with all financial data and assumptions.
        output_path: Output file path. If None, auto-generates.

    Returns:
        str: Path to saved Excel file.
    """
    C = config

    # ── Normalize WACC inputs ──
    # Beta: clamp to [0.6, 1.5]; outside range → sector-standard 1.0
    raw_beta = C.get("beta", 1.0)
    if not raw_beta or raw_beta < 0.6 or raw_beta > 1.5:
        C["beta"] = 1.0
    # Size Premium: auto-determine from market cap (JPY mn) unless explicitly overridden
    if "size_premium" not in C.get("_override_keys", set()):
        try:
            mkt_cap = C["current_price"] * C["shares_outstanding"] / 1_000_000
        except (KeyError, TypeError):
            mkt_cap = 0
        if mkt_cap >= 1_000_000:        # >= 1 trillion JPY
            C["size_premium"] = 0.0
        elif mkt_cap >= 100_000:         # >= 100 billion JPY
            C["size_premium"] = 0.015
        else:
            C["size_premium"] = 0.03

    # Restore flat arrays from Base scenario
    _base = C["scenarios"]["Base"]
    if "revenue_growth" in _base:
        C["revenue_growth"] = _base["revenue_growth"]
    if "cogs_pct" in _base:
        C["cogs_pct"] = _base["cogs_pct"]
    C["sga_pct"] = _base["sga_pct"]

    # When segments exist without scenario-level revenue_growth/cogs_pct,
    # compute implied values from segment data for potential downstream use
    if C.get("segments") and "revenue_growth" not in C:
        _segs = C["segments"]
        _n_proj = C.get("projection_years", 5)
        _base_revs = []
        for yr in range(_n_proj):
            total = 0
            for seg in _segs:
                hist_rev = seg.get("historical", {}).get("revenue", [])
                base_rev = ([v for v in hist_rev if v is not None] or [0])[-1]
                growths = seg.get("projections", {}).get("revenue_growth", [])
                r = base_rev
                for y in range(yr + 1):
                    g = growths[y] if y < len(growths) else 0
                    r = r * (1 + g)
                total += r
            _base_revs.append(total)
        _by_rev = C.get("base_year_revenue", 1)
        C["revenue_growth"] = [(_base_revs[0] / _by_rev) - 1 if _by_rev else 0]
        for i in range(1, len(_base_revs)):
            C["revenue_growth"].append(
                (_base_revs[i] / _base_revs[i - 1]) - 1 if _base_revs[i - 1] else 0
            )
    if C.get("segments") and "cogs_pct" not in C:
        # Derive implied COGS% from segment OP margins and SGA%
        C["cogs_pct"] = [0.78] * C.get("projection_years", 5)  # reasonable default

    USE_EV_SALES = (C.get("primary_multiple", "EV/EBITDA") == "EV/Sales")

    # ── Pre-calculate Segment Analysis row numbers (needed by DCF Model) ──
    has_segments = bool(C.get("segments"))
    # Check that segments use revenue_growth (v2 format)
    if has_segments:
        _has_seg_growth = any(
            s.get("projections", {}).get("revenue_growth")
            for s in C["segments"]
        )
        if not _has_seg_growth:
            print("WARNING: segments lack 'revenue_growth' — falling back to legacy mode")
            has_segments = False  # disable segment-linked mode

    seg_info = None
    if has_segments:
        _segments = C["segments"]
        _n_hist_seg = 0
        for _seg in _segments:
            _hr = _seg.get("historical", {}).get("revenue", [])
            if len(_hr) > _n_hist_seg:
                _n_hist_seg = len(_hr)
        # Display area: per segment 5 rows (header+rev+op+opm+blank)
        # Total block: header(1) + rev(1) + op(1) + opm(1) + blank(1) = 5
        _seg_display_rows = len(_segments) * 5
        # total_rev_row = 5 + seg_display_rows + 1 (header) → first data row
        _total_rev_row = 5 + _seg_display_rows + 1
        _total_op_row = _total_rev_row + 1

        # Pre-calculate Consolidated Inputs row positions (SGA%, NWC%)
        # Matrix layout: section header(1) + year headers(1)
        #   per segment: rev_growth block(7) + opm block(7) = 14 rows
        _matrix_start = 5 + _seg_display_rows + 5 + 5 + 1  # display+total+recon+gap
        _matrix_cur = _matrix_start + 2  # +2 for section header + year headers
        _matrix_cur += len(_segments) * 14  # 14 rows per segment

        _consolidated_start = _matrix_cur + 1  # gap
        _scenarios_dict = C.get("scenarios", {})

        _has_cogs_pct = any("cogs_pct" in _scenarios_dict.get(sn, {}) for sn in SCENARIO_NAMES)
        if _has_cogs_pct:
            _cogs_sub_header = _consolidated_start + 2
            _cogs_scenario_rows = [_cogs_sub_header + 1 + s for s in range(NUM_SCENARIOS)]
            _sga_sub_header = _cogs_sub_header + 1 + NUM_SCENARIOS + 1
        else:
            _cogs_scenario_rows = None
            _sga_sub_header = _consolidated_start + 2
        _sga_scenario_rows = [_sga_sub_header + 1 + s for s in range(NUM_SCENARIOS)]
        _has_nwc_pct = (
            C.get("nwc_method") == "revenue_pct"
            and all(
                "nwc_pct" in _scenarios_dict[sn]
                for sn in SCENARIO_NAMES if sn in _scenarios_dict
            )
        )
        _nwc_scenario_rows = None
        if _has_nwc_pct:
            _nwc_sub_header = _sga_sub_header + 1 + NUM_SCENARIOS + 1
            _nwc_scenario_rows = [_nwc_sub_header + 1 + s for s in range(NUM_SCENARIOS)]

        seg_info = {
            "total_rev_row": _total_rev_row,
            "total_op_row": _total_op_row,
            "n_hist": _n_hist_seg,
            "cogs_scenario_rows": _cogs_scenario_rows,
            "sga_scenario_rows": _sga_scenario_rows,
            "nwc_scenario_rows": _nwc_scenario_rows,
        }

    wb = openpyxl.Workbook()

    # =====================================================================
    # SHEET 1: Executive Summary
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = "003366"

    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22
    ws1.column_dimensions["E"].width = 22

    # Disclaimer
    set_cell(ws1, 1, 2,
        "DISCLAIMER: This is a sample analysis for demonstration purposes only. "
        "It does not constitute investment advice.",
        font=GREY_FONT)
    ws1.merge_cells("B1:E1")

    # Title
    set_cell(ws1, 3, 2, f'{C["company_name"]} ({C["ticker"]})', font=TITLE_FONT)
    ws1.merge_cells("B3:D3")
    set_cell(ws1, 4, 2, "Equity Research Report", font=SUB_FONT)

    # Company info
    set_cell(ws1, 6, 2, "Company", font=BOLD_FONT)
    set_cell(ws1, 6, 3, C["company_name"])
    set_cell(ws1, 7, 2, "Ticker", font=BOLD_FONT)
    set_cell(ws1, 7, 3, f'{C["ticker"]} ({C["exchange"]})')
    set_cell(ws1, 8, 2, "Sector", font=BOLD_FONT)
    set_cell(ws1, 8, 3, C["sector"])
    set_cell(ws1, 9, 2, "Current Price", font=BOLD_FONT)
    set_cell(ws1, 9, 3, C["current_price"], font=BLUE_FONT, fmt=FMT_YEN)

    # Target Price = average of 4 methods (C16:C19)
    set_cell(ws1, 10, 2, "Target Price (Mid)", font=BOLD_FONT)
    set_cell(ws1, 10, 3, "=ROUND(AVERAGE(C16:C19),0)", font=BLACK_FONT, fmt=FMT_YEN)

    # Recommendation
    set_cell(ws1, 11, 2, "Recommendation", font=BOLD_FONT)
    set_cell(ws1, 11, 3, '=IF(C12>0.15,"BUY",IF(C12>0.05,"HOLD","SELL"))', font=BLACK_FONT)

    # Upside / Downside
    set_cell(ws1, 12, 2, "Upside / Downside", font=BOLD_FONT)
    set_cell(ws1, 12, 3, "=(C10-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

    # Valuation Summary section
    c = set_cell(ws1, 14, 2, "Valuation Summary", font=SUB_FONT)
    c.fill = LIGHT_FILL
    for col_idx in range(3, 5):
        ws1.cell(row=14, column=col_idx).fill = LIGHT_FILL

    header_row(ws1, 15, 2, 4, ["Methodology", "Implied Value (JPY)", "vs Current Price"])

    # DCF - Perpetuity Growth
    set_cell(ws1, 16, 2, "DCF - Perpetuity Growth")
    set_cell(ws1, 16, 3, f"='DCF Model'!C{R_PRICE_PGM}", font=GREEN_FONT, fmt=FMT_YEN)
    set_cell(ws1, 16, 4, "=(C16-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

    # DCF - Exit Multiple
    set_cell(ws1, 17, 2, "DCF - Exit Multiple")
    set_cell(ws1, 17, 3, f"='DCF Model'!C{R_PRICE_EXIT}", font=GREEN_FONT, fmt=FMT_YEN)
    set_cell(ws1, 17, 4, "=(C17-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

    # Comps
    if USE_EV_SALES:
        set_cell(ws1, 18, 2, "Comps - EV/Sales Median")
    else:
        set_cell(ws1, 18, 2, "Comps - EV/EBITDA Median")
    set_cell(ws1, 18, 3, "='Comps Analysis'!C27", font=GREEN_FONT, fmt=FMT_YEN)
    set_cell(ws1, 18, 4, "=(C18-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

    set_cell(ws1, 19, 2, "Comps - PER Median")
    set_cell(ws1, 19, 3, "='Comps Analysis'!C28", font=GREEN_FONT, fmt=FMT_YEN)
    set_cell(ws1, 19, 4, "=(C19-C9)/C9", font=BLACK_FONT, fmt=FMT_PCT)

    # Integrated Valuation Range
    set_cell(ws1, 21, 2, "Integrated Valuation Range", font=BOLD_FONT)
    set_cell(ws1, 21, 3, '=MIN(C16:C19)&" - "&MAX(C16:C19)', font=BLACK_FONT)

    # Investment Thesis
    c = set_cell(ws1, 23, 2, "Key Investment Thesis", font=SUB_FONT)
    c.fill = LIGHT_FILL
    for col_idx in range(3, 5):
        ws1.cell(row=23, column=col_idx).fill = LIGHT_FILL
    for i, line in enumerate(C["investment_thesis"]):
        set_cell(ws1, 24 + i, 2, line)

    # Key Risks
    c = set_cell(ws1, 28, 2, "Key Risks", font=SUB_FONT)
    c.fill = LIGHT_FILL
    for col_idx in range(3, 5):
        ws1.cell(row=28, column=col_idx).fill = LIGHT_FILL
    for i, line in enumerate(C["key_risks"]):
        set_cell(ws1, 29 + i, 2, line)

    # =====================================================================
    # SHEET 2: Financial Statements (V3 — Full PL Waterfall + BS Highlights)
    # =====================================================================
    ws2 = wb.create_sheet("Financial Statements")
    ws2.sheet_properties.tabColor = "003366"

    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 32
    n_hist = len(C["hist_years"])
    for _ci in range(n_hist):
        ws2.column_dimensions[col_letter(3 + _ci)].width = 18

    set_cell(ws2, 2, 2, f'{C["company_name"]} - Historical Financials (JPY mn)', font=TITLE_FONT)

    header_row(ws2, 4, 3, 3 + n_hist - 1, C["hist_years"])

    # ── Income Statement (V3 full waterfall) ──
    section_title(ws2, 5, 2, "Income Statement")

    set_cell(ws2, 6, 2, "Revenue", font=BOLD_FONT)
    set_cell(ws2, 7, 2, "COGS", font=BOLD_FONT)
    set_cell(ws2, 8, 2, "Gross Profit", font=BOLD_FONT)
    set_cell(ws2, 9, 2, "Gross Margin")
    set_cell(ws2, 10, 2, "SGA", font=BOLD_FONT)
    set_cell(ws2, 11, 2, "Operating Income", font=BOLD_FONT)
    set_cell(ws2, 12, 2, "Net Income", font=BOLD_FONT)
    set_cell(ws2, 13, 2, "Operating Margin")
    set_cell(ws2, 14, 2, "Net Margin")
    set_cell(ws2, 15, 2, "Revenue Growth (YoY)")
    set_cell(ws2, 16, 2, "Operating Income Growth (YoY)")

    for i in range(n_hist):
        col = 3 + i
        cl = col_letter(col)

        # Revenue
        set_cell(ws2, 6, col, C["hist_revenue"][i], font=BLUE_FONT, fmt=FMT_YEN)
        # COGS
        cogs_val = C["hist_cogs"][i] if C["hist_cogs"] and i < len(C["hist_cogs"]) else None
        set_cell(ws2, 7, col, cogs_val, font=BLUE_FONT, fmt=FMT_YEN)
        # Gross Profit = Revenue - COGS
        set_cell(ws2, 8, col, f"={cl}6-{cl}7", font=BLACK_FONT, fmt=FMT_YEN)
        # Gross Margin = GP / Revenue
        set_cell(ws2, 9, col, f"={cl}8/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        # SGA
        sga_val = C["hist_sga"][i] if C["hist_sga"] and i < len(C["hist_sga"]) else None
        set_cell(ws2, 10, col, sga_val, font=BLUE_FONT, fmt=FMT_YEN)
        # Operating Income
        set_cell(ws2, 11, col, C["hist_operating_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
        # Net Income
        set_cell(ws2, 12, col, C["hist_net_income"][i], font=BLUE_FONT, fmt=FMT_YEN)
        # Operating Margin
        set_cell(ws2, 13, col, f"={cl}11/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        # Net Margin
        set_cell(ws2, 14, col, f"={cl}12/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        # YoY Growth
        if i == 0:
            set_cell(ws2, 15, col, "n/a")
            set_cell(ws2, 16, col, "n/a")
        else:
            prev_cl = col_letter(col - 1)
            set_cell(ws2, 15, col, f"=({cl}6-{prev_cl}6)/{prev_cl}6", font=BLACK_FONT, fmt=FMT_PCT)
            set_cell(ws2, 16, col, f"=({cl}11-{prev_cl}11)/{prev_cl}11", font=BLACK_FONT, fmt=FMT_PCT)

    # ── Cash Flow Statement ──
    section_title(ws2, 18, 2, "Cash Flow Statement")
    set_cell(ws2, 19, 2, "Operating Cash Flow", font=BOLD_FONT)
    set_cell(ws2, 20, 2, "Capex", font=BOLD_FONT)
    set_cell(ws2, 21, 2, "Free Cash Flow", font=BOLD_FONT)
    set_cell(ws2, 22, 2, "FCF Margin")
    set_cell(ws2, 23, 2, "Capex / Revenue")

    for i in range(n_hist):
        col = 3 + i
        cl = col_letter(col)
        ocf_val = C["hist_ocf"][i] if C["hist_ocf"] and i < len(C["hist_ocf"]) else None
        capex_val = C["hist_capex"][i] if C["hist_capex"] and i < len(C["hist_capex"]) else None
        set_cell(ws2, 19, col, ocf_val, font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws2, 20, col, capex_val, font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws2, 21, col, f"={cl}19-{cl}20", font=BLACK_FONT, fmt=FMT_YEN)
        set_cell(ws2, 22, col, f"={cl}21/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)
        set_cell(ws2, 23, col, f"={cl}20/{cl}6", font=BLACK_FONT, fmt=FMT_PCT)

    # ── Balance Sheet Highlights ──
    section_title(ws2, 25, 2, "Balance Sheet Highlights")
    set_cell(ws2, 26, 2, "Cash & Deposits", font=BOLD_FONT)
    set_cell(ws2, 27, 2, "Short-term Debt", font=BOLD_FONT)
    set_cell(ws2, 28, 2, "Net Debt (Cash)", font=BOLD_FONT)

    for i in range(n_hist):
        col = 3 + i
        cl = col_letter(col)
        cash_val = C["hist_cash"][i] if C["hist_cash"] and i < len(C["hist_cash"]) else None
        debt_val = C["hist_debt"][i] if C["hist_debt"] and i < len(C["hist_debt"]) else None
        set_cell(ws2, 26, col, cash_val, font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws2, 27, col, debt_val if debt_val is not None else 0, font=BLUE_FONT, fmt=FMT_YEN)
        # Net Debt = Debt - Cash (negative = net cash)
        set_cell(ws2, 28, col, f"={cl}27-{cl}26", font=BLACK_FONT, fmt=FMT_YEN)

    # =====================================================================
    # SHEET 3: DCF Model (V3 — Full Waterfall)
    # =====================================================================
    ws3 = wb.create_sheet("DCF Model")
    ws3.sheet_properties.tabColor = "006600"

    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 32
    for letter in ["C", "D", "E", "F", "G"]:
        ws3.column_dimensions[letter].width = 16

    set_cell(ws3, 2, 2, f'DCF Valuation Model - {C["company_name"]}', font=TITLE_FONT)

    # ── Assumptions ──
    c = section_title(ws3, 4, 2, "Assumptions")
    c.fill = LIGHT_FILL
    for col_idx in range(3, 8):
        ws3.cell(row=4, column=col_idx).fill = LIGHT_FILL

    # Determine effective Capex/D&A assumption values for display
    _capex_method = C.get("capex_method", "revenue_pct")
    _da_method = C.get("da_method", "revenue_pct")
    _capex_pct_display = C["capex_pct"]
    _da_pct_display = C["da_pct"]
    if _capex_method == "direct":
        _cp = C.get("capex_direct", {}).get("projections", [])
        if _cp:
            _capex_pct_display = sum(c for c in _cp if c is not None) / len(_cp) / C["base_year_revenue"] if C["base_year_revenue"] else 0
    if _da_method == "direct":
        _dp = C.get("da_direct", {}).get("projections", [])
        if _dp:
            _da_pct_display = sum(d for d in _dp if d is not None) / len(_dp) / C["base_year_revenue"] if C["base_year_revenue"] else 0

    assumptions = [
        # Revenue Growth Rate and COGS % removed — now in per-year driver rows
        ("Capex / Revenue",            _capex_pct_display,        FMT_PCT),      # C5
        ("Effective Tax Rate",         C["tax_rate"],             FMT_PCT),      # C6
        ("Risk-Free Rate",             C["risk_free"],            FMT_PCT),      # C7
        ("Beta",                       C["beta"],                 "0.00"),       # C8
        ("Equity Risk Premium",        C["erp"],                  FMT_PCT),      # C9
        ("Size Premium",               C["size_premium"],         FMT_PCT),      # C10
        ("After-tax Cost of Debt",     C["cost_of_debt_at"],      FMT_PCT),      # C11
        ("D/E Ratio",                  C["de_ratio"],             "0.000"),      # C12
        ("Terminal Growth Rate",       C["terminal_growth"],      FMT_PCT),      # C13
        ("Exit Multiple (EV/EBITDA)",  C["exit_multiple"],        FMT_RATIO),    # C14
        ("Fully Diluted Shares",       C["shares_outstanding"],   FMT_INT),      # C15
        ("Net Debt (JPY mn)",          C["net_debt"],             FMT_YEN),      # C16
        ("Base Year Revenue (JPY mn)", C["base_year_revenue"],    FMT_YEN),      # C17
        ("D&A / Revenue",              _da_pct_display,           FMT_PCT),      # C18
        ("Stub Fraction (yr remaining)", C.get("stub_fraction", 1.0), "0.00"),  # C19
        ("LTM Revenue (JPY mn)",         C.get("ltm_revenue", C["base_year_revenue"]), FMT_YEN),  # C20
    ]
    for i, (label, val, fmt) in enumerate(assumptions):
        r = 5 + i
        set_cell(ws3, r, 2, label, font=BOLD_FONT)
        set_cell(ws3, r, 3, val, font=BLUE_FONT, fmt=fmt)

    # ── WACC Calculation ──
    c = section_title(ws3, 22, 2, "WACC Calculation")
    c.fill = LIGHT_FILL
    for col_idx in range(3, 8):
        ws3.cell(row=22, column=col_idx).fill = LIGHT_FILL

    set_cell(ws3, 23, 2, "Cost of Equity (Ke)", font=BOLD_FONT)
    set_cell(ws3, 23, 3, "=C7+C8*C9+C10", font=BLACK_FONT, fmt=FMT_PCT2)

    set_cell(ws3, 24, 2, "Weight of Equity", font=BOLD_FONT)
    set_cell(ws3, 24, 3, "=1/(1+C12)", font=BLACK_FONT, fmt=FMT_PCT2)

    set_cell(ws3, 25, 2, "Weight of Debt", font=BOLD_FONT)
    set_cell(ws3, 25, 3, "=C12/(1+C12)", font=BLACK_FONT, fmt=FMT_PCT2)

    set_cell(ws3, 26, 2, "WACC", font=BOLD_FONT)
    set_cell(ws3, 26, 3, "=C23*C24+C11*C25", font=BLACK_FONT, fmt=FMT_PCT2)

    # ── Active Scenario Selector (Row 27) ──
    set_cell(ws3, 27, 2, "Active Scenario", font=BOLD_FONT)
    set_cell(ws3, 27, 3, "Base", font=BLUE_FONT, border=INPUT_BORDER)
    # D27 = MATCH index (1-5) for CHOOSE formula
    if has_segments:
        # No local matrix — match against array constant
        set_cell(ws3, 27, 4,
                 '=MATCH(C27,{"Base","Upside","Management","Downside 1","Downside 2"},0)',
                 font=BLACK_FONT)
    else:
        set_cell(ws3, 27, 4,
                 f"=MATCH(C27,B{R_SCEN_BLK_GROWTH + 1}:B{R_SCEN_BLK_GROWTH + NUM_SCENARIOS},0)",
                 font=BLACK_FONT)

    dv_scenario = DataValidation(
        type="list",
        formula1='"Base,Upside,Management,Downside 1,Downside 2"',
        allow_blank=False,
        showDropDown=False,   # openpyxl quirk: False = show dropdown
    )
    dv_scenario.add("C27")
    ws3.add_data_validation(dv_scenario)

    # ── Projected FCF (V3 Full Waterfall) ──
    c = section_title(ws3, 28, 2, "Projected Free Cash Flow")
    c.fill = LIGHT_FILL
    for col_idx in range(3, 8):
        ws3.cell(row=28, column=col_idx).fill = LIGHT_FILL

    proj_years = C["projection_years"]
    year_labels = [f"Year {y}" for y in range(1, proj_years + 1)]
    # Use actual FY labels if projection_start_fy is set
    if C.get("projection_start_fy"):
        import re as _re
        _m = _re.search(r"FY(\d+)", C["projection_start_fy"])
        if _m:
            _base_fy = int(_m.group(1))
            year_labels = [f"FY{_base_fy + y}(E)" for y in range(proj_years)]
    header_row(ws3, 29, 3, 3 + proj_years - 1, year_labels)

    # Driver row labels
    row_labels_drv = [
        ("Revenue Growth (YoY)",          R_DRV_GROWTH),
        ("COGS % of Revenue",             R_DRV_COGS),
        ("SGA % of Revenue",              R_DRV_SGA),
    ]
    for label, r in row_labels_drv:
        set_cell(ws3, r, 2, label, font=BOLD_FONT)

    # V3: FCF row labels — full waterfall
    row_labels_fcf = [
        ("Revenue",                       R_REVENUE),
        ("COGS",                          R_COGS),
        ("Gross Profit",                  R_GROSS_PROFIT),
        ("Gross Margin",                  R_GROSS_MARGIN),
        ("SGA Expense",                   R_SGA),
        ("Implied Operating Margin",      R_OP_M_IMPL),
        ("Operating Income (EBIT)",       R_EBIT),
        ("Less: Tax",                     R_TAX),
        ("NOPAT",                         R_NOPAT),
        ("Plus: D&A",                     R_DA),
        ("Less: Capex",                   R_CAPEX),
        ("Change in NWC",                 R_CHG_NWC),
        ("Unlevered Free Cash Flow",      R_UFCF),
        ("Discount Factor",               R_DISC),
        ("PV of FCF",                     R_PV_FCF),
    ]
    for label, r in row_labels_fcf:
        set_cell(ws3, r, 2, label, font=BOLD_FONT)

    # V3: Year-by-year projection loop — full waterfall with driver rows
    for yr in range(proj_years):
        col = 3 + yr
        cl = col_letter(col)
        prev_cl = col_letter(col - 1) if yr > 0 else None

        if has_segments:
            # ── Segment-linked mode: Revenue & EBIT from Segment Analysis ──
            seg_cl = col_letter(3 + seg_info["n_hist"] + yr)  # Segment sheet column

            # Revenue Growth — back-calculated display
            if yr == 0:
                set_cell(ws3, R_DRV_GROWTH, col, f"=({cl}{R_REVENUE}/C17)-1",
                         font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)
            else:
                set_cell(ws3, R_DRV_GROWTH, col, f"=({cl}{R_REVENUE}/{prev_cl}{R_REVENUE})-1",
                         font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)

            # COGS% — CHOOSE from Segment Analysis Consolidated Inputs (or back-calc if no cogs override)
            if seg_info.get("cogs_scenario_rows"):
                cogs_refs = [f"'Segment Analysis'!{seg_cl}{r}" for r in seg_info["cogs_scenario_rows"]]
                set_cell(ws3, R_DRV_COGS, col,
                         f"=CHOOSE($D$27,{','.join(cogs_refs)})",
                         font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)
            else:
                set_cell(ws3, R_DRV_COGS, col, f"=IFERROR({cl}{R_COGS}/{cl}{R_REVENUE},0)",
                         font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)

            # SGA% — CHOOSE from Segment Analysis Consolidated Inputs
            sga_refs = [f"'Segment Analysis'!{seg_cl}{r}" for r in seg_info["sga_scenario_rows"]]
            set_cell(ws3, R_DRV_SGA, col,
                     f"=CHOOSE($D$27,{','.join(sga_refs)})",
                     font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)

            # Revenue — from Segment Analysis total
            set_cell(ws3, R_REVENUE, col,
                     f"='Segment Analysis'!{seg_cl}{seg_info['total_rev_row']}",
                     font=BLACK_FONT, fmt=FMT_YEN)

            # COGS — from COGS% driver when override exists, otherwise back-calc
            if seg_info.get("cogs_scenario_rows"):
                set_cell(ws3, R_COGS, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_COGS}",
                         font=BLACK_FONT, fmt=FMT_YEN)
            else:
                set_cell(ws3, R_COGS, col, f"={cl}{R_REVENUE}-{cl}{R_SGA}-{cl}{R_EBIT}",
                         font=BLACK_FONT, fmt=FMT_YEN)

            # Gross Profit = Revenue - COGS
            set_cell(ws3, R_GROSS_PROFIT, col, f"={cl}{R_REVENUE}-{cl}{R_COGS}", font=BLACK_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER)

            # Gross Margin = GP / Revenue
            set_cell(ws3, R_GROSS_MARGIN, col, f"={cl}{R_GROSS_PROFIT}/{cl}{R_REVENUE}", font=BLACK_FONT, fmt=FMT_PCT)

            # SGA Expense = Revenue * SGA% driver
            set_cell(ws3, R_SGA, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_SGA}", font=BLACK_FONT, fmt=FMT_YEN)

            # EBIT — derived from COGS/SGA overrides when cogs_scenario_rows exists,
            # otherwise linked from Segment Analysis total OP
            if seg_info.get("cogs_scenario_rows"):
                set_cell(ws3, R_EBIT, col,
                         f"={cl}{R_GROSS_PROFIT}-{cl}{R_SGA}",
                         font=BLACK_FONT, fmt=FMT_YEN, border=SUBTOTAL_BORDER)
            else:
                set_cell(ws3, R_EBIT, col,
                         f"='Segment Analysis'!{seg_cl}{seg_info['total_op_row']}",
                         font=BLACK_FONT, fmt=FMT_YEN, border=SUBTOTAL_BORDER)

            # Implied Operating Margin = EBIT / Revenue
            set_cell(ws3, R_OP_M_IMPL, col,
                     f"={cl}{R_EBIT}/{cl}{R_REVENUE}",
                     font=BLACK_FONT, fmt=FMT_PCT)

        else:
            # ── Legacy mode: top-down Revenue Growth × Base Year ──

            # Driver rows (CHOOSE formulas referencing scenario matrix)
            set_cell(ws3, R_DRV_GROWTH, col, choose_formula(R_SCEN_BLK_GROWTH, cl),
                     font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)
            set_cell(ws3, R_DRV_COGS, col, choose_formula(R_SCEN_BLK_COGS, cl),
                     font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)
            set_cell(ws3, R_DRV_SGA, col, choose_formula(R_SCEN_BLK_SGA, cl),
                     font=BLACK_FONT, fmt=FMT_PCT, fill=LIGHT_FILL)

            # Revenue — Year 1 grows from Base Year Revenue (latest FY actual, C17)
            if yr == 0:
                set_cell(ws3, R_REVENUE, col, f"=C17*(1+{cl}{R_DRV_GROWTH})", font=BLACK_FONT, fmt=FMT_YEN)
            else:
                set_cell(ws3, R_REVENUE, col, f"={prev_cl}{R_REVENUE}*(1+{cl}{R_DRV_GROWTH})", font=BLACK_FONT, fmt=FMT_YEN)

            # COGS = Revenue * COGS% driver
            set_cell(ws3, R_COGS, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_COGS}", font=BLACK_FONT, fmt=FMT_YEN)

            # Gross Profit = Revenue - COGS
            set_cell(ws3, R_GROSS_PROFIT, col, f"={cl}{R_REVENUE}-{cl}{R_COGS}", font=BLACK_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER)

            # Gross Margin = GP / Revenue
            set_cell(ws3, R_GROSS_MARGIN, col, f"={cl}{R_GROSS_PROFIT}/{cl}{R_REVENUE}", font=BLACK_FONT, fmt=FMT_PCT)

            # SGA Expense = Revenue * SGA% driver
            set_cell(ws3, R_SGA, col, f"={cl}{R_REVENUE}*{cl}{R_DRV_SGA}", font=BLACK_FONT, fmt=FMT_YEN)

            # Implied Operating Margin = (GP - SGA) / Revenue
            set_cell(ws3, R_OP_M_IMPL, col,
                     f"=({cl}{R_GROSS_PROFIT}-{cl}{R_SGA})/{cl}{R_REVENUE}",
                     font=BLACK_FONT, fmt=FMT_PCT)

            # EBIT = Gross Profit - SGA
            set_cell(ws3, R_EBIT, col, f"={cl}{R_GROSS_PROFIT}-{cl}{R_SGA}", font=BLACK_FONT, fmt=FMT_YEN,
                     border=SUBTOTAL_BORDER)

        # Tax with NOPAT floor (no tax benefit when EBIT < 0)
        set_cell(ws3, R_TAX, col, f"=MAX(0,{cl}{R_EBIT}*C6)", font=BLACK_FONT, fmt=FMT_YEN)

        # NOPAT
        set_cell(ws3, R_NOPAT, col, f"={cl}{R_EBIT}-{cl}{R_TAX}", font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER)
        # D&A
        if _da_method == "direct":
            _da_arr = C.get("da_direct", {}).get("projections", [])
            _da_val = _da_arr[yr] if yr < len(_da_arr) and _da_arr[yr] is not None else None
            if _da_val is not None:
                set_cell(ws3, R_DA, col, _da_val, font=BLUE_FONT, fmt=FMT_YEN)
            else:
                set_cell(ws3, R_DA, col, f"={cl}{R_REVENUE}*C18", font=BLACK_FONT, fmt=FMT_YEN)
        else:
            set_cell(ws3, R_DA, col, f"={cl}{R_REVENUE}*C18", font=BLACK_FONT, fmt=FMT_YEN)
        # Capex
        if _capex_method == "direct":
            _cx_arr = C.get("capex_direct", {}).get("projections", [])
            _cx_val = _cx_arr[yr] if yr < len(_cx_arr) and _cx_arr[yr] is not None else None
            if _cx_val is not None:
                set_cell(ws3, R_CAPEX, col, _cx_val, font=BLUE_FONT, fmt=FMT_YEN)
            else:
                set_cell(ws3, R_CAPEX, col, f"={cl}{R_REVENUE}*C5", font=BLACK_FONT, fmt=FMT_YEN)
        else:
            set_cell(ws3, R_CAPEX, col, f"={cl}{R_REVENUE}*C5", font=BLACK_FONT, fmt=FMT_YEN)
        # Change in NWC (linked from NWC Schedule; NWC col = DCF col + 1)
        nwc_col_letter = col_letter(col + 1)
        set_cell(ws3, R_CHG_NWC, col,
                 f"='NWC Schedule'!{nwc_col_letter}{NWC_R_CHG_NWC}",
                 font=BLACK_FONT, fmt=FMT_YEN)
        # UFCF = NOPAT + D&A - Capex - Change in NWC
        set_cell(ws3, R_UFCF, col, f"={cl}{R_NOPAT}+{cl}{R_DA}-{cl}{R_CAPEX}-{cl}{R_CHG_NWC}", font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        # Discount Factor
        set_cell(ws3, R_DISC, col, f"=1/(1+C26)^(C19+{yr})", font=BLACK_FONT, fmt="0.0000")
        # PV of FCF
        set_cell(ws3, R_PV_FCF, col, f"={cl}{R_UFCF}*{cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN,
                 border=SUBTOTAL_BORDER)

    # ── Valuation - Perpetuity Growth Method ──
    c = section_title(ws3, R_PGM_SEC, 2, "Valuation - Perpetuity Growth Method")
    c.fill = LIGHT_GREEN
    for col_idx in range(3, 8):
        ws3.cell(row=R_PGM_SEC, column=col_idx).fill = LIGHT_GREEN

    last_cl = col_letter(3 + proj_years - 1)  # G for 5 years

    set_cell(ws3, R_SUM_PV, 2, "Sum of PV of FCFs", font=BOLD_FONT)
    set_cell(ws3, R_SUM_PV, 3, f"=SUM(C{R_PV_FCF}:{last_cl}{R_PV_FCF})", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_TV_PGM, 2, "Terminal Value (PGM)", font=BOLD_FONT)
    set_cell(ws3, R_TV_PGM, 3, f"={last_cl}{R_UFCF}*(1+C13)/(C26-C13)", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_PV_TV_PGM, 2, "PV of Terminal Value", font=BOLD_FONT)
    set_cell(ws3, R_PV_TV_PGM, 3, f"=C{R_TV_PGM}*{last_cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_EV_PGM, 2, "Enterprise Value", font=BOLD_FONT)
    set_cell(ws3, R_EV_PGM, 3, f"=C{R_SUM_PV}+C{R_PV_TV_PGM}", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_EQ_PGM, 2, "Equity Value", font=BOLD_FONT)
    set_cell(ws3, R_EQ_PGM, 3, f"=C{R_EV_PGM}-C16", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_PRICE_PGM, 2, "Implied Share Price (PGM)", font=BOLD_FONT)
    set_cell(ws3, R_PRICE_PGM, 3, f"=ROUND(C{R_EQ_PGM}*1000000/C15,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)

    # ── Valuation - Exit Multiple Method ──
    c = section_title(ws3, R_EXIT_SEC, 2, "Valuation - Exit Multiple Method")
    c.fill = LIGHT_GREEN
    for col_idx in range(3, 8):
        ws3.cell(row=R_EXIT_SEC, column=col_idx).fill = LIGHT_GREEN

    set_cell(ws3, R_SUM_PV_EX, 2, "Sum of PV of FCFs", font=BOLD_FONT)
    set_cell(ws3, R_SUM_PV_EX, 3, f"=C{R_SUM_PV}", font=BLACK_FONT, fmt=FMT_YEN)

    # Year 5 EBITDA = EBIT + D&A
    set_cell(ws3, R_YR5_EBITDA, 2, "Year 5 EBITDA", font=BOLD_FONT)
    set_cell(ws3, R_YR5_EBITDA, 3, f"={last_cl}{R_EBIT}+{last_cl}{R_DA}", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_TV_EXIT, 2, "Terminal Value (Exit Multiple)", font=BOLD_FONT)
    set_cell(ws3, R_TV_EXIT, 3, f"=C{R_YR5_EBITDA}*C14", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_PV_TV_EXIT, 2, "PV of Terminal Value", font=BOLD_FONT)
    set_cell(ws3, R_PV_TV_EXIT, 3, f"=C{R_TV_EXIT}*{last_cl}{R_DISC}", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_EV_EXIT, 2, "Enterprise Value", font=BOLD_FONT)
    set_cell(ws3, R_EV_EXIT, 3, f"=C{R_SUM_PV_EX}+C{R_PV_TV_EXIT}", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_EQ_EXIT, 2, "Equity Value", font=BOLD_FONT)
    set_cell(ws3, R_EQ_EXIT, 3, f"=C{R_EV_EXIT}-C16", font=BLACK_FONT, fmt=FMT_YEN)

    set_cell(ws3, R_PRICE_EXIT, 2, "Implied Share Price (Exit Multiple)", font=BOLD_FONT)
    set_cell(ws3, R_PRICE_EXIT, 3, f"=ROUND(C{R_EQ_EXIT}*1000000/C15,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)

    # ── Scenario Input Matrix ──
    if has_segments:
        # All scenario inputs are in Segment Analysis sheet — show note only
        c = section_title(ws3, R_SCEN_SEC, 2,
                          "All scenario inputs are in Segment Analysis sheet.")
        c.font = GREY_FONT
    else:
        c = section_title(ws3, R_SCEN_SEC, 2, "Scenario Input Matrix")
        c.fill = LIGHT_GREEN
        for col_idx in range(3, 8):
            ws3.cell(row=R_SCEN_SEC, column=col_idx).fill = LIGHT_GREEN

        # Year headers for scenario matrix
        for yr in range(proj_years):
            set_cell(ws3, R_SCEN_YEARS, 3 + yr, f"Year {yr + 1}",
                     font=HEADER_FONT, fill=HEADER_FILL,
                     alignment=Alignment(horizontal="center"))

        driver_blocks = [
            ("Revenue Growth (YoY)", "revenue_growth",    FMT_PCT, R_SCEN_BLK_GROWTH),
            ("COGS % of Revenue",    "cogs_pct",          FMT_PCT, R_SCEN_BLK_COGS),
            ("SGA % of Revenue",     "sga_pct",           FMT_PCT, R_SCEN_BLK_SGA),
        ]

        for drv_label, drv_key, drv_fmt, blk_start in driver_blocks:
            # Sub-header row
            section_title(ws3, blk_start, 2, drv_label)
            # 5 scenario rows
            for s, scen_name in enumerate(SCENARIO_NAMES):
                r = blk_start + 1 + s
                set_cell(ws3, r, 2, scen_name, font=BOLD_FONT)
                scen_data = config["scenarios"][scen_name][drv_key]
                for yr in range(proj_years):
                    set_cell(ws3, r, 3 + yr, scen_data[yr],
                             font=BLUE_FONT, fmt=drv_fmt, border=INPUT_BORDER)

    # =====================================================================
    # SHEET 4: NWC Schedule (DSO/DIH/DPO)
    # =====================================================================
    ws_nwc = wb.create_sheet("NWC Schedule")
    ws_nwc.sheet_properties.tabColor = "CC6600"

    ws_nwc.column_dimensions["A"].width = 3
    ws_nwc.column_dimensions["B"].width = 28
    ws_nwc.column_dimensions["C"].width = 16
    for letter in ["D", "E", "F", "G", "H"]:
        ws_nwc.column_dimensions[letter].width = 16

    set_cell(ws_nwc, 2, 2, f'NWC Schedule - {C["company_name"]}', font=TITLE_FONT)

    # ── Headers: Base Year + FY labels (matching DCF Model sheet) ──
    nwc_proj_labels = [f"Year {y}" for y in range(1, proj_years + 1)]
    if C.get("projection_start_fy"):
        import re as _re
        _m = _re.search(r"FY(\d+)", C["projection_start_fy"])
        if _m:
            _base_fy = int(_m.group(1))
            nwc_proj_labels = [f"FY{_base_fy + y}(E)" for y in range(proj_years)]
    nwc_headers = ["Base Year"] + nwc_proj_labels
    header_row(ws_nwc, 4, 3, 3 + proj_years, nwc_headers)

    # ── Determine NWC method ──
    nwc_method = C.get("nwc_method", "days")

    if nwc_method == "revenue_pct":
        # ================================================================
        # REVENUE PERCENTAGE METHOD
        # Row layout is identical to days method; only cell content differs.
        # ================================================================

        # ── Working Capital Drivers (% of Revenue) ──
        c = section_title(ws_nwc, NWC_R_DSO - 1, 2, "Working Capital Drivers (% of Revenue)")
        c.fill = LIGHT_FILL
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_DSO - 1, column=col_idx).fill = LIGHT_FILL

        set_cell(ws_nwc, NWC_R_DSO, 2, "NWC % of Revenue", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_DIH, 2, "\u2014", font=GREY_FONT)
        set_cell(ws_nwc, NWC_R_DPO, 2, "\u2014", font=GREY_FONT)

        # Base Year NWC%: =C18/C9
        set_cell(ws_nwc, NWC_R_DSO, 3, f"=C{NWC_R_NWC}/C{NWC_R_REV}",
                 font=BLACK_FONT, fmt=FMT_PCT)
        set_cell(ws_nwc, NWC_R_DIH, 3, "\u2014", font=GREY_FONT)
        set_cell(ws_nwc, NWC_R_DPO, 3, "\u2014", font=GREY_FONT)

        # Projected NWC%
        for yr in range(proj_years):
            nwc_col = 4 + yr
            cl = col_letter(nwc_col)
            if has_segments and seg_info and seg_info.get("nwc_scenario_rows"):
                # Reference Segment Analysis Consolidated Inputs NWC% rows
                nwc_seg_cl = col_letter(3 + seg_info["n_hist"] + yr)
                nwc_refs = [f"'Segment Analysis'!{nwc_seg_cl}{r}"
                            for r in seg_info["nwc_scenario_rows"]]
                set_cell(ws_nwc, NWC_R_DSO, nwc_col,
                         f"=CHOOSE('DCF Model'!$D$27,{','.join(nwc_refs)})",
                         font=BLACK_FONT, fmt=FMT_PCT)
            else:
                set_cell(ws_nwc, NWC_R_DSO, nwc_col,
                         nwc_choose_formula(NWC_R_SCEN_BLK_DSO, cl),
                         font=BLACK_FONT, fmt=FMT_PCT)
            set_cell(ws_nwc, NWC_R_DIH, nwc_col, "\u2014", font=GREY_FONT)
            set_cell(ws_nwc, NWC_R_DPO, nwc_col, "\u2014", font=GREY_FONT)

        # ── Revenue & COGS ──
        c = section_title(ws_nwc, NWC_R_REV - 1, 2, "P&L Reference (JPY mn)")
        c.fill = LIGHT_FILL
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_REV - 1, column=col_idx).fill = LIGHT_FILL

        set_cell(ws_nwc, NWC_R_REV, 2, "Revenue", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_COGS, 2, "\u2014", font=GREY_FONT)

        # Base Year Revenue
        set_cell(ws_nwc, NWC_R_REV, 3, C["base_year_revenue"], font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws_nwc, NWC_R_COGS, 3, "\u2014", font=GREY_FONT)

        # Projected Revenue (linked to DCF Model)
        for yr in range(proj_years):
            nwc_col = 4 + yr
            dcf_col_letter = col_letter(3 + yr)
            set_cell(ws_nwc, NWC_R_REV, nwc_col,
                     f"='DCF Model'!{dcf_col_letter}{R_REVENUE}",
                     font=BLACK_FONT, fmt=FMT_YEN)
            set_cell(ws_nwc, NWC_R_COGS, nwc_col, "\u2014", font=GREY_FONT)

        # ── Working Capital Items: all show "—" ──
        c = section_title(ws_nwc, NWC_R_AR - 1, 2, "Working Capital Items (JPY mn)")
        c.fill = LIGHT_FILL
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_AR - 1, column=col_idx).fill = LIGHT_FILL

        for _row in [NWC_R_AR, NWC_R_INV, NWC_R_CA, NWC_R_AP, NWC_R_CL]:
            set_cell(ws_nwc, _row, 2, "\u2014", font=GREY_FONT)
            for _ci in range(3, 3 + proj_years + 1):
                set_cell(ws_nwc, _row, _ci, "\u2014", font=GREY_FONT)

        # ── NWC Summary ──
        c = section_title(ws_nwc, NWC_R_NWC - 1, 2, "Net Working Capital (JPY mn)")
        c.fill = LIGHT_GREEN
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_NWC - 1, column=col_idx).fill = LIGHT_GREEN

        set_cell(ws_nwc, NWC_R_NWC, 2, "Net Working Capital", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_CHG_NWC, 2, "Change in NWC", font=BOLD_FONT)

        # Base Year NWC: hardcoded from computed value
        set_cell(ws_nwc, NWC_R_NWC, 3, C.get("base_year_nwc", 0),
                 font=BLUE_FONT, fmt=FMT_YEN, border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        set_cell(ws_nwc, NWC_R_CHG_NWC, 3, "n/a", font=BLACK_FONT)

        # Projected NWC = Revenue × NWC%, Change = ΔNWC
        for yr in range(proj_years):
            nwc_col = 4 + yr
            cl = col_letter(nwc_col)
            prev_cl = col_letter(nwc_col - 1)
            set_cell(ws_nwc, NWC_R_NWC, nwc_col,
                     f"={cl}{NWC_R_REV}*{cl}{NWC_R_DSO}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
            set_cell(ws_nwc, NWC_R_CHG_NWC, nwc_col,
                     f"={cl}{NWC_R_NWC}-{prev_cl}{NWC_R_NWC}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

        # ── Apply consistent borders ──
        _nwc_data_rows = (
            list(range(NWC_R_DSO, NWC_R_DPO + 1))
            + list(range(NWC_R_REV, NWC_R_COGS + 1))
            + list(range(NWC_R_AR, NWC_R_CL + 1))
            + [NWC_R_NWC, NWC_R_CHG_NWC]
        )
        _nwc_col_end = 3 + proj_years
        for _r in _nwc_data_rows:
            for _ci in range(2, _nwc_col_end + 1):
                _cell = ws_nwc.cell(row=_r, column=_ci)
                _has_border = (_cell.border and any([
                    getattr(_cell.border.top, 'style', None),
                    getattr(_cell.border.bottom, 'style', None),
                    getattr(_cell.border.left, 'style', None),
                    getattr(_cell.border.right, 'style', None),
                ]))
                if not _has_border:
                    _cell.border = NWC_DATA_BORDER

        # ── Scenario Input Matrix: NWC % of Revenue ──
        if has_segments and seg_info and seg_info.get("nwc_scenario_rows"):
            # NWC inputs are in Segment Analysis Consolidated Inputs
            c = section_title(ws_nwc, NWC_R_SCEN_SEC, 2,
                              "NWC inputs are in Segment Analysis sheet.")
            c.font = GREY_FONT
        else:
            c = section_title(ws_nwc, NWC_R_SCEN_SEC, 2,
                              "Scenario Input Matrix (NWC % of Revenue)")
            c.fill = LIGHT_GREEN
            for col_idx in range(3, 3 + proj_years + 1):
                ws_nwc.cell(row=NWC_R_SCEN_SEC, column=col_idx).fill = LIGHT_GREEN

            for yr in range(proj_years):
                _scen_yr_label = nwc_proj_labels[yr] if yr < len(nwc_proj_labels) else f"Year {yr + 1}"
                set_cell(ws_nwc, NWC_R_SCEN_YEARS, 4 + yr, _scen_yr_label,
                         font=HEADER_FONT, fill=HEADER_FILL,
                         alignment=Alignment(horizontal="center"))

            # Single block: NWC % of Revenue at NWC_R_SCEN_BLK_DSO (row 24)
            section_title(ws_nwc, NWC_R_SCEN_BLK_DSO, 2, "NWC % of Revenue")
            for s, scen_name in enumerate(SCENARIO_NAMES):
                r = NWC_R_SCEN_BLK_DSO + 1 + s
                set_cell(ws_nwc, r, 2, scen_name, font=BOLD_FONT)
                scen_data = config["scenarios"][scen_name]["nwc_pct"]
                for yr in range(proj_years):
                    set_cell(ws_nwc, r, 4 + yr, scen_data[yr],
                             font=BLUE_FONT, fmt=FMT_PCT, border=INPUT_BORDER)

    else:
        # ================================================================
        # DAYS METHOD (default) — existing code, unchanged
        # ================================================================

        # ── Working Capital Drivers ──
        c = section_title(ws_nwc, NWC_R_DSO - 1, 2, "Working Capital Drivers (Days)")
        c.fill = LIGHT_FILL
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_DSO - 1, column=col_idx).fill = LIGHT_FILL

        set_cell(ws_nwc, NWC_R_DSO, 2, "DSO (Days Sales Outstanding)", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_DIH, 2, "DIH (Days Inventory Held)", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_DPO, 2, "DPO (Days Payable Outstanding)", font=BOLD_FONT)

        # Base Year DSO/DIH/DPO (computed from actuals)
        set_cell(ws_nwc, NWC_R_DSO, 3, f"=C{NWC_R_AR}/C{NWC_R_REV}*365",
                 font=BLACK_FONT, fmt=FMT_DAYS)
        set_cell(ws_nwc, NWC_R_DIH, 3, f"=C{NWC_R_INV}/C{NWC_R_COGS}*365",
                 font=BLACK_FONT, fmt=FMT_DAYS)
        set_cell(ws_nwc, NWC_R_DPO, 3, f"=C{NWC_R_AP}/C{NWC_R_COGS}*365",
                 font=BLACK_FONT, fmt=FMT_DAYS)

        # Projected DSO/DIH/DPO (CHOOSE from scenario matrix)
        for yr in range(proj_years):
            nwc_col = 4 + yr
            cl = col_letter(nwc_col)
            set_cell(ws_nwc, NWC_R_DSO, nwc_col,
                     nwc_choose_formula(NWC_R_SCEN_BLK_DSO, cl),
                     font=BLACK_FONT, fmt=FMT_DAYS)
            set_cell(ws_nwc, NWC_R_DIH, nwc_col,
                     nwc_choose_formula(NWC_R_SCEN_BLK_DIH, cl),
                     font=BLACK_FONT, fmt=FMT_DAYS)
            set_cell(ws_nwc, NWC_R_DPO, nwc_col,
                     nwc_choose_formula(NWC_R_SCEN_BLK_DPO, cl),
                     font=BLACK_FONT, fmt=FMT_DAYS)

        # ── Revenue & COGS (linked from DCF Model) ──
        c = section_title(ws_nwc, NWC_R_REV - 1, 2, "P&L Reference (JPY mn)")
        c.fill = LIGHT_FILL
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_REV - 1, column=col_idx).fill = LIGHT_FILL

        set_cell(ws_nwc, NWC_R_REV, 2, "Revenue", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_COGS, 2, "COGS", font=BOLD_FONT)

        # Base Year
        set_cell(ws_nwc, NWC_R_REV, 3, C["base_year_revenue"], font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws_nwc, NWC_R_COGS, 3, C["base_year_cogs"], font=BLUE_FONT, fmt=FMT_YEN)

        # Projected (linked to DCF Model; NWC col D = DCF col C, offset +1)
        for yr in range(proj_years):
            nwc_col = 4 + yr
            dcf_col_letter = col_letter(3 + yr)
            set_cell(ws_nwc, NWC_R_REV, nwc_col,
                     f"='DCF Model'!{dcf_col_letter}{R_REVENUE}",
                     font=BLACK_FONT, fmt=FMT_YEN)
            set_cell(ws_nwc, NWC_R_COGS, nwc_col,
                     f"='DCF Model'!{dcf_col_letter}{R_COGS}",
                     font=BLACK_FONT, fmt=FMT_YEN)

        # ── Working Capital Items ──
        c = section_title(ws_nwc, NWC_R_AR - 1, 2, "Working Capital Items (JPY mn)")
        c.fill = LIGHT_FILL
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_AR - 1, column=col_idx).fill = LIGHT_FILL

        set_cell(ws_nwc, NWC_R_AR, 2, "Accounts Receivable", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_INV, 2, "Inventory", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_CA, 2, "Current Assets (AR + Inv)", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_AP, 2, "Accounts Payable", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_CL, 2, "Current Liabilities (AP)", font=BOLD_FONT)

        # Base Year actuals
        set_cell(ws_nwc, NWC_R_AR, 3, C["base_year_ar"], font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws_nwc, NWC_R_INV, 3, C["base_year_inv"], font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws_nwc, NWC_R_CA, 3, f"=C{NWC_R_AR}+C{NWC_R_INV}", font=BLACK_FONT, fmt=FMT_YEN)
        set_cell(ws_nwc, NWC_R_AP, 3, C["base_year_ap"], font=BLUE_FONT, fmt=FMT_YEN)
        set_cell(ws_nwc, NWC_R_CL, 3, f"=C{NWC_R_AP}", font=BLACK_FONT, fmt=FMT_YEN)

        # Projected WC items
        for yr in range(proj_years):
            nwc_col = 4 + yr
            cl = col_letter(nwc_col)
            set_cell(ws_nwc, NWC_R_AR, nwc_col,
                     f"={cl}{NWC_R_REV}*{cl}{NWC_R_DSO}/365", font=BLACK_FONT, fmt=FMT_YEN)
            set_cell(ws_nwc, NWC_R_INV, nwc_col,
                     f"={cl}{NWC_R_COGS}*{cl}{NWC_R_DIH}/365", font=BLACK_FONT, fmt=FMT_YEN)
            set_cell(ws_nwc, NWC_R_CA, nwc_col,
                     f"={cl}{NWC_R_AR}+{cl}{NWC_R_INV}", font=BLACK_FONT, fmt=FMT_YEN)
            set_cell(ws_nwc, NWC_R_AP, nwc_col,
                     f"={cl}{NWC_R_COGS}*{cl}{NWC_R_DPO}/365", font=BLACK_FONT, fmt=FMT_YEN)
            set_cell(ws_nwc, NWC_R_CL, nwc_col,
                     f"={cl}{NWC_R_AP}", font=BLACK_FONT, fmt=FMT_YEN)

        # ── NWC Summary ──
        c = section_title(ws_nwc, NWC_R_NWC - 1, 2, "Net Working Capital (JPY mn)")
        c.fill = LIGHT_GREEN
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_NWC - 1, column=col_idx).fill = LIGHT_GREEN

        set_cell(ws_nwc, NWC_R_NWC, 2, "Net Working Capital", font=BOLD_FONT)
        set_cell(ws_nwc, NWC_R_CHG_NWC, 2, "Change in NWC", font=BOLD_FONT)

        # Base Year NWC
        set_cell(ws_nwc, NWC_R_NWC, 3, f"=C{NWC_R_CA}-C{NWC_R_CL}",
                 font=BLACK_FONT, fmt=FMT_YEN, border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
        set_cell(ws_nwc, NWC_R_CHG_NWC, 3, "n/a", font=BLACK_FONT)

        # Projected NWC & Change
        for yr in range(proj_years):
            nwc_col = 4 + yr
            cl = col_letter(nwc_col)
            prev_cl = col_letter(nwc_col - 1)
            set_cell(ws_nwc, NWC_R_NWC, nwc_col,
                     f"={cl}{NWC_R_CA}-{cl}{NWC_R_CL}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=SUBTOTAL_BORDER, fill=SUBTOTAL_FILL)
            set_cell(ws_nwc, NWC_R_CHG_NWC, nwc_col,
                     f"={cl}{NWC_R_NWC}-{prev_cl}{NWC_R_NWC}",
                     font=BLACK_FONT, fmt=FMT_YEN, border=TOP_BOTTOM)

        # ── Apply consistent borders to all NWC data rows ──
        _nwc_data_rows = (
            list(range(NWC_R_DSO, NWC_R_DPO + 1))       # Drivers: DSO, DIH, DPO
            + list(range(NWC_R_REV, NWC_R_COGS + 1))     # P&L Reference: Revenue, COGS
            + list(range(NWC_R_AR, NWC_R_CL + 1))         # WC Items: AR, Inv, CA, AP, CL
            + [NWC_R_NWC, NWC_R_CHG_NWC]                  # NWC Summary
        )
        _nwc_col_end = 3 + proj_years  # last data column
        for _r in _nwc_data_rows:
            for _ci in range(2, _nwc_col_end + 1):
                _cell = ws_nwc.cell(row=_r, column=_ci)
                # Preserve existing meaningful borders (SUBTOTAL_BORDER, TOP_BOTTOM, etc.)
                _has_border = (_cell.border and any([
                    getattr(_cell.border.top, 'style', None),
                    getattr(_cell.border.bottom, 'style', None),
                    getattr(_cell.border.left, 'style', None),
                    getattr(_cell.border.right, 'style', None),
                ]))
                if not _has_border:
                    _cell.border = NWC_DATA_BORDER

        # ── Scenario Input Matrix (DSO, DIH, DPO) ──
        c = section_title(ws_nwc, NWC_R_SCEN_SEC, 2, "Scenario Input Matrix (Working Capital Days)")
        c.fill = LIGHT_GREEN
        for col_idx in range(3, 3 + proj_years + 1):
            ws_nwc.cell(row=NWC_R_SCEN_SEC, column=col_idx).fill = LIGHT_GREEN

        for yr in range(proj_years):
            _scen_yr_label = nwc_proj_labels[yr] if yr < len(nwc_proj_labels) else f"Year {yr + 1}"
            set_cell(ws_nwc, NWC_R_SCEN_YEARS, 4 + yr, _scen_yr_label,
                     font=HEADER_FONT, fill=HEADER_FILL,
                     alignment=Alignment(horizontal="center"))

        nwc_driver_blocks = [
            ("DSO (Days)", "dso_days", FMT_DAYS, NWC_R_SCEN_BLK_DSO),
            ("DIH (Days)", "dih_days", FMT_DAYS, NWC_R_SCEN_BLK_DIH),
            ("DPO (Days)", "dpo_days", FMT_DAYS, NWC_R_SCEN_BLK_DPO),
        ]

        for drv_label, drv_key, drv_fmt, blk_start in nwc_driver_blocks:
            section_title(ws_nwc, blk_start, 2, drv_label)
            for s, scen_name in enumerate(SCENARIO_NAMES):
                r = blk_start + 1 + s
                set_cell(ws_nwc, r, 2, scen_name, font=BOLD_FONT)
                scen_data = config["scenarios"][scen_name][drv_key]
                for yr in range(proj_years):
                    set_cell(ws_nwc, r, 4 + yr, scen_data[yr],
                             font=BLUE_FONT, fmt=drv_fmt, border=INPUT_BORDER)

    # =====================================================================
    # SHEET 5: Comps Analysis
    # =====================================================================
    ws4 = wb.create_sheet("Comps Analysis")
    ws4.sheet_properties.tabColor = "006600"

    ws4.column_dimensions["A"].width = 3
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 10
    for letter in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        ws4.column_dimensions[letter].width = 12

    set_cell(ws4, 2, 2, "Comparable Company Analysis", font=TITLE_FONT)

    # Header row
    comp_headers = [
        "Company", "Ticker", "Mkt Cap\n(JPY mn)", "EV\n(JPY mn)",
        "Revenue\n(JPY mn)", "EBITDA\n(JPY mn)", "Op Income\n(JPY mn)",
        "Net Income\n(JPY mn)", "EV/EBITDA", "EV/Revenue", "PER",
        "PBR", "Op Margin", "ROE"
    ]
    header_row(ws4, 4, 2, 15, comp_headers)

    # Company data rows (rows 5-10 for 6 comps)
    comps = C["comps"]
    for i, comp in enumerate(comps):
        r = 5 + i

        _na = lambda ws, r, c: set_cell(ws, r, c, "N/A", font=BLACK_FONT, border=THIN_BORDER,
                                         alignment=Alignment(horizontal="right"))

        set_cell(ws4, r, 2, comp["name"], font=BOLD_FONT)
        set_cell(ws4, r, 3, comp["ticker"])

        # Mkt Cap & EV: may be None if yfinance fetch failed
        if comp["mkt_cap"] is None:
            _na(ws4, r, 4)
        else:
            set_cell(ws4, r, 4, comp["mkt_cap"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
        if comp["ev"] is None:
            _na(ws4, r, 5)
        else:
            set_cell(ws4, r, 5, comp["ev"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

        set_cell(ws4, r, 6, comp["revenue"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
        set_cell(ws4, r, 7, comp["ebitda"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
        set_cell(ws4, r, 8, comp["op_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)
        set_cell(ws4, r, 9, comp["net_income"], font=BLUE_FONT, fmt=FMT_YEN, border=THIN_BORDER)

        # EV/EBITDA
        if comp["ev"] is None or comp["ebitda"] <= 0:
            _na(ws4, r, 10)
        else:
            set_cell(ws4, r, 10, f"=E{r}/G{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

        # EV/Revenue
        if comp["ev"] is None or comp["revenue"] <= 0:
            _na(ws4, r, 11)
        else:
            set_cell(ws4, r, 11, f"=E{r}/F{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

        # PER
        if comp["mkt_cap"] is None or comp["net_income"] <= 0:
            _na(ws4, r, 12)
        else:
            set_cell(ws4, r, 12, f"=D{r}/I{r}", font=BLACK_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

        # PBR
        if comp["pbr"] is None:
            _na(ws4, r, 13)
        else:
            set_cell(ws4, r, 13, comp["pbr"], font=BLUE_FONT, fmt=FMT_RATIO, border=THIN_BORDER)

        set_cell(ws4, r, 14, f"=H{r}/F{r}", font=BLACK_FONT, fmt=FMT_PCT, border=THIN_BORDER)

        # ROE
        if comp["roe"] is None:
            _na(ws4, r, 15)
        else:
            set_cell(ws4, r, 15, comp["roe"], font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)

    last_comp_row = 5 + len(comps) - 1

    # ── Statistics ──
    section_title(ws4, 14, 2, "Statistics")

    stat_labels = ["25th Percentile", "Median (50th)", "75th Percentile"]
    stat_rows = [15, 16, 17]

    stat_col_map = [
        (4, 10),  # EV/EBITDA
        (5, 11),  # EV/Revenue
        (6, 12),  # PER
        (7, 13),  # PBR
        (8, 14),  # Op Margin
        (9, 15),  # ROE
    ]

    for stat_idx, (label, r) in enumerate(zip(stat_labels, stat_rows)):
        set_cell(ws4, r, 2, label, font=BOLD_FONT)

        for dst_col, src_col in stat_col_map:
            src_letter = col_letter(src_col)
            rng = f"{src_letter}5:{src_letter}{last_comp_row}"

            if src_col in (14, 15):
                fmt = FMT_PCT
            else:
                fmt = FMT_RATIO

            if stat_idx == 0:
                formula = f"=PERCENTILE({rng},0.25)"
            elif stat_idx == 1:
                formula = f"=MEDIAN({rng})"
            else:
                formula = f"=PERCENTILE({rng},0.75)"

            set_cell(ws4, r, dst_col, formula, font=BLACK_FONT, fmt=fmt, border=THIN_BORDER)

    # ── Implied Valuation ──
    c = section_title(ws4, 19, 2, f'Implied Valuation for {C["company_name"]}')
    c.fill = LIGHT_GREEN
    for col_idx in range(3, 10):
        ws4.cell(row=19, column=col_idx).fill = LIGHT_GREEN

    section_title(ws4, 20, 2, f'{C["company_name"]} Financials')

    if USE_EV_SALES:
        set_cell(ws4, 21, 2, "Revenue (JPY mn)", font=BOLD_FONT)
        set_cell(ws4, 21, 3, C["base_year_revenue"], font=BLUE_FONT, fmt=FMT_YEN)
    else:
        set_cell(ws4, 21, 2, "EBITDA (JPY mn)", font=BOLD_FONT)
        set_cell(ws4, 21, 3, C["core_ebitda"], font=BLUE_FONT, fmt=FMT_YEN)

    set_cell(ws4, 22, 2, "Net Income (JPY mn)", font=BOLD_FONT)
    set_cell(ws4, 22, 3, C["core_net_income"], font=BLUE_FONT, fmt=FMT_YEN)
    set_cell(ws4, 23, 2, "Shares Outstanding", font=BOLD_FONT)
    set_cell(ws4, 23, 3, C["shares_outstanding"], font=BLUE_FONT, fmt=FMT_INT)
    set_cell(ws4, 24, 2, "Net Debt (JPY mn)", font=BOLD_FONT)
    set_cell(ws4, 24, 3, C["net_debt"], font=BLUE_FONT, fmt=FMT_YEN)

    section_title(ws4, 26, 2, "Implied Share Price (Median Multiples)")

    if USE_EV_SALES:
        set_cell(ws4, 27, 2, "Via EV/Sales (Median)", font=BOLD_FONT)
        set_cell(ws4, 27, 3, "=ROUND((C21*E16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
                 border=TOP_BOTTOM)
    else:
        set_cell(ws4, 27, 2, "Via EV/EBITDA (Median)", font=BOLD_FONT)
        set_cell(ws4, 27, 3, "=ROUND((C21*D16-C24)*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
                 border=TOP_BOTTOM)

    set_cell(ws4, 28, 2, "Via PER (Median)", font=BOLD_FONT)
    set_cell(ws4, 28, 3, "=ROUND(C22*F16*1000000/C23,0)", font=BLACK_FONT, fmt=FMT_YEN,
             border=TOP_BOTTOM)

    # =====================================================================
    # SHEET 6: Sensitivity Analysis (Dynamic Excel formulas)
    # =====================================================================
    ws5 = wb.create_sheet("Sensitivity Analysis")
    ws5.sheet_properties.tabColor = "996600"

    ws5.column_dimensions["A"].width = 3
    ws5.column_dimensions["B"].width = 24
    for letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws5.column_dimensions[letter].width = 14

    set_cell(ws5, 2, 2, "Sensitivity Analysis", font=TITLE_FONT)

    # ── Current values reference ──
    set_cell(ws5, 3, 2, "Current WACC:", font=BOLD_FONT)
    set_cell(ws5, 3, 3, "='DCF Model'!C26", font=BLACK_FONT, fmt=FMT_PCT2)
    set_cell(ws5, 3, 5, "Terminal g:", font=BOLD_FONT)
    set_cell(ws5, 3, 6, "='DCF Model'!C13", font=BLACK_FONT, fmt=FMT_PCT2)
    set_cell(ws5, 3, 8, "Exit Multiple:", font=BOLD_FONT)
    set_cell(ws5, 3, 9, "='DCF Model'!C14", font=BLACK_FONT, fmt=FMT_RATIO)

    # ── Dynamic formula builders ──
    _DCF = "'DCF Model'"
    _SHARES = f"{_DCF}!C15"
    _NET_DEBT = f"{_DCF}!C16"
    _last_cl = col_letter(3 + proj_years - 1)
    _ufcf_cells = [f"{_DCF}!{col_letter(3 + yr)}{R_UFCF}" for yr in range(proj_years)]
    _stub_ref = f"{_DCF}!C{R_STUB_FRACTION}"

    def _build_pgm_formula(wacc_ref, tg_ref):
        pv_parts = [f"{_ufcf_cells[yr]}/(1+{wacc_ref})^({_stub_ref}+{yr})" for yr in range(proj_years)]
        last_ufcf = _ufcf_cells[proj_years - 1]
        pv_tv = f"{last_ufcf}*(1+{tg_ref})/({wacc_ref}-{tg_ref})/(1+{wacc_ref})^({_stub_ref}+{proj_years-1})"
        return f'=IFERROR(ROUND(({"+".join(pv_parts)}+{pv_tv}-{_NET_DEBT})*1000000/{_SHARES},0),"")'

    def _build_exit_formula(wacc_ref, mult_ref):
        pv_parts = [f"{_ufcf_cells[yr]}/(1+{wacc_ref})^({_stub_ref}+{yr})" for yr in range(proj_years)]
        yr5_ebitda = f"({_DCF}!{_last_cl}{R_EBIT}+{_DCF}!{_last_cl}{R_DA})"
        pv_tv = f"{yr5_ebitda}*{mult_ref}/(1+{wacc_ref})^({_stub_ref}+{proj_years-1})"
        return f'=IFERROR(ROUND(({"+".join(pv_parts)}+{pv_tv}-{_NET_DEBT})*1000000/{_SHARES},0),"")'

    # ── Dynamic header helpers ──
    _N_GRID = 7
    _CENTER_IDX = 3  # 4th position (0-based)
    _WACC_STEP = 0.005
    _TG_STEP   = 0.0025
    _EXIT_STEP = 1.0
    _ANCHOR_WACC = "$C$3"
    _ANCHOR_TG   = "$F$3"
    _ANCHOR_EXIT = "$I$3"

    def _offset_formula(anchor, offset_val):
        if offset_val == 0:
            return f"={anchor}"
        elif offset_val > 0:
            return f"={anchor}+{offset_val}"
        else:
            return f"={anchor}-{abs(offset_val)}"

    # ── Table 1: WACC vs Terminal Growth Rate (PGM) ──
    T1_TITLE = 5
    T1_HDR = 6
    T1_DATA = 7

    section_title(ws5, T1_TITLE, 2,
                  "Table 1: WACC vs Terminal Growth Rate (PGM - Implied Share Price, JPY)")

    # Column headers: TG (dynamic, centered on F3)
    set_cell(ws5, T1_HDR, 2, "WACC \\ Terminal g", font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
    for j in range(_N_GRID):
        offset = round((j - _CENTER_IDX) * _TG_STEP, 6)
        set_cell(ws5, T1_HDR, 3 + j, _offset_formula(_ANCHOR_TG, offset),
                 font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_PCT,
                 alignment=Alignment(horizontal="center"), border=THIN_BORDER)

    # Row headers: WACC (dynamic, centered on C3) + data formulas
    for i in range(_N_GRID):
        r = T1_DATA + i
        offset = round((i - _CENTER_IDX) * _WACC_STEP, 6)
        set_cell(ws5, r, 2, _offset_formula(_ANCHOR_WACC, offset),
                 font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
        for j in range(_N_GRID):
            col = 3 + j
            cl = col_letter(col)
            formula = _build_pgm_formula(f"$B{r}", f"{cl}${T1_HDR}")
            set_cell(ws5, r, col, formula, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # ── Table 2: WACC vs Exit Multiple ──
    T2_TITLE = T1_DATA + _N_GRID + 2
    T2_HDR = T2_TITLE + 1
    T2_DATA = T2_HDR + 1

    section_title(ws5, T2_TITLE, 2,
                  "Table 2: WACC vs Exit Multiple (Exit Multiple - Implied Share Price, JPY)")

    # Column headers: Exit Multiple (dynamic, centered on I3)
    set_cell(ws5, T2_HDR, 2, "WACC \\ Exit Multiple", font=HEADER_FONT, fill=HEADER_FILL,
             alignment=Alignment(horizontal="center", wrap_text=True), border=THIN_BORDER)
    for j in range(_N_GRID):
        offset = round((j - _CENTER_IDX) * _EXIT_STEP, 6)
        set_cell(ws5, T2_HDR, 3 + j, _offset_formula(_ANCHOR_EXIT, offset),
                 font=HEADER_FONT, fill=HEADER_FILL, fmt=FMT_RATIO,
                 alignment=Alignment(horizontal="center"), border=THIN_BORDER)

    # Row headers: WACC (dynamic) + data formulas
    for i in range(_N_GRID):
        r = T2_DATA + i
        offset = round((i - _CENTER_IDX) * _WACC_STEP, 6)
        set_cell(ws5, r, 2, _offset_formula(_ANCHOR_WACC, offset),
                 font=BLUE_FONT, fmt=FMT_PCT, border=THIN_BORDER)
        for j in range(_N_GRID):
            col = 3 + j
            cl = col_letter(col)
            formula = _build_exit_formula(f"$B{r}", f"{cl}${T2_HDR}")
            set_cell(ws5, r, col, formula, font=BLACK_FONT, fmt=FMT_YEN, border=THIN_BORDER)

    # ── Note ──
    _note_row = T2_DATA + _N_GRID + 1
    set_cell(ws5, _note_row, 2,
             "All values dynamically linked to DCF Model. "
             "Headers auto-center on current WACC / Terminal g / Exit Multiple.",
             font=GREY_FONT)
    ws5.merge_cells(start_row=_note_row, start_column=2,
                    end_row=_note_row, end_column=9)

    # =====================================================================
    # SHEET 7 & 8: Segment / Driver Analysis (optional)
    # =====================================================================
    segments = C.get("segments")
    if segments:
        proj_years = C["projection_years"]
        _year_labels = [f"Year {y}" for y in range(1, proj_years + 1)]
        if C.get("projection_start_fy"):
            import re as _re
            _m = _re.search(r"FY(\d+)", C["projection_start_fy"])
            if _m:
                _base_fy = int(_m.group(1))
                _year_labels = [f"FY{_base_fy + y}(E)" for y in range(proj_years)]

        _create_segment_sheet(wb, C, segments, proj_years, _year_labels)
        _create_driver_sheet(wb, C, segments, proj_years, _year_labels)

    # =====================================================================
    # SAVE & VERIFY
    # =====================================================================

    if output_path is None:
        ticker_safe = C["ticker"].replace(".", "")
        output_path = f"{ticker_safe}_Equity_Research_V3.xlsx"

    wb.save(output_path)
    print(f"\nSaved: {output_path}")

    # Run recalc.py for verification
    recalc_script = os.path.join("scripts", "recalc.py")
    if os.path.exists(recalc_script):
        print(f"\nRunning verification: python {recalc_script} {output_path}")
        result = subprocess.run([sys.executable, recalc_script, output_path],
                                capture_output=True, text=True)
        print(result.stdout)
        if result.stderr:
            print("STDERR:", result.stderr)

    return output_path


# =====================================================================
# STANDALONE EXECUTION (backward compatibility)
# =====================================================================
if __name__ == "__main__":
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "scripts")))
    from pdf_parser import extract_all_financials
    from comps_fetcher import get_comps_data

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _pdf_result = extract_all_financials(_script_dir)

    _FALLBACK = {
    "hist_revenue": [271, 332, 490, 517],
    "hist_operating_income": [-433, -598, -527, -800],
    "hist_net_income": [-2237, -413, -69, -801],
    "hist_cogs": [147, 156, 52, 177],
    "hist_sga": [558, 775, 966, 1141],
    "hist_ocf": [-515, -619, -491, -815],
    "hist_capex": [137, 20, 433, 162],
    "hist_cash": [604, 852, 1720, 2594],
    "hist_debt": [None, 200, 200, 200],
    "latest_net_debt": -2394,
    }

    if _pdf_result is None:
        print("FATAL: PDF extraction failed. Using fallback values.")
        _pdf_result = _FALLBACK

    def _get(key):
        val = _pdf_result.get(key)
        if val is None:
            return _FALLBACK.get(key)
        return val

    config = {
    # ── Company Info ──
    "company_name": "TEMPLATE COMPANY",
    "ticker": "0000.T",
    "exchange": "TSE Growth",
    "sector": "Information & Communication",
    "current_price": 1000,
    "shares_outstanding": 10_000_000,
    "net_debt": _get("latest_net_debt") or -1000,  # JPY mn (negative = net cash), from BS
    
    # ── Historical Financials (JPY mn) — all from PDFs ──
    "hist_years": ["FY2022 (Mar-22)", "FY2023 (Mar-23)", "FY2024 (Mar-24)", "FY2025 (Mar-25)"],
    "hist_revenue":          _get("hist_revenue"),
    "hist_operating_income": _get("hist_operating_income"),
    "hist_net_income":       _get("hist_net_income"),
    "hist_cogs":             _get("hist_cogs"),
    "hist_sga":              _get("hist_sga"),
    "hist_ocf":              _get("hist_ocf"),
    "hist_capex":            _get("hist_capex"),
    "hist_cash":             _get("hist_cash"),
    "hist_debt":             _get("hist_debt"),
    
    # ── DCF Assumptions — Future Projections ──
    "scenarios": {
    "Base":       {"revenue_growth": [0.10,0.08,0.07,0.06,0.05], "cogs_pct": [0.70,0.70,0.70,0.70,0.70], "sga_pct": [0.13,0.13,0.13,0.13,0.13], "dso_days": [60,60,60,60,60], "dih_days": [30,30,30,30,30], "dpo_days": [45,45,45,45,45]},
    "Upside":     {"revenue_growth": [0.10,0.12,0.18,0.20,0.15], "cogs_pct": [0.71,0.705,0.70,0.70,0.68], "sga_pct": [0.12,0.12,0.12,0.12,0.12], "dso_days": [55,55,55,55,55], "dih_days": [28,28,28,28,28], "dpo_days": [48,48,48,48,48]},
    "Management": {"revenue_growth": [0.10,0.10,0.10,0.10,0.10], "cogs_pct": [0.70,0.70,0.70,0.70,0.70], "sga_pct": [0.13,0.13,0.13,0.13,0.13], "dso_days": [60,60,60,60,60], "dih_days": [30,30,30,30,30], "dpo_days": [45,45,45,45,45]},
    "Downside 1": {"revenue_growth": [0.02,0.02,0.02,0.02,0.02], "cogs_pct": [0.73,0.73,0.74,0.75,0.73], "sga_pct": [0.14,0.14,0.14,0.14,0.14], "dso_days": [65,65,65,65,65], "dih_days": [33,33,33,33,33], "dpo_days": [42,42,42,42,42]},
    "Downside 2": {"revenue_growth": [0.00,0.00,0.00,0.00,0.00], "cogs_pct": [0.76,0.76,0.76,0.76,0.76], "sga_pct": [0.15,0.15,0.15,0.15,0.15], "dso_days": [70,70,70,70,70], "dih_days": [35,35,35,35,35], "dpo_days": [40,40,40,40,40]},
    },
    "capex_pct": 0.03,
    "da_pct": 0.015,
    "tax_rate": 0.30,
    "risk_free": 0.022,
    "beta": 1.75,
    "erp": 0.060,
    "size_premium": 0.050,
    "cost_of_debt_at": 0.015,
    "de_ratio": 0.05,
    "terminal_growth": 0.020,
    "exit_multiple": 15.0,
    "projection_years": 5,
    "base_year_revenue": (_get("hist_revenue") or [517])[-1],
    "base_year_cogs": (_get("hist_cogs") or [177])[-1],
    
    # ── NWC Base Year Actuals (JPY mn) — edit for each company ──
    "base_year_ar":   85,    # Accounts Receivable
    "base_year_inv":  14,    # Inventory
    "base_year_ap":   22,    # Accounts Payable
    
    # ── Comparable Companies (loaded dynamically from CSV) ──
    "comps": get_comps_data(os.path.join(_script_dir, "comps_input.csv")),
    
    # ── Kudan Comps Data (for implied valuation) ──
    "core_ebitda": (_get("hist_operating_income") or [-800])[-1] + 8,
    "core_net_income": (_get("hist_net_income") or [-801])[-1],
    
    # ── Investment Thesis & Risks ──
    "investment_thesis": [
    "1. Global leader in Artificial Perception (SLAM) technology",
    "2. High leverage on revenue growth due to fixed-cost intensive IP licensing model",
    "3. Transitioning from R&D phase to commercial scaling phase",
    ],
    "key_risks": [
    "1. Prolonged losses and negative free cash flow burning cash runway",
    "2. Long sales cycles converting PoC (Proof of Concept) to commercial licenses",
    "3. High WACC (17.7%) depressing present value heavily",
    ],
    
    # ── V3 Settings ──
    "primary_multiple": "EV/Sales",  # "EV/EBITDA" or "EV/Sales"
    }
    
    # Restore flat arrays from Base scenario (backward compatibility for sensitivity analysis)
    _base = config["scenarios"]["Base"]
    config["revenue_growth"]    = _base["revenue_growth"]
    config["cogs_pct"]          = _base["cogs_pct"]
    config["sga_pct"]           = _base["sga_pct"]
    
    # =====================================================================

    config["current_price"], config["shares_outstanding"] = get_live_market_data(
        config.get("ticker", ""),
        config.get("current_price", 0),
        config.get("shares_outstanding", 0)
    )

    generate_dcf_workbook(config)
